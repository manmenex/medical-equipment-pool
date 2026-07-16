"""
Excel report renderer (FUNCTION 6).

Produces a workbook with:
    - "Summary" sheet: headline counts (Total FDA Recalls, Total ECRI
      Alerts, Hospital Matches, Possible Matches, No Match).
    - "FDA Recalls" / "ECRI Alerts" sheets: full detail rows with an
      Excel AutoFilter enabled on the header row so the hospital can
      filter by Manufacturer / Department / Recall Class / Date directly
      in Excel.
    - "Matches" sheet: every match result with its confidence score and
      reasoning.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.report_data import ReportData
from app.utils.exceptions import ReportGenerationError
from app.utils.logging_config import get_audit_logger, get_logger

logger = get_logger(__name__)
audit_export = get_audit_logger("export")

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws: Worksheet, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws: Worksheet, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _write_table(ws: Worksheet, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_header(ws)
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "A2"
    _autosize(ws)


def generate_excel_report(data: ReportData, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb = Workbook()

        summary_ws = wb.active
        summary_ws.title = "Summary"
        summary_ws.append(["Medical Device Recall Report"])
        summary_ws["A1"].font = Font(size=16, bold=True)
        summary_ws.append([f"Period: {data.period.label} ({data.period.start_date} to {data.period.end_date})"])
        summary_ws.append([])
        summary_ws.append(["Metric", "Count"])
        _style_header(summary_ws, row=4)
        summary_rows = [
            ("Total FDA Recalls", data.total_fda),
            ("Total ECRI Alerts", data.total_ecri),
            ("Hospital Matches", data.total_matched),
            ("Possible Matches", data.total_possible),
            ("No Match", data.total_not_matched),
        ]
        for label, count in summary_rows:
            summary_ws.append([label, count])
        _autosize(summary_ws)

        fda_ws = wb.create_sheet("FDA Recalls")
        _write_table(
            fda_ws,
            [
                "Recall Number", "Recall Class", "Product", "Manufacturer", "Reason",
                "Date Initiated", "Status", "Distribution", "Quantity", "Model Number",
                "Catalog Number", "UDI", "Recall URL",
            ],
            [
                [
                    r.recall_number, r.recall_class, r.product, r.manufacturer, r.reason,
                    r.date_initiated.isoformat() if r.date_initiated else "",
                    r.status, r.distribution, r.quantity, r.model_number, r.catalog_number,
                    r.udi, r.recall_url,
                ]
                for r in data.fda_recalls
            ],
        )

        ecri_ws = wb.create_sheet("ECRI Alerts")
        _write_table(
            ecri_ws,
            [
                "Alert Number", "Title", "Manufacturer", "Product", "Model Number",
                "Catalog Number", "Hazard Level", "Date Issued", "Status", "Alert URL",
            ],
            [
                [
                    a.alert_number, a.title, a.manufacturer, a.product, a.model_number,
                    a.catalog_number, a.hazard_level,
                    a.date_issued.isoformat() if a.date_issued else "",
                    a.status, a.alert_url,
                ]
                for a in data.ecri_alerts
            ],
        )

        matches_ws = wb.create_sheet("Matches")
        _write_table(
            matches_ws,
            [
                "Recall Source", "Recall Number", "Hospital Asset Number", "Manufacturer",
                "Model", "Confidence Score", "Status", "Method", "Reason",
            ],
            [
                [
                    m.recall_source.value, m.recall_number, m.hospital_asset_number, m.manufacturer,
                    m.model, m.confidence_score, m.status.value, m.method.value, m.reason,
                ]
                for m in data.match_results
            ],
        )

        wb.save(output_path)
    except Exception as exc:
        raise ReportGenerationError(f"Failed to generate Excel report: {exc}") from exc

    audit_export.info("Excel report generated: %s", output_path)
    return output_path

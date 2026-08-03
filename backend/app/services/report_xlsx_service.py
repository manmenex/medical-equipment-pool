"""Roadmap PR18E (docs/design/PR18_PRINTING_EXPORT_PLAN.md §11 "Excel
Strategy"): the backend Excel adapter. Turns an already-built
`app.schemas.report_export.ExportDocument` into `.xlsx` bytes using
`openpyxl` -- this module owns no report/query/eligibility logic of its own
and never mutates the `ExportDocument` it is given, matching
`app.services.report_pdf_service`'s own boundary for the PDF adapter.

Branding (design §16, interim neutral fallback -- Owner Decision #2 remains
open, not decided here): no hospital name, no department name, no logo. Same
product-neutral Thai title and "Medical Equipment Pool" secondary label
`report_pdf_service`/Browser Print already use.

Library choice: `openpyxl` (already an existing, vetted dependency --
`app.services.report_service`'s legacy exporter and `app.services.
import_service`'s `.xlsx` import parser both already depend on it). See this
PR's description for the full comparison against `xlsxwriter` and the
rationale for not adding a new dependency.

Streaming: deliberately NOT using `openpyxl`'s write-only/streaming mode.
Design §11/§18 say streaming should be used "where compatible with the
required formatting" -- write-only mode cannot reliably combine a
metadata block, a frozen header row, an autofilter range, per-cell number
formats, and formula-injection-safe text in one worksheet the way this
adapter requires, and at the approved `MAX_EXPORT_ROWS = 5000` row bound
(`app.services.report_export_service`), standard (buffered) `openpyxl`
generation is fast and low-memory in practice -- the same reasoning the
pre-existing legacy `report_service.export_xlsx` already relies on for up to
50,000 rows.

Admission control (Roadmap PR18E review round 1, H2): design §18's "explicit
time, memory, and concurrency bounds" applies to every renderer, not only
PDF. `build_workbook_bounded` below reuses the exact same proven
protection model `app.services.report_pdf_service.render_pdf_bounded`
implements -- a bounded semaphore, one total deadline covering both queue
wait and active generation, and renderer-lifetime (not request-lifetime)
concurrency accounting via a `Task.add_done_callback` release -- rather
than inventing a second, unrelated concurrency framework. The concrete
`RENDER_TIMEOUT_SECONDS`/`MAX_CONCURRENT_RENDERS` values are lower/higher
than PDF's own (see their docstrings below): `openpyxl` at this row bound
has none of WeasyPrint's native font-shaping/page-layout cost, so a
shorter timeout and a higher concurrency ceiling are both still
generous-but-finite bounds for this renderer's actual resource profile.
"""

import asyncio
import io
from datetime import date, datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.exceptions import XlsxRenderTimeoutError
from app.schemas.report_export import ExportDocument, ExportMetadata, ExportValue, ExportValueType, ReportIdentity

# Roadmap PR18E review round 1 (H2): mirrors
# `report_pdf_service.RENDER_TIMEOUT_SECONDS`/`MAX_CONCURRENT_RENDERS`'s own
# design §18 rationale -- `RENDER_TIMEOUT_SECONDS` is the *total* budget for
# one `.xlsx` export request (time queued for renderer capacity plus the
# active generation itself, never just the latter); `MAX_CONCURRENT_RENDERS`
# bounds how many `openpyxl` workbook generations this worker process admits
# at once. Not derived from `MAX_EXPORT_ROWS` -- deliberately generous-but-
# finite constants for this renderer's own resource profile, distinct from
# (and not shared with) PDF's semaphore: `openpyxl` at the approved 5,000-row
# bound has none of WeasyPrint's native font-shaping/page-layout cost, so a
# shorter timeout and a higher concurrency ceiling are still safely bounded.
RENDER_TIMEOUT_SECONDS = 15
MAX_CONCURRENT_RENDERS = 8
_render_semaphore = asyncio.Semaphore(MAX_CONCURRENT_RENDERS)

# Roadmap PR18E design §11: "Cells whose text begins with spreadsheet formula
# prefixes (=, +, -, @) must be written as literal text." A leading single
# quote is the standard, portable mitigation (OWASP's documented CSV/Excel
# formula-injection defense) -- it forces the cell to display and store as
# literal text starting with that character, never as a formula, in Excel and
# in any other spreadsheet tool that might import this file.
_FORMULA_INJECTION_PREFIXES = ("=", "+", "-", "@")

# English worksheet-tab names (design §11 "worksheet title") -- ASCII, well
# under Excel's 31-character sheet-name limit, and stable regardless of the
# document's Thai display name (which appears inside the sheet as the title
# row instead, matching PDF/Browser Print's own metadata block).
_WORKSHEET_TITLES: dict[ReportIdentity, str] = {
    ReportIdentity.RECEIVE_REPORT: "Receive Report",
    ReportIdentity.ISSUE_REPORT: "Issue Report",
    ReportIdentity.EQUIPMENT_VERIFY_CHECKLIST: "Equipment Verify Checklist",
}

_TITLE_FONT = Font(bold=True, size=14)
_SECONDARY_LABEL_FONT = Font(size=9, color="64748B")
_META_LABEL_FONT = Font(size=9, bold=True)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", start_color="1E293B", end_color="1E293B")
_WRAP_TOP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# Column-width bounds for the "column sizing strategy" requirement (design
# §11) -- wide enough for a typical value, capped so one unusually long cell
# cannot blow out the whole sheet's usable width.
_MIN_COLUMN_WIDTH = 12
_MAX_COLUMN_WIDTH = 40


def _sanitize_string_cell(value: str) -> str:
    """Roadmap PR18E design §11/security: literal-text guard against
    spreadsheet formula injection. Called for every string this module
    writes into any worksheet cell -- see `_write_cell` below, the single
    call site every write in this module goes through (Roadmap PR18E
    review round 1, H1)."""
    if value.startswith(_FORMULA_INJECTION_PREFIXES):
        return f"'{value}"
    return value


def _write_cell(ws: Worksheet, *, row: int, column: int, value: ExportValue, font=None):
    """Roadmap PR18E review round 1 (H1): the single, centralized call site
    for writing *any* value into *any* worksheet cell in this module --
    report rows, the metadata block (title, secondary label, generated-at/
    generated-by/report-identity/template-version/row-count, applied-filter
    labels and values), and the column header row all go through this one
    function, never `ws.cell(...)` directly. Any string value is routed
    through `_sanitize_string_cell` unconditionally, regardless of whether
    its source is a hardcoded label or dynamic, potentially operator-
    editable data (a user's display name, a ward/category/equipment/
    operator name resolved into a filter summary, a future worksheet
    string) -- centralizing here means there is exactly one place that can
    forget to sanitize, not one per call site scattered through the
    module. Non-string values (numbers, dates, datetimes, `None`) pass
    through unchanged; `openpyxl`'s own type system, not this guard, is
    what keeps those cell types native."""
    if isinstance(value, str):
        value = _sanitize_string_cell(value)
    cell = ws.cell(row=row, column=column, value=value)
    if font is not None:
        cell.font = font
    return cell


def _format_datetime_in_timezone(value: datetime, timezone_name: str) -> datetime:
    """Converts a timezone-aware `datetime` into the document's own declared
    display timezone and strips `tzinfo` -- `openpyxl`/Excel's datetime cell
    type has no timezone concept at all (`wb.save()` raises `TypeError` on a
    timezone-aware value), so the wall-clock value in the report's own
    timezone (matching `report_pdf_service._format_datetime_in_timezone` and
    `PrintDocumentView.tsx`'s own conversion) is what a reader actually
    wants, not a bare UTC instant."""
    try:
        zone = ZoneInfo(timezone_name)
    except Exception:
        zone = ZoneInfo("UTC")
    return value.astimezone(zone).replace(tzinfo=None)


def _cell_value_for(value: ExportValue, value_type: ExportValueType, *, timezone_name: str) -> ExportValue:
    """Maps one `ExportRow` value to the actual Python value that should be
    written for its declared `value_type` -- native numeric/date/datetime
    types wherever the schema allows (design §11: "actual Excel date/time
    cell types, not preformatted ISO text"), a Thai-labeled string for
    `"boolean"` (matching `report_pdf_service._format_value`'s own
    convention, since a raw Excel TRUE/FALSE would be English-only), and
    `None` left as a genuinely empty cell (not a "-" placeholder) so
    spreadsheet features like COUNTBLANK/filtering behave correctly.
    Deliberately does *not* apply the formula-injection guard itself
    (Roadmap PR18E review round 1, H1) -- `_write_cell` is the single,
    centralized place that guard runs, for every string this module
    writes, not just row values."""
    if value is None:
        return None
    if value_type == "string":
        assert isinstance(value, str)
        return value
    if value_type == "boolean":
        return "ใช่" if value else "ไม่ใช่"
    if value_type == "datetime":
        assert isinstance(value, datetime)
        return _format_datetime_in_timezone(value, timezone_name)
    if value_type == "date":
        assert isinstance(value, date)
        return value
    # "integer"/"decimal": native numeric passthrough -- Excel's own number
    # type, sortable/summable, never ambiguous the way formatted text is.
    return value


def _number_format_for(value_type: ExportValueType) -> str | None:
    if value_type == "date":
        return "yyyy-mm-dd"
    if value_type == "datetime":
        return "yyyy-mm-dd hh:mm"
    return None


def _write_metadata_block(ws: Worksheet, metadata: ExportMetadata) -> int:
    """Writes the workbook metadata/header section (design §11: "a metadata/
    header section with report display name and identity; applied-filter
    summary; generated timestamp, timezone, generated-by display name,
    template version, and row count") starting at row 1. Returns the first
    row index available for the data table below it."""
    row = 1
    _write_cell(ws, row=row, column=1, value="Medical Equipment Pool", font=_SECONDARY_LABEL_FONT)
    row += 1
    _write_cell(ws, row=row, column=1, value=metadata.display_name_th, font=_TITLE_FONT)
    row += 2

    generated_at_display = _format_datetime_in_timezone(metadata.generated_at, metadata.timezone)
    meta_lines = [
        ("สร้างเมื่อ", f"{generated_at_display.isoformat(sep=' ', timespec='minutes')} ({metadata.timezone})"),
        ("สร้างโดย", metadata.generated_by_display_name),
        ("รหัสรายงาน", metadata.report_identity.value),
        ("เวอร์ชันเอกสาร", metadata.template_version),
        ("จำนวนรายการ", str(metadata.row_count)),
    ]
    for label, value in meta_lines:
        _write_cell(ws, row=row, column=1, value=f"{label}:", font=_META_LABEL_FONT)
        _write_cell(ws, row=row, column=2, value=value)
        row += 1

    if metadata.applied_filters:
        row += 1
        _write_cell(ws, row=row, column=1, value="ตัวกรองที่ใช้:", font=_META_LABEL_FONT)
        row += 1
        for filter_summary in metadata.applied_filters:
            _write_cell(ws, row=row, column=1, value=f"{filter_summary.label_th}:", font=_META_LABEL_FONT)
            _write_cell(ws, row=row, column=2, value=filter_summary.value)
            row += 1

    return row + 1


def build_workbook_sync(document: ExportDocument) -> bytes:
    """Renders one `ExportDocument` to `.xlsx` bytes. CPU-bound and
    synchronous by design -- callers must run this via `asyncio.to_thread`
    (or an equivalent executor) so it never blocks the event loop, exactly
    as `app.services.import_service._parse_workbook_sync` and
    `app.services.report_pdf_service.render_pdf` already do for their own
    CPU-bound work.

    Single worksheet only (design §22 PR18E non-goal: "no multiple
    worksheets unless approved by design"). No macros, no formulas from
    document content, no charts, no pivot tables, no external links.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _WORKSHEET_TITLES[document.metadata.report_identity]

    header_row = _write_metadata_block(worksheet, document.metadata)

    for column_index, column in enumerate(document.columns, start=1):
        cell = _write_cell(worksheet, row=header_row, column=column_index, value=column.label_th, font=_HEADER_FONT)
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_TOP_ALIGNMENT

    for row_offset, row in enumerate(document.rows, start=1):
        for column_index, column in enumerate(document.columns, start=1):
            cell = _write_cell(
                worksheet,
                row=header_row + row_offset,
                column=column_index,
                value=_cell_value_for(row.values[column.key], column.value_type, timezone_name=document.metadata.timezone),
            )
            cell.alignment = _WRAP_TOP_ALIGNMENT
            number_format = _number_format_for(column.value_type)
            if number_format is not None:
                cell.number_format = number_format

    last_row = header_row + len(document.rows)
    last_column_letter = get_column_letter(max(len(document.columns), 1))

    # "frozen header row" (design §11): everything from row 1 through the
    # header row stays visible while the data below scrolls -- the standard
    # openpyxl idiom for freezing every row above a given cell.
    worksheet.freeze_panes = f"A{header_row + 1}"

    # "auto filter" (design §11): scoped to the header + data rows only,
    # never the metadata block above it.
    worksheet.auto_filter.ref = f"A{header_row}:{last_column_letter}{last_row}"

    for column_index, column in enumerate(document.columns, start=1):
        width = min(max(len(column.label_th) + 2, _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _release_render_slot_when_done(task: "asyncio.Task[bytes]") -> None:
    """Roadmap PR18E review round 1 (H2): mirrors
    `report_pdf_service._release_render_slot_when_done`'s exact rationale --
    the semaphore models *renderer* concurrency, not *request* concurrency.
    This callback is the only place `_render_semaphore` is released, and it
    only ever runs once the underlying `asyncio.to_thread(build_workbook_
    sync, ...)` task has actually finished (successfully, or with an
    exception), never merely because a caller's `wait_for` gave up waiting
    on it. Retrieving the task's exception (if any) here, even though
    nothing further is done with it, prevents asyncio's "Task exception was
    never retrieved" warning for a generation that finishes after its
    caller already received `XlsxRenderTimeoutError`."""
    if not task.cancelled():
        task.exception()
    _render_semaphore.release()


def _render_timeout_error() -> XlsxRenderTimeoutError:
    return XlsxRenderTimeoutError(
        f"Excel generation did not complete within {RENDER_TIMEOUT_SECONDS} seconds. "
        "Narrow the applied filters and try again."
    )


async def build_workbook_bounded(document: ExportDocument) -> bytes:
    """Roadmap PR18E review round 1 (H2): the one, single call site that
    enforces design §18's time and concurrency bounds around
    `build_workbook_sync` -- `build_workbook_sync` itself stays a plain,
    directly-unit-testable synchronous function; callers (the API route)
    must call this wrapper, never `asyncio.to_thread(build_workbook_sync,
    ...)` directly, so no call site can accidentally run an unbounded
    generation. This reuses, call for call, the exact protection model
    already proven by `report_pdf_service.render_pdf_bounded` (including
    its round 2/round 3 corrections) rather than a second, independently
    invented concurrency framework:

    - One total `deadline` (via the running loop's own monotonic clock)
      covers both the wait for renderer capacity and the active
      generation itself -- a request that spends its whole budget queued
      behind other generations fails with the same structured timeout
      error a request that spent the whole budget actually generating
      would get; the render is never started at all if the deadline has
      already passed once a slot is acquired.
    - The semaphore models *renderer* concurrency, not *request*
      concurrency: generation runs as its own `asyncio.Task`
      (`render_task` below), and `asyncio.wait_for` wraps
      `asyncio.shield(render_task)`, so a caller-facing timeout cancels
      only its own wrapper -- `render_task` keeps running in the
      background exactly as it would have without the timeout. The
      semaphore is released solely by `render_task`'s own done-callback
      (`_release_render_slot_when_done`), so a request that times out
      from the *caller's* perspective does not free the slot until
      generation has *actually* finished.
    - `_render_semaphore.acquire()` is itself wrapped in `asyncio.
      wait_for` while queuing for a slot; `asyncio.Semaphore.acquire` is
      cancellation-safe, so a queue-timeout never leaks or misroutes a
      permit.

    `build_workbook_sync` holds no external resource of its own (no temp
    file, no open connection or socket -- the workbook is built entirely
    in memory and returned as bytes), so accounting for the slot
    correctly is the entire cleanup obligation here.
    """
    loop = asyncio.get_running_loop()
    deadline = loop.time() + RENDER_TIMEOUT_SECONDS

    remaining = deadline - loop.time()
    if remaining <= 0:
        raise _render_timeout_error()
    try:
        await asyncio.wait_for(_render_semaphore.acquire(), timeout=remaining)
    except asyncio.TimeoutError as exc:
        raise _render_timeout_error() from exc

    remaining = deadline - loop.time()
    if remaining <= 0:
        # The entire budget was consumed just queuing for a slot -- fail
        # now, before ever starting generation, rather than letting it
        # begin with no time left for it.
        _render_semaphore.release()
        raise _render_timeout_error()

    try:
        render_task: "asyncio.Task[bytes]" = asyncio.ensure_future(asyncio.to_thread(build_workbook_sync, document))
    except BaseException:
        # Task creation itself failed before any generation work could
        # start -- nothing for the done-callback to ever fire for, so
        # release the slot directly here instead of leaking it.
        _render_semaphore.release()
        raise
    render_task.add_done_callback(_release_render_slot_when_done)

    try:
        return await asyncio.wait_for(asyncio.shield(render_task), timeout=remaining)
    except asyncio.TimeoutError as exc:
        raise _render_timeout_error() from exc

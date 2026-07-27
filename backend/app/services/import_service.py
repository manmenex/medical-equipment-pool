"""Inventory Import (Roadmap PR12).

See docs/audits/04-consolidated-implementation-plan.md Part D ("PR12 --
Inventory import") and Part F ("Inventory Import Plan", §10) for the full
authoritative design this module implements, and ADR-002
(knowledge/adr/ADR-002-identifier-model.md) for the identifier model this
module must not violate.

Design summary (do not restate elsewhere -- read the plan for rationale):

  - Roadmap PR12 review findings PR12-H1/H1R (ADR-002) + Repository Owner
    decision: this workflow is UPDATE-ONLY. BCM Code is the sole match
    key used to find the existing equipment record a row refers to; a row
    whose BCM Code does not match an existing record fails validation
    rather than creating one, because the import source file has no
    Asset Number column and equipment.asset_number is NOT NULL/UNIQUE
    real inventory metadata this workflow must never generate, derive,
    or synthesize (ADR-002: "Not merged with, or inferred from, BCM Code
    or Item No"). Create-from-import is deferred to a future Roadmap
    item. Item No, Asset ID, and Serial Number are validated as
    uniqueness constraints against *other* records, never used to look
    up "is this the same equipment" (per the Repository Owner's explicit
    approved treatment).
  - `process_import` is the single entry point for both preview
    (`commit=False`, read-only, zero database writes including audit
    writes) and commit (`commit=True`, revalidates from the raw uploaded
    file bytes every time -- it never trusts a prior preview result
    supplied by the frontend, since preview and commit are two
    independent calls to this same function).
  - "Update existing" mode only ever touches the approved master-data
    field set (Asset ID, Manufacturer/brand, Model, and the raw
    Location/Receive Date/Register Date/Purchase Year provenance fields
    inside `equipment_metadata`) -- it never writes `status`,
    `legacy_status`, `asset_number`, `item_no`, `serial_number`, or
    `equipment_name` on an existing record (Part F.3).
  - Unexpected failures during the commit write phase roll back the
    entire batch (no partial commit); expected per-row validation
    failures never reach the write phase at all (every bounded/persisted
    field is length- and duplicate-checked during validation, not left to
    fail at the database), so they coexist freely with rows that do
    succeed.
  - Review finding PR12-H2 / PR12-H2R: uploads are bounded (compressed
    size, row count, filename, extension) before any parsing; the .xlsx
    ZIP container's central directory is inspected (entry count, per-
    entry and aggregate uncompressed size, compression ratio, permitted
    entry paths) before a single byte of member content is decompressed,
    so a small, highly-compressible archive cannot expand past these
    bounds inside openpyxl. Worksheet count and header column count are
    likewise capped. The CPU-bound XLSX parse runs off the event loop via
    `asyncio.to_thread`. Database duplicate lookups are bulk `IN(...)`
    queries (a handful of queries per batch), not one query per row.
"""

import asyncio
import io
import logging
import re
import uuid
import zipfile
from dataclasses import dataclass
from enum import Enum

from fastapi import Request, UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import AUDIT_ACTION_IMPORT, AUDIT_ENTITY_EQUIPMENT, record_audit_event
from app.core.exceptions import DomainError, InvalidInputError
from app.crud import equipment as equipment_crud
from app.models.equipment import EquipmentStatus
from app.services.identifiers import normalize_bcm_code, normalize_item_no

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Upload bounds (review finding PR12-H2). Conservative for the confirmed
# project scale assumption ("low hundreds of equipment records", Part A) --
# generous enough for a real hospital import batch, small enough that an
# oversized/pathological upload is rejected before it can consume
# unbounded memory or event-loop time.
# ---------------------------------------------------------------------------
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MiB
MAX_IMPORT_ROWS = 5000
MAX_FILENAME_LENGTH = 255
ALLOWED_FILENAME_EXTENSIONS = (".xlsx",)
_UPLOAD_READ_CHUNK_BYTES = 65536

# ---------------------------------------------------------------------------
# Decompressed-archive bounds (review finding PR12-H2R). An .xlsx file is a
# ZIP container -- the compressed MAX_UPLOAD_BYTES cap above bounds what is
# read off the wire, but says nothing about how large the *decompressed*
# content is. A small, highly-compressible upload (a classic zip-bomb
# pattern: repetitive XML/shared-string content) can still stay under
# MAX_UPLOAD_BYTES while expanding to a size that exhausts memory/CPU once
# `openpyxl.load_workbook` decompresses it. These bounds are enforced by
# reading the ZIP central directory only (`ZipFile.infolist()` never
# decompresses member data) before openpyxl -- or any other library -- is
# allowed to touch the archive's actual content.
# ---------------------------------------------------------------------------
MAX_ZIP_ENTRIES = 100
MAX_ZIP_ENTRY_UNCOMPRESSED_BYTES = 25 * 1024 * 1024  # 25 MiB per member
MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES = 60 * 1024 * 1024  # 60 MiB aggregate
MAX_ZIP_COMPRESSION_RATIO = 200  # uncompressed:compressed; classic zip bombs are far higher
MAX_WORKSHEET_COUNT = 25
MAX_HEADER_COLUMNS = 200

# Only entries under these top-level OOXML package paths are permitted --
# an .xlsx member outside this list (path traversal, an embedded macro/
# object type this application never reads, an unrelated payload smuggled
# into the archive) is rejected outright rather than silently ignored.
_ALLOWED_ZIP_ENTRY_PATTERN = re.compile(
    r"^(\[Content_Types\]\.xml|_rels/|docProps/|xl/|customXml/)"
)

# ---------------------------------------------------------------------------
# Part F.1 step 2: the confirmed source spreadsheet headers. Matched
# case-insensitively, trimmed -- but the canonical label used in every
# error message and header-index lookup is exactly this text.
# ---------------------------------------------------------------------------
REQUIRED_HEADERS: tuple[str, ...] = (
    "Item No.",
    "ID CODE",
    "Asset ID",
    "Equipment Name",
    "Manufacturer",
    "Model",
    "Serial Number",
    "Location",
    "Receive Date",
    "Register Date",
    "Purchase Year",
    "Asset Status",
)

# Part F.2: illustrative Asset Status mapping, pending confirmation against
# a real hospital inventory export (§14 open question 4). Keys are
# lowercased for case-insensitive exact-match lookup; an unrecognized
# value is a per-row failure, never a guess (Part F.1 step 3).
ASSET_STATUS_MAPPING: dict[str, EquipmentStatus] = {
    "active": EquipmentStatus.AVAILABLE_AT_POOL,
    "in use": EquipmentStatus.AVAILABLE_AT_POOL,
    "in service": EquipmentStatus.AVAILABLE_AT_POOL,
    "defective": EquipmentStatus.UNAVAILABLE_DEFECTIVE,
    "faulty": EquipmentStatus.UNAVAILABLE_DEFECTIVE,
    "under repair": EquipmentStatus.UNAVAILABLE_DEFECTIVE,
    "decommissioned": EquipmentStatus.DECOMMISSIONED,
    "disposed": EquipmentStatus.DECOMMISSIONED,
    "written off": EquipmentStatus.DECOMMISSIONED,
}

# Review finding PR12-H4: every bounded/persisted field's maximum length,
# copied from backend/app/models/equipment.py's column widths. Checked
# during validation (preview and commit alike), so an overlength value is
# always reported as a per-row failure -- never a PostgreSQL DataError
# discovered only at commit time, after preview already reported success.
FIELD_MAX_LENGTHS: dict[str, int] = {
    "Asset ID": 100,
    "Serial Number": 100,
    "Equipment Name": 255,
    "Manufacturer": 100,
    "Model": 100,
    "Asset Status": 100,  # raw_source_status
}

# Review findings PR12-H1 / PR12-H1R (ADR-002) + Repository Owner decision:
# equipment.asset_number is NOT NULL + UNIQUE and used throughout the
# application (dashboards, reports, PM/Calibration notifications, API
# responses) as a real, hospital-assigned business identifier. The import
# source file has no "Asset Number" column, and both an attempted
# BCM-derived value and a random placeholder token were independently
# rejected as fabricating that identifier, contrary to ADR-002 ("Not
# merged with, or inferred from, BCM Code or Item No" / Asset Number is
# assigned inventory metadata, not application-generated).
#
# Resolution (Repository Owner decision): Roadmap PR12 is update-only.
# Import never creates new equipment and never generates, derives, or
# synthesizes an Asset Number. A row whose BCM Code does not match an
# existing equipment record fails validation, instructing the operator to
# create that equipment through the standard Equipment Master workflow
# first, then re-import to update it. Create-from-import is deferred to a
# future Roadmap item that must first define the real hospital Asset
# Number assignment workflow.
NO_MATCHING_EQUIPMENT_REASON = (
    "No existing equipment matches this BCM Code. Roadmap PR12 Inventory "
    "Import is update-only and does not create new equipment records or "
    "assign Asset Numbers. Create this equipment through the standard "
    "Equipment Master workflow first, then re-import this file to update it."
)


class ImportRowStatus(str, Enum):
    SUCCESS = "success"
    FAILED = "failed"
    # Review finding PR12-H1R / Repository Owner decision: Roadmap PR12 is
    # update-only and `update_existing=false` is now rejected outright at
    # process_import's entry point (never reaching per-row validation), so
    # no code path can currently produce SKIPPED. Kept in the API contract
    # (never emitted) rather than removed, since narrowing the response
    # shape is a separate decision from this fix.
    SKIPPED = "skipped"


class ImportRowAction(str, Enum):
    UPDATE = "update"


@dataclass
class ImportRowResult:
    row_number: int
    bcm_code: str | None
    item_no: str | None
    asset_id: str | None
    status: ImportRowStatus
    action: ImportRowAction | None = None
    reason: str | None = None
    equipment_id: str | None = None


@dataclass
class ImportBatchResult:
    filename: str
    total_rows: int
    succeeded: int
    failed: int
    skipped: int
    rows: list[ImportRowResult]
    audit_log_id: str | None = None


@dataclass
class _RowPlan:
    result: ImportRowResult
    equipment_id: uuid.UUID | None = None
    update_data: dict | None = None
    import_source_metadata: dict | None = None


def _cell_to_text(value: object) -> str:
    """Best-effort text form of a raw openpyxl cell value, preserving BCM
    Code / Item No / Asset ID / Serial Number as strings throughout.

    Excel-side numeric coercion (a source column typed/formatted as a
    number in the spreadsheet itself, losing leading zeros before this
    file is even read) cannot be recovered here -- this only avoids
    *compounding* that loss (e.g. rendering an integer-valued float as
    "1.0" instead of "1").
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _cell_to_text_verbatim(value: object) -> str:
    """Like `_cell_to_text`, but never strips or otherwise normalizes a
    string value -- used only for the Asset Status source cell, whose
    exact original text (leading/trailing whitespace, casing, internal
    whitespace) must be preserved byte-for-byte in the `raw_source_status`
    audit column. Business validation (the Asset Status mapping lookup)
    normalizes its own copy of this value at the point of use; this
    function's output is never itself normalized, since it is what gets
    persisted verbatim.

    A non-string cell still needs the same type coercion `_cell_to_text`
    applies -- there is no "original whitespace" to lose for a numeric
    cell, since Excel's own numeric type never carries surrounding text.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value)


# ---------------------------------------------------------------------------
# Upload validation (review finding PR12-H2). All of this happens before a
# single byte is handed to openpyxl.
# ---------------------------------------------------------------------------


def _validate_filename(filename: str | None) -> str:
    name = (filename or "").strip()
    if not name:
        return "upload.xlsx"
    if len(name) > MAX_FILENAME_LENGTH:
        raise InvalidInputError(f"Filename exceeds the maximum length of {MAX_FILENAME_LENGTH} characters.")
    if not name.lower().endswith(ALLOWED_FILENAME_EXTENSIONS):
        raise InvalidInputError("Only .xlsx files are supported.")
    return name


async def _read_upload_bounded(file: UploadFile, *, max_bytes: int) -> bytes:
    """Reads the upload in fixed-size chunks, aborting as soon as the
    running total exceeds `max_bytes` -- an oversized upload is rejected
    without ever buffering the whole thing into memory."""
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await file.read(_UPLOAD_READ_CHUNK_BYTES)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            raise InvalidInputError(
                f"Uploaded file exceeds the maximum allowed size of {max_bytes // (1024 * 1024)} MB."
            )
        chunks.append(chunk)
    return b"".join(chunks)


# ---------------------------------------------------------------------------
# Synchronous parsing (review finding PR12-H2: run via asyncio.to_thread so
# CPU-bound XLSX decompression/parsing never blocks the event loop).
# ---------------------------------------------------------------------------


def _validate_zip_archive_bounds(content: bytes) -> None:
    """Review finding PR12-H2R: inspect the .xlsx ZIP container's central
    directory -- which `ZipFile.infolist()` reads without decompressing
    any member -- and reject an unsafe archive before openpyxl (or
    anything else) is allowed to expand a single byte of member content.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
    except zipfile.BadZipFile as exc:
        raise InvalidInputError(
            "The uploaded file could not be read as a valid Excel (.xlsx) spreadsheet."
        ) from exc

    if len(infos) > MAX_ZIP_ENTRIES:
        raise InvalidInputError(
            f"The uploaded file's internal structure contains more than {MAX_ZIP_ENTRIES} "
            "parts, exceeding what a normal Excel (.xlsx) file contains."
        )

    total_uncompressed = 0
    for info in infos:
        name = info.filename.replace("\\", "/")
        if ".." in name.split("/") or name.startswith("/") or not _ALLOWED_ZIP_ENTRY_PATTERN.match(name):
            raise InvalidInputError(
                "The uploaded file contains an unexpected internal component and could not "
                "be accepted as a valid Excel (.xlsx) spreadsheet."
            )
        if info.file_size > MAX_ZIP_ENTRY_UNCOMPRESSED_BYTES:
            raise InvalidInputError(
                "The uploaded file's internal content is too large to process safely."
            )
        if info.compress_size == 0:
            ratio = float(info.file_size) if info.file_size else 0.0
        else:
            ratio = info.file_size / info.compress_size
        if ratio > MAX_ZIP_COMPRESSION_RATIO:
            raise InvalidInputError(
                "The uploaded file's internal content is compressed in a way that could not "
                "be accepted as a valid Excel (.xlsx) spreadsheet."
            )
        total_uncompressed += info.file_size
        if total_uncompressed > MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES:
            raise InvalidInputError(
                "The uploaded file's internal content is too large to process safely."
            )


def _load_worksheet(content: bytes) -> Worksheet:
    _validate_zip_archive_bounds(content)
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise InvalidInputError(
            "The uploaded file could not be read as a valid Excel (.xlsx) spreadsheet."
        ) from exc
    if len(workbook.sheetnames) > MAX_WORKSHEET_COUNT:
        raise InvalidInputError(
            f"The uploaded file contains more than {MAX_WORKSHEET_COUNT} worksheets, "
            "exceeding what a normal inventory import file contains."
        )
    worksheet = workbook.active
    if worksheet is None:
        raise InvalidInputError("The uploaded spreadsheet has no worksheets.")
    return worksheet


def _validate_headers(worksheet: Worksheet) -> dict[str, int]:
    """Confirm every required header is present (Part F.1 step 2); reject
    the file outright with a clear message rather than a partial/guessed
    mapping. Returns {canonical_header: 0-based_column_index}."""
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise InvalidInputError("The uploaded spreadsheet is empty; no header row was found.")
    if len(header_row) > MAX_HEADER_COLUMNS:
        raise InvalidInputError(
            f"The uploaded file's header row contains more than {MAX_HEADER_COLUMNS} columns, "
            "exceeding what a normal inventory import file contains."
        )

    found: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        text = _cell_to_text(raw)
        if not text:
            continue
        found[text.strip().lower()] = idx

    header_index: dict[str, int] = {}
    missing: list[str] = []
    for header in REQUIRED_HEADERS:
        idx = found.get(header.lower())
        if idx is None:
            missing.append(header)
        else:
            header_index[header] = idx
    if missing:
        raise InvalidInputError(
            "The uploaded spreadsheet is missing required column(s): " + ", ".join(missing)
        )
    return header_index


@dataclass
class _ParsedRow:
    row_number: int
    item_no_raw: str
    bcm_code_raw: str
    asset_id_raw: str
    equipment_name_raw: str
    brand_raw: str
    model_raw: str
    serial_number_raw: str
    location_raw: str
    receive_date_raw: str
    register_date_raw: str
    purchase_year_raw: str
    # Review finding PR12-#2: the exact, unmodified Asset Status source
    # cell text (no strip/case/whitespace normalization) -- this is what
    # gets persisted verbatim into equipment.raw_source_status. Business
    # validation (the status mapping lookup) normalizes its own copy at
    # the point of use; see _validate_rows.
    asset_status_verbatim: str


def _parse_rows(worksheet: Worksheet, header_index: dict[str, int]) -> list[_ParsedRow]:
    def cell(values: tuple, header: str) -> str:
        idx = header_index[header]
        return _cell_to_text(values[idx]) if idx < len(values) else ""

    def cell_verbatim(values: tuple, header: str) -> str:
        idx = header_index[header]
        return _cell_to_text_verbatim(values[idx]) if idx < len(values) else ""

    parsed: list[_ParsedRow] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if values is None or all(v is None or _cell_to_text(v) == "" for v in values):
            continue  # trailing/blank row -- not counted, not reported
        if len(parsed) >= MAX_IMPORT_ROWS:
            raise InvalidInputError(
                f"The uploaded file contains more than {MAX_IMPORT_ROWS} data rows, "
                "exceeding the maximum a single import batch may contain."
            )
        parsed.append(
            _ParsedRow(
                row_number=row_number,
                item_no_raw=cell(values, "Item No."),
                bcm_code_raw=cell(values, "ID CODE"),
                asset_id_raw=cell(values, "Asset ID"),
                equipment_name_raw=cell(values, "Equipment Name"),
                brand_raw=cell(values, "Manufacturer"),
                model_raw=cell(values, "Model"),
                serial_number_raw=cell(values, "Serial Number"),
                location_raw=cell(values, "Location"),
                receive_date_raw=cell(values, "Receive Date"),
                register_date_raw=cell(values, "Register Date"),
                purchase_year_raw=cell(values, "Purchase Year"),
                asset_status_verbatim=cell_verbatim(values, "Asset Status"),
            )
        )
    return parsed


def _parse_workbook_sync(content: bytes) -> list[_ParsedRow]:
    """Everything CPU-bound and synchronous about turning raw file bytes
    into parsed rows, run as one unit via asyncio.to_thread (review
    finding PR12-H2) so it never blocks the event loop that serves every
    other concurrent request."""
    worksheet = _load_worksheet(content)
    header_index = _validate_headers(worksheet)
    return _parse_rows(worksheet, header_index)


def _find_in_file_duplicates(values: dict[int, str]) -> set[int]:
    """values: {row_number: normalized_value} for rows where the value is
    non-empty. Returns the set of row_numbers whose value appears more
    than once -- ALL occurrences fail, per Part F.1 step 3's recommended
    "flag all duplicates within a file" policy."""
    seen: dict[str, list[int]] = {}
    for row_number, value in values.items():
        seen.setdefault(value, []).append(row_number)
    duplicated_rows: set[int] = set()
    for rows in seen.values():
        if len(rows) > 1:
            duplicated_rows.update(rows)
    return duplicated_rows


def _length_violation(field_label: str, value: str | None) -> str | None:
    max_length = FIELD_MAX_LENGTHS[field_label]
    if value is not None and len(value) > max_length:
        return f"{field_label} exceeds the maximum length of {max_length} characters."
    return None


async def _validate_rows(db: AsyncSession, parsed_rows: list[_ParsedRow]) -> list[_RowPlan]:
    plans: dict[int, _RowPlan] = {}
    normalized_bcm: dict[int, str] = {}
    normalized_item_no: dict[int, str] = {}
    normalized_asset_id: dict[int, str] = {}
    normalized_serial: dict[int, str] = {}

    def fail(row_number: int, bcm_code: str | None, item_no: str | None, asset_id: str | None, reason: str) -> None:
        plans[row_number] = _RowPlan(
            result=ImportRowResult(
                row_number=row_number,
                bcm_code=bcm_code,
                item_no=item_no,
                asset_id=asset_id,
                status=ImportRowStatus.FAILED,
                reason=reason,
            )
        )

    # Pass 1: structural normalization + length validation (review finding
    # PR12-H4). A row that fails here is final -- it never reaches a
    # database check.
    for row in parsed_rows:
        if not row.bcm_code_raw:
            fail(row.row_number, None, row.item_no_raw or None, row.asset_id_raw or None, "Missing BCM Code (ID CODE column).")
            continue
        try:
            bcm_code = normalize_bcm_code(row.bcm_code_raw)
        except InvalidInputError as exc:
            fail(row.row_number, None, row.item_no_raw or None, row.asset_id_raw or None, f"Invalid BCM Code: {exc.message}")
            continue

        item_no: str | None = None
        if row.item_no_raw:
            try:
                item_no = normalize_item_no(row.item_no_raw)
            except InvalidInputError as exc:
                fail(row.row_number, bcm_code, None, row.asset_id_raw or None, f"Invalid Item No: {exc.message}")
                continue

        length_error = (
            _length_violation("Asset ID", row.asset_id_raw or None)
            or _length_violation("Serial Number", row.serial_number_raw or None)
            or _length_violation("Equipment Name", row.equipment_name_raw or None)
            or _length_violation("Manufacturer", row.brand_raw or None)
            or _length_violation("Model", row.model_raw or None)
            or _length_violation("Asset Status", row.asset_status_verbatim or None)
        )
        if length_error is not None:
            fail(row.row_number, bcm_code, item_no, row.asset_id_raw or None, length_error)
            continue

        normalized_bcm[row.row_number] = bcm_code
        if item_no:
            normalized_item_no[row.row_number] = item_no
        if row.asset_id_raw:
            normalized_asset_id[row.row_number] = row.asset_id_raw
        if row.serial_number_raw:
            normalized_serial[row.row_number] = row.serial_number_raw

    # Pass 2: in-file duplicate detection -- flags ALL occurrences.
    dup_bcm = _find_in_file_duplicates(normalized_bcm)
    dup_item_no = _find_in_file_duplicates(normalized_item_no)
    dup_asset_id = _find_in_file_duplicates(normalized_asset_id)
    dup_serial = _find_in_file_duplicates(normalized_serial)

    for row in parsed_rows:
        if row.row_number in plans:
            continue
        bcm_code = normalized_bcm[row.row_number]
        item_no = normalized_item_no.get(row.row_number)
        asset_id = normalized_asset_id.get(row.row_number) or None

        if row.row_number in dup_bcm:
            fail(row.row_number, bcm_code, item_no, asset_id, "Duplicate BCM Code within the uploaded file.")
        elif row.row_number in dup_item_no:
            fail(row.row_number, bcm_code, item_no, asset_id, "Duplicate Item No within the uploaded file.")
        elif row.row_number in dup_asset_id:
            fail(row.row_number, bcm_code, item_no, asset_id, "Duplicate Asset ID within the uploaded file.")
        elif row.row_number in dup_serial:
            fail(row.row_number, bcm_code, item_no, asset_id, "Duplicate Serial Number within the uploaded file.")

    # Pass 3: bulk database lookups (review finding PR12-H2) -- a fixed
    # handful of IN(...) queries for the whole batch, never one query per
    # row.
    surviving_rows = [row for row in parsed_rows if row.row_number not in plans]
    bcm_codes_to_check = {normalized_bcm[row.row_number] for row in surviving_rows}
    item_nos_to_check = {normalized_item_no[row.row_number] for row in surviving_rows if row.row_number in normalized_item_no}
    serials_to_check = {normalized_serial[row.row_number] for row in surviving_rows if row.row_number in normalized_serial}
    asset_ids_to_check = {normalized_asset_id[row.row_number] for row in surviving_rows if row.row_number in normalized_asset_id}

    existing_by_bcm = await equipment_crud.get_by_bcm_codes(db, list(bcm_codes_to_check))
    existing_by_item_no = await equipment_crud.get_by_item_nos(db, list(item_nos_to_check))
    existing_by_serial = await equipment_crud.get_by_serial_numbers(db, list(serials_to_check))
    existing_by_asset_id = await equipment_crud.get_by_asset_ids(db, list(asset_ids_to_check))

    # Pass 4: per-row resolution, purely in-memory against the bulk-fetched
    # lookup tables above -- no further database round trips in this loop.
    #
    # Repository Owner decision (resolving review findings PR12-H1/H1R):
    # Roadmap PR12 is update-only. A row whose BCM Code does not match an
    # existing equipment record fails validation rather than creating a
    # new record -- the import source file has no Asset Number column,
    # and equipment.asset_number is NOT NULL/UNIQUE real inventory
    # metadata that this workflow must never generate, derive, or
    # synthesize (ADR-002). Create-from-import is deferred to a future
    # Roadmap item.
    for row in surviving_rows:
        bcm_code = normalized_bcm[row.row_number]
        item_no = normalized_item_no.get(row.row_number)
        asset_id = normalized_asset_id.get(row.row_number) or None
        serial_number = normalized_serial.get(row.row_number) or None

        existing = existing_by_bcm.get(bcm_code)
        if existing is None:
            fail(row.row_number, bcm_code, item_no, asset_id, NO_MATCHING_EQUIPMENT_REASON)
            continue

        # Asset Status: validated for every row about to be updated (Part
        # F.1 step 3's per-row rule) -- even though update never writes
        # `status`, an unrecognized source value still fails the row as a
        # data-quality signal, and raw_source_status is only ever
        # persisted for rows whose source value could be classified at
        # all.
        #
        # Review finding PR12-#2: raw_source_status is an audit field and
        # must preserve the exact source-cell text, so business
        # validation here reads from a separately normalized copy
        # (trimmed, lowercased) -- row.asset_status_verbatim itself is
        # never mutated and is what gets persisted below, unmodified.
        normalized_source_status = row.asset_status_verbatim.strip().lower()
        mapped_status = ASSET_STATUS_MAPPING.get(normalized_source_status)
        if mapped_status is None:
            fail(
                row.row_number,
                bcm_code,
                item_no,
                asset_id,
                f"Unrecognized Asset Status value: '{row.asset_status_verbatim}'.",
            )
            continue

        import_source_metadata = {
            "location": row.location_raw or None,
            "receive_date": row.receive_date_raw or None,
            "register_date": row.register_date_raw or None,
            "purchase_year": row.purchase_year_raw or None,
        }

        # UPDATE path (Part F.3): only the approved master-data field set
        # -- never status/legacy_status/asset_number/item_no/
        # serial_number/equipment_name/bcm_code. Review finding PR12-H3:
        # Item No and Serial Number are still validated (against a
        # *different* equipment record) even though this path never
        # writes them, so a source row whose identifiers point at a
        # different physical device is never reported as a clean update.
        conflict_reason = _find_update_conflict(
            existing_id=existing.id,
            item_no=item_no,
            serial_number=serial_number,
            asset_id=asset_id,
            existing_by_item_no=existing_by_item_no,
            existing_by_serial=existing_by_serial,
            existing_by_asset_id=existing_by_asset_id,
        )
        if conflict_reason is not None:
            fail(row.row_number, bcm_code, item_no, asset_id, conflict_reason)
            continue

        update_data = {
            "asset_id": asset_id,
            "brand": row.brand_raw or None,
            "model": row.model_raw or None,
            # Verbatim source text, never the normalized value used for
            # the mapping lookup above (review finding PR12-#2).
            "raw_source_status": row.asset_status_verbatim or None,
        }
        plans[row.row_number] = _RowPlan(
            result=ImportRowResult(
                row_number=row.row_number,
                bcm_code=bcm_code,
                item_no=item_no,
                asset_id=asset_id,
                status=ImportRowStatus.SUCCESS,
                action=ImportRowAction.UPDATE,
            ),
            equipment_id=existing.id,
            update_data=update_data,
            import_source_metadata=import_source_metadata,
        )

    # Preserve original row order.
    return [plans[row.row_number] for row in parsed_rows]


def _find_update_conflict(
    *,
    existing_id: uuid.UUID,
    item_no: str | None,
    serial_number: str | None,
    asset_id: str | None,
    existing_by_item_no: dict,
    existing_by_serial: dict,
    existing_by_asset_id: dict,
) -> str | None:
    """Review finding PR12-H3: an update-mode row must not be reported as
    a clean update when its Item No or Serial Number belongs to a
    *different* equipment record than the one BCM Code matched -- even
    though update mode never writes either field, accepting such a row
    as "success" would misrepresent that the uploaded data is internally
    consistent. A value that belongs to the same record (or is absent)
    is not a conflict."""
    other_item_no = existing_by_item_no.get(item_no) if item_no is not None else None
    if other_item_no is not None and other_item_no.id != existing_id:
        return "Item No already belongs to a different equipment record."
    other_serial = existing_by_serial.get(serial_number) if serial_number is not None else None
    if other_serial is not None and other_serial.id != existing_id:
        return "Serial Number already belongs to a different equipment record."
    other_asset_id = existing_by_asset_id.get(asset_id) if asset_id is not None else None
    if other_asset_id is not None and other_asset_id.id != existing_id:
        return "Asset ID already belongs to a different equipment record."
    return None


def _summarize(filename: str, plans: list[_RowPlan]) -> ImportBatchResult:
    rows = [plan.result for plan in plans]
    return ImportBatchResult(
        filename=filename,
        total_rows=len(rows),
        succeeded=sum(1 for r in rows if r.status == ImportRowStatus.SUCCESS),
        failed=sum(1 for r in rows if r.status == ImportRowStatus.FAILED),
        skipped=sum(1 for r in rows if r.status == ImportRowStatus.SKIPPED),
        rows=rows,
    )


class ImportCommitFailedError(DomainError):
    """Unexpected failure during the commit write phase (Roadmap PR12) --
    the entire batch is rolled back, never partially committed. Distinct
    from any per-row validation failure, which never reaches the write
    phase at all and does not raise. Review findings PR12-H4/"keep the
    entire import transactional": every bounded/persisted field is
    length- and duplicate-checked during validation, so this branch is
    reserved for genuinely unexpected failures (e.g. a dropped database
    connection), not a foreseeable validation error -- the client always
    sees a safe, generic message here, never a raw database exception."""

    code = "IMPORT_COMMIT_FAILED"
    status_code = 500


async def _commit_rows(
    db: AsyncSession,
    *,
    filename: str,
    plans: list[_RowPlan],
    update_existing: bool,
    actor_user_id: uuid.UUID | None,
    request: Request | None,
) -> ImportBatchResult:
    try:
        for plan in plans:
            if plan.result.status != ImportRowStatus.SUCCESS:
                continue
            # Roadmap PR12 is update-only (see _validate_rows): every
            # SUCCESS row's action is ImportRowAction.UPDATE by
            # construction -- there is no create path to branch on here.
            existing = await equipment_crud.get_by_id(db, plan.equipment_id)
            if existing is None:
                raise ImportCommitFailedError(
                    f"Equipment {plan.equipment_id} no longer exists mid-commit."
                )
            merged_metadata = dict(existing.equipment_metadata or {})
            merged_metadata["import_source"] = plan.import_source_metadata
            update_payload = dict(plan.update_data)
            update_payload["equipment_metadata"] = merged_metadata
            await equipment_crud.update(db, existing, data=update_payload)
            plan.result.equipment_id = str(existing.id)

        summary = _summarize(filename, plans)
        audit_log = await record_audit_event(
            db,
            actor_user_id=actor_user_id,
            action=AUDIT_ACTION_IMPORT,
            entity_type=AUDIT_ENTITY_EQUIPMENT,
            entity_id=None,
            after={
                "filename": filename,
                "total_rows": summary.total_rows,
                "succeeded": summary.succeeded,
                "failed": summary.failed,
                "skipped": summary.skipped,
                "update_existing": update_existing,
            },
            request=request,
        )
        await db.commit()
    except Exception as exc:
        await db.rollback()
        logger.exception("Inventory import commit failed unexpectedly; batch rolled back (filename=%s)", filename)
        if isinstance(exc, ImportCommitFailedError):
            raise
        raise ImportCommitFailedError(
            "Import commit failed unexpectedly due to a system error. No rows were saved; please retry."
        ) from exc

    summary.audit_log_id = str(audit_log.id)
    return summary


async def process_import(
    db: AsyncSession,
    *,
    file: UploadFile,
    update_existing: bool,
    commit: bool,
    actor_user_id: uuid.UUID | None = None,
    request: Request | None = None,
) -> ImportBatchResult:
    """Single entry point for both preview (`commit=False`) and commit
    (`commit=True`). Always re-parses and re-validates the uploaded file
    from scratch -- commit never trusts a prior preview result, so a
    stale or tampered client-supplied preview payload cannot influence
    what actually gets written.
    """
    # Review finding PR12-H1R / Repository Owner decision: Roadmap PR12 is
    # update-only. `update_existing` is kept in the request shape for API
    # compatibility, but `false` no longer has a coherent meaning -- there
    # is no create path for it to select instead, so a request that sends
    # it is rejected clearly and immediately, rather than silently
    # accepted and processed into a batch where every row can only ever
    # fail or (formerly) skip.
    if not update_existing:
        raise InvalidInputError(
            "Roadmap PR12 Inventory Import is update-only; update_existing must be true. "
            "Import never creates new equipment -- create the equipment through the "
            "standard Equipment Master workflow first, then re-import this file to update it."
        )

    filename = _validate_filename(file.filename)
    content = await _read_upload_bounded(file, max_bytes=MAX_UPLOAD_BYTES)
    # Review finding PR12-H2: CPU-bound XLSX parsing runs in a worker
    # thread, never blocking the event loop this coroutine shares with
    # every other concurrent request.
    parsed_rows = await asyncio.to_thread(_parse_workbook_sync, content)
    plans = await _validate_rows(db, parsed_rows)

    if not commit:
        return _summarize(filename, plans)

    return await _commit_rows(
        db,
        filename=filename,
        plans=plans,
        update_existing=update_existing,
        actor_user_id=actor_user_id,
        request=request,
    )

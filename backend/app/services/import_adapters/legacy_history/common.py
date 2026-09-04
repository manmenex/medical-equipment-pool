"""Roadmap PR21B. Shared, dataset-agnostic parsing/resolution primitives
-- safe workbook access, timestamp normalization, Ward resolution,
blank-row detection, finding construction (§30's "good shared
candidates" list). No ISSUE/RECEIVE-specific logic lives here (§31)."""

from __future__ import annotations

import io
import uuid
from datetime import date, datetime, time, timezone
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import InvalidInputError
from app.models.legacy_history import LegacyWardAlias
from app.models.master_data import Ward
# §32: reuse the repository's own established workbook-security hardening
# rather than re-implement a second copy of it (the same discipline
# `identifiers.normalize_bcm_code` already enforces for BCM
# normalization) -- these three checks are what stand between this
# package and a hostile .xlsx (macro-enabled OOXML structure, embedded
# VBA parts, and zip-bomb/entry-count bounds), unchanged from Roadmap
# PR20C/PR93-H3R.
from app.services.import_adapters.equipment_master import (
    _reject_macro_ooxml_structure,
    _reject_macro_parts,
)
from app.services.import_service import _validate_zip_archive_bounds

# PR21 fix P1: the generic `import_service.MAX_WORKSHEET_COUNT` (25) is
# tuned for the older/generic import flows (e.g. Equipment Master),
# whose source topology is a single working sheet plus a handful of
# incidental ones. The Owner-approved PR21 legacy-history workbook is
# independently verified to contain 28 worksheets (manifest SHA-256
# `8657cfc6c23036c64ea601dcc64c2b2e9d4fc5b51321534098d7a9ff1d84b00c`),
# so reusing the generic cap rejects the real authoritative artifact
# before its two canonical sheets can even be selected. This package
# therefore defines its own bounded allowance instead of raising the
# generic constant (which would silently relax the unrelated import
# datasets that still rely on it) -- 32 gives a narrow four-sheet
# tolerance above the verified 28-sheet shape (room for a harmless
# helper/report sheet without recoupling the cap to today's exact
# count) while still rejecting pathological worksheet-count abuse. The
# zip-bomb/macro/archive-bounds hardening below is unaffected --
# dataset-specific policy applies only to worksheet *count*, never to
# the security bounds themselves.
PR21_MAX_WORKSHEET_COUNT = 32

# Roadmap PR21D1 / Owner Decision Closure Round 3 (GitHub PR #106, design
# doc §55.7): the generic `import_service.MAX_UPLOAD_BYTES` (10 MiB) is
# tuned for the older/generic import flows and rejects the Owner-approved
# PR21 workbook (~19.7 MiB, 20,690,045 bytes) before its bytes ever reach
# this package's own loader. Mirrors `PR21_MAX_WORKSHEET_COUNT`'s own
# precedent exactly: a narrow, dataset-specific allowance instead of
# raising the generic constant (which would silently relax the unrelated
# Equipment Master upload path) -- 32 MiB gives bounded headroom above
# the verified ~19.7 MiB artifact. This constant is consulted only by the
# upload endpoint (`app.api.v1.import_sessions.upload_source`), selected
# by `ImportSession.dataset_type` *before* any byte is read off the wire
# (never a client-supplied value) -- it does not, by itself, relax any
# zip-bomb/decompression/entry-count bound below, all of which remain
# unconditional regardless of which cap admitted the compressed upload.
PR21_MAX_UPLOAD_BYTES = 32 * 1024 * 1024

# PR21D1 fix (P1 review, GitHub PR #107): the framework's own generic
# `import_adapter.MAX_IMPORT_ROWS` (5,000) is tuned for the older/generic
# single-sheet import flows and rejects the combined Issue+Receive record
# count `LegacyTransactionHistoryAdapter.parse()` produces for the
# Owner-approved PR21 workbook before validation can even run. The exact,
# evidence-derived combined non-blank data-row count for the four
# canonical sheets, per `docs/evidence/pr21/equipment-pool-workbook-
# manifest.json` (workbook SHA-256
# `8657cfc6c23036c64ea601dcc64c2b2e9d4fc5b51321534098d7a9ff1d84b00c`,
# `non_empty_row_count - 1` per sheet, i.e. excluding the header row,
# which is exactly what `row_is_blank()`-filtered parsing produces):
#   Orders ยืมเครื่อง (Issue header):      5,685
#   ข้อมูลส่งเครื่องมือ (Issue line):        19,871
#   Orders คืนเครื่อง (Receive header):     6,158
#   ข้อมูลรับเครื่องมือ (Receive line):       19,750
#   combined total:                        51,464
# No merged governance decision (PR #106/§55, DECISION_LOG) names an
# exact numeric row-admission bound for this dataset, so -- mirroring
# `PR21_MAX_WORKSHEET_COUNT`/`PR21_MAX_UPLOAD_BYTES`'s own precedent of a
# narrow, evidence-justified, bounded operational allowance rather than
# an unbounded or arbitrarily large one -- this sets a bound with
# deliberately modest headroom above the verified 51,464 combined count
# (roughly the same ~15-20% margin `PR21_MAX_WORKSHEET_COUNT` uses above
# its own verified 28-sheet count), never `sys.maxsize` or an
# unjustified round number. Consulted only via
# `LegacyTransactionHistoryAdapter.max_import_rows` -- the framework's
# own `MAX_IMPORT_ROWS` constant and every other adapter's default
# admission bound are completely unaffected.
PR21_MAX_IMPORT_RECORDS = 60_000

# §22/§10.1: the workbook's `วันที่`/`เวลา` cells carry no embedded
# timezone (openpyxl returns naive `datetime`/`time` values). Asia/
# Bangkok is this repository's own established business timezone for
# every existing date/shift computation (`app.core.reporting_time`'s
# `_BANGKOK_TZ`, DST-free UTC+7) -- reused here for the identical reason
# reporting_time.py already commits to it, not re-derived.
BANGKOK_TZ = ZoneInfo("Asia/Bangkok")


def load_workbook_bytes(content: bytes) -> Workbook:
    """§21/§32: validates archive bounds and rejects macro-enabled OOXML
    structure/parts before ever handing the bytes to `openpyxl`, exactly
    like `equipment_master._load_sheet1` -- this function returns the
    whole `Workbook` (not one pre-selected sheet), since this package
    must read two named sheets, not one."""
    _validate_zip_archive_bounds(content)
    _reject_macro_ooxml_structure(content)
    _reject_macro_parts(content)
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise InvalidInputError(
            "The uploaded file could not be read as a valid Excel (.xlsx) spreadsheet."
        ) from exc
    if len(workbook.sheetnames) > PR21_MAX_WORKSHEET_COUNT:
        raise InvalidInputError(
            f"The uploaded file contains more than {PR21_MAX_WORKSHEET_COUNT} worksheets, "
            "exceeding the PR21 legacy history import's bounded worksheet-count allowance."
        )
    return workbook


def require_sheet(workbook: Workbook, sheet_name: str):
    """§4: a required canonical sheet must be present, selected by exact
    name -- never `workbook.active`, never a positional guess. Any other
    sheet present (SDC sheets, report/print sheets, Equipment Master,
    etc.) is simply never looked at -- their mere presence in the same
    workbook is not itself an error (§4: "must NOT cause this bounded
    parser to silently ingest them", not "must reject a workbook that
    also contains them")."""
    if sheet_name not in workbook.sheetnames:
        raise InvalidInputError(
            f"The uploaded file has no worksheet named '{sheet_name}'. "
            "Canonical Issue import requires both required worksheets to be present by name."
        )
    return workbook[sheet_name]


def require_governed_headers(worksheet, governed_headers: dict[str, str | None], *, max_header_columns: int) -> dict[str, int]:
    """Closed-world header validation, mirrored from
    `equipment_master._validate_headers` and generalized for reuse by
    any canonical sheet (this slice's two, and a future PR21C Receive
    sheet): header is exactly row 1, column order is not authoritative
    (exact-name based). Every governed header must be present exactly
    once; missing/duplicate/unknown-extra are each a structural `ERROR`
    -- fail-closed on schema drift, never a partial/best-effort read."""
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise InvalidInputError("The uploaded worksheet is empty; no header row was found.")
    if len(header_row) > max_header_columns:
        raise InvalidInputError(
            f"The uploaded worksheet's header row contains more than {max_header_columns} columns, "
            "exceeding what a normal legacy history export contains."
        )

    seen: dict[str, list[int]] = {}
    for idx, raw in enumerate(header_row):
        text = cell_text(raw)
        if text is None:
            continue
        seen.setdefault(text, []).append(idx)

    missing = [header for header in governed_headers if header not in seen]
    if missing:
        raise InvalidInputError("The uploaded worksheet is missing required column(s): " + ", ".join(missing))
    duplicated = [header for header, idxs in seen.items() if header in governed_headers and len(idxs) > 1]
    if duplicated:
        raise InvalidInputError("The uploaded worksheet has duplicate column(s): " + ", ".join(sorted(duplicated)))
    unknown = sorted(header for header in seen if header not in governed_headers)
    if unknown:
        raise InvalidInputError(
            "The uploaded worksheet contains unrecognized column(s) not part of the approved "
            "canonical schema: " + ", ".join(unknown)
        )

    return {header: seen[header][0] for header in governed_headers}


def cell_text(value: object) -> str | None:
    """Trim-only extraction (§7/§8's general cell-normalization rule,
    mirrored from `equipment_master._cell_text`): an empty string
    normalizes to `None`. Not applied to `ME.Code`/`ลำดับ`, which have
    their own stricter cell-acceptance rules."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip() or None


def row_is_blank(values: tuple) -> bool:
    """§28: a purely formatted/trailing blank row (every cell empty once
    trimmed) is ignored entirely -- never counted as a data row, never
    the subject of a finding. A partially populated row is NOT blank and
    proceeds to normal validation, where a missing required value
    becomes its own finding."""
    return all(cell_text(v) is None for v in values)


def combine_bangkok_datetime(date_value: object, time_value: object) -> datetime | None:
    """§22: combines the workbook's separate `วันที่` (date) and `เวลา`
    (time) cells into one aware-UTC `datetime`. Returns `None` for any
    input that cannot be deterministically combined -- malformed/missing
    date or time is always the caller's `ERROR` to report, never
    silently defaulted (§18/§28: "no synthetic defaults", "do not
    fabricate missing date or time"). Accepts `datetime.datetime` (a
    date-only cell, mirroring openpyxl's own representation),
    `datetime.date`, and `datetime.time`/`datetime.datetime` (for the
    time cell) -- any other cell type is treated as malformed."""
    if isinstance(date_value, datetime):
        the_date = date_value.date()
    elif isinstance(date_value, date):
        the_date = date_value
    else:
        return None

    if isinstance(time_value, datetime):
        the_time = time_value.time()
    elif isinstance(time_value, time):
        the_time = time_value
    else:
        return None

    local_naive = datetime.combine(the_date, the_time)
    local_aware = local_naive.replace(tzinfo=BANGKOK_TZ)
    return local_aware.astimezone(timezone.utc)


class WardResolution:
    """Outcome of resolving one raw legacy Ward text value (OD-PR21-4,
    §16). `outcome` is exactly one of `"resolved"`/`"unknown"`/
    `"ambiguous"` -- `ward_id` is set only for `"resolved"`."""

    __slots__ = ("outcome", "ward_id")

    def __init__(self, outcome: str, ward_id: uuid.UUID | None = None) -> None:
        self.outcome = outcome
        self.ward_id = ward_id


def resolve_ward(
    raw_text: str, *, ward_by_exact: dict[str, set[uuid.UUID]], ward_alias_by_text: dict[str, uuid.UUID]
) -> WardResolution:
    """OD-PR21-4's two allowed deterministic paths, in order: (A) exact
    canonical Ward match (`Ward.code` or `Ward.name`, case-sensitive --
    "no case-insensitive guessing unless existing canonical Ward
    normalization explicitly defines it", and none does), then (B) an
    explicit persisted `LegacyWardAlias` mapping. No fuzzy/contains/
    case-insensitive fallback exists at any step. If the exact-match
    lookup itself is ambiguous (the same raw text simultaneously names
    two different Wards -- e.g. one Ward's `code` collides with a
    different Ward's `name`), that is `"ambiguous"`, never an arbitrary
    pick -- checked and reported *before* ever consulting the alias
    table, so an ambiguous exact match is never silently masked by a
    coincidentally-present alias row."""
    exact_ids = ward_by_exact.get(raw_text)
    if exact_ids:
        if len(exact_ids) > 1:
            return WardResolution("ambiguous")
        return WardResolution("resolved", next(iter(exact_ids)))
    alias_id = ward_alias_by_text.get(raw_text)
    if alias_id is not None:
        return WardResolution("resolved", alias_id)
    return WardResolution("unknown")


async def load_ward_exact_lookup(db: AsyncSession, texts: set[str]) -> dict[str, set[uuid.UUID]]:
    """Bulk exact-match lookup against `Ward.code` and `Ward.name`
    together (§16: OD-PR21-4 does not specify which canonical column a
    raw legacy Ward string matches against, so both are checked).
    Returns every distinct `Ward.id` a raw text value matched, whether
    one (resolvable) or more than one (ambiguous, per `resolve_ward`'s
    own docstring) -- never collapsed here, so ambiguity is never lost
    before the caller can classify it."""
    if not texts:
        return {}
    rows = (await db.execute(select(Ward.id, Ward.code, Ward.name).where((Ward.code.in_(texts)) | (Ward.name.in_(texts))))).all()
    candidates: dict[str, set[uuid.UUID]] = {}
    for ward_id, code, name in rows:
        if code in texts:
            candidates.setdefault(code, set()).add(ward_id)
        if name in texts:
            candidates.setdefault(name, set()).add(ward_id)
    return candidates


async def load_ward_alias_lookup(db: AsyncSession, texts: set[str]) -> dict[str, uuid.UUID]:
    """Bulk exact-match lookup against `LegacyWardAlias.raw_alias`
    (§16). `raw_alias` is already database-unique
    (`uq_legacy_ward_aliases_raw_alias`), so no ambiguity handling is
    needed here."""
    if not texts:
        return {}
    rows = (
        await db.execute(select(LegacyWardAlias.raw_alias, LegacyWardAlias.ward_id).where(LegacyWardAlias.raw_alias.in_(texts)))
    ).all()
    return {raw_alias: ward_id for raw_alias, ward_id in rows}

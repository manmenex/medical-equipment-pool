"""Roadmap PR21B -- Canonical Issue Parser + Validation (bounded slice).

Parses and validates exactly the two approved canonical Issue sheets --
`Orders ยืมเครื่อง` (order header, one row per order) and
`ข้อมูลส่งเครื่องมือ` (equipment line item, one row per historical
ISSUE event) -- into typed `LegacyIssueCandidate` objects plus
`LegacyIssueFinding` validation findings (§1/§9). This module never
inserts a `LegacyEquipmentEvent` row, never touches Receive sheets, and
never parses/classifies the SDC sheets (`ข้อมูลการส่ง SDC`/
`ข้อมูลการรับ SDC`) -- their field-contract ambiguity (§6.1) remains
open and is not decided here (§4/§54).

**Not a registered `ImportAdapter`** -- see this package's own
`__init__.py` docstring for why, and how that keeps a real PR21
`ImportSession` structurally unable to reach `validated` from this
module alone."""

from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import InvalidInputError
from app.models.equipment import Equipment
from app.services.identifiers import normalize_bcm_code
from app.services.import_adapter import RawImportRecord
from app.services.import_adapters.legacy_history.common import (
    cell_text,
    combine_bangkok_datetime,
    load_ward_alias_lookup,
    load_ward_exact_lookup,
    load_workbook_bytes,
    require_governed_headers,
    require_sheet,
    resolve_ward,
    row_is_blank,
)
from app.services.import_adapters.legacy_history.types import LegacyIssueCandidate, LegacyIssueFinding, SourceCoordinate
from app.services.import_service import MAX_HEADER_COLUMNS

# ---------------------------------------------------------------------------
# §4/§6.1: the two approved canonical Issue sheets. A closed-world pair --
# this module never reads any other worksheet, regardless of what else the
# workbook contains (SDC sheets, report/print views, Equipment Master,
# helper sheets).
# ---------------------------------------------------------------------------
HEADER_SHEET_NAME = "Orders ยืมเครื่อง"
LINE_SHEET_NAME = "ข้อมูลส่งเครื่องมือ"
EVENT_TYPE = "ISSUE"

# §7: `Orders ยืมเครื่อง`'s 8 governed columns. `None` marks a column that
# is governed (its header presence is still required, closed-world) but
# never extracted into a typed field -- `หมายเหตุ` (never extracted,
# OD-PR21-6) and the two columns §22/§23 classify as ignored
# (`ผู้ส่งเครื่องยืม (User)`: provenance-only free text, never promoted to
# a typed field this slice reads; `จำนวนเครื่องส่งทั้งหมด`: a count,
# never promoted into schema).
_HEADER_GOVERNED_HEADERS: dict[str, str | None] = {
    "วันที่": "date",
    "เลขที่ใบยืม": "order_ref",
    "แผนกที่ยืม": "ward_text",
    "ผู้ส่งเครื่องยืม (User)": None,
    "ผู้ส่งเครื่องยืม (BME)": "bme_name",
    "เวลา": "time",
    "จำนวนเครื่องส่งทั้งหมด": None,
    "หมายเหตุ": None,
}

# §8: `ข้อมูลส่งเครื่องมือ`'s 20 governed columns. `SCAN CODE ส่ง`/
# `Barcode ส่งเครื่อง` are governed (header presence required) but never
# extracted -- §14/§10.1: no safe deterministic extractor for either
# exists anywhere in this repository yet, so no cross-check against
# `ME.Code` is implemented (`ME.Code` remains solely governing; see
# `_validate_equipment` below for the explicit non-implementation note).
# `Equipment`/`Brand`/`Model`/`Serial no.` are provenance-only (§22, not
# equipment identity). `รูปเครื่อง`/`จำนวน`/the three checklist columns
# are ignored (§23). `หมายเหตุ` is never extracted (OD-PR21-6, §21).
# `ชื่อ (User)` is provenance-only free text, never promoted (§20).
_LINE_GOVERNED_HEADERS: dict[str, str | None] = {
    "ลำดับ": "row_key",
    "วันที่": "date",
    "เลขที่ใบส่ง": "order_ref",
    "SCAN CODE ส่ง": None,
    "ME.Code": "me_code",
    "Barcode ส่งเครื่อง": None,
    "Equipment": None,
    "Brand": None,
    "Model": None,
    "Serial no.": None,
    "รูปเครื่อง": None,
    "แผนกที่ส่ง": "ward_text",
    "ตัวเครื่องหน้าจอไม่แตกร้าว": None,
    "สายชาร์จ (สายไม่ขาด)": None,
    "Pole Clamp (ไม่แตกหักมไม่ง้อ)": None,
    "จำนวน": None,
    "หมายเหตุ": None,
    "เวลา": "time",
    "ชื่อ BME": "bme_name",
    "ชื่อ (User)": None,
}

# ---------------------------------------------------------------------------
# §25: centralized, stable finding codes for this module. Every one of
# §25's listed categories that genuinely applies to the Issue side is
# represented below; "unknown BME name" is deliberately absent -- see
# `_validate_bme` for why no code is emitted for it in this slice.
# ---------------------------------------------------------------------------
CODE_HEADER_ORDER_REFERENCE_MISSING = "LEGACY_ISSUE_HEADER_ORDER_REFERENCE_MISSING"
CODE_HEADER_ORDER_REFERENCE_DUPLICATE = "LEGACY_ISSUE_HEADER_ORDER_REFERENCE_DUPLICATE"
CODE_ORDER_REFERENCE_MISSING = "LEGACY_ISSUE_ORDER_REFERENCE_MISSING"
CODE_ORPHAN_ORDER_REFERENCE = "LEGACY_ISSUE_ORPHAN_ORDER_REFERENCE"
CODE_AMBIGUOUS_ORDER_REFERENCE = "LEGACY_ISSUE_AMBIGUOUS_ORDER_REFERENCE"
CODE_SOURCE_ROW_KEY_MISSING = "LEGACY_ISSUE_SOURCE_ROW_KEY_MISSING"
CODE_SOURCE_ROW_KEY_DUPLICATE = "LEGACY_ISSUE_SOURCE_ROW_KEY_DUPLICATE"
CODE_ME_CODE_MISSING = "LEGACY_ISSUE_ME_CODE_MISSING"
CODE_ME_CODE_INVALID = "LEGACY_ISSUE_ME_CODE_INVALID"
CODE_EQUIPMENT_NOT_FOUND = "LEGACY_ISSUE_EQUIPMENT_NOT_FOUND"
CODE_INVALID_TIMESTAMP = "LEGACY_ISSUE_INVALID_TIMESTAMP"
CODE_HEADER_LINE_TIMESTAMP_CONFLICT = "LEGACY_ISSUE_HEADER_LINE_TIMESTAMP_CONFLICT"
CODE_WARD_NOT_FOUND = "LEGACY_ISSUE_WARD_NOT_FOUND"
CODE_WARD_AMBIGUOUS = "LEGACY_ISSUE_WARD_AMBIGUOUS"
CODE_HEADER_LINE_WARD_CONFLICT = "LEGACY_ISSUE_HEADER_LINE_WARD_CONFLICT"


def _extract_me_code_cell(value: object) -> tuple[str, str | None]:
    """`ME.Code` acceptance rule mirrors `equipment_master`'s BCM cell
    rule exactly (§13: normalized via the same `normalize_bcm_code()`,
    which itself expects a string) -- text-only; any numeric-typed cell
    is a blocking `ERROR`, never silently coerced. Returns
    `(outcome, value)` where `outcome` is `"blank"`/`"valid"`/
    `"invalid"`."""
    if value is None:
        return "blank", None
    if isinstance(value, str):
        text = value.strip()
        return ("blank", None) if not text else ("valid", text)
    return "invalid", None


def _extract_row_key_cell(value: object) -> tuple[str, str | None]:
    """`ลำดับ` acceptance rule: text or a losslessly-convertible
    integer-numeric cell (mirrors `equipment_master`'s Item No. cell
    rule, §10 -- Excel commonly stores a sequence number as a numeric
    cell). Preserved as its original string form, never re-derived from
    worksheet row position (§10: "do not use worksheet row number as
    logical event identity")."""
    if value is None:
        return "blank", None
    if isinstance(value, str):
        text = value.strip()
        return ("blank", None) if not text else ("valid", text)
    if isinstance(value, bool):
        return "invalid", None
    if isinstance(value, int):
        return "valid", str(value)
    if isinstance(value, float) and value.is_integer():
        return "valid", str(int(value))
    return "invalid", None


def _parse_sheet(worksheet: Worksheet, governed_headers: dict[str, str | None]) -> list[RawImportRecord]:
    """§28: blank/purely-formatted rows are skipped, never counted as
    data. §7-pattern data-smuggling guard: any non-blank cell in a
    column outside the governed set is rejected, not silently ignored."""
    header_index = require_governed_headers(worksheet, governed_headers, max_header_columns=MAX_HEADER_COLUMNS)
    known_indices = set(header_index.values())

    def cell_at(values: tuple, header: str) -> object:
        idx = header_index[header]
        return values[idx] if idx < len(values) else None

    records: list[RawImportRecord] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if values is None or row_is_blank(values):
            continue
        for idx, value in enumerate(values):
            if idx not in known_indices and cell_text(value) is not None:
                raise InvalidInputError(
                    f"Row {row_number} contains data in a column outside the approved canonical "
                    "Issue schema (an unnamed/extra column is not permitted)."
                )
        fields: dict[str, Any] = {}
        for header, role in governed_headers.items():
            if role is None:
                continue
            raw = cell_at(values, header)
            if role in ("date", "time"):
                fields[role] = raw
            elif role == "me_code":
                fields["me_code_cell"] = _extract_me_code_cell(raw)
            elif role == "row_key":
                fields["row_key_cell"] = _extract_row_key_cell(raw)
            else:
                fields[role] = cell_text(raw)
        records.append(RawImportRecord(row_number=row_number, fields=fields))
    return records


def parse_workbook(content: bytes) -> tuple[list[RawImportRecord], list[RawImportRecord]]:
    """§7/§8: synchronous, CPU-bound parse of both required canonical
    sheets into typed records, in source-file row order. Returns
    `(header_records, line_records)`."""
    workbook = load_workbook_bytes(content)
    header_ws = require_sheet(workbook, HEADER_SHEET_NAME)
    line_ws = require_sheet(workbook, LINE_SHEET_NAME)
    header_records = _parse_sheet(header_ws, _HEADER_GOVERNED_HEADERS)
    line_records = _parse_sheet(line_ws, _LINE_GOVERNED_HEADERS)
    return header_records, line_records


def verify_workbook_authority(*, workbook_checksum: str, approved_workbook_sha256: str) -> None:
    """§5: the workbook-authority gate. Raises if the active source's
    own checksum does not match the `LegacyMigrationAuthority`'s
    immutably-approved checksum -- a validated candidate must never be
    produced under an authority/source mismatch. Deliberately a plain,
    synchronous, side-effect-free comparison -- the actual `ImportSource`
    <-> `LegacyMigrationAuthority` *persistence* relationship is PR21A's
    own schema (`LegacyEquipmentEvent.migration_authority_id`); this
    function only enforces that this parser is never invoked to produce
    candidates for a workbook whose bytes do not match the approved
    snapshot, regardless of which orchestration layer calls it."""
    if workbook_checksum != approved_workbook_sha256:
        raise InvalidInputError(
            "The active import source's checksum does not match the approved "
            "LegacyMigrationAuthority workbook snapshot; refusing to parse."
        )


# ---------------------------------------------------------------------------
# §6.3/§13 pattern: exactly one bulk-lookup pass for the whole batch (never
# one query per row).
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class IssueValidationContext:
    equipment_by_me_code: dict[str, Equipment] = field(default_factory=dict)
    ward_by_exact: dict[str, set[uuid.UUID]] = field(default_factory=dict)
    ward_alias_by_text: dict[str, uuid.UUID] = field(default_factory=dict)
    header_by_ref: dict[str, list[int]] = field(default_factory=dict)
    duplicate_row_key_rows: frozenset[int] = frozenset()


def _rows_with_duplicate_values(values_by_row: dict[int, str]) -> frozenset[int]:
    """§12: ALL rows sharing a duplicated value are flagged -- never
    merely the second occurrence, never first/last-wins (mirrors
    `equipment_master._rows_with_duplicate_values` exactly)."""
    seen: dict[str, list[int]] = {}
    for row_number, value in values_by_row.items():
        seen.setdefault(value, []).append(row_number)
    duplicated: set[int] = set()
    for rows in seen.values():
        if len(rows) > 1:
            duplicated.update(rows)
    return frozenset(duplicated)


async def preload_business_context(
    db: AsyncSession, header_records: list[RawImportRecord], line_records: list[RawImportRecord]
) -> IssueValidationContext:
    me_codes: set[str] = set()
    normalized_by_row: dict[int, str] = {}
    for record in line_records:
        outcome, value = record.fields["me_code_cell"]
        if outcome == "valid":
            try:
                normalized = normalize_bcm_code(value)
            except InvalidInputError:
                continue  # re-validated (and reported) in validate_and_build_candidates
            normalized_by_row[record.row_number] = normalized
            me_codes.add(normalized)
    # Stashed back so validate_and_build_candidates never re-derives it.
    for record in line_records:
        record.fields["me_code_normalized"] = normalized_by_row.get(record.row_number)

    row_keys: dict[int, str] = {}
    for record in line_records:
        outcome, value = record.fields["row_key_cell"]
        if outcome == "valid":
            row_keys[record.row_number] = value  # type: ignore[assignment]
    duplicate_row_key_rows = _rows_with_duplicate_values(row_keys)

    equipment_rows = (await db.execute(select(Equipment).where(Equipment.bcm_code.in_(me_codes)))).scalars().all()
    equipment_by_me_code = {e.bcm_code: e for e in equipment_rows if e.bcm_code is not None}

    ward_texts: set[str] = set()
    for record in header_records:
        if record.fields.get("ward_text"):
            ward_texts.add(record.fields["ward_text"])
    for record in line_records:
        if record.fields.get("ward_text"):
            ward_texts.add(record.fields["ward_text"])
    ward_by_exact = await load_ward_exact_lookup(db, ward_texts)
    ward_alias_by_text = await load_ward_alias_lookup(db, ward_texts)

    header_by_ref: dict[str, list[int]] = {}
    for record in header_records:
        ref = record.fields.get("order_ref")
        if ref:
            header_by_ref.setdefault(ref, []).append(record.row_number)

    return IssueValidationContext(
        equipment_by_me_code=equipment_by_me_code,
        ward_by_exact=ward_by_exact,
        ward_alias_by_text=ward_alias_by_text,
        header_by_ref=header_by_ref,
        duplicate_row_key_rows=duplicate_row_key_rows,
    )


def _validate_headers_structural(
    header_records: list[RawImportRecord], header_by_ref: dict[str, list[int]]
) -> list[LegacyIssueFinding]:
    """§27/§29: a header row missing its own `เลขที่ใบยืม`, or one whose
    key collides with another header row's, is a malformed-header
    finding on the header row itself -- independent of whether any line
    ever references it. `header_by_ref` is `IssueValidationContext`'s
    own grouping (already computed once in `preload_business_context`),
    reused here rather than regrouped a second time."""
    findings: list[LegacyIssueFinding] = []
    for record in header_records:
        if not record.fields.get("order_ref"):
            findings.append(
                LegacyIssueFinding(
                    sheet_name=HEADER_SHEET_NAME,
                    source_row_number=record.row_number,
                    field="order_ref",
                    error_code=CODE_HEADER_ORDER_REFERENCE_MISSING,
                    message="เลขที่ใบยืม is required on every order header row.",
                )
            )
    for ref, rows in header_by_ref.items():
        if len(rows) > 1:
            for row_number in rows:
                findings.append(
                    LegacyIssueFinding(
                        sheet_name=HEADER_SHEET_NAME,
                        source_row_number=row_number,
                        field="order_ref",
                        error_code=CODE_HEADER_ORDER_REFERENCE_DUPLICATE,
                        message=f"เลขที่ใบยืม '{ref}' appears on more than one order header row.",
                    )
                )
    return findings


def _validate_equipment(record: RawImportRecord, context: IssueValidationContext) -> tuple[list[LegacyIssueFinding], Equipment | None]:
    """§13/§14: `ME.Code` is the sole governing legacy Equipment
    identifier for the Issue side. `SCAN CODE ส่ง`/`Barcode ส่งเครื่อง`
    are NOT cross-checked against it -- no safe, deterministic extractor
    for either exists anywhere in this repository (§14/§10.1's own
    explicit finding); inventing ad hoc URL/barcode parsing heuristics
    here would violate the "never silently prefer one" rule by
    fabricating a comparison this slice cannot actually verify is
    correct. This is a documented non-implementation, not an oversight."""
    findings: list[LegacyIssueFinding] = []
    outcome, value = record.fields["me_code_cell"]
    if outcome == "blank":
        findings.append(
            LegacyIssueFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=record.row_number,
                field="me_code",
                error_code=CODE_ME_CODE_MISSING,
                message="ME.Code is required.",
            )
        )
        return findings, None
    if outcome == "invalid":
        findings.append(
            LegacyIssueFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=record.row_number,
                field="me_code",
                error_code=CODE_ME_CODE_INVALID,
                message="ME.Code must be a text cell, not a numeric cell.",
            )
        )
        return findings, None

    normalized = record.fields.get("me_code_normalized")
    if normalized is None:
        try:
            normalize_bcm_code(value)  # re-raise for the message
        except InvalidInputError as exc:
            findings.append(
                LegacyIssueFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=record.row_number,
                    field="me_code",
                    error_code=CODE_ME_CODE_INVALID,
                    message=f"Invalid ME.Code: {exc.message}",
                )
            )
        return findings, None

    equipment = context.equipment_by_me_code.get(normalized)
    if equipment is None:
        findings.append(
            LegacyIssueFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=record.row_number,
                field="me_code",
                error_code=CODE_EQUIPMENT_NOT_FOUND,
                message=f"No Equipment record matches ME.Code '{normalized}'.",
            )
        )
        return findings, None
    return findings, equipment


def _validate_ward(
    header: RawImportRecord | None, line: RawImportRecord, context: IssueValidationContext
) -> tuple[list[LegacyIssueFinding], str | None, uuid.UUID | None]:
    """§16/§17: OD-PR21-4's exact-match/alias resolution, plus the
    header/line precedence rule this slice must decide (§9's own flagged
    unresolved item). Returns `(findings, raw_text_used, resolved_ward_id)`
    -- `raw_text_used` is the winning raw text preserved on the
    candidate regardless of which side it came from."""
    findings: list[LegacyIssueFinding] = []
    header_text = header.fields.get("ward_text") if header is not None else None
    line_text = line.fields.get("ward_text")

    def _resolve(text: str, sheet_name: str, row_number: int) -> uuid.UUID | None:
        resolution = resolve_ward(text, ward_by_exact=context.ward_by_exact, ward_alias_by_text=context.ward_alias_by_text)
        if resolution.outcome == "resolved":
            return resolution.ward_id
        code = CODE_WARD_AMBIGUOUS if resolution.outcome == "ambiguous" else CODE_WARD_NOT_FOUND
        findings.append(
            LegacyIssueFinding(
                sheet_name=sheet_name,
                source_row_number=row_number,
                field="ward_text",
                error_code=code,
                message=f"Ward text '{text}' could not be unambiguously resolved.",
            )
        )
        return None

    header_ward_id: uuid.UUID | None = None
    if header_text:
        header_ward_id = _resolve(header_text, HEADER_SHEET_NAME, header.row_number if header is not None else -1)
    line_ward_id: uuid.UUID | None = None
    if line_text:
        line_ward_id = _resolve(line_text, LINE_SHEET_NAME, line.row_number)

    if findings:
        # Either side failed to resolve -- never fall back to the side
        # that did resolve (that would be a silent preference the
        # design's own precedence rule does not authorize).
        return findings, None, None

    if header_text and line_text:
        if header_ward_id == line_ward_id:
            return findings, line_text, line_ward_id
        findings.append(
            LegacyIssueFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=line.row_number,
                field="ward_text",
                error_code=CODE_HEADER_LINE_WARD_CONFLICT,
                message=f"Header Ward '{header_text}' and line Ward '{line_text}' resolve to different Wards.",
            )
        )
        return findings, None, None
    if line_text:
        return findings, line_text, line_ward_id
    if header_text:
        return findings, header_text, header_ward_id
    findings.append(
        LegacyIssueFinding(
            sheet_name=LINE_SHEET_NAME,
            source_row_number=line.row_number,
            field="ward_text",
            error_code=CODE_WARD_NOT_FOUND,
            message="No Ward text is present on either the order header or the line item.",
        )
    )
    return findings, None, None


def _validate_timestamp(header: RawImportRecord | None, line: RawImportRecord) -> tuple[list[LegacyIssueFinding], datetime | None]:
    """§18/§22: `occurred_at` is fundamentally a line-item fact (§7:
    "Events are line-item based") -- the line's own date+time is the
    required, primary source. The header's date+time, when both are
    present and combinable, is cross-checked for consistency only; a
    malformed/incomplete HEADER timestamp does not itself block a row
    whose own line timestamp combines cleanly, since nothing on the
    candidate is ever derived from the header timestamp directly."""
    findings: list[LegacyIssueFinding] = []
    line_dt = combine_bangkok_datetime(line.fields.get("date"), line.fields.get("time"))
    if line_dt is None:
        findings.append(
            LegacyIssueFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=line.row_number,
                field="date_time",
                error_code=CODE_INVALID_TIMESTAMP,
                message="วันที่/เวลา could not be combined into a valid timestamp.",
            )
        )
        return findings, None

    if header is not None:
        header_dt = combine_bangkok_datetime(header.fields.get("date"), header.fields.get("time"))
        if header_dt is not None and header_dt != line_dt:
            findings.append(
                LegacyIssueFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="date_time",
                    error_code=CODE_HEADER_LINE_TIMESTAMP_CONFLICT,
                    message="Header and line date/time do not represent the same event time.",
                )
            )
            return findings, None

    assert isinstance(line_dt, datetime)
    return findings, line_dt


def validate_and_build_candidates(
    header_records: list[RawImportRecord],
    line_records: list[RawImportRecord],
    context: IssueValidationContext,
    *,
    migration_authority_id: uuid.UUID,
    import_session_id: uuid.UUID,
    import_source_id: uuid.UUID,
) -> tuple[list[LegacyIssueCandidate], list[LegacyIssueFinding]]:
    """§6/§9/§11/§24: header-line resolution, identity/duplicate rules,
    Equipment/Ward/timestamp validation, and typed candidate
    construction. Every finding is returned regardless of row outcome
    (mirrors `ImportAdapter.validate_business_rules`'s own contract) --
    the caller decides `validated` vs `validation_failed` for the whole
    batch from the aggregate presence of any `severity == "error"`
    finding (§24's all-or-nothing gate), exactly as
    `import_validation_service.run_validation` already does for every
    other adapter."""
    findings: list[LegacyIssueFinding] = list(_validate_headers_structural(header_records, context.header_by_ref))
    header_by_row_number = {r.row_number: r for r in header_records}

    ambiguous_refs = {ref for ref, rows in context.header_by_ref.items() if len(rows) > 1}
    unambiguous_header_row: dict[str, int] = {
        ref: rows[0] for ref, rows in context.header_by_ref.items() if len(rows) == 1
    }

    candidates: list[LegacyIssueCandidate] = []
    for line in line_records:
        row_findings: list[LegacyIssueFinding] = []

        row_key_outcome, row_key_value = line.fields["row_key_cell"]
        if row_key_outcome != "valid":
            row_findings.append(
                LegacyIssueFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="row_key",
                    error_code=CODE_SOURCE_ROW_KEY_MISSING,
                    message="ลำดับ is required on every data-bearing line-item row.",
                )
            )
        elif line.row_number in context.duplicate_row_key_rows:
            row_findings.append(
                LegacyIssueFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="row_key",
                    error_code=CODE_SOURCE_ROW_KEY_DUPLICATE,
                    message=f"ลำดับ '{row_key_value}' appears on more than one line-item row.",
                )
            )

        order_ref = line.fields.get("order_ref")
        header_record: RawImportRecord | None = None
        if not order_ref:
            row_findings.append(
                LegacyIssueFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="order_ref",
                    error_code=CODE_ORDER_REFERENCE_MISSING,
                    message="เลขที่ใบส่ง is required on every line-item row.",
                )
            )
        elif order_ref in ambiguous_refs:
            row_findings.append(
                LegacyIssueFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="order_ref",
                    error_code=CODE_AMBIGUOUS_ORDER_REFERENCE,
                    message=f"เลขที่ใบส่ง '{order_ref}' matches more than one order header row.",
                )
            )
        elif order_ref not in unambiguous_header_row:
            row_findings.append(
                LegacyIssueFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="order_ref",
                    error_code=CODE_ORPHAN_ORDER_REFERENCE,
                    message=f"เลขที่ใบส่ง '{order_ref}' does not match any order header row.",
                )
            )
        else:
            header_record = header_by_row_number[unambiguous_header_row[order_ref]]

        equipment_findings, equipment = _validate_equipment(line, context)
        row_findings.extend(equipment_findings)

        ward_findings, ward_text, resolved_ward_id = _validate_ward(header_record, line, context)
        row_findings.extend(ward_findings)

        timestamp_findings, occurred_at = _validate_timestamp(header_record, line)
        row_findings.extend(timestamp_findings)

        findings.extend(row_findings)

        has_error = any(f.severity == "error" for f in row_findings)
        if has_error or header_record is None or equipment is None or occurred_at is None:
            continue

        bme_name = line.fields.get("bme_name") or (header_record.fields.get("bme_name") if header_record else None)

        candidates.append(
            LegacyIssueCandidate(
                event_type=EVENT_TYPE,
                legacy_source_row_key=row_key_value,  # type: ignore[arg-type]
                legacy_order_reference=order_ref,
                equipment_id=equipment.id,
                occurred_at=occurred_at,
                legacy_ward_text=ward_text,
                resolved_ward_id=resolved_ward_id,
                legacy_bme_name=bme_name,
                migration_authority_id=migration_authority_id,
                import_session_id=import_session_id,
                import_source_id=import_source_id,
                header_source_ref=SourceCoordinate(sheet_name=HEADER_SHEET_NAME, source_row_number=header_record.row_number),
                line_source_ref=SourceCoordinate(sheet_name=LINE_SHEET_NAME, source_row_number=line.row_number),
            )
        )

    return candidates, findings

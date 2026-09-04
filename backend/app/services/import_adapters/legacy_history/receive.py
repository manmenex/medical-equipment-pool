"""Roadmap PR21C -- Canonical Receive Parser + Validation (bounded slice).

Parses and validates exactly the two approved canonical Receive sheets --
`Orders คืนเครื่อง` (order header, one row per order) and
`ข้อมูลรับเครื่องมือ` (equipment line item, one row per historical
RECEIVE event) -- into typed `LegacyReceiveCandidate` objects plus
`LegacyReceiveFinding` validation findings. This module never inserts a
`LegacyEquipmentEvent` row, never pairs a RECEIVE candidate with any
ISSUE candidate (PR22's own future responsibility), and never
parses/classifies the SDC sheets (`ข้อมูลการรับ SDC`) -- their
field-contract ambiguity remains open and is not decided here.

This module is the Receive-side mirror of `issue.py` (PR21B): same
header/line topology rules, same identifier/Ward/timestamp/privacy
policy, same finding-code style with a `LEGACY_RECEIVE_` prefix instead
of `LEGACY_ISSUE_`. Column names below are taken from the Owner-approved
workbook evidence (`docs/evidence/pr21/equipment-pool-workbook-manifest.json`),
not re-derived or guessed.

**Not a registered `ImportAdapter`** -- see this package's own
`__init__.py` docstring for why, and how that keeps a real PR21
`ImportSession` structurally unable to reach `validated` from this
module (or `issue.py`) alone. Adding this Receive parser does NOT by
itself authorize registering a combined final adapter: the SDC
full-scope Owner Decision remains open, and that registration is a
separate, later, explicitly-gated step."""

from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import InvalidInputError
from app.models.equipment import Equipment
from app.services.identifiers import normalize_bcm_code
from app.services.import_adapter import RawImportRecord
from app.services.import_adapters.legacy_history.common import (
    cell_text,
    combine_bangkok_datetime,
    load_ward_alias_lookup,
    load_ward_exact_lookup,
    load_workbook_bytes,
    require_governed_headers,
    require_sheet,
    resolve_ward,
    row_is_blank,
)
from app.services.import_adapters.legacy_history.types import LegacyReceiveCandidate, LegacyReceiveFinding, SourceCoordinate
from app.services.import_service import MAX_HEADER_COLUMNS

# ---------------------------------------------------------------------------
# The two approved canonical Receive sheets, verified against the approved
# workbook evidence manifest. A closed-world pair -- this module never
# reads any other worksheet, regardless of what else the workbook
# contains (SDC sheets, report/print views, Equipment Master, helper
# sheets, or the Issue-side sheets).
# ---------------------------------------------------------------------------
HEADER_SHEET_NAME = "Orders คืนเครื่อง"
LINE_SHEET_NAME = "ข้อมูลรับเครื่องมือ"
EVENT_TYPE = "RECEIVE"

# `Orders คืนเครื่อง`'s 8 governed columns (evidence-verified header_names,
# equipment-pool-workbook-manifest.json). `None` marks a column that is
# governed (its header presence is still required, closed-world) but
# never extracted into a typed field -- `หมายเหตุ` (never extracted,
# OD-PR21-6) and the two columns classified ignored the same way as the
# Issue header (`ผู้ส่งเครื่องคืน (User)`: provenance-only free text,
# never promoted to a typed field this slice reads;
# `จำนวนเครื่องรับคืนทั้งหมด`: a count, never promoted into schema).
_HEADER_GOVERNED_HEADERS: dict[str, str | None] = {
    "วันที่": "date",
    "เลขที่ใบคืน": "order_ref",
    "แผนกที่คืน": "ward_text",
    "ผู้ส่งเครื่องคืน (User)": None,
    "ผู้รับเครื่องคืน (BME)": "bme_name",
    "เวลา": "time",
    "จำนวนเครื่องรับคืนทั้งหมด": None,
    "หมายเหตุ": None,
}

# `ข้อมูลรับเครื่องมือ`'s 20 governed columns (evidence-verified
# header_names, same manifest). `SCAN CODE รับ`/`Barcode รับเครื่อง` are
# governed (header presence required) but never extracted -- same policy
# as the Issue side's `SCAN CODE ส่ง`/`Barcode ส่งเครื่อง`: no safe
# deterministic extractor for either exists anywhere in this repository,
# so no cross-check against `ME.Code` is implemented here either
# (`ME.Code` remains solely governing on both sides -- see
# `_validate_equipment` below for the explicit non-implementation note,
# kept symmetric with Issue per this slice's own instruction not to
# create inconsistent validation rules between sides).
# `Equipment`/`Brand`/`Model`/`Serial no.` are provenance-only (not
# equipment identity). `รูปเครื่อง`/`จำนวน`/the three checklist columns
# are ignored. `หมายเหตุ` is never extracted (OD-PR21-6). `ชื่อ (User)`
# is provenance-only free text, never promoted.
_LINE_GOVERNED_HEADERS: dict[str, str | None] = {
    "ลำดับ": "row_key",
    "วันที่": "date",
    "เลขที่ใบรับเครื่อง": "order_ref",
    "SCAN CODE รับ": None,
    "ME.Code": "me_code",
    "Barcode รับเครื่อง": None,
    "Equipment": None,
    "Brand": None,
    "Model": None,
    "Serial no.": None,
    "แผนกที่รับ": "ward_text",
    "รูปเครื่อง": None,
    "ตัวเครื่องหน้าจอไม่แตกร้าว": None,
    "สายชาร์จ (สายไม่ขาด)": None,
    "Pole Clamp (ไม่แตกหักมไม่ง้อ)": None,
    "จำนวน": None,
    "หมายเหตุ": None,
    "เวลา": "time",
    "ชื่อ BME": "bme_name",
    "ชื่อ (User)": None,
}

# ---------------------------------------------------------------------------
# Centralized, stable finding codes for this module -- symmetric with
# `issue.py`'s own `LEGACY_ISSUE_*` codes, prefixed `LEGACY_RECEIVE_`
# instead. "Unknown BME name" is deliberately absent, same reason as
# Issue: no BME roster data source exists within these two canonical
# sheets.
# ---------------------------------------------------------------------------
CODE_HEADER_ORDER_REFERENCE_MISSING = "LEGACY_RECEIVE_HEADER_ORDER_REFERENCE_MISSING"
CODE_HEADER_ORDER_REFERENCE_DUPLICATE = "LEGACY_RECEIVE_HEADER_ORDER_REFERENCE_DUPLICATE"
CODE_ORDER_REFERENCE_MISSING = "LEGACY_RECEIVE_ORDER_REFERENCE_MISSING"
CODE_ORPHAN_ORDER_REFERENCE = "LEGACY_RECEIVE_ORPHAN_ORDER_REFERENCE"
CODE_AMBIGUOUS_ORDER_REFERENCE = "LEGACY_RECEIVE_AMBIGUOUS_ORDER_REFERENCE"
CODE_SOURCE_ROW_KEY_MISSING = "LEGACY_RECEIVE_SOURCE_ROW_KEY_MISSING"
CODE_SOURCE_ROW_KEY_DUPLICATE = "LEGACY_RECEIVE_SOURCE_ROW_KEY_DUPLICATE"
CODE_ME_CODE_MISSING = "LEGACY_RECEIVE_ME_CODE_MISSING"
CODE_ME_CODE_INVALID = "LEGACY_RECEIVE_ME_CODE_INVALID"
CODE_EQUIPMENT_NOT_FOUND = "LEGACY_RECEIVE_EQUIPMENT_NOT_FOUND"
CODE_INVALID_TIMESTAMP = "LEGACY_RECEIVE_INVALID_TIMESTAMP"
CODE_HEADER_LINE_TIMESTAMP_CONFLICT = "LEGACY_RECEIVE_HEADER_LINE_TIMESTAMP_CONFLICT"
CODE_WARD_NOT_FOUND = "LEGACY_RECEIVE_WARD_NOT_FOUND"
CODE_WARD_AMBIGUOUS = "LEGACY_RECEIVE_WARD_AMBIGUOUS"
CODE_HEADER_LINE_WARD_CONFLICT = "LEGACY_RECEIVE_HEADER_LINE_WARD_CONFLICT"


def _extract_me_code_cell(value: object) -> tuple[str, str | None]:
    """`ME.Code` acceptance rule, identical to `issue._extract_me_code_cell`
    -- text-only; any numeric-typed cell is a blocking `ERROR`, never
    silently coerced. Returns `(outcome, value)` where `outcome` is
    `"blank"`/`"valid"`/`"invalid"`."""
    if value is None:
        return "blank", None
    if isinstance(value, str):
        text = value.strip()
        return ("blank", None) if not text else ("valid", text)
    return "invalid", None


def _extract_row_key_cell(value: object) -> tuple[str, str | None]:
    """`ลำดับ` acceptance rule, identical to `issue._extract_row_key_cell`
    -- text or a losslessly-convertible integer-numeric cell. Preserved
    as its original string form, never re-derived from worksheet row
    position."""
    if value is None:
        return "blank", None
    if isinstance(value, str):
        text = value.strip()
        return ("blank", None) if not text else ("valid", text)
    if isinstance(value, bool):
        return "invalid", None
    if isinstance(value, int):
        return "valid", str(value)
    if isinstance(value, float) and value.is_integer():
        return "valid", str(int(value))
    return "invalid", None


def _parse_sheet(worksheet: Worksheet, governed_headers: dict[str, str | None]) -> list[RawImportRecord]:
    """Blank/purely-formatted rows are skipped, never counted as data.
    Data-smuggling guard: any non-blank cell in a column outside the
    governed set is rejected, not silently ignored -- identical policy
    to `issue._parse_sheet`."""
    header_index = require_governed_headers(worksheet, governed_headers, max_header_columns=MAX_HEADER_COLUMNS)
    known_indices = set(header_index.values())

    def cell_at(values: tuple, header: str) -> object:
        idx = header_index[header]
        return values[idx] if idx < len(values) else None

    records: list[RawImportRecord] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if values is None or row_is_blank(values):
            continue
        for idx, value in enumerate(values):
            if idx not in known_indices and cell_text(value) is not None:
                raise InvalidInputError(
                    f"Row {row_number} contains data in a column outside the approved canonical "
                    "Receive schema (an unnamed/extra column is not permitted)."
                )
        fields: dict[str, Any] = {}
        for header, role in governed_headers.items():
            if role is None:
                continue
            raw = cell_at(values, header)
            if role in ("date", "time"):
                fields[role] = raw
            elif role == "me_code":
                fields["me_code_cell"] = _extract_me_code_cell(raw)
            elif role == "row_key":
                fields["row_key_cell"] = _extract_row_key_cell(raw)
            else:
                fields[role] = cell_text(raw)
        records.append(RawImportRecord(row_number=row_number, fields=fields))
    return records


def parse_workbook(content: bytes) -> tuple[list[RawImportRecord], list[RawImportRecord]]:
    """Synchronous, CPU-bound parse of both required canonical Receive
    sheets into typed records, in source-file row order. Returns
    `(header_records, line_records)`. Reuses `common.load_workbook_bytes`
    (and its `PR21_MAX_WORKSHEET_COUNT` allowance) unchanged -- no
    Receive-specific worksheet-count policy exists or is needed."""
    workbook = load_workbook_bytes(content)
    header_ws = require_sheet(workbook, HEADER_SHEET_NAME)
    line_ws = require_sheet(workbook, LINE_SHEET_NAME)
    header_records = _parse_sheet(header_ws, _HEADER_GOVERNED_HEADERS)
    line_records = _parse_sheet(line_ws, _LINE_GOVERNED_HEADERS)
    return header_records, line_records


def verify_workbook_authority(*, workbook_checksum: str, approved_workbook_sha256: str) -> None:
    """The workbook-authority gate, identical to `issue.verify_workbook_authority`.
    Raises if the active source's own checksum does not match the
    `LegacyMigrationAuthority`'s immutably-approved checksum -- a
    validated candidate must never be produced under an
    authority/source mismatch. There is no checksum-based bypass of any
    other structural check (worksheet-count cap, security hardening) --
    this function only enforces the authority match itself."""
    if workbook_checksum != approved_workbook_sha256:
        raise InvalidInputError(
            "The active import source's checksum does not match the approved "
            "LegacyMigrationAuthority workbook snapshot; refusing to parse."
        )


# ---------------------------------------------------------------------------
# Exactly one bulk-lookup pass for the whole batch (never one query per
# row) -- identical discipline to `issue.preload_business_context`.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ReceiveValidationContext:
    equipment_by_me_code: dict[str, Equipment] = field(default_factory=dict)
    ward_by_exact: dict[str, set[uuid.UUID]] = field(default_factory=dict)
    ward_alias_by_text: dict[str, uuid.UUID] = field(default_factory=dict)
    header_by_ref: dict[str, list[int]] = field(default_factory=dict)
    duplicate_row_key_rows: frozenset[int] = frozenset()


def _rows_with_duplicate_values(values_by_row: dict[int, str]) -> frozenset[int]:
    """ALL rows sharing a duplicated value are flagged -- never merely
    the second occurrence, never first/last-wins (identical to
    `issue._rows_with_duplicate_values`)."""
    seen: dict[str, list[int]] = {}
    for row_number, value in values_by_row.items():
        seen.setdefault(value, []).append(row_number)
    duplicated: set[int] = set()
    for rows in seen.values():
        if len(rows) > 1:
            duplicated.update(rows)
    return frozenset(duplicated)


async def preload_business_context(
    db: AsyncSession, header_records: list[RawImportRecord], line_records: list[RawImportRecord]
) -> ReceiveValidationContext:
    me_codes: set[str] = set()
    normalized_by_row: dict[int, str] = {}
    for record in line_records:
        outcome, value = record.fields["me_code_cell"]
        if outcome == "valid":
            try:
                normalized = normalize_bcm_code(value)
            except InvalidInputError:
                continue  # re-validated (and reported) in validate_and_build_candidates
            normalized_by_row[record.row_number] = normalized
            me_codes.add(normalized)
    # Stashed back so validate_and_build_candidates never re-derives it.
    for record in line_records:
        record.fields["me_code_normalized"] = normalized_by_row.get(record.row_number)

    row_keys: dict[int, str] = {}
    for record in line_records:
        outcome, value = record.fields["row_key_cell"]
        if outcome == "valid":
            row_keys[record.row_number] = value  # type: ignore[assignment]
    duplicate_row_key_rows = _rows_with_duplicate_values(row_keys)

    equipment_rows = (await db.execute(select(Equipment).where(Equipment.bcm_code.in_(me_codes)))).scalars().all()
    equipment_by_me_code = {e.bcm_code: e for e in equipment_rows if e.bcm_code is not None}

    ward_texts: set[str] = set()
    for record in header_records:
        if record.fields.get("ward_text"):
            ward_texts.add(record.fields["ward_text"])
    for record in line_records:
        if record.fields.get("ward_text"):
            ward_texts.add(record.fields["ward_text"])
    ward_by_exact = await load_ward_exact_lookup(db, ward_texts)
    ward_alias_by_text = await load_ward_alias_lookup(db, ward_texts)

    header_by_ref: dict[str, list[int]] = {}
    for record in header_records:
        ref = record.fields.get("order_ref")
        if ref:
            header_by_ref.setdefault(ref, []).append(record.row_number)

    return ReceiveValidationContext(
        equipment_by_me_code=equipment_by_me_code,
        ward_by_exact=ward_by_exact,
        ward_alias_by_text=ward_alias_by_text,
        header_by_ref=header_by_ref,
        duplicate_row_key_rows=duplicate_row_key_rows,
    )


def _validate_headers_structural(
    header_records: list[RawImportRecord], header_by_ref: dict[str, list[int]]
) -> list[LegacyReceiveFinding]:
    """A header row missing its own `เลขที่ใบคืน`, or one whose key
    collides with another header row's, is a malformed-header finding on
    the header row itself -- independent of whether any line ever
    references it. `header_by_ref` is `ReceiveValidationContext`'s own
    grouping (already computed once in `preload_business_context`),
    reused here rather than regrouped a second time."""
    findings: list[LegacyReceiveFinding] = []
    for record in header_records:
        if not record.fields.get("order_ref"):
            findings.append(
                LegacyReceiveFinding(
                    sheet_name=HEADER_SHEET_NAME,
                    source_row_number=record.row_number,
                    field="order_ref",
                    error_code=CODE_HEADER_ORDER_REFERENCE_MISSING,
                    message="เลขที่ใบคืน is required on every order header row.",
                )
            )
    for ref, rows in header_by_ref.items():
        if len(rows) > 1:
            for row_number in rows:
                findings.append(
                    LegacyReceiveFinding(
                        sheet_name=HEADER_SHEET_NAME,
                        source_row_number=row_number,
                        field="order_ref",
                        error_code=CODE_HEADER_ORDER_REFERENCE_DUPLICATE,
                        message=f"เลขที่ใบคืน '{ref}' appears on more than one order header row.",
                    )
                )
    return findings


def _validate_equipment(
    record: RawImportRecord, context: ReceiveValidationContext
) -> tuple[list[LegacyReceiveFinding], Equipment | None]:
    """`ME.Code` is the sole governing legacy Equipment identifier for
    the Receive side, symmetric with Issue. `SCAN CODE รับ`/
    `Barcode รับเครื่อง` are NOT cross-checked against it -- no safe,
    deterministic extractor for either exists anywhere in this
    repository (same finding as Issue's `SCAN CODE ส่ง`/
    `Barcode ส่งเครื่อง`); inventing ad hoc heuristics here would create
    an inconsistent validation rule between the two sides, which this
    slice's own instructions explicitly forbid. This is a documented
    non-implementation, not an oversight.

    Current Equipment status/version is never consulted here -- a
    historical RECEIVE event's validity does not depend on Equipment's
    *current* live state (AVAILABLE_AT_POOL, ISSUED_TO_WARD,
    UNAVAILABLE_DEFECTIVE, DECOMMISSIONED are all equally acceptable);
    only identity resolution is checked, and Equipment is never
    mutated."""
    findings: list[LegacyReceiveFinding] = []
    outcome, value = record.fields["me_code_cell"]
    if outcome == "blank":
        findings.append(
            LegacyReceiveFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=record.row_number,
                field="me_code",
                error_code=CODE_ME_CODE_MISSING,
                message="ME.Code is required.",
            )
        )
        return findings, None
    if outcome == "invalid":
        findings.append(
            LegacyReceiveFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=record.row_number,
                field="me_code",
                error_code=CODE_ME_CODE_INVALID,
                message="ME.Code must be a text cell, not a numeric cell.",
            )
        )
        return findings, None

    normalized = record.fields.get("me_code_normalized")
    if normalized is None:
        try:
            normalize_bcm_code(value)  # re-raise for the message
        except InvalidInputError as exc:
            findings.append(
                LegacyReceiveFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=record.row_number,
                    field="me_code",
                    error_code=CODE_ME_CODE_INVALID,
                    message=f"Invalid ME.Code: {exc.message}",
                )
            )
        return findings, None

    equipment = context.equipment_by_me_code.get(normalized)
    if equipment is None:
        findings.append(
            LegacyReceiveFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=record.row_number,
                field="me_code",
                error_code=CODE_EQUIPMENT_NOT_FOUND,
                message=f"No Equipment record matches ME.Code '{normalized}'.",
            )
        )
        return findings, None
    return findings, equipment


def _validate_ward(
    header: RawImportRecord | None, line: RawImportRecord, context: ReceiveValidationContext
) -> tuple[list[LegacyReceiveFinding], str | None, uuid.UUID | None]:
    """OD-PR21-4's exact-match/alias resolution, plus the header/line
    precedence rule -- identical policy to `issue._validate_ward`, since
    the Receive source structure shows no material difference in header
    vs. line Ward roles that would justify a Receive-specific
    precedence. Returns `(findings, raw_text_used, resolved_ward_id)` --
    `raw_text_used` is the winning raw text preserved on the candidate
    regardless of which side it came from."""
    findings: list[LegacyReceiveFinding] = []
    header_text = header.fields.get("ward_text") if header is not None else None
    line_text = line.fields.get("ward_text")

    def _resolve(text: str, sheet_name: str, row_number: int) -> uuid.UUID | None:
        resolution = resolve_ward(text, ward_by_exact=context.ward_by_exact, ward_alias_by_text=context.ward_alias_by_text)
        if resolution.outcome == "resolved":
            return resolution.ward_id
        code = CODE_WARD_AMBIGUOUS if resolution.outcome == "ambiguous" else CODE_WARD_NOT_FOUND
        findings.append(
            LegacyReceiveFinding(
                sheet_name=sheet_name,
                source_row_number=row_number,
                field="ward_text",
                error_code=code,
                message=f"Ward text '{text}' could not be unambiguously resolved.",
            )
        )
        return None

    header_ward_id: uuid.UUID | None = None
    if header_text:
        header_ward_id = _resolve(header_text, HEADER_SHEET_NAME, header.row_number if header is not None else -1)
    line_ward_id: uuid.UUID | None = None
    if line_text:
        line_ward_id = _resolve(line_text, LINE_SHEET_NAME, line.row_number)

    if findings:
        # Either side failed to resolve -- never fall back to the side
        # that did resolve (that would be a silent preference the
        # design's own precedence rule does not authorize).
        return findings, None, None

    if header_text and line_text:
        if header_ward_id == line_ward_id:
            return findings, line_text, line_ward_id
        findings.append(
            LegacyReceiveFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=line.row_number,
                field="ward_text",
                error_code=CODE_HEADER_LINE_WARD_CONFLICT,
                message=f"Header Ward '{header_text}' and line Ward '{line_text}' resolve to different Wards.",
            )
        )
        return findings, None, None
    if line_text:
        return findings, line_text, line_ward_id
    if header_text:
        return findings, header_text, header_ward_id
    findings.append(
        LegacyReceiveFinding(
            sheet_name=LINE_SHEET_NAME,
            source_row_number=line.row_number,
            field="ward_text",
            error_code=CODE_WARD_NOT_FOUND,
            message="No Ward text is present on either the order header or the line item.",
        )
    )
    return findings, None, None


def _validate_timestamp(header: RawImportRecord | None, line: RawImportRecord) -> tuple[list[LegacyReceiveFinding], datetime | None]:
    """`occurred_at` is fundamentally a line-item fact, identical policy
    to `issue._validate_timestamp` -- the line's own date+time is the
    required, primary source. The header's date+time, when both are
    present and combinable, is cross-checked for consistency only; a
    malformed/incomplete HEADER timestamp does not itself block a row
    whose own line timestamp combines cleanly, since nothing on the
    candidate is ever derived from the header timestamp directly."""
    findings: list[LegacyReceiveFinding] = []
    line_dt = combine_bangkok_datetime(line.fields.get("date"), line.fields.get("time"))
    if line_dt is None:
        findings.append(
            LegacyReceiveFinding(
                sheet_name=LINE_SHEET_NAME,
                source_row_number=line.row_number,
                field="date_time",
                error_code=CODE_INVALID_TIMESTAMP,
                message="วันที่/เวลา could not be combined into a valid timestamp.",
            )
        )
        return findings, None

    if header is not None:
        header_dt = combine_bangkok_datetime(header.fields.get("date"), header.fields.get("time"))
        if header_dt is not None and header_dt != line_dt:
            findings.append(
                LegacyReceiveFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="date_time",
                    error_code=CODE_HEADER_LINE_TIMESTAMP_CONFLICT,
                    message="Header and line date/time do not represent the same event time.",
                )
            )
            return findings, None

    assert isinstance(line_dt, datetime)
    return findings, line_dt


def validate_and_build_candidates(
    header_records: list[RawImportRecord],
    line_records: list[RawImportRecord],
    context: ReceiveValidationContext,
    *,
    migration_authority_id: uuid.UUID,
    import_session_id: uuid.UUID,
    import_source_id: uuid.UUID,
) -> tuple[list[LegacyReceiveCandidate], list[LegacyReceiveFinding]]:
    """Header-line resolution, identity/duplicate rules, Equipment/Ward/
    timestamp validation, and typed candidate construction -- identical
    architecture to `issue.validate_and_build_candidates`. Every finding
    is returned regardless of row outcome; the caller decides
    `validated` vs `validation_failed` for the whole batch from the
    aggregate presence of any `severity == "error"` finding, exactly as
    `import_validation_service.run_validation` already does for every
    other adapter (this module produces no batch-level decision itself).
    No ISSUE candidate is ever consulted, matched, or paired here --
    RECEIVE candidates are built entirely independently of any ISSUE
    output (PR22 owns future Issue<->Receive reconciliation)."""
    findings: list[LegacyReceiveFinding] = list(_validate_headers_structural(header_records, context.header_by_ref))
    header_by_row_number = {r.row_number: r for r in header_records}

    ambiguous_refs = {ref for ref, rows in context.header_by_ref.items() if len(rows) > 1}
    unambiguous_header_row: dict[str, int] = {
        ref: rows[0] for ref, rows in context.header_by_ref.items() if len(rows) == 1
    }

    candidates: list[LegacyReceiveCandidate] = []
    for line in line_records:
        row_findings: list[LegacyReceiveFinding] = []

        row_key_outcome, row_key_value = line.fields["row_key_cell"]
        if row_key_outcome != "valid":
            row_findings.append(
                LegacyReceiveFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="row_key",
                    error_code=CODE_SOURCE_ROW_KEY_MISSING,
                    message="ลำดับ is required on every data-bearing line-item row.",
                )
            )
        elif line.row_number in context.duplicate_row_key_rows:
            row_findings.append(
                LegacyReceiveFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="row_key",
                    error_code=CODE_SOURCE_ROW_KEY_DUPLICATE,
                    message=f"ลำดับ '{row_key_value}' appears on more than one line-item row.",
                )
            )

        order_ref = line.fields.get("order_ref")
        header_record: RawImportRecord | None = None
        if not order_ref:
            row_findings.append(
                LegacyReceiveFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="order_ref",
                    error_code=CODE_ORDER_REFERENCE_MISSING,
                    message="เลขที่ใบรับเครื่อง is required on every line-item row.",
                )
            )
        elif order_ref in ambiguous_refs:
            row_findings.append(
                LegacyReceiveFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="order_ref",
                    error_code=CODE_AMBIGUOUS_ORDER_REFERENCE,
                    message=f"เลขที่ใบรับเครื่อง '{order_ref}' matches more than one order header row.",
                )
            )
        elif order_ref not in unambiguous_header_row:
            row_findings.append(
                LegacyReceiveFinding(
                    sheet_name=LINE_SHEET_NAME,
                    source_row_number=line.row_number,
                    field="order_ref",
                    error_code=CODE_ORPHAN_ORDER_REFERENCE,
                    message=f"เลขที่ใบรับเครื่อง '{order_ref}' does not match any order header row.",
                )
            )
        else:
            header_record = header_by_row_number[unambiguous_header_row[order_ref]]

        equipment_findings, equipment = _validate_equipment(line, context)
        row_findings.extend(equipment_findings)

        ward_findings, ward_text, resolved_ward_id = _validate_ward(header_record, line, context)
        row_findings.extend(ward_findings)

        timestamp_findings, occurred_at = _validate_timestamp(header_record, line)
        row_findings.extend(timestamp_findings)

        findings.extend(row_findings)

        has_error = any(f.severity == "error" for f in row_findings)
        if has_error or header_record is None or equipment is None or occurred_at is None:
            continue

        bme_name = line.fields.get("bme_name") or (header_record.fields.get("bme_name") if header_record else None)

        candidates.append(
            LegacyReceiveCandidate(
                event_type=EVENT_TYPE,
                legacy_source_row_key=row_key_value,  # type: ignore[arg-type]
                legacy_order_reference=order_ref,
                equipment_id=equipment.id,
                occurred_at=occurred_at,
                legacy_ward_text=ward_text,
                resolved_ward_id=resolved_ward_id,
                legacy_bme_name=bme_name,
                migration_authority_id=migration_authority_id,
                import_session_id=import_session_id,
                import_source_id=import_source_id,
                header_source_ref=SourceCoordinate(sheet_name=HEADER_SHEET_NAME, source_row_number=header_record.row_number),
                line_source_ref=SourceCoordinate(sheet_name=LINE_SHEET_NAME, source_row_number=line.row_number),
            )
        )

    return candidates, findings

"""Equipment Master import adapter -- Roadmap PR20C (Parse + Normalize +
Validate only).

Authoritative design: `docs/design/PR20_EQUIPMENT_MASTER_IMPORT_PLAN.md`
§6.3 (adapter shape), §7 (source workbook contract), §8 (32-column field
mapping), §9 OD-1/OD-2/OD-3/OD-4 (Owner Decisions, all RESOLVED), all
merged via GitHub PR #92 (squash SHA
120319afb44f12340790a74dfaf53fa5068591ee). Do not restate the rationale
here -- read the design doc for it. This module implements exactly that
contract.

**Scope boundary (§9 OD-2, strictly enforced by this module's shape):**
this adapter classifies each row (structurally sound / CREATE candidate /
UPDATE candidate / conflict) via `ValidationFinding` rows only. It never
mutates `equipment` -- `plan_dry_run`/`execute` are intentionally left at
the base `ImportAdapter` class's `NotImplementedError` default (§24:
PR20D/PR20E's own scope, not this module's).

**OD-4 (never fabricate `asset_number`):** the current 32-column
`export_template.xlsx` contract supplies no `asset_number` source. Every
row that would otherwise be a potential CREATE candidate therefore always
receives the blocking `ASSET_NUMBER_REQUIRED_FOR_CREATE` finding (§9
OD-4) -- this is the correct, designed behavior, not a bug: parser/
validation readiness (this module) is deliberately distinct from
CREATE-execution readiness (blocked until the Repository Owner supplies
an authoritative Asset Number source).
"""

from __future__ import annotations

import asyncio
import io
import math
import uuid
import zipfile
import xml.parsers.expat as expat
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import AUDIT_ACTION_IMPORT_DRY_RUN_PLAN_CREATED, AUDIT_ENTITY_IMPORT_SESSION, record_audit_event
from app.core.db_errors import IntegrityViolationKind, classify_integrity_error
from app.core.exceptions import ImportNoConfirmedPlanError, InvalidInputError
from app.crud import equipment as equipment_crud
from app.crud import import_dry_run_plan as import_dry_run_plan_crud
from app.models.equipment import Equipment, EquipmentStatus
from app.models.import_session import EquipmentMasterDryRunPlanRow, ImportSession
from app.services.identifiers import normalize_bcm_code, normalize_item_no
from app.services.import_adapter import (
    MAX_IMPORT_ROWS,
    AdapterExecutionConflict,
    DryRunPlan,
    FieldError,
    ImportAdapter,
    RawImportRecord,
    register_adapter,
)
from app.services.import_adapter_context import get_adapter_invocation_context, record_resolved_execution_resource
from app.services.import_service import (
    MAX_HEADER_COLUMNS,
    MAX_UPLOAD_BYTES,
    MAX_WORKSHEET_COUNT,
    _validate_zip_archive_bounds,
)
from app.services.import_source_reader import VerifiedSourceContent

DATASET_TYPE = "equipment_master"

# ---------------------------------------------------------------------------
# §7/§9 OD-1: the authoritative worksheet/header contract, verbatim from the
# Owner-supplied export_template.xlsx evidence. A closed-world list -- the
# parser must not invent, assume, or accept a column beyond this set.
# ---------------------------------------------------------------------------
WORKSHEET_NAME = "Sheet1"

# Header text -> internal role. Every key below must be present in row 1
# exactly once (§7's required-structural-columns rule); a role of `None`
# means the column is intentionally-ignored/deferred (§8) -- its header
# presence is still governed, but no per-row cell value is ever extracted
# for it, since "PR20C must not depend on a deferred field" (§8).
_GOVERNED_HEADERS: dict[str, str | None] = {
    "Item No.": "item_no",
    "ID CODE": "bcm_code",
    "Asset ID": "asset_id",
    "ปีที่ซื้อ": None,
    "วันที่ลงทะเบียน": None,
    "วันที่รับ": None,
    "วันเริ่มประกัน": None,
    "วันหมดประกัน": None,
    "ชื่อไทย": "equipment_name",
    "ชื่ออังกฤษ": None,
    "Ownership": None,
    "กลุ่มโรงพยาบาล": None,
    "โรงพยาบาล": None,
    "หน่วยงาน": None,
    "อาคาร": None,
    "ชั้น": None,
    "ห้อง": None,
    "ประเภทเครื่องมือ": None,
    "ชนิดเครื่องมือ": None,
    "ยี่ห้อ": "brand",
    "รุ่น": "model",
    "S/N": "serial_number",
    "ราคาซื้อ": None,
    "ผู้ขาย": None,
    "ชื่อผู้ติดต่อ": None,
    "เบอร์โทรผู้ติดต่อ": None,
    "สถานะเครื่องมือ": "status",
    "อยู่ในประกัน": None,
    "Life Expect": None,
    "ความเสี่ยง": None,
    "Classification": None,
    "TOR": None,
}
assert len(_GOVERNED_HEADERS) == 32  # §9 OD-1: exactly 32 governed columns

# §8: Equipment column widths this adapter must bound against, matching
# `backend/app/models/equipment.py` exactly -- so an overlength value is
# always a reported finding, never left to fail only once a later slice
# actually attempts the write.
_FIELD_MAX_LENGTHS: dict[str, int] = {
    "asset_id": 100,
    "equipment_name": 255,
    "brand": 100,
    "model": 100,
    "serial_number": 100,
}

# §10, PR93-H1R: authoritative legacy-status mapping, verified against the
# real Repository-Owner-supplied `export_template.xlsx` (§9 OD-1's own
# implementation-time task, now closed). The `สถานะเครื่องมือ` column's
# entire observed vocabulary across all 4,729 source records is exactly
# these four values -- no other string was observed:
#
#   Active        3,873 rows -> AVAILABLE_AT_POOL
#   Decommission    734 rows -> DECOMMISSIONED
#   Defective        84 rows -> UNAVAILABLE_DEFECTIVE
#   Missing          38 rows -> unmappable (blocking ERROR; not a new state)
#
# No guessed/illustrative alias is retained (e.g. "available", "faulty",
# "disposed", "decommissioned", or any of their Thai equivalents this
# module previously carried) -- none of those strings were observed in the
# authoritative source, so accepting them would silently widen this
# mapping beyond what §10's implementation-time evidence requirement
# actually verified. Keys are lowercased/trimmed for case-insensitive
# exact-match lookup; any value not listed here (including "Missing") is
# unmappable by design -- a blocking finding, never a guess.
STATUS_MAPPING: dict[str, EquipmentStatus] = {
    "active": EquipmentStatus.AVAILABLE_AT_POOL,
    "defective": EquipmentStatus.UNAVAILABLE_DEFECTIVE,
    "decommission": EquipmentStatus.DECOMMISSIONED,
    # "Missing" is deliberately absent -- confirmed present in the source
    # (38 rows) but the Repository Owner has resolved it as UNMAPPABLE /
    # blocking ERROR, not a silent equivalence to any of the four states.
    #
    # ISSUED_TO_WARD is intentionally absent -- §9 OD-2's Legacy Lifecycle
    # Policy: PR20 never initializes a CREATE candidate directly into
    # ISSUED_TO_WARD (no synthetic BorrowTransaction), and the authoritative
    # source itself contains no value the Repository Owner has mapped to
    # it. A legacy value that appears to mean "currently issued" is
    # unmappable-by-design and falls through to the blocking
    # STATUS_UNMAPPABLE finding below, per §11.
}

# ---------------------------------------------------------------------------
# §12: centralized, stable finding codes. `ASSET_NUMBER_REQUIRED_FOR_CREATE`
# is used verbatim as specified by §9 OD-4 -- every other code follows this
# module's own `EQUIPMENT_MASTER_<CONDITION>` convention (§12's own
# illustrative naming), since no exact code beyond the OD-4 one is fixed by
# the merged design.
# ---------------------------------------------------------------------------
CODE_BCM_MISSING = "EQUIPMENT_MASTER_BCM_MISSING"
CODE_BCM_INVALID = "EQUIPMENT_MASTER_BCM_INVALID"
CODE_BCM_DUPLICATE_IN_SOURCE = "EQUIPMENT_MASTER_BCM_DUPLICATE_IN_SOURCE"
CODE_ITEM_NO_MISSING = "EQUIPMENT_MASTER_ITEM_NO_MISSING"
CODE_ITEM_NO_INVALID = "EQUIPMENT_MASTER_ITEM_NO_INVALID"
CODE_ITEM_NO_DUPLICATE_IN_SOURCE = "EQUIPMENT_MASTER_ITEM_NO_DUPLICATE_IN_SOURCE"
CODE_IDENTITY_CONFLICT = "EQUIPMENT_MASTER_IDENTITY_CONFLICT"
CODE_EQUIPMENT_NAME_MISSING = "EQUIPMENT_MASTER_EQUIPMENT_NAME_MISSING"
CODE_FIELD_TOO_LONG = "EQUIPMENT_MASTER_FIELD_TOO_LONG"
CODE_SERIAL_NUMBER_CONFLICT = "EQUIPMENT_MASTER_SERIAL_NUMBER_CONFLICT"
CODE_ASSET_ID_CONFLICT = "EQUIPMENT_MASTER_ASSET_ID_CONFLICT"
CODE_STATUS_MISSING = "EQUIPMENT_MASTER_STATUS_MISSING"
CODE_STATUS_UNMAPPABLE = "EQUIPMENT_MASTER_STATUS_UNMAPPABLE"
CODE_STATUS_MISMATCH = "EQUIPMENT_MASTER_STATUS_MISMATCH"
# §9 OD-4: this exact code is the merged design's own stable identifier --
# not prefixed with this module's own convention, since OD-4 names it
# directly.
CODE_ASSET_NUMBER_REQUIRED_FOR_CREATE = "ASSET_NUMBER_REQUIRED_FOR_CREATE"


# ---------------------------------------------------------------------------
# §7: cell-level extraction for the two governed identifiers. Distinct
# acceptance rules per §7/§9 OD-1's identifier cell contract: BCM (`ID
# CODE`) is text-only; Item Number (`Item No.`) accepts text or a
# losslessly-convertible integer-numeric cell. Blank vs. invalid-but-present
# are reported as distinct outcomes, on purpose -- both currently prevent
# CREATE/UPDATE (§9 OD-3's blank/null matrix), but a distinct code is more
# diagnostically useful and does not itself change OD-3's blocking policy.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class _CellResult:
    outcome: str  # "blank" | "valid" | "invalid"
    value: str | None = None
    detail: str | None = None


def _extract_bcm_cell(value: object) -> _CellResult:
    if value is None:
        return _CellResult("blank")
    if isinstance(value, str):
        text = value.strip()
        return _CellResult("blank") if not text else _CellResult("valid", value=text)
    # §7: BCM (`ID CODE`) accepts a text-typed cell only -- any numeric-typed
    # (including bool, a numeric subtype) or other non-text cell is a
    # blocking ERROR, never silently coerced to text.
    return _CellResult("invalid", detail="ID CODE must be a text cell, not a numeric cell.")


def _extract_item_no_cell(value: object) -> _CellResult:
    if value is None:
        return _CellResult("blank")
    if isinstance(value, str):
        text = value.strip()
        return _CellResult("blank") if not text else _CellResult("valid", value=text)
    if isinstance(value, bool):
        return _CellResult("invalid", detail="Item No. must not be a boolean-typed cell.")
    if isinstance(value, int):
        # §7: converted losslessly to its canonical decimal string
        # representation -- never scientific notation, never a
        # floating-point artifact. No leading zero is ever reconstructed
        # (§7's PR92-H1R correction): this is the observed value only.
        return _CellResult("valid", value=str(value))
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return _CellResult("invalid", detail="Item No. numeric cell is not a finite number.")
        if not value.is_integer():
            return _CellResult("invalid", detail="Item No. numeric cell must be a whole number.")
        return _CellResult("valid", value=str(int(value)))
    return _CellResult("invalid", detail="Item No. cell has an unsupported type.")


def _cell_text(value: object) -> str | None:
    """Trim-only extraction for a governed non-identifier text column
    (§8's general cell-normalization rule: "trim whitespace; an empty
    string normalizes to null"). Never applied to BCM/Item No, which have
    their own distinct cell-type rules above."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip() or None


# ---------------------------------------------------------------------------
# §7: workbook/worksheet/header structural parsing (synchronous, CPU-bound
# -- always invoked via the framework's own `asyncio.to_thread`, never
# directly, per `import_validation_service.run_validation`).
# ---------------------------------------------------------------------------


# PR93-H3R: substring markers checked against every OOXML *content type*
# and *relationship type* string this module actually parses out of
# `[Content_Types].xml` / the package-level and workbook-level `.rels`
# parts (never against arbitrary file content) -- covers macro-enabled
# workbook/template/add-in main parts (any format variant: xlsm/xlsb/xltm
# all declare a content type containing "macroEnabled"), the VBA project
# part/relationship, VBA digital-signature parts, and legacy macro-sheet
# content/relationship types. Matching is substring/case-insensitive by
# design: the OOXML spec's actual macro-related content-type and
# relationship-type strings all reliably contain one of these tokens, and
# a broader net here is the correct fail-closed choice over an exact,
# brittle enumeration that a variant spelling could slip past.
_FORBIDDEN_OOXML_METADATA_MARKERS = (
    "macroenabled",
    "vbaproject",
    "vbasignature",
    "macrosheet",
)

# The package-level and workbook-level metadata parts this module actually
# inspects for §7's structural classification -- never any other part, and
# never anything these parts might reference externally (§21: "no
# external... network access").
_OOXML_METADATA_PARTS = (
    "[Content_Types].xml",
    "_rels/.rels",
    "xl/_rels/workbook.xml.rels",
)


class _OOXMLMetadataSignals:
    __slots__ = ("values",)

    def __init__(self) -> None:
        self.values: set[str] = set()


def _collect_ooxml_metadata_signals(part_name: str, data: bytes) -> set[str]:
    """Parses one OOXML package-metadata XML part with `xml.parsers.expat`
    (stdlib, no new dependency) and returns every `ContentType`/`Type`
    attribute value found on an `Override`/`Default`/`Relationship`
    element. Namespace-*un*aware on purpose: none of these attribute names
    are ever namespace-prefixed in a conformant OOXML package, so a plain
    `ParserCreate()` is sufficient and avoids pulling in a namespace layer
    this check doesn't need.

    Fails closed on anything that would make classification unreliable:
    a DOCTYPE declaration is rejected outright (§21's "never execute
    macros" extends to never *resolving* an entity either -- disallowing
    DOCTYPE entirely defeats both external-entity and entity-expansion
    ["billion laughs"] attacks without needing a parser flag that stdlib
    `expat` doesn't expose, and no legitimate `[Content_Types].xml` or
    `.rels` part ever carries one), and malformed XML raises the same
    structured `InvalidInputError` §7 already uses for every other
    workbook-contract failure -- never "could not classify, continue
    anyway.\""""
    signals = _OOXMLMetadataSignals()

    def _reject_doctype(*_args: object) -> None:
        raise InvalidInputError(
            f"The uploaded file's internal '{part_name}' package metadata is malformed and "
            "cannot be safely classified as a plain Excel (.xlsx) spreadsheet."
        )

    def _start_element(name: str, attrs: dict[str, str]) -> None:
        local = name.rsplit(":", 1)[-1]
        if local in ("Override", "Default"):
            content_type = attrs.get("ContentType")
            if content_type:
                signals.values.add(content_type)
        elif local == "Relationship":
            rel_type = attrs.get("Type")
            if rel_type:
                signals.values.add(rel_type)

    parser = expat.ParserCreate()
    parser.StartDoctypeDeclHandler = _reject_doctype
    parser.StartElementHandler = _start_element
    try:
        parser.Parse(data, True)
    except InvalidInputError:
        raise
    except expat.ExpatError as exc:
        raise InvalidInputError(
            f"The uploaded file's internal '{part_name}' package metadata is malformed and "
            "cannot be safely classified as a plain Excel (.xlsx) spreadsheet."
        ) from exc
    return signals.values


def _reject_macro_ooxml_structure(content: bytes) -> None:
    """PR93-H3R: filename-substring detection alone (`_reject_macro_parts`
    below) misses a macro-enabled OOXML package that declares a
    macro-enabled *content type* or a VBA *relationship* without ever
    naming a part `vbaProject...` -- e.g. `[Content_Types].xml` declaring
    `application/vnd.ms-excel.sheet.macroEnabled.main+xml` for the main
    workbook part, with no part named `vbaProject` anywhere in the
    archive. This inspects the package's own declared structure (§21:
    "reject macro-enabled OOXML STRUCTURE, not merely suspicious
    filenames") -- `[Content_Types].xml` and the package-level/workbook
    relationship files -- before `load_workbook` ever touches the
    archive. A client-supplied filename or MIME type is never consulted
    here; only the archive's own internal package metadata is."""
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = set(archive.namelist())
            for part_name in _OOXML_METADATA_PARTS:
                if part_name not in names:
                    continue  # optional part legitimately absent from some workbooks
                data = archive.read(part_name)
                for value in _collect_ooxml_metadata_signals(part_name, data):
                    lowered = value.lower()
                    if any(marker in lowered for marker in _FORBIDDEN_OOXML_METADATA_MARKERS):
                        raise InvalidInputError(
                            "The uploaded file declares macro-enabled or VBA-related workbook "
                            "content and cannot be accepted as a plain Excel (.xlsx) spreadsheet."
                        )
    except zipfile.BadZipFile as exc:
        raise InvalidInputError(
            "The uploaded file could not be read as a valid Excel (.xlsx) spreadsheet."
        ) from exc


def _reject_macro_parts(content: bytes) -> None:
    """§21: "Never execute workbook formulas or macros." Defense-in-depth
    alongside `_reject_macro_ooxml_structure` above (which inspects the
    package's own declared content types/relationships): the shared PR12
    `_validate_zip_archive_bounds` bounds the archive's *shape* (entry
    count/size/ratio, allowed top-level paths) but its `xl/` prefix
    allowlist does not itself distinguish an ordinary `.xlsx` from a
    macro-enabled workbook smuggled in under an `.xlsx` filename/content-
    type -- `xl/vbaProject.bin` (and its optional digital-signature
    sibling) is a legitimate path under `xl/`. Never merely relying on
    "openpyxl doesn't execute it" is insufficient defense-in-depth for
    hostile input (§21): any such part is rejected outright, before
    `load_workbook` ever touches the archive."""
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = archive.namelist()
    except zipfile.BadZipFile as exc:
        raise InvalidInputError(
            "The uploaded file could not be read as a valid Excel (.xlsx) spreadsheet."
        ) from exc
    for name in names:
        normalized = name.replace("\\", "/").lower()
        if "vbaproject" in normalized:
            raise InvalidInputError(
                "The uploaded file contains a macro-enabled component (VBA project) and cannot "
                "be accepted as a plain Excel (.xlsx) spreadsheet."
            )


def _load_sheet1(content: bytes) -> Worksheet:
    _validate_zip_archive_bounds(content)
    _reject_macro_ooxml_structure(content)
    _reject_macro_parts(content)
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise InvalidInputError(
            "The uploaded file could not be read as a valid Excel (.xlsx) spreadsheet."
        ) from exc
    if len(workbook.sheetnames) > MAX_WORKSHEET_COUNT:
        raise InvalidInputError(
            f"The uploaded file contains more than {MAX_WORKSHEET_COUNT} worksheets, "
            "exceeding what a normal Equipment Master export contains."
        )
    # §7: the parser MUST select `Sheet1` by name -- never `workbook.active`
    # (an arbitrary first/default worksheet). Any other worksheet present
    # is ignored, not interpreted, per the V1 contract.
    if WORKSHEET_NAME not in workbook.sheetnames:
        raise InvalidInputError(
            f"The uploaded file has no worksheet named '{WORKSHEET_NAME}'. "
            "Equipment Master import requires the authoritative worksheet to be present by name."
        )
    return workbook[WORKSHEET_NAME]


def _validate_headers(worksheet: Worksheet) -> dict[str, int]:
    """§7: header is exactly row 1; column order is not authoritative
    (exact-header-name based). All 32 governed headers must be present
    exactly once; missing/duplicate/unknown-extra are each a structural
    `ERROR` (fail-closed on schema drift, §7)."""
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise InvalidInputError("The uploaded spreadsheet is empty; no header row was found.")
    if len(header_row) > MAX_HEADER_COLUMNS:
        raise InvalidInputError(
            f"The uploaded file's header row contains more than {MAX_HEADER_COLUMNS} columns, "
            "exceeding what a normal Equipment Master export contains."
        )

    seen: dict[str, list[int]] = {}
    for idx, raw in enumerate(header_row):
        text = _cell_text(raw)
        if text is None:
            continue
        seen.setdefault(text, []).append(idx)

    missing = [header for header in _GOVERNED_HEADERS if header not in seen]
    if missing:
        raise InvalidInputError(
            "The uploaded spreadsheet is missing required column(s): " + ", ".join(missing)
        )
    duplicated = [header for header, idxs in seen.items() if header in _GOVERNED_HEADERS and len(idxs) > 1]
    if duplicated:
        raise InvalidInputError(
            "The uploaded spreadsheet has duplicate column(s): " + ", ".join(sorted(duplicated))
        )
    unknown = sorted(header for header in seen if header not in _GOVERNED_HEADERS)
    if unknown:
        raise InvalidInputError(
            "The uploaded spreadsheet contains unrecognized column(s) not part of the approved "
            "Equipment Master schema: " + ", ".join(unknown)
        )

    return {header: seen[header][0] for header in _GOVERNED_HEADERS}


def _parse_rows(worksheet: Worksheet, header_index: dict[str, int]) -> list[RawImportRecord]:
    def cell_at(values: tuple, header: str) -> object:
        idx = header_index[header]
        return values[idx] if idx < len(values) else None

    # §7's closed-world header contract must also bind the data rows, not
    # merely row 1: `_validate_headers` only inspects non-blank header
    # cells, so a column whose row-1 cell is blank (no name at all) is
    # invisible to that check and would otherwise let data smuggle through
    # a column outside the approved 32-column schema. Any non-blank cell
    # at a column index that isn't one of the 32 governed indices is
    # therefore rejected here, structurally, the first time it's seen.
    known_indices = set(header_index.values())

    records: list[RawImportRecord] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if values is None or all(_cell_text(v) is None for v in values):
            continue  # §7: blank rows skipped, not counted toward total_rows
        for idx, value in enumerate(values):
            if idx not in known_indices and _cell_text(value) is not None:
                raise InvalidInputError(
                    f"Row {row_number} contains data in a column outside the approved Equipment "
                    "Master schema (an unnamed/extra column is not permitted)."
                )
        if len(records) >= MAX_IMPORT_ROWS:
            raise InvalidInputError(
                f"The uploaded file contains more than {MAX_IMPORT_ROWS} data rows, "
                "exceeding the maximum a single import batch may contain."
            )
        records.append(
            RawImportRecord(
                row_number=row_number,
                fields={
                    "bcm": _extract_bcm_cell(cell_at(values, "ID CODE")),
                    "item_no": _extract_item_no_cell(cell_at(values, "Item No.")),
                    "asset_id": _cell_text(cell_at(values, "Asset ID")),
                    "equipment_name": _cell_text(cell_at(values, "ชื่อไทย")),
                    "brand": _cell_text(cell_at(values, "ยี่ห้อ")),
                    "model": _cell_text(cell_at(values, "รุ่น")),
                    "serial_number": _cell_text(cell_at(values, "S/N")),
                    "status": _cell_text(cell_at(values, "สถานะเครื่องมือ")),
                },
            )
        )
    return records


def _parse_workbook_sync(content: bytes) -> list[RawImportRecord]:
    """Everything CPU-bound about turning verified source bytes into
    typed records, run as one unit via the framework's own
    `asyncio.to_thread(adapter.parse, raw_input)` (never a second,
    adapter-owned concurrency mechanism, §10.1)."""
    worksheet = _load_sheet1(content)
    header_index = _validate_headers(worksheet)
    return _parse_rows(worksheet, header_index)


# ---------------------------------------------------------------------------
# §6.3/§13: bulk business context -- exactly two-to-four `IN(...)` queries
# for the whole batch (never one query per row), plus in-workbook duplicate
# detection (§9 OD-3 cases 6/7, §13(B)/(C)).
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class EquipmentMasterContext:
    existing_by_bcm: dict[str, Equipment] = field(default_factory=dict)
    existing_by_item_no: dict[str, Equipment] = field(default_factory=dict)
    existing_by_serial: dict[str, Equipment] = field(default_factory=dict)
    existing_by_asset_id: dict[str, Equipment] = field(default_factory=dict)
    duplicate_bcm_rows: frozenset[int] = frozenset()
    duplicate_item_no_rows: frozenset[int] = frozenset()


def _rows_with_duplicate_values(values_by_row: dict[int, str]) -> frozenset[int]:
    """§9 OD-3 cases 6/7, §13: ALL rows sharing a duplicated value are
    flagged -- never merely the second occurrence, never "last row wins"."""
    seen: dict[str, list[int]] = {}
    for row_number, value in values_by_row.items():
        seen.setdefault(value, []).append(row_number)
    duplicated: set[int] = set()
    for rows in seen.values():
        if len(rows) > 1:
            duplicated.update(rows)
    return frozenset(duplicated)


# ---------------------------------------------------------------------------
# Roadmap PR20D (docs/design/PR20_EQUIPMENT_MASTER_IMPORT_PLAN.md §14, §15.1,
# §8): `plan_dry_run` row-classification helpers. Field selection mirrors §8's
# CREATE/UPDATE write-column table exactly -- `bcm_code`/`item_no` are
# identity fields (CREATE-only, immutable on UPDATE); `status` is CREATE-only
# (§9 OD-2's Legacy Lifecycle Policy: never overwrites current live status on
# UPDATE).
# ---------------------------------------------------------------------------


def _finding_payload(finding: FieldError) -> dict[str, Any]:
    return {"field": finding.field, "error_code": finding.error_code, "message": finding.message, "severity": finding.severity}


def _create_normalized_values(record: RawImportRecord) -> dict[str, Any]:
    values: dict[str, Any] = {
        "bcm_code": record.fields.get("bcm_normalized"),
        "item_no": record.fields.get("item_no_normalized"),
        "equipment_name": record.fields.get("equipment_name"),
        "brand": record.fields.get("brand"),
        "model": record.fields.get("model"),
        "serial_number": record.fields.get("serial_number"),
        "asset_id": record.fields.get("asset_id"),
    }
    status_raw = record.fields.get("status")
    if status_raw is not None:
        mapped = STATUS_MAPPING.get(status_raw.strip().lower())
        if mapped is not None:
            values["status"] = mapped.value
    return values


def _update_normalized_values(record: RawImportRecord) -> dict[str, Any]:
    return {
        "equipment_name": record.fields.get("equipment_name"),
        "brand": record.fields.get("brand"),
        "model": record.fields.get("model"),
        "serial_number": record.fields.get("serial_number"),
        "asset_id": record.fields.get("asset_id"),
    }


class EquipmentMasterAdapter(ImportAdapter):
    """Roadmap PR20C/PR20D. Implements `parse`/`preload_business_context`/
    `validate_business_rules` (PR20C) and `plan_dry_run`/
    `persist_dry_run_plan` (PR20D) -- `execute` remains at the base
    class's `NotImplementedError` default; `precheck_execute`/
    `on_execution_failure`/`on_execution_recovery` (design §14.4a/§14.4b/
    §14.4c) are PR20E's own future addition to `ImportAdapter`, not yet
    present (§24's PR20E scope boundary). This adapter never mutates
    `equipment` -- `plan_dry_run` is a read-only evaluation and
    `persist_dry_run_plan` writes only to the two PR20D-owned planning
    tables."""

    dataset_type = DATASET_TYPE
    ruleset_version = "1"

    def parse(self, raw_input: Any) -> list[RawImportRecord]:
        # §6.5: `raw_input` is always a checksum/length-verified
        # `VerifiedSourceContent` for this adapter -- PR20A's framework
        # code (never this adapter) resolves and hands it in, since
        # `equipment_master` always uses the byte-storage upload path
        # (§6.2's registration-endpoint guard rejects the metadata-only
        # path for this dataset_type before this adapter is ever reached).
        if not isinstance(raw_input, VerifiedSourceContent):
            raise InvalidInputError(
                "Equipment Master import requires a verified, byte-backed source; "
                "none was supplied for this session."
            )
        content = raw_input.content
        if len(content) > MAX_UPLOAD_BYTES:
            raise InvalidInputError(
                f"Stored source content exceeds the maximum allowed size of "
                f"{MAX_UPLOAD_BYTES // (1024 * 1024)} MB."
            )
        return _parse_workbook_sync(content)

    async def preload_business_context(
        self, db: AsyncSession, records: list[RawImportRecord]
    ) -> EquipmentMasterContext:
        bcm_values: dict[int, str] = {}
        item_no_values: dict[int, str] = {}
        for record in records:
            bcm_cell: _CellResult = record.fields["bcm"]
            if bcm_cell.outcome == "valid":
                bcm_values[record.row_number] = bcm_cell.value  # type: ignore[assignment]
            item_no_cell: _CellResult = record.fields["item_no"]
            if item_no_cell.outcome == "valid":
                item_no_values[record.row_number] = item_no_cell.value  # type: ignore[assignment]

        # §9 OD-3's blank/null matrix and identity matrix both operate on
        # the *normalized* form (ADR-002 canonicalization) -- normalize
        # here, once, rather than duplicating this in validate_business_rules.
        normalized_bcm: dict[int, str] = {}
        for row_number, raw in bcm_values.items():
            try:
                normalized_bcm[row_number] = normalize_bcm_code(raw)
            except InvalidInputError:
                continue  # re-validated (and reported) in validate_business_rules
        normalized_item_no: dict[int, str] = {}
        for row_number, raw in item_no_values.items():
            try:
                normalized_item_no[row_number] = normalize_item_no(raw)
            except InvalidInputError:
                continue

        duplicate_bcm_rows = _rows_with_duplicate_values(normalized_bcm)
        duplicate_item_no_rows = _rows_with_duplicate_values(normalized_item_no)

        bcm_codes_to_check = set(normalized_bcm.values())
        item_nos_to_check = set(normalized_item_no.values())
        serials_to_check = {
            record.fields["serial_number"] for record in records if record.fields["serial_number"]
        }
        asset_ids_to_check = {
            record.fields["asset_id"] for record in records if record.fields["asset_id"]
        }

        existing_by_bcm = await equipment_crud.get_by_bcm_codes(db, list(bcm_codes_to_check))
        existing_by_item_no = await equipment_crud.get_by_item_nos(db, list(item_nos_to_check))
        existing_by_serial = await equipment_crud.get_by_serial_numbers(db, list(serials_to_check))
        existing_by_asset_id = await equipment_crud.get_by_asset_ids(db, list(asset_ids_to_check))

        # Stash the normalized values back so validate_business_rules never
        # re-derives them (single source of truth for what "normalized"
        # means for this batch).
        for record in records:
            row_number = record.row_number
            record.fields["bcm_normalized"] = normalized_bcm.get(row_number)
            record.fields["item_no_normalized"] = normalized_item_no.get(row_number)

        return EquipmentMasterContext(
            existing_by_bcm=existing_by_bcm,
            existing_by_item_no=existing_by_item_no,
            existing_by_serial=existing_by_serial,
            existing_by_asset_id=existing_by_asset_id,
            duplicate_bcm_rows=duplicate_bcm_rows,
            duplicate_item_no_rows=duplicate_item_no_rows,
        )

    def validate_business_rules(
        self, record: RawImportRecord, context: object
    ) -> list[FieldError]:
        assert isinstance(context, EquipmentMasterContext)
        findings: list[FieldError] = []

        bcm_cell: _CellResult = record.fields["bcm"]
        item_no_cell: _CellResult = record.fields["item_no"]

        # §9 OD-3: BCM cell-type/blank checks.
        bcm_blocking = False
        if bcm_cell.outcome == "blank":
            findings.append(
                FieldError(field="bcm_code", error_code=CODE_BCM_MISSING, message="ID CODE is required.")
            )
            bcm_blocking = True
        elif bcm_cell.outcome == "invalid":
            findings.append(
                FieldError(field="bcm_code", error_code=CODE_BCM_INVALID, message=bcm_cell.detail or "Invalid ID CODE.")
            )
            bcm_blocking = True
        else:
            normalized_bcm = record.fields.get("bcm_normalized")
            if normalized_bcm is None:
                # Failed normalize_bcm_code() in preload_business_context.
                try:
                    normalize_bcm_code(bcm_cell.value)  # re-raise for the message
                except InvalidInputError as exc:
                    findings.append(
                        FieldError(field="bcm_code", error_code=CODE_BCM_INVALID, message=f"Invalid ID CODE: {exc.message}")
                    )
                bcm_blocking = True
            elif record.row_number in context.duplicate_bcm_rows:
                findings.append(
                    FieldError(
                        field="bcm_code",
                        error_code=CODE_BCM_DUPLICATE_IN_SOURCE,
                        message=f"ID CODE '{normalized_bcm}' appears on more than one row in this workbook.",
                    )
                )
                bcm_blocking = True

        # §9 OD-3: Item Number cell-type/blank checks.
        item_no_blocking = False
        if item_no_cell.outcome == "blank":
            findings.append(
                FieldError(field="item_no", error_code=CODE_ITEM_NO_MISSING, message="Item No. is required.")
            )
            item_no_blocking = True
        elif item_no_cell.outcome == "invalid":
            findings.append(
                FieldError(
                    field="item_no", error_code=CODE_ITEM_NO_INVALID, message=item_no_cell.detail or "Invalid Item No."
                )
            )
            item_no_blocking = True
        else:
            normalized_item_no = record.fields.get("item_no_normalized")
            if normalized_item_no is None:
                try:
                    normalize_item_no(item_no_cell.value)
                except InvalidInputError as exc:
                    findings.append(
                        FieldError(
                            field="item_no", error_code=CODE_ITEM_NO_INVALID, message=f"Invalid Item No.: {exc.message}"
                        )
                    )
                item_no_blocking = True
            elif record.row_number in context.duplicate_item_no_rows:
                findings.append(
                    FieldError(
                        field="item_no",
                        error_code=CODE_ITEM_NO_DUPLICATE_IN_SOURCE,
                        message=f"Item No. '{normalized_item_no}' appears on more than one row in this workbook.",
                    )
                )
                item_no_blocking = True

        # §9 OD-3: "The blank/null cases above are mandatory preconditions"
        # -- evaluated before the identity matrix; a row failing either
        # never reaches identity resolution, and no fallback/derived
        # identifier is ever substituted.
        if bcm_blocking or item_no_blocking:
            return findings

        bcm = record.fields["bcm_normalized"]
        item_no = record.fields["item_no_normalized"]
        equip_by_bcm = context.existing_by_bcm.get(bcm)
        equip_by_item_no = context.existing_by_item_no.get(item_no)

        if equip_by_bcm is None and equip_by_item_no is None:
            # §9 OD-3 case 1: CREATE candidate.
            findings.extend(self._validate_create_candidate(record, context))
        elif equip_by_bcm is not None and equip_by_item_no is not None and equip_by_bcm.id == equip_by_item_no.id:
            # §9 OD-3 case 2: UPDATE candidate.
            findings.extend(self._validate_update_candidate(record, context, equip_by_bcm))
        else:
            # §9 OD-3 cases 3/4/5: every remaining combination is a
            # blocking cross-identity conflict -- never auto-replace,
            # never auto-choose, never create a duplicate logical
            # Equipment.
            findings.append(
                FieldError(
                    field=None,
                    error_code=CODE_IDENTITY_CONFLICT,
                    message=(
                        f"BCM '{bcm}' and Item No. '{item_no}' do not consistently identify the same "
                        "existing Equipment record (or the same new record)."
                    ),
                )
            )

        return findings

    def _validate_create_candidate(
        self, record: RawImportRecord, context: EquipmentMasterContext
    ) -> list[FieldError]:
        findings: list[FieldError] = []

        equipment_name = record.fields["equipment_name"]
        if equipment_name is None:
            findings.append(
                FieldError(
                    field="equipment_name",
                    error_code=CODE_EQUIPMENT_NAME_MISSING,
                    message="ชื่อไทย (Thai name) is required to create a new Equipment record.",
                )
            )
        else:
            findings.extend(self._length_findings({"equipment_name": equipment_name}))

        findings.extend(
            self._length_findings(
                {
                    key: record.fields[key]
                    for key in ("asset_id", "brand", "model", "serial_number")
                    if record.fields[key] is not None
                }
            )
        )

        serial_number = record.fields["serial_number"]
        if serial_number is not None and serial_number in context.existing_by_serial:
            # Unlike the UPDATE case, a CREATE candidate has no "self" record
            # to exempt -- any existing match is a blocking conflict, since
            # Equipment.serial_number is DB-unique.
            findings.append(
                FieldError(
                    field="serial_number",
                    error_code=CODE_SERIAL_NUMBER_CONFLICT,
                    message="S/N already belongs to an existing Equipment record.",
                )
            )

        status_raw = record.fields["status"]
        if status_raw is None:
            findings.append(
                FieldError(
                    field="status",
                    error_code=CODE_STATUS_MISSING,
                    message="สถานะเครื่องมือ (equipment status) is required to create a new Equipment record.",
                )
            )
        elif status_raw.strip().lower() not in STATUS_MAPPING:
            findings.append(
                FieldError(
                    field="status",
                    error_code=CODE_STATUS_UNMAPPABLE,
                    message=f"Unrecognized equipment status value: '{status_raw}'.",
                )
            )

        # §9 OD-4: every potential CREATE candidate is blocked -- the
        # currently-approved 32-column contract supplies no authoritative
        # asset_number source. This finding is unconditional for every
        # CREATE candidate, by design.
        findings.append(
            FieldError(
                field="asset_number",
                error_code=CODE_ASSET_NUMBER_REQUIRED_FOR_CREATE,
                message=(
                    "An authoritative Asset Number source is required to create a new Equipment "
                    "record; none is available in the current Equipment Master workbook contract. "
                    "This row cannot become executable CREATE work until the Repository Owner "
                    "supplies one (Roadmap PR20 §9 OD-4)."
                ),
            )
        )

        return findings

    def _validate_update_candidate(
        self, record: RawImportRecord, context: EquipmentMasterContext, target: Equipment
    ) -> list[FieldError]:
        findings: list[FieldError] = []

        findings.extend(
            self._length_findings(
                {
                    key: record.fields[key]
                    for key in ("asset_id", "equipment_name", "brand", "model", "serial_number")
                    if record.fields[key] is not None
                }
            )
        )

        serial_number = record.fields["serial_number"]
        if serial_number is not None:
            other = context.existing_by_serial.get(serial_number)
            if other is not None and other.id != target.id:
                findings.append(
                    FieldError(
                        field="serial_number",
                        error_code=CODE_SERIAL_NUMBER_CONFLICT,
                        message="S/N already belongs to a different Equipment record.",
                    )
                )

        asset_id = record.fields["asset_id"]
        if asset_id is not None:
            other = context.existing_by_asset_id.get(asset_id)
            if other is not None and other.id != target.id:
                findings.append(
                    FieldError(
                        field="asset_id",
                        error_code=CODE_ASSET_ID_CONFLICT,
                        message="Asset ID already belongs to a different Equipment record.",
                        severity="warning",
                    )
                )

        # §9 OD-2 Legacy Lifecycle Policy / §10: legacy status is read for
        # cross-check only -- never applied to the existing record's live
        # status. §10 states the unmappable-status fallback generally ("any
        # legacy status this design cannot safely map produces a blocking
        # ERROR"), not only for CREATE -- a non-blank UPDATE-row status
        # value that fails to map to one of the four states is therefore
        # the same blocking CODE_STATUS_UNMAPPABLE as a CREATE candidate's,
        # never silently ignored. A *mapped* value that simply differs from
        # the live status is a separate, non-blocking WARNING only (§8's
        # UPDATE column: "never overwrites current live status").
        status_raw = record.fields["status"]
        if status_raw is not None:
            mapped = STATUS_MAPPING.get(status_raw.strip().lower())
            if mapped is None:
                findings.append(
                    FieldError(
                        field="status",
                        error_code=CODE_STATUS_UNMAPPABLE,
                        message=f"Unrecognized equipment status value: '{status_raw}'.",
                    )
                )
            elif mapped != target.status:
                findings.append(
                    FieldError(
                        field="status",
                        error_code=CODE_STATUS_MISMATCH,
                        message=(
                            f"Legacy status '{status_raw}' maps to '{mapped.value}', which differs from "
                            f"the current live status '{target.status.value}'. Not applied -- current "
                            "operational status remains authoritative."
                        ),
                        severity="warning",
                    )
                )

        return findings

    @staticmethod
    def _length_findings(values: dict[str, str]) -> list[FieldError]:
        findings: list[FieldError] = []
        for key, value in values.items():
            max_length = _FIELD_MAX_LENGTHS[key]
            if len(value) > max_length:
                findings.append(
                    FieldError(
                        field=key,
                        error_code=CODE_FIELD_TOO_LONG,
                        message=f"{key} exceeds the maximum length of {max_length} characters.",
                    )
                )
        return findings

    async def plan_dry_run(self, db: AsyncSession) -> DryRunPlan:
        """Roadmap PR20D (design §14, §15.1). Re-runs `parse` ->
        `preload_business_context` -> `validate_business_rules` against
        the same frozen, verified source content and the current database
        state (§6.3) -- entirely read-only; this method itself never
        writes. Every row resolves to exactly one of CREATE/UPDATE/SKIP
        (§6): a row carrying any blocking (`severity == "error"`) finding
        is always SKIP (never an executable CREATE/UPDATE, closing §9
        OD-4's CREATE Asset Number gate and every other blocking
        validation finding the same way); a CREATE candidate with no
        blocking finding would be plan-ed as CREATE, but under the
        current 32-column source contract this branch is unreachable in
        practice, since `_validate_create_candidate` unconditionally
        emits the blocking `ASSET_NUMBER_REQUIRED_FOR_CREATE` finding for
        every CREATE candidate (§9 OD-4) -- kept so a future, richer
        source contract can light this path up without a further
        `plan_dry_run` change. An UPDATE row captures
        `matched_equipment.version` (PR20B) *now*, at dry-run time, into
        `expected_equipment_version` -- the only place this value is ever
        captured (§15.1); it is never re-read at execute time (PR20E)."""
        ctx = get_adapter_invocation_context()
        records = await asyncio.to_thread(self.parse, ctx.verified_source_content)
        context = await self.preload_business_context(db, records)

        rows: list[dict[str, Any]] = []
        creates = updates = skips = warning_rows = blocking_conflicts = 0
        for record in records:
            findings = self.validate_business_rules(record, context)
            blocking_findings = [f for f in findings if f.severity == "error"]
            warning_findings = [f for f in findings if f.severity == "warning"]

            bcm = record.fields.get("bcm_normalized")
            item_no = record.fields.get("item_no_normalized")
            matched_identity_fields: dict[str, Any] = {}
            if bcm:
                matched_identity_fields["bcm_code"] = bcm
            if item_no:
                matched_identity_fields["item_no"] = item_no

            equip_by_bcm = context.existing_by_bcm.get(bcm) if bcm else None
            equip_by_item_no = context.existing_by_item_no.get(item_no) if item_no else None
            is_create_candidate = equip_by_bcm is None and equip_by_item_no is None
            is_update_candidate = (
                equip_by_bcm is not None
                and equip_by_item_no is not None
                and equip_by_bcm.id == equip_by_item_no.id
            )

            if blocking_findings:
                action = "SKIP"
                target_equipment_id = None
                expected_equipment_version = None
                normalized_values: dict[str, Any] = {}
                row_warnings = [_finding_payload(f) for f in findings]
                skips += 1
                if any(f.error_code == CODE_IDENTITY_CONFLICT for f in blocking_findings):
                    blocking_conflicts += 1
            elif is_create_candidate:
                action = "CREATE"
                target_equipment_id = None
                expected_equipment_version = None
                normalized_values = _create_normalized_values(record)
                row_warnings = [_finding_payload(f) for f in warning_findings]
                creates += 1
            elif is_update_candidate:
                assert equip_by_bcm is not None
                action = "UPDATE"
                target_equipment_id = equip_by_bcm.id
                expected_equipment_version = equip_by_bcm.version
                normalized_values = _update_normalized_values(record)
                row_warnings = [_finding_payload(f) for f in warning_findings]
                updates += 1
            else:
                # Defensive fallback mirroring validate_business_rules'
                # own cases 3/4/5 identity-conflict branch -- unreachable
                # today, since that branch always emits a blocking
                # CODE_IDENTITY_CONFLICT finding already handled above.
                # Handled explicitly rather than assumed, so a future
                # validate_business_rules change can never silently leave
                # this row unclassified.
                action = "SKIP"
                target_equipment_id = None
                expected_equipment_version = None
                normalized_values = {}
                row_warnings = [_finding_payload(f) for f in findings]
                skips += 1
                blocking_conflicts += 1

            if warning_findings:
                warning_rows += 1

            rows.append(
                {
                    "source_row_number": record.row_number,
                    "action": action,
                    "target_equipment_id": target_equipment_id,
                    "normalized_values": normalized_values,
                    "matched_identity_fields": matched_identity_fields,
                    "expected_equipment_version": expected_equipment_version,
                    "warnings": row_warnings,
                }
            )

        summary = {
            "total_rows": len(rows),
            "creates": creates,
            "updates": updates,
            "skips": skips,
            "warnings": warning_rows,
            "blocking_conflicts": blocking_conflicts,
        }
        return DryRunPlan(summary={"rows": rows, **summary})

    async def persist_dry_run_plan(self, db: AsyncSession, plan: DryRunPlan) -> None:
        """Roadmap PR20D (design §14.2, §14.3). Called by the framework on
        the normal writable session, in the same transaction as the
        session's `dry_run_completed` fenced-completion write (§14.3's
        `db.commit()` covers both). Marks any prior `active` plan for this
        session `superseded`, then inserts the new, `active`,
        **unconfirmed** (`confirmed_at IS NULL`, §14.4a) plan and its rows
        -- one immutable artifact per successful dry-run attempt, never
        updated in place (§10).

        Fix round 2 (PR94-H1): locks the owning `ImportSession` row
        (`SELECT ... FOR UPDATE`) *before* touching the plan tables below.
        `app.crud.import_dry_run_plan.confirm_plan` also locks
        Session-then-Plan, in that same order, before its own conditional
        `UPDATE` -- without this lock here, this function would instead
        lock Plan-then-Session, the exact opposite order, a genuine
        circular-wait deadlock risk against a concurrent confirmation.

        Fix round 3 (PR94, TX1 atomicity/lock-order): the caller
        (`import_execution_service.run_dry_run`) now calls
        `import_job_crud.fenced_phase_success` -- which locks the Job row
        then the Session row -- *before* this function, in the same TX1,
        so by the time this function's own `SELECT ... FOR UPDATE` above
        runs, the Session row is normally already locked by this same
        transaction (a cheap, non-blocking re-affirmation, never a new
        wait). This keeps the *acquisition* order Job -> Session -> Plan
        throughout TX1 -- matching `claim_stale_job`/
        `transition_session_for_recovery`'s own Job-then-Session order for
        stale-job recovery -- so dry-run completion publication and a
        concurrent recovery attempt can never deadlock against each
        other. This function still takes the lock explicitly itself
        (rather than assuming the caller already did) so it remains
        correct even if ever reached via a path that hasn't already
        locked the session."""
        ctx = get_adapter_invocation_context()
        summary = plan.summary
        rows_data = summary.get("rows", [])

        session_lock_stmt = select(ImportSession.id).where(ImportSession.id == ctx.import_session_id)
        if db.get_bind().dialect.name == "postgresql":
            session_lock_stmt = session_lock_stmt.with_for_update()
        await db.execute(session_lock_stmt)

        await import_dry_run_plan_crud.supersede_active_plan(db, import_session_id=ctx.import_session_id)
        plan_row = await import_dry_run_plan_crud.insert_plan(
            db,
            import_session_id=ctx.import_session_id,
            import_source_id=ctx.import_source_id,
            source_checksum=ctx.source_checksum,
            accepted_validation_job_id=ctx.accepted_validation_job_id,
            dry_run_job_id=ctx.dry_run_job_id,
            ruleset_version=ctx.ruleset_version,
            summary_total_rows=summary["total_rows"],
            summary_creates=summary["creates"],
            summary_updates=summary["updates"],
            summary_skips=summary["skips"],
            summary_warnings=summary["warnings"],
            summary_blocking_conflicts=summary["blocking_conflicts"],
        )

        row_models = [
            EquipmentMasterDryRunPlanRow(
                dry_run_plan_id=plan_row.id,
                source_row_number=row["source_row_number"],
                action=row["action"],
                target_equipment_id=row["target_equipment_id"],
                normalized_values=row["normalized_values"],
                matched_identity_fields=row["matched_identity_fields"],
                expected_equipment_version=row["expected_equipment_version"],
                warnings=row["warnings"],
            )
            for row in rows_data
        ]
        await import_dry_run_plan_crud.bulk_insert_plan_rows(db, row_models)

        # §26: audit references identity/summary only -- never raw source
        # values (no row content, no equipment_name/brand/serial_number).
        await record_audit_event(
            db,
            actor_user_id=ctx.actor_user_id,
            action=AUDIT_ACTION_IMPORT_DRY_RUN_PLAN_CREATED,
            entity_type=AUDIT_ENTITY_IMPORT_SESSION,
            entity_id=ctx.import_session_id,
            after={
                "dry_run_plan_id": str(plan_row.id),
                "import_source_id": str(ctx.import_source_id),
                "source_checksum": ctx.source_checksum,
                "summary_total_rows": plan_row.summary_total_rows,
                "summary_creates": plan_row.summary_creates,
                "summary_updates": plan_row.summary_updates,
                "summary_skips": plan_row.summary_skips,
                "summary_warnings": plan_row.summary_warnings,
                "summary_blocking_conflicts": plan_row.summary_blocking_conflicts,
            },
        )

    async def precheck_execute(self, db: AsyncSession) -> None:
        """Roadmap PR20E (design §14.4a). Called by the framework
        **before** `admit_phase_job` runs -- a read-only rejection, no
        session state touched. Verifies the session has an `active`,
        **confirmed** plan; if not (never confirmed, or confirmed but
        since superseded by a newer, not-yet-confirmed dry-run), raises
        `ImportNoConfirmedPlanError` (`409 IMPORT_NO_CONFIRMED_PLAN`) --
        a cheap, retryable rejection, never a terminal execute failure."""
        ctx = get_adapter_invocation_context()
        plan = await import_dry_run_plan_crud.get_active_confirmed_for_session(
            db, import_session_id=ctx.import_session_id
        )
        if plan is None:
            raise ImportNoConfirmedPlanError(
                f"Import session '{ctx.import_session_id}' has no confirmed dry-run plan to execute. "
                "Confirm the current plan (POST .../dry-run-plan/{plan_id}/confirm) first."
            )

    async def execute(self, db: AsyncSession) -> int:
        """Roadmap PR20E (design §14.4, §14.4a, §15, §15.1). Resolves the
        session's `active`, confirmed plan internally -- **no plan
        identity is threaded through the invocation context** (design
        §14.4, fix round 4 H6) -- re-running the identical query
        `precheck_execute` already used. The two calls are separated only
        by `admit_phase_job`'s own atomic admission, so a `None` result
        here indicates a framework-invariant violation, never a
        client-correctable input error (§14.4).

        Applies every plan row's already-persisted action, exactly as
        planned: `CREATE` via a plain insert guarded by Equipment's
        existing unique constraints (§16) -- a row lacking a genuine,
        non-fabricated `asset_number` (§9 OD-4) or an approved
        `status` mapping (§9 OD-2, never `ISSUED_TO_WARD` -- creating
        Equipment directly into that state without a `BorrowTransaction`
        would violate the ledger invariant, so this is refused exactly
        like every other unauthorized CREATE) is refused, never
        fabricated. `UPDATE` via the row's own persisted
        `expected_equipment_version` compare-and-swap (§15.1) -- never a
        freshly re-read version. A same-value UPDATE (no writable field
        actually differs from the record's current value) is a no-op:
        no write, no version bump (§9 OD-2, reusing `equipment_crud.
        update()`'s own empty-`data` guard).

        Any conflict -- missing plan, source-binding mismatch, a CREATE
        row with no authoritative `asset_number`/status, a unique-
        constraint violation, a missing UPDATE target, or a stale
        `expected_equipment_version` -- raises `AdapterExecutionConflict`
        carrying the resolved plan's own id, so the framework's TX2
        failure path can mark the plan `failed` (§14.4b) using only that
        primitive. Never calls `db.commit()`/`db.rollback()` itself --
        the caller's TX1 owns the boundary (§15): if any row conflicts,
        the entire attempt rolls back, never partially applying the
        other, non-conflicting rows. On success, marks the plan
        `consumed` in the same transaction and returns the count of rows
        that resulted in a genuine Equipment mutation (skips and no-op
        updates are not "imported")."""
        ctx = get_adapter_invocation_context()
        plan = await import_dry_run_plan_crud.get_active_confirmed_for_session(
            db, import_session_id=ctx.import_session_id
        )
        if plan is None:
            raise AdapterExecutionConflict(
                f"Import session '{ctx.import_session_id}' has no confirmed active plan at execute time "
                "despite precheck_execute succeeding -- a framework-invariant violation.",
                resolved_resource_id=None,
            )
        # Captured as a primitive immediately, before any write is
        # attempted below: on PostgreSQL, a failed flush deactivates the
        # whole Session's transaction (an aborted transaction rejects any
        # further command), which implicitly expires every already-loaded
        # object's attributes -- so `plan.id` itself becomes unsafe to
        # read from an exception handler after that point (it would
        # trigger a reload that raises `PendingRollbackError`, masking
        # the real conflict being reported). Every conflict raised below
        # uses this primitive, never a live `plan.id` attribute access.
        plan_id = plan.id
        # Roadmap PR20E fix round 2 (PR #95 review, §H2): recorded
        # unconditionally, before any mutation is attempted, so the
        # framework can retain this primitive for the entire remainder
        # of TX1 -- including a failure that happens *after* this
        # method has already returned successfully (fenced completion,
        # `on_execution_success` below, audit, commit) -- and for TX2
        # failure publication, not only when this method itself raises
        # `AdapterExecutionConflict`.
        record_resolved_execution_resource(plan_id)
        # §14.4: defensive, post-admission re-check -- structurally
        # guaranteed by §6.2's immutable-once-frozen source, but checked,
        # not assumed.
        if plan.import_source_id != ctx.import_source_id or plan.source_checksum != ctx.source_checksum:
            raise AdapterExecutionConflict(
                f"Persisted plan '{plan_id}' source binding does not match the current session's frozen source.",
                resolved_resource_id=plan_id,
            )

        try:
            imported_rows = await self._apply_plan_rows(db, plan_id)
        except AdapterExecutionConflict:
            raise
        except Exception as exc:
            # Any other exception while a live worker is applying a known
            # plan must still mark that plan `failed` (§14.4: "it must
            # never be left `active` once its owning session has
            # terminally failed") -- only a genuine hard crash (the
            # process dies mid-transaction, never reaching this handler
            # at all) falls through to §14.4c's on_execution_recovery gap.
            raise AdapterExecutionConflict(
                f"Plan '{plan_id}' execution failed: {exc}", resolved_resource_id=plan_id
            ) from exc

        # Roadmap PR20E fix round 2 (PR #95 review, §H1): plan
        # consumption deliberately does NOT happen here -- doing so
        # would lock the Plan row *before* the framework has obtained
        # its own Job -> Session completion fence, the exact Plan ->
        # Job -> Session lock-order violation PR95-H1 identified
        # (recovery, and this adapter's own `on_execution_failure`, both
        # acquire Job -> Session -> Plan). `on_execution_success` below
        # -- called by the framework only *after* `fenced_phase_success`
        # has already succeeded -- performs that transition instead, in
        # the same TX1, under the correct global lock order.
        return imported_rows

    async def _apply_plan_rows(self, db: AsyncSession, plan_id: uuid.UUID) -> int:
        rows = await import_dry_run_plan_crud.list_all_plan_rows(db, plan_id=plan_id)

        update_target_ids = {row.target_equipment_id for row in rows if row.action == "UPDATE"}
        # PR95 review fix round, §H3: `for_update=True` locks every
        # UPDATE target (ordered by id, so two concurrent executions
        # touching overlapping rows always acquire in the same order)
        # for the rest of this transaction -- `_apply_update_row` below
        # relies on this same snapshot to validate freshness *before*
        # ever treating a row as a no-op, not only when it performs a
        # genuine CAS write.
        current_by_id = (
            await equipment_crud.get_by_ids(db, sorted(update_target_ids), for_update=True)
            if update_target_ids
            else {}
        )

        imported_rows = 0
        for row in rows:
            if row.action == "SKIP":
                continue
            if row.action == "CREATE":
                imported_rows += await self._apply_create_row(db, plan_id, row)
            elif row.action == "UPDATE":
                imported_rows += await self._apply_update_row(db, plan_id, row, current_by_id)
        return imported_rows

    async def _apply_create_row(self, db: AsyncSession, plan_id: uuid.UUID, row: EquipmentMasterDryRunPlanRow) -> int:
        # Captured as a primitive before the `create` flush below, for the
        # same reason `plan_id` is captured in `execute()` -- once that
        # flush fails with `IntegrityError`, PostgreSQL has aborted the
        # whole transaction, and re-reading a live ORM attribute (even one
        # already loaded, like `row.id`) requires a reload that raises
        # `PendingRollbackError` instead, masking the real conflict.
        row_id = row.id
        values = row.normalized_values or {}
        # §9 OD-4: this codebase's own PR20C parser never persists this
        # key today (no authoritative source exists under the current
        # 32-column contract, §9 OD-4) -- kept as a real, enforced check
        # (never assumed unreachable) so a future, richer source contract
        # can light this path up safely, and so a plan row that somehow
        # reached this point despite upstream validation fails closed
        # rather than fabricating a placeholder.
        asset_number = values.get("asset_number")
        if not asset_number:
            raise AdapterExecutionConflict(
                f"Plan row {row_id} is a CREATE action with no authoritative asset_number -- refusing to "
                "fabricate one.",
                resolved_resource_id=plan_id,
            )
        status_raw = values.get("status")
        try:
            status = EquipmentStatus(status_raw) if status_raw is not None else None
        except ValueError:
            status = None
        # §9 OD-2 Legacy Lifecycle Policy / §22: ISSUED_TO_WARD is
        # deliberately excluded here -- creating Equipment directly into
        # that state with no corresponding BorrowTransaction would
        # violate the invariant that ISSUED_TO_WARD equipment always has
        # an active transaction. STATUS_MAPPING (above) already never
        # produces this value for a real parsed row; this is enforced
        # again here, defensively, for any row reaching execute() by
        # another path.
        if status not in _CREATE_ALLOWED_STATUSES:
            raise AdapterExecutionConflict(
                f"Plan row {row_id} is a CREATE action with an unauthorized or missing status value.",
                resolved_resource_id=plan_id,
            )

        create_data = {
            "asset_number": asset_number,
            "item_no": values.get("item_no"),
            "bcm_code": values.get("bcm_code"),
            "equipment_name": values.get("equipment_name"),
            "brand": values.get("brand"),
            "model": values.get("model"),
            "serial_number": values.get("serial_number"),
            "asset_id": values.get("asset_id"),
            "status": status,
            # §9 OD-2 Location Policy: category_id/department_owner_id/
            # current_location_id are deliberately omitted -- left at
            # the model's own NULL default, never populated on CREATE
            # (no approved legacy-name-to-UUID matching algorithm).
        }
        try:
            await equipment_crud.create(db, data=create_data)
        except IntegrityError as exc:
            kind = classify_integrity_error(exc)
            if kind is not IntegrityViolationKind.UNIQUE:
                raise
            raise AdapterExecutionConflict(
                f"Plan row {row_id}'s CREATE conflicts with an existing Equipment identifier.",
                resolved_resource_id=plan_id,
            ) from exc
        return 1

    async def _apply_update_row(
        self, db: AsyncSession, plan_id: uuid.UUID, row: EquipmentMasterDryRunPlanRow, current_by_id: dict
    ) -> int:
        row_id = row.id
        target_equipment_id = row.target_equipment_id
        current = current_by_id.get(target_equipment_id)
        if current is None:
            raise AdapterExecutionConflict(
                f"Plan row {row_id}'s target Equipment '{target_equipment_id}' no longer exists.",
                resolved_resource_id=plan_id,
            )
        # PR95 review fix round, §H3: freshness MUST be validated before
        # a no-op decision is ever made -- "no-op" means "no version
        # bump", never "skip the optimistic-concurrency check". Without
        # this, a plan whose writable descriptive fields happen to
        # already match the row's *current* values (e.g. another
        # legitimate mutation changed a protected field, or changed and
        # then reverted a writable one) would be silently accepted and
        # the plan consumed, even though the row is genuinely stale
        # relative to what this plan was dry-run against. `current` was
        # loaded (and, on PostgreSQL, row-locked) in the SAME transaction
        # by `_apply_plan_rows`'s bulk `get_by_ids(..., for_update=True)`
        # call -- never re-read here, so this comparison is against the
        # exact snapshot the rest of this row's processing also uses.
        current_version = current.version
        expected_version = row.expected_equipment_version
        if current_version != expected_version:
            raise AdapterExecutionConflict(
                f"Plan row {row_id}'s target Equipment '{target_equipment_id}' has changed since the "
                "dry-run -- expected_equipment_version no longer matches. A new dry-run is required.",
                resolved_resource_id=plan_id,
            )

        candidate = row.normalized_values or {}
        # §9 OD-2: only the approved UPDATE-writable fields are ever
        # considered -- protected fields (identity, version, lifecycle/
        # status, current operational location) are never in this set,
        # so they can never appear in `diff` below regardless of what
        # `normalized_values` happens to contain.
        diff = {
            field_name: value
            for field_name, value in candidate.items()
            if field_name in _UPDATE_WRITABLE_FIELDS and getattr(current, field_name) != value
        }
        if not diff:
            # §9 OD-2 / PR91-H1: a same-value update is a no-op -- no
            # write, no version bump. Freshness has already been
            # validated above -- this is a genuine no-op, never a stale
            # plan silently accepted.
            return 0

        updated = await equipment_crud.update_with_cas(
            db,
            equipment_id=target_equipment_id,
            expected_version=expected_version,
            data=diff,
        )
        if updated is None:
            raise AdapterExecutionConflict(
                f"Plan row {row_id}'s target Equipment '{target_equipment_id}' has changed since the "
                "dry-run -- expected_equipment_version no longer matches. A new dry-run is required.",
                resolved_resource_id=plan_id,
            )
        return 1

    async def on_execution_success(self, db: AsyncSession, resolved_resource_id) -> None:
        """Roadmap PR20E fix round 2 (PR #95 review, §H1). Called by the
        framework on TX1's own session, only *after*
        `fenced_phase_success` has already locked Job then Session --
        marks the plan `consumed` here, never inside `execute()` itself,
        so the Plan row is only ever touched after that fence succeeds
        (Job -> Session -> Plan, the same order recovery and
        `on_execution_failure` below use). `resolved_resource_id` is the
        plan's own id, recorded by `execute()` via
        `record_resolved_execution_resource` -- `None` only for the
        framework-invariant-violation case where no plan could be
        resolved at all (harmless no-op)."""
        if resolved_resource_id is None:
            return
        await import_dry_run_plan_crud.mark_plan_consumed(db, plan_id=resolved_resource_id)

    async def on_execution_failure(self, db: AsyncSession, resolved_resource_id) -> None:
        """Roadmap PR20E (design §14.4b). Called by the framework inside
        TX2, only *after* `fenced_phase_failure` has already succeeded
        (fix round 2, §H1: Job -> Session -> Plan, the same order
        `on_execution_success` above and recovery use -- a worker whose
        own failure publication was itself fenced out must never touch
        this plan; whichever path actually won the fence owns that
        transition instead). Called whenever `execute()` raised,
        regardless of whether it raised `AdapterExecutionConflict`
        itself or the failure happened later in the framework's own
        TX1 work (fix round 2, §H2) -- `resolved_resource_id` is always
        the value `execute()` recorded via
        `record_resolved_execution_resource`, or `None` for the
        framework-invariant-violation case where no plan could be
        resolved at all (harmless no-op)."""
        if resolved_resource_id is None:
            return
        await import_dry_run_plan_crud.mark_plan_failed(db, plan_id=resolved_resource_id)

    async def on_execution_recovery(self, db: AsyncSession, session_id) -> None:
        """Roadmap PR20E (design §14.4c). Called by `recover_session()`
        only when the recovered job's `job_type` was `'execute'` --
        reconciles the case §14.4b's exception-based hook cannot cover: a
        hard worker crash that never raised anything at all."""
        await import_dry_run_plan_crud.mark_active_plan_failed_for_session(db, import_session_id=session_id)


# §9 OD-2 Legacy Lifecycle Policy / §22: ISSUED_TO_WARD is deliberately
# excluded -- see EquipmentMasterAdapter._apply_create_row's docstring.
_CREATE_ALLOWED_STATUSES = frozenset(
    {EquipmentStatus.AVAILABLE_AT_POOL, EquipmentStatus.UNAVAILABLE_DEFECTIVE, EquipmentStatus.DECOMMISSIONED}
)

# §9 OD-2: the exact UPDATE-writable field list, mirroring
# `_update_normalized_values`'s own key set above -- never bcm_code/
# item_no (identity), status (lifecycle), current_location_id/
# department_owner_id (operational location), or version (handled only
# by `equipment_crud.update_with_cas`'s own CAS predicate).
_UPDATE_WRITABLE_FIELDS = frozenset({"equipment_name", "brand", "model", "serial_number", "asset_id"})


register_adapter(EquipmentMasterAdapter())

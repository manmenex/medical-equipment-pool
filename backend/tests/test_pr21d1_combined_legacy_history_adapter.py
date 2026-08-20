"""Roadmap PR21D1 -- Combined Canonical Adapter + Source Admission.

Authoritative design: this PR's own `combined.py` module docstring plus
`docs/design/PR21_LEGACY_TRANSACTION_HISTORY_IMPORT_PLAN.md` and Owner
Decision Closure Round 3 (GitHub PR #106). Tests the registered
`LegacyTransactionHistoryAdapter` (`app.services.import_adapters.
legacy_history.combined`) directly at the adapter-method level --
`parse`/`preload_business_context`/`validate_business_rules`/
`plan_dry_run` -- mirroring `test_pr20d_dry_run_plan.py`'s own direct-
adapter-invocation conventions for `plan_dry_run`. HTTP-level
source-admission/persistence/execution-block tests live in the sibling
file `test_pr21d1_source_admission.py`.

Synthetic workbooks only -- no production workbook data is read or
committed anywhere in this file."""

import hashlib
import uuid
from datetime import date, time
from io import BytesIO

import pytest
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import InvalidInputError
from app.models.equipment import Equipment, EquipmentStatus
from app.models.legacy_history import LegacyMigrationAuthority
from app.models.master_data import Ward
from app.models.transaction import BorrowTransaction
from app.models.user import User
from app.services.identifiers import normalize_bcm_code
from app.services.import_adapter import ImportAdapter, get_adapter
from app.services.import_adapter_context import AdapterInvocationContext, adapter_invocation_context
from app.services.import_adapters.legacy_history import common as common_module
from app.services.import_adapters.legacy_history import issue as issue_module
from app.services.import_adapters.legacy_history import receive as receive_module
from app.services.import_adapters.legacy_history.combined import LegacyTransactionHistoryAdapter
from app.services.import_plan_providers.legacy_history import DATASET_TYPE, LegacyHistoryDryRunPlanProvider
from app.services.import_source_reader import SourceDescriptor, VerifiedSourceContent

# pytest.ini sets asyncio_mode = auto -- async test functions below need no
# explicit marker (mirrors test_pr21a/test_pr21b/test_pr21c's own convention).

ISSUE_HEADER_SHEET = issue_module.HEADER_SHEET_NAME
ISSUE_LINE_SHEET = issue_module.LINE_SHEET_NAME
RECEIVE_HEADER_SHEET = receive_module.HEADER_SHEET_NAME
RECEIVE_LINE_SHEET = receive_module.LINE_SHEET_NAME

ISSUE_HEADER_COLUMNS = list(issue_module._HEADER_GOVERNED_HEADERS.keys())
ISSUE_LINE_COLUMNS = list(issue_module._LINE_GOVERNED_HEADERS.keys())
RECEIVE_HEADER_COLUMNS = list(receive_module._HEADER_GOVERNED_HEADERS.keys())
RECEIVE_LINE_COLUMNS = list(receive_module._LINE_GOVERNED_HEADERS.keys())


# ---------------------------------------------------------------------------
# Workbook construction helpers.
# ---------------------------------------------------------------------------


def _build_combined_workbook(
    *,
    issue_headers: list[dict] | None = None,
    issue_lines: list[dict] | None = None,
    receive_headers: list[dict] | None = None,
    receive_lines: list[dict] | None = None,
    extra_sheets: dict[str, list] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = ISSUE_HEADER_SHEET
    ws.append(ISSUE_HEADER_COLUMNS)
    for row in issue_headers or []:
        ws.append([row.get(col) for col in ISSUE_HEADER_COLUMNS])

    ws = wb.create_sheet(ISSUE_LINE_SHEET)
    ws.append(ISSUE_LINE_COLUMNS)
    for row in issue_lines or []:
        ws.append([row.get(col) for col in ISSUE_LINE_COLUMNS])

    ws = wb.create_sheet(RECEIVE_HEADER_SHEET)
    ws.append(RECEIVE_HEADER_COLUMNS)
    for row in receive_headers or []:
        ws.append([row.get(col) for col in RECEIVE_HEADER_COLUMNS])

    ws = wb.create_sheet(RECEIVE_LINE_SHEET)
    ws.append(RECEIVE_LINE_COLUMNS)
    for row in receive_lines or []:
        ws.append([row.get(col) for col in RECEIVE_LINE_COLUMNS])

    if extra_sheets:
        for name, rows in extra_sheets.items():
            ws = wb.create_sheet(name)
            for row in rows:
                ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _issue_header_row(*, order_ref="ORD-0001", ward="Ward 1", bme="BME Header", d=date(2024, 1, 10), t=time(9, 0, 0), notes=None):
    return {
        "วันที่": d,
        "เลขที่ใบยืม": order_ref,
        "แผนกที่ยืม": ward,
        "ผู้ส่งเครื่องยืม (User)": "header-user",
        "ผู้ส่งเครื่องยืม (BME)": bme,
        "เวลา": t,
        "จำนวนเครื่องส่งทั้งหมด": 1,
        "หมายเหตุ": notes,
    }


def _issue_line_row(*, row_key="1", order_ref="ORD-0001", me_code="ME001", ward="Ward 1", bme="BME Line", d=date(2024, 1, 10), t=time(9, 0, 0), notes=None):
    return {
        "ลำดับ": row_key,
        "วันที่": d,
        "เลขที่ใบส่ง": order_ref,
        "SCAN CODE ส่ง": "SCAN-1",
        "ME.Code": me_code,
        "Barcode ส่งเครื่อง": "BARCODE-1",
        "Equipment": "Infusion Pump",
        "Brand": "BrandX",
        "Model": "ModelY",
        "Serial no.": "SN-1",
        "รูปเครื่อง": None,
        "แผนกที่ส่ง": ward,
        "ตัวเครื่องหน้าจอไม่แตกร้าว": "OK",
        "สายชาร์จ (สายไม่ขาด)": "OK",
        "Pole Clamp (ไม่แตกหักมไม่ง้อ)": "OK",
        "จำนวน": 1,
        "หมายเหตุ": notes,
        "เวลา": t,
        "ชื่อ BME": bme,
        "ชื่อ (User)": "line-user",
    }


def _receive_header_row(*, order_ref="RET-0001", ward="Ward 1", bme="BME Header", d=date(2024, 1, 11), t=time(10, 0, 0), notes=None):
    return {
        "วันที่": d,
        "เลขที่ใบคืน": order_ref,
        "แผนกที่คืน": ward,
        "ผู้ส่งเครื่องคืน (User)": "header-user",
        "ผู้รับเครื่องคืน (BME)": bme,
        "เวลา": t,
        "จำนวนเครื่องรับคืนทั้งหมด": 1,
        "หมายเหตุ": notes,
    }


def _receive_line_row(*, row_key="1", order_ref="RET-0001", me_code="ME001", ward="Ward 1", bme="BME Line", d=date(2024, 1, 11), t=time(10, 0, 0), notes=None):
    return {
        "ลำดับ": row_key,
        "วันที่": d,
        "เลขที่ใบรับเครื่อง": order_ref,
        "SCAN CODE รับ": "SCAN-1",
        "ME.Code": me_code,
        "Barcode รับเครื่อง": "BARCODE-1",
        "Equipment": "Infusion Pump",
        "Brand": "BrandX",
        "Model": "ModelY",
        "Serial no.": "SN-1",
        "แผนกที่รับ": ward,
        "รูปเครื่อง": None,
        "ตัวเครื่องหน้าจอไม่แตกร้าว": "OK",
        "สายชาร์จ (สายไม่ขาด)": "OK",
        "Pole Clamp (ไม่แตกหักมไม่ง้อ)": "OK",
        "จำนวน": 1,
        "หมายเหตุ": notes,
        "เวลา": t,
        "ชื่อ BME": bme,
        "ชื่อ (User)": "line-user",
    }


def _default_workbook(**overrides) -> bytes:
    kwargs = dict(
        issue_headers=[_issue_header_row()],
        issue_lines=[_issue_line_row()],
        receive_headers=[_receive_header_row()],
        receive_lines=[_receive_line_row()],
    )
    kwargs.update(overrides)
    return _build_combined_workbook(**kwargs)


async def _seed_equipment(db_session: AsyncSession, *, me_code: str = "ME001", **kwargs) -> Equipment:
    defaults = dict(
        asset_number=f"AN-{uuid.uuid4().hex[:10]}",
        equipment_name="Legacy Test Equipment",
        status=EquipmentStatus.AVAILABLE_AT_POOL,
        bcm_code=normalize_bcm_code(me_code),
    )
    defaults.update(kwargs)
    eq = Equipment(**defaults)
    db_session.add(eq)
    await db_session.commit()
    await db_session.refresh(eq)
    return eq


async def _seed_ward(db_session: AsyncSession, *, code: str = "Ward 1", name: str | None = None) -> Ward:
    ward = Ward(code=code, name=name or code)
    db_session.add(ward)
    await db_session.commit()
    await db_session.refresh(ward)
    return ward


async def _seed_authority(db_session: AsyncSession, *, checksum: str, actor_id: uuid.UUID) -> LegacyMigrationAuthority:
    authority = LegacyMigrationAuthority(
        scope="pr21d1-test", approved_workbook_sha256=checksum, approved_by_user_id=actor_id
    )
    db_session.add(authority)
    await db_session.commit()
    await db_session.refresh(authority)
    return authority


async def _get_actor_id(db_session: AsyncSession) -> uuid.UUID:
    row = (await db_session.execute(select(User.id).limit(1))).first()
    return row[0]


def _verified_content(content: bytes) -> VerifiedSourceContent:
    checksum = hashlib.sha256(content).hexdigest()
    session_id = uuid.uuid4()
    source_id = uuid.uuid4()
    return VerifiedSourceContent(
        content=content,
        source_descriptor=SourceDescriptor(
            import_source_id=source_id,
            import_session_id=session_id,
            dataset_type=DATASET_TYPE,
            expected_checksum=checksum,
            expected_byte_size=len(content),
            content_type=None,
            original_filename=None,
            registration_status="frozen",
        ),
    )


def _invocation_context(content: bytes, *, checksum: str | None = None) -> AdapterInvocationContext:
    verified = _verified_content(content)
    return AdapterInvocationContext(
        import_session_id=verified.source_descriptor.import_session_id,
        import_source_id=verified.source_descriptor.import_source_id,
        dataset_type=DATASET_TYPE,
        source_checksum=checksum or hashlib.sha256(content).hexdigest(),
        source_fingerprint="fp",
        ruleset_version="1",
        verified_source_content=verified,
        dry_run_job_id=uuid.uuid4(),
        accepted_validation_job_id=uuid.uuid4(),
        actor_user_id=None,
    )


async def _run_validate(adapter: LegacyTransactionHistoryAdapter, db_session: AsyncSession, content: bytes):
    """Mirrors what `import_validation_service.run_validation` does with
    an adapter's `parse`/`preload_business_context`/`validate_business_rules`
    -- exercised directly here since this file tests the adapter, not the
    full session/job pipeline (that's `test_pr21d1_source_admission.py`)."""
    raw_records = adapter.parse(_verified_content(content))
    context = await adapter.preload_business_context(db_session, raw_records)
    findings_by_row = {rec.row_number: adapter.validate_business_rules(rec, context) for rec in raw_records}
    return raw_records, findings_by_row


# ---------------------------------------------------------------------------
# A. Registration + execution-block guard.
# ---------------------------------------------------------------------------


def test_adapter_registered_for_legacy_transaction_history():
    adapter = get_adapter(DATASET_TYPE)
    assert adapter is not None
    assert isinstance(adapter, LegacyTransactionHistoryAdapter)


def test_execute_not_overridden_stays_structurally_blocked():
    adapter = get_adapter(DATASET_TYPE)
    assert type(adapter).execute is ImportAdapter.execute
    assert type(adapter).precheck_execute is ImportAdapter.precheck_execute


# ---------------------------------------------------------------------------
# B. parse(): flattening, defense-in-depth checks.
# ---------------------------------------------------------------------------


def test_parse_rejects_non_verified_source_content():
    adapter = LegacyTransactionHistoryAdapter()
    with pytest.raises(InvalidInputError):
        adapter.parse(b"not-a-verified-source-content")


def test_parse_rejects_content_over_pr21_upload_cap(monkeypatch):
    adapter = LegacyTransactionHistoryAdapter()
    monkeypatch.setattr(common_module, "PR21_MAX_UPLOAD_BYTES", 10)
    content = _default_workbook()
    assert len(content) > 10
    with pytest.raises(InvalidInputError):
        adapter.parse(_verified_content(content))


def test_parse_flattens_all_four_sheets_with_synthetic_row_numbers():
    adapter = LegacyTransactionHistoryAdapter()
    content = _default_workbook()
    records = adapter.parse(_verified_content(content))
    assert len(records) == 4
    kinds = [r.fields["_kind"] for r in records]
    assert kinds == ["issue_header", "issue_line", "receive_header", "receive_line"]
    row_numbers = [r.row_number for r in records]
    assert row_numbers == [1, 2, 3, 4], "synthetic row numbers must be globally unique across all four sheets"
    for rec in records:
        # Row 1 is the header row on every sheet (openpyxl-native, 1-based
        # including the header row) -- each sheet's own first *data* row
        # is therefore source_row_number 2, not 1.
        assert rec.fields["_source_row_number"] == 2


def test_parse_ignores_sdc_and_unknown_extra_sheets():
    content = _default_workbook(extra_sheets={"SDC Sheet": [["ignored"]], "Some Report": [["also ignored"]]})
    adapter = LegacyTransactionHistoryAdapter()
    records = adapter.parse(_verified_content(content))
    assert len(records) == 4, "SDC/unknown sheets must not be parsed, counted, or cause rejection"


def test_parse_missing_canonical_sheet_raises():
    wb = Workbook()
    ws = wb.active
    ws.title = ISSUE_HEADER_SHEET
    ws.append(ISSUE_HEADER_COLUMNS)
    ws = wb.create_sheet(ISSUE_LINE_SHEET)
    ws.append(ISSUE_LINE_COLUMNS)
    ws = wb.create_sheet(RECEIVE_HEADER_SHEET)
    ws.append(RECEIVE_HEADER_COLUMNS)
    # RECEIVE_LINE_SHEET deliberately omitted.
    buf = BytesIO()
    wb.save(buf)
    adapter = LegacyTransactionHistoryAdapter()
    with pytest.raises(InvalidInputError):
        adapter.parse(_verified_content(buf.getvalue()))


@pytest.mark.parametrize("extra_sheet_count,should_raise", [(24, False), (28, False), (29, True)])
def test_worksheet_cap_boundary(extra_sheet_count, should_raise):
    """4 canonical sheets + N extra sheets, at/around
    `common.PR21_MAX_WORKSHEET_COUNT` (32): 28 extra -> 32 total (at cap,
    allowed); 29 extra -> 33 total (over cap, rejected)."""
    extra_sheets = {f"Extra{i}": [["x"]] for i in range(extra_sheet_count)}
    content = _default_workbook(extra_sheets=extra_sheets)
    adapter = LegacyTransactionHistoryAdapter()
    if should_raise:
        with pytest.raises(InvalidInputError):
            adapter.parse(_verified_content(content))
    else:
        records = adapter.parse(_verified_content(content))
        assert len(records) == 4


def test_malformed_file_rejected_not_a_zip():
    adapter = LegacyTransactionHistoryAdapter()
    with pytest.raises(InvalidInputError):
        adapter.parse(_verified_content(b"this is not a valid xlsx file at all"))


# ---------------------------------------------------------------------------
# C. preload_business_context()/validate_business_rules(): combined
#    all-or-nothing validate-phase findings.
# ---------------------------------------------------------------------------


async def test_combined_success_both_sides_valid_produces_no_findings(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    adapter = LegacyTransactionHistoryAdapter()
    content = _default_workbook()
    _records, findings_by_row = await _run_validate(adapter, db_session, content)
    all_findings = [f for findings in findings_by_row.values() for f in findings]
    assert all_findings == []


async def test_issue_side_failure_produces_findings_only_for_issue_rows(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    adapter = LegacyTransactionHistoryAdapter()
    # Issue line references a ME.Code that resolves to no Equipment ->
    # LEGACY_ISSUE_EQUIPMENT_NOT_FOUND. Receive side stays fully valid.
    content = _default_workbook(issue_lines=[_issue_line_row(me_code="ME_DOES_NOT_EXIST")])
    records, findings_by_row = await _run_validate(adapter, db_session, content)

    issue_line_rec = next(r for r in records if r.fields["_kind"] == "issue_line")
    receive_line_rec = next(r for r in records if r.fields["_kind"] == "receive_line")
    assert findings_by_row[issue_line_rec.row_number] != []
    assert findings_by_row[receive_line_rec.row_number] == []
    error_codes = {f.error_code for f in findings_by_row[issue_line_rec.row_number]}
    assert issue_module.CODE_EQUIPMENT_NOT_FOUND in error_codes


async def test_receive_side_failure_produces_findings_only_for_receive_rows(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    adapter = LegacyTransactionHistoryAdapter()
    content = _default_workbook(receive_lines=[_receive_line_row(me_code="ME_DOES_NOT_EXIST")])
    records, findings_by_row = await _run_validate(adapter, db_session, content)

    issue_line_rec = next(r for r in records if r.fields["_kind"] == "issue_line")
    receive_line_rec = next(r for r in records if r.fields["_kind"] == "receive_line")
    assert findings_by_row[issue_line_rec.row_number] == []
    assert findings_by_row[receive_line_rec.row_number] != []
    error_codes = {f.error_code for f in findings_by_row[receive_line_rec.row_number]}
    assert receive_module.CODE_EQUIPMENT_NOT_FOUND in error_codes


async def test_both_sides_failure_produces_findings_on_both(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    adapter = LegacyTransactionHistoryAdapter()
    content = _default_workbook(
        issue_lines=[_issue_line_row(me_code="ME_MISSING_A")],
        receive_lines=[_receive_line_row(me_code="ME_MISSING_B")],
    )
    records, findings_by_row = await _run_validate(adapter, db_session, content)
    issue_line_rec = next(r for r in records if r.fields["_kind"] == "issue_line")
    receive_line_rec = next(r for r in records if r.fields["_kind"] == "receive_line")
    assert findings_by_row[issue_line_rec.row_number] != []
    assert findings_by_row[receive_line_rec.row_number] != []


async def test_finding_field_and_message_carry_sheet_identity(db_session: AsyncSession):
    """§15 of the task: `ImportRowError` has no `sheet_name` column, so
    sheet identity must survive in `field`/`message` instead."""
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    adapter = LegacyTransactionHistoryAdapter()
    content = _default_workbook(issue_lines=[_issue_line_row(me_code="ME_DOES_NOT_EXIST")])
    records, findings_by_row = await _run_validate(adapter, db_session, content)
    issue_line_rec = next(r for r in records if r.fields["_kind"] == "issue_line")
    finding = findings_by_row[issue_line_rec.row_number][0]
    assert finding.field is not None and finding.field.startswith(f"{ISSUE_LINE_SHEET}:")
    assert finding.message.startswith(f"{ISSUE_LINE_SHEET} row 2:")


async def test_sdc_and_extra_sheets_present_do_not_affect_validation(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    adapter = LegacyTransactionHistoryAdapter()
    content = _default_workbook(extra_sheets={"SDC": [["x"]]})
    _records, findings_by_row = await _run_validate(adapter, db_session, content)
    all_findings = [f for findings in findings_by_row.values() for f in findings]
    assert all_findings == []


# ---------------------------------------------------------------------------
# D. plan_dry_run(): checksum/authority gate, candidate construction,
#    all-or-nothing re-validation, no mutation, no pairing, no notes.
# ---------------------------------------------------------------------------


async def test_plan_dry_run_requires_migration_authority(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    adapter = LegacyTransactionHistoryAdapter()
    content = _default_workbook()
    ctx = _invocation_context(content)
    with adapter_invocation_context(ctx):
        with pytest.raises(InvalidInputError):
            await adapter.plan_dry_run(db_session)


async def test_plan_dry_run_rejects_checksum_mismatch_authority(db_session: AsyncSession, seeded_users):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    actor_id = await _get_actor_id(db_session)
    await _seed_authority(db_session, checksum="f" * 64, actor_id=actor_id)
    adapter = LegacyTransactionHistoryAdapter()
    content = _default_workbook()
    ctx = _invocation_context(content)  # real content checksum != "f" * 64
    with adapter_invocation_context(ctx):
        with pytest.raises(InvalidInputError):
            await adapter.plan_dry_run(db_session)


async def test_plan_dry_run_success_builds_candidates_for_both_sides(db_session: AsyncSession, seeded_users):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    actor_id = await _get_actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    equipment_before = (await db_session.execute(select(Equipment))).scalars().all()
    borrow_before = (await db_session.execute(select(BorrowTransaction))).scalars().all()
    users_before = (await db_session.execute(select(User))).scalars().all()

    adapter = LegacyTransactionHistoryAdapter()
    ctx = _invocation_context(content, checksum=checksum)
    with adapter_invocation_context(ctx):
        plan = await adapter.plan_dry_run(db_session)

    assert plan.summary["issue_events"] == 1
    assert plan.summary["receive_events"] == 1
    rows = plan.summary["rows"]
    assert len(rows) == 2
    event_types = {r["event_type"] for r in rows}
    assert event_types == {"ISSUE", "RECEIVE"}

    for row in rows:
        assert row["legacy_source_row_key"] == "1"
        nv = row["normalized_values"]
        assert nv["equipment_id"] is not None
        assert nv["occurred_at"] is not None
        assert nv["legacy_ward_text"] == "Ward 1"
        assert nv["resolved_ward_id"] is not None
        assert "header_source_ref" in nv and "line_source_ref" in nv
        assert "notes" not in nv, "no notes content may ever be carried in a plan row (OD-PR21-6)"
        assert row["warnings"] == []

    # No pairing: Issue and Receive rows are independent, own coordinates.
    issue_row = next(r for r in rows if r["event_type"] == "ISSUE")
    receive_row = next(r for r in rows if r["event_type"] == "RECEIVE")
    assert issue_row["normalized_values"]["line_source_ref"]["sheet_name"] == ISSUE_LINE_SHEET
    assert receive_row["normalized_values"]["line_source_ref"]["sheet_name"] == RECEIVE_LINE_SHEET

    # Read-only: plan_dry_run must never mutate Equipment/BorrowTransaction/User.
    equipment_after = (await db_session.execute(select(Equipment))).scalars().all()
    borrow_after = (await db_session.execute(select(BorrowTransaction))).scalars().all()
    users_after = (await db_session.execute(select(User))).scalars().all()
    assert len(equipment_after) == len(equipment_before)
    assert len(borrow_after) == len(borrow_before) == 0
    assert len(users_after) == len(users_before)


async def test_plan_dry_run_blocks_on_issue_side_error_even_if_receive_valid(db_session: AsyncSession, seeded_users):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    actor_id = await _get_actor_id(db_session)
    content = _default_workbook(issue_lines=[_issue_line_row(me_code="ME_DOES_NOT_EXIST")])
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    adapter = LegacyTransactionHistoryAdapter()
    ctx = _invocation_context(content, checksum=checksum)
    with adapter_invocation_context(ctx):
        with pytest.raises(InvalidInputError):
            await adapter.plan_dry_run(db_session)


async def test_plan_dry_run_blocks_on_receive_side_error_even_if_issue_valid(db_session: AsyncSession, seeded_users):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    actor_id = await _get_actor_id(db_session)
    content = _default_workbook(receive_lines=[_receive_line_row(me_code="ME_DOES_NOT_EXIST")])
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    adapter = LegacyTransactionHistoryAdapter()
    ctx = _invocation_context(content, checksum=checksum)
    with adapter_invocation_context(ctx):
        with pytest.raises(InvalidInputError):
            await adapter.plan_dry_run(db_session)


async def test_plan_dry_run_header_one_to_n_provenance_preserved_both_sides(db_session: AsyncSession, seeded_users):
    """One header row, two line rows on each side -- each line row's own
    plan row keeps its own `line_source_ref`, all sharing the same header
    coordinate on their own side."""
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    actor_id = await _get_actor_id(db_session)
    content = _default_workbook(
        issue_lines=[_issue_line_row(row_key="1"), _issue_line_row(row_key="2")],
        receive_lines=[_receive_line_row(row_key="1"), _receive_line_row(row_key="2")],
    )
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    adapter = LegacyTransactionHistoryAdapter()
    ctx = _invocation_context(content, checksum=checksum)
    with adapter_invocation_context(ctx):
        plan = await adapter.plan_dry_run(db_session)

    rows = plan.summary["rows"]
    assert len(rows) == 4
    issue_rows = [r for r in rows if r["event_type"] == "ISSUE"]
    receive_rows = [r for r in rows if r["event_type"] == "RECEIVE"]
    assert len(issue_rows) == 2
    assert len(receive_rows) == 2
    for r in issue_rows:
        assert r["normalized_values"]["header_source_ref"]["source_row_number"] == 2
    for r in receive_rows:
        assert r["normalized_values"]["header_source_ref"]["source_row_number"] == 2


async def test_plan_dry_run_ward_alias_and_authority_snapshot_retained(db_session: AsyncSession, seeded_users):
    """OD-PR21-4 Ward alias resolution and the migration authority id are
    captured in the plan at dry-run time -- immutable thereafter."""
    from app.models.legacy_history import LegacyWardAlias

    actor_id = await _get_actor_id(db_session)
    ward = await _seed_ward(db_session, code="CanonicalWard")
    await _seed_ward(db_session, code="Ward 1")  # resolves the Receive side's default rows
    db_session.add(LegacyWardAlias(raw_alias="Legacy Ward Text", ward_id=ward.id, created_by_user_id=actor_id))
    await db_session.commit()
    await _seed_equipment(db_session)

    content = _default_workbook(
        issue_headers=[_issue_header_row(ward="Legacy Ward Text")],
        issue_lines=[_issue_line_row(ward="Legacy Ward Text")],
    )
    checksum = hashlib.sha256(content).hexdigest()
    authority = await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    adapter = LegacyTransactionHistoryAdapter()
    ctx = _invocation_context(content, checksum=checksum)
    with adapter_invocation_context(ctx):
        plan = await adapter.plan_dry_run(db_session)

    assert plan.summary["migration_authority_id"] == str(authority.id)
    issue_row = next(r for r in plan.summary["rows"] if r["event_type"] == "ISSUE")
    assert issue_row["normalized_values"]["resolved_ward_id"] == str(ward.id)


# ---------------------------------------------------------------------------
# E. persist_dry_run_plan()/provider: exercised via the provider directly
#    (mirrors PR21A's own test convention -- the generic `GET/POST
#    .../dry-run-plan` HTTP endpoints are Equipment-Master-specific, a
#    pre-existing gap unrelated to this slice; see this PR's own final
#    report and `test_pr21d1_source_admission.py`'s discovery test).
# ---------------------------------------------------------------------------


async def test_provider_registered_and_reachable_by_dataset_type():
    from app.services.import_plan_provider import get_plan_provider

    provider = get_plan_provider(DATASET_TYPE)
    assert isinstance(provider, LegacyHistoryDryRunPlanProvider)

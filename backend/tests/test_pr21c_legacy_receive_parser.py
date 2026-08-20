"""Roadmap PR21C -- Canonical Receive Parser + Validation (bounded slice).

Authoritative design: `docs/design/PR21_LEGACY_TRANSACTION_HISTORY_IMPORT_PLAN.md`,
merged foundation GitHub PR #103 / PR21A (squash SHA
`28f0f5eabb64cf4b27294fd3df251e90b167de0a`) and GitHub PR #104 / PR21B
canonical Issue subset (squash SHA
`a8ae9fbfc571f74bad2100abf8f90bbd22a68e74`). Tests the Receive-only
canonical parser/validator in `app.services.import_adapters.legacy_history.receive`
directly (not through the `ImportAdapter`/`ImportSession` pipeline --
this package is deliberately never registered there, see §J below and
that package's own `__init__.py` docstring).

Synthetic workbooks only -- no production workbook data is read or
committed anywhere in this file. Column names mirror
`docs/evidence/pr21/equipment-pool-workbook-manifest.json`'s verified
`header_names` for `Orders คืนเครื่อง`/`ข้อมูลรับเครื่องมือ`."""

import uuid
from dataclasses import fields as dataclass_fields
from datetime import date, time
from io import BytesIO

import pytest
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import InvalidInputError
from app.models.equipment import Equipment, EquipmentStatus
from app.models.legacy_history import LegacyWardAlias
from app.models.master_data import Ward
from app.models.transaction import BorrowTransaction
from app.models.user import User
from app.services.identifiers import normalize_bcm_code
from app.services.import_adapter import get_adapter
from app.services.import_adapters.legacy_history import receive as receive_module
from app.services.import_adapters.legacy_history.types import LegacyReceiveCandidate
from app.services.import_plan_providers.legacy_history import DATASET_TYPE

HEADER_SHEET = receive_module.HEADER_SHEET_NAME
LINE_SHEET = receive_module.LINE_SHEET_NAME
HEADER_COLUMNS = list(receive_module._HEADER_GOVERNED_HEADERS.keys())
LINE_COLUMNS = list(receive_module._LINE_GOVERNED_HEADERS.keys())

MIGRATION_AUTHORITY_ID = uuid.uuid4()
IMPORT_SESSION_ID = uuid.uuid4()
IMPORT_SOURCE_ID = uuid.uuid4()


def _build_workbook(header_rows: list[dict], line_rows: list[dict], *, extra_sheets: dict[str, list] | None = None) -> bytes:
    wb = Workbook()
    ws_header = wb.active
    ws_header.title = HEADER_SHEET
    ws_header.append(HEADER_COLUMNS)
    for row in header_rows:
        ws_header.append([row.get(col) for col in HEADER_COLUMNS])

    ws_line = wb.create_sheet(LINE_SHEET)
    ws_line.append(LINE_COLUMNS)
    for row in line_rows:
        ws_line.append([row.get(col) for col in LINE_COLUMNS])

    if extra_sheets:
        for name, rows in extra_sheets.items():
            ws = wb.create_sheet(name)
            for row in rows:
                ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_row(*, order_ref="RET-0001", ward="Ward 1", bme="BME Header", d=date(2024, 1, 10), t=time(9, 0, 0), notes=None):
    return {
        "วันที่": d,
        "เลขที่ใบคืน": order_ref,
        "แผนกที่คืน": ward,
        "ผู้ส่งเครื่องคืน (User)": "header-user",
        "ผู้รับเครื่องคืน (BME)": bme,
        "เวลา": t,
        "จำนวนเครื่องรับคืนทั้งหมด": 1,
        "หมายเหตุ": notes,
    }


def _line_row(*, row_key="1", order_ref="RET-0001", me_code="ME001", ward="Ward 1", bme="BME Line", d=date(2024, 1, 10), t=time(9, 0, 0), notes=None):
    return {
        "ลำดับ": row_key,
        "วันที่": d,
        "เลขที่ใบรับเครื่อง": order_ref,
        "SCAN CODE รับ": "SCAN-1",
        "ME.Code": me_code,
        "Barcode รับเครื่อง": "BARCODE-1",
        "Equipment": "Infusion Pump",
        "Brand": "BrandX",
        "Model": "ModelY",
        "Serial no.": "SN-1",
        "แผนกที่รับ": ward,
        "รูปเครื่อง": None,
        "ตัวเครื่องหน้าจอไม่แตกร้าว": "OK",
        "สายชาร์จ (สายไม่ขาด)": "OK",
        "Pole Clamp (ไม่แตกหักมไม่ง้อ)": "OK",
        "จำนวน": 1,
        "หมายเหตุ": notes,
        "เวลา": t,
        "ชื่อ BME": bme,
        "ชื่อ (User)": "line-user",
    }


async def _seed_equipment(db_session: AsyncSession, *, me_code: str = "ME001", **kwargs) -> Equipment:
    defaults = dict(
        asset_number=f"AN-{uuid.uuid4().hex[:10]}",
        equipment_name="Legacy Test Equipment",
        status=EquipmentStatus.AVAILABLE_AT_POOL,
        bcm_code=normalize_bcm_code(me_code),
    )
    defaults.update(kwargs)
    eq = Equipment(**defaults)
    db_session.add(eq)
    await db_session.commit()
    await db_session.refresh(eq)
    return eq


async def _seed_ward(db_session: AsyncSession, *, code: str, name: str | None = None) -> Ward:
    ward = Ward(code=code, name=name or code)
    db_session.add(ward)
    await db_session.commit()
    await db_session.refresh(ward)
    return ward


async def _run(db_session: AsyncSession, content: bytes):
    header_records, line_records = receive_module.parse_workbook(content)
    context = await receive_module.preload_business_context(db_session, header_records, line_records)
    candidates, findings = receive_module.validate_and_build_candidates(
        header_records,
        line_records,
        context,
        migration_authority_id=MIGRATION_AUTHORITY_ID,
        import_session_id=IMPORT_SESSION_ID,
        import_source_id=IMPORT_SOURCE_ID,
    )
    return header_records, line_records, candidates, findings


def _codes(findings) -> set[str]:
    return {f.error_code for f in findings}


# ---------------------------------------------------------------------------
# A. Sheet contract
# ---------------------------------------------------------------------------


async def test_both_required_sheets_found(db_session: AsyncSession):
    content = _build_workbook([_header_row()], [_line_row()])
    header_records, line_records = receive_module.parse_workbook(content)
    assert len(header_records) == 1
    assert len(line_records) == 1


def test_missing_orders_sheet_raises():
    wb = Workbook()
    ws = wb.active
    ws.title = LINE_SHEET
    ws.append(LINE_COLUMNS)
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(InvalidInputError):
        receive_module.parse_workbook(buf.getvalue())


def test_missing_line_item_sheet_raises():
    wb = Workbook()
    ws = wb.active
    ws.title = HEADER_SHEET
    ws.append(HEADER_COLUMNS)
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(InvalidInputError):
        receive_module.parse_workbook(buf.getvalue())


def test_derived_sdc_sheets_ignored():
    """Presence of the Receive-side SDC sheet must not cause silent
    ingestion -- the parser must simply never look at it."""
    content = _build_workbook(
        [_header_row()],
        [_line_row()],
        extra_sheets={"ข้อมูลการรับ SDC": [["some", "unrelated", "sdc", "columns"], ["a", "b", "c", "d"]]},
    )
    header_records, line_records = receive_module.parse_workbook(content)
    assert len(header_records) == 1
    assert len(line_records) == 1


def _dummy_extra_sheets(count: int, *, prefix: str = "Dummy") -> dict[str, list]:
    return {f"{prefix}{i}": [["placeholder"]] for i in range(count)}


def test_28_sheet_workbook_shape_accepted_by_shared_loader():
    """The Receive parser reuses `common.load_workbook_bytes` unchanged
    -- the PR21-wide `PR21_MAX_WORKSHEET_COUNT` (32) allowance applies
    here exactly as it does for Issue, with no Receive-specific cap."""
    content = _build_workbook([_header_row()], [_line_row()], extra_sheets=_dummy_extra_sheets(26))
    header_records, line_records = receive_module.parse_workbook(content)
    assert len(header_records) == 1
    assert len(line_records) == 1


def test_33_sheet_workbook_rejected_by_shared_loader():
    from app.services.import_adapters.legacy_history import common as common_module

    extra = common_module.PR21_MAX_WORKSHEET_COUNT - 2 + 1
    content = _build_workbook([_header_row()], [_line_row()], extra_sheets=_dummy_extra_sheets(extra))
    with pytest.raises(InvalidInputError):
        receive_module.parse_workbook(content)


# ---------------------------------------------------------------------------
# B. Topology
# ---------------------------------------------------------------------------


async def test_one_header_multiple_lines_succeeds(db_session: AsyncSession):
    equipment_a = await _seed_equipment(db_session, me_code="ME001", asset_number="AN-A")
    equipment_b = await _seed_equipment(db_session, me_code="ME002", asset_number="AN-B")
    await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook(
        [_header_row(order_ref="RET-0001")],
        [
            _line_row(row_key="1", order_ref="RET-0001", me_code="ME001"),
            _line_row(row_key="2", order_ref="RET-0001", me_code="ME002"),
        ],
    )
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    assert len(candidates) == 2
    assert {c.equipment_id for c in candidates} == {equipment_a.id, equipment_b.id}
    assert candidates[0].header_source_ref.source_row_number == candidates[1].header_source_ref.source_row_number
    assert candidates[0].line_source_ref.source_row_number != candidates[1].line_source_ref.source_row_number


async def test_orphan_line_reference_errors(db_session: AsyncSession):
    await _seed_equipment(db_session)
    content = _build_workbook([_header_row(order_ref="RET-0001")], [_line_row(order_ref="RET-DOES-NOT-EXIST")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_ORPHAN_ORDER_REFERENCE in _codes(findings)


async def test_duplicate_ambiguous_header_ref_errors(db_session: AsyncSession):
    await _seed_equipment(db_session)
    content = _build_workbook(
        [_header_row(order_ref="RET-0001"), _header_row(order_ref="RET-0001")],
        [_line_row(order_ref="RET-0001")],
    )
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    codes = _codes(findings)
    assert receive_module.CODE_HEADER_ORDER_REFERENCE_DUPLICATE in codes
    assert receive_module.CODE_AMBIGUOUS_ORDER_REFERENCE in codes
    duplicate_findings = [f for f in findings if f.error_code == receive_module.CODE_HEADER_ORDER_REFERENCE_DUPLICATE]
    assert len(duplicate_findings) == 2  # both header rows flagged, not merely one


# ---------------------------------------------------------------------------
# C. Row keys
# ---------------------------------------------------------------------------


async def test_valid_text_row_key_accepted(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook([_header_row()], [_line_row(row_key="42")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    assert candidates[0].legacy_source_row_key == "42"


def test_valid_integer_numeric_row_key_accepted():
    wb = Workbook()
    ws = wb.active
    ws.title = HEADER_SHEET
    ws.append(HEADER_COLUMNS)
    ws.append([_header_row().get(col) for col in HEADER_COLUMNS])
    ws_line = wb.create_sheet(LINE_SHEET)
    ws_line.append(LINE_COLUMNS)
    row = _line_row(row_key=None)
    row["ลำดับ"] = 7  # numeric cell, losslessly integer
    ws_line.append([row.get(col) for col in LINE_COLUMNS])
    buf = BytesIO()
    wb.save(buf)
    _, line_records = receive_module.parse_workbook(buf.getvalue())
    assert line_records[0].fields["row_key_cell"] == ("valid", "7")


async def test_blank_data_bearing_row_key_errors(db_session: AsyncSession):
    await _seed_equipment(db_session)
    content = _build_workbook([_header_row()], [_line_row(row_key=None)])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_SOURCE_ROW_KEY_MISSING in _codes(findings)


async def test_duplicate_row_key_errors(db_session: AsyncSession):
    await _seed_equipment(db_session, me_code="ME001", asset_number="AN-A")
    await _seed_equipment(db_session, me_code="ME002", asset_number="AN-B")
    content = _build_workbook(
        [_header_row()],
        [_line_row(row_key="1", me_code="ME001"), _line_row(row_key="1", me_code="ME002")],
    )
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    dup_findings = [f for f in findings if f.error_code == receive_module.CODE_SOURCE_ROW_KEY_DUPLICATE]
    assert len(dup_findings) == 2


def test_blank_formatting_rows_ignored():
    wb = Workbook()
    ws_header = wb.active
    ws_header.title = HEADER_SHEET
    ws_header.append(HEADER_COLUMNS)
    ws_header.append([_header_row().get(col) for col in HEADER_COLUMNS])
    ws_header.append([None] * len(HEADER_COLUMNS))  # trailing blank/formatted row

    ws_line = wb.create_sheet(LINE_SHEET)
    ws_line.append(LINE_COLUMNS)
    ws_line.append([_line_row().get(col) for col in LINE_COLUMNS])
    ws_line.append([None] * len(LINE_COLUMNS))

    buf = BytesIO()
    wb.save(buf)
    header_records, line_records = receive_module.parse_workbook(buf.getvalue())
    assert len(header_records) == 1
    assert len(line_records) == 1


# ---------------------------------------------------------------------------
# D. Equipment
# ---------------------------------------------------------------------------


async def test_valid_me_code_resolves(db_session: AsyncSession):
    equipment = await _seed_equipment(db_session, me_code="ME001")
    await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook([_header_row()], [_line_row(me_code="ME001")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    assert candidates[0].equipment_id == equipment.id


async def test_blank_me_code_errors(db_session: AsyncSession):
    content = _build_workbook([_header_row()], [_line_row(me_code=None)])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_ME_CODE_MISSING in _codes(findings)


def test_typed_numeric_me_code_rejected():
    wb = Workbook()
    ws = wb.active
    ws.title = HEADER_SHEET
    ws.append(HEADER_COLUMNS)
    ws.append([_header_row().get(col) for col in HEADER_COLUMNS])
    ws_line = wb.create_sheet(LINE_SHEET)
    ws_line.append(LINE_COLUMNS)
    row = _line_row()
    row["ME.Code"] = 12345  # numeric cell, never coerced
    ws_line.append([row.get(col) for col in LINE_COLUMNS])
    buf = BytesIO()
    wb.save(buf)
    _, line_records = receive_module.parse_workbook(buf.getvalue())
    assert line_records[0].fields["me_code_cell"] == ("invalid", None)


async def test_malformed_bcm_me_code_errors(db_session: AsyncSession):
    content = _build_workbook([_header_row()], [_line_row(me_code="   ")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    # An all-whitespace string strips to blank, same as None.
    assert receive_module.CODE_ME_CODE_MISSING in _codes(findings)


async def test_unknown_equipment_errors(db_session: AsyncSession):
    content = _build_workbook([_header_row()], [_line_row(me_code="NOTREGISTERED")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_EQUIPMENT_NOT_FOUND in _codes(findings)


@pytest.mark.parametrize(
    "status",
    [
        EquipmentStatus.AVAILABLE_AT_POOL,
        EquipmentStatus.ISSUED_TO_WARD,
        EquipmentStatus.UNAVAILABLE_DEFECTIVE,
        EquipmentStatus.DECOMMISSIONED,
    ],
)
async def test_receive_validation_succeeds_regardless_of_current_equipment_status(db_session: AsyncSession, status):
    """A historical RECEIVE event may describe a past state inconsistent
    with current live Equipment status -- current status alone must
    never invalidate a historical event, and must never be mutated by
    validation."""
    equipment = await _seed_equipment(db_session, me_code="ME001", status=status)
    await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook([_header_row()], [_line_row(me_code="ME001")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    assert len(candidates) == 1
    assert candidates[0].equipment_id == equipment.id
    await db_session.refresh(equipment)
    assert equipment.status == status


async def test_equipment_status_and_version_unchanged(db_session: AsyncSession):
    equipment = await _seed_equipment(db_session, me_code="ME001")
    await _seed_ward(db_session, code="Ward 1")
    status_before, version_before = equipment.status, equipment.version
    content = _build_workbook([_header_row()], [_line_row(me_code="ME001")])
    await _run(db_session, content)
    await db_session.refresh(equipment)
    assert equipment.status == status_before
    assert equipment.version == version_before


async def test_no_open_borrow_transaction_required(db_session: AsyncSession):
    """Receive validation must not query or require any current
    `BorrowTransaction` row -- a valid historical RECEIVE candidate is
    constructible even when no `BorrowTransaction` exists at all for the
    Equipment (in production, matching/reconciliation is PR22's
    responsibility, not this parser's)."""
    equipment = await _seed_equipment(db_session, me_code="ME001")
    await _seed_ward(db_session, code="Ward 1")
    open_borrows = (
        await db_session.execute(select(BorrowTransaction).where(BorrowTransaction.equipment_id == equipment.id))
    ).scalars().all()
    assert open_borrows == []
    content = _build_workbook([_header_row()], [_line_row(me_code="ME001")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    assert len(candidates) == 1


# ---------------------------------------------------------------------------
# E. Ward
# ---------------------------------------------------------------------------


async def test_ward_exact_match_succeeds(db_session: AsyncSession):
    await _seed_equipment(db_session)
    ward = await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook([_header_row(ward="Ward 1")], [_line_row(ward="Ward 1")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    assert candidates[0].resolved_ward_id == ward.id


async def test_ward_alias_succeeds(db_session: AsyncSession, seeded_users):
    await _seed_equipment(db_session)
    ward = await _seed_ward(db_session, code="Ward 1")
    actor_id = (await db_session.execute(select(User.id).limit(1))).scalar_one()
    db_session.add(LegacyWardAlias(raw_alias="W1-ALIAS", ward_id=ward.id, created_by_user_id=actor_id))
    await db_session.commit()
    content = _build_workbook([_header_row(ward="W1-ALIAS")], [_line_row(ward="W1-ALIAS")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    assert candidates[0].resolved_ward_id == ward.id


async def test_unknown_ward_errors(db_session: AsyncSession):
    await _seed_equipment(db_session)
    content = _build_workbook([_header_row(ward="NoSuchWard")], [_line_row(ward="NoSuchWard")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_WARD_NOT_FOUND in _codes(findings)


async def test_ambiguous_ward_errors(db_session: AsyncSession):
    """A raw text that simultaneously names two different Wards (one's
    `code`, a different one's `name`) must be `"ambiguous"`, never an
    arbitrary pick."""
    await _seed_equipment(db_session)
    await _seed_ward(db_session, code="AMBIG", name="Ward Alpha")
    await _seed_ward(db_session, code="ZZZ", name="AMBIG")
    content = _build_workbook([_header_row(ward="AMBIG")], [_line_row(ward="AMBIG")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_WARD_AMBIGUOUS in _codes(findings)


async def test_header_line_ward_mismatch_errors(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session, code="Ward 1")
    await _seed_ward(db_session, code="Ward 2")
    content = _build_workbook([_header_row(ward="Ward 1")], [_line_row(ward="Ward 2")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_HEADER_LINE_WARD_CONFLICT in _codes(findings)


async def test_ward_resolution_does_not_mutate_current_location(db_session: AsyncSession):
    """Ward resolution is read-only identity lookup -- no Ward row is
    ever written to by this validation."""
    await _seed_equipment(db_session)
    ward = await _seed_ward(db_session, code="Ward 1")
    name_before, code_before = ward.name, ward.code
    content = _build_workbook([_header_row(ward="Ward 1")], [_line_row(ward="Ward 1")])
    await _run(db_session, content)
    await db_session.refresh(ward)
    assert ward.name == name_before
    assert ward.code == code_before


# ---------------------------------------------------------------------------
# F. Timestamps
# ---------------------------------------------------------------------------


async def test_valid_timestamp_combines_to_aware_utc(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook([_header_row(d=date(2024, 1, 10), t=time(9, 0, 0))], [_line_row(d=date(2024, 1, 10), t=time(9, 0, 0))])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    occurred_at = candidates[0].occurred_at
    assert occurred_at.tzinfo is not None
    assert occurred_at.utcoffset().total_seconds() == 0
    # Asia/Bangkok is UTC+7 with no DST -- 09:00 local == 02:00 UTC.
    assert occurred_at.hour == 2


async def test_malformed_date_errors(db_session: AsyncSession):
    await _seed_equipment(db_session)
    content = _build_workbook([_header_row()], [_line_row(d="not-a-date")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_INVALID_TIMESTAMP in _codes(findings)


async def test_malformed_time_errors(db_session: AsyncSession):
    await _seed_equipment(db_session)
    content = _build_workbook([_header_row()], [_line_row(t="not-a-time")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_INVALID_TIMESTAMP in _codes(findings)


async def test_header_line_timestamp_conflict_errors(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook(
        [_header_row(d=date(2024, 1, 10), t=time(9, 0, 0))],
        [_line_row(d=date(2024, 1, 11), t=time(9, 0, 0))],
    )
    _, _, candidates, findings = await _run(db_session, content)
    assert not candidates
    assert receive_module.CODE_HEADER_LINE_TIMESTAMP_CONFLICT in _codes(findings)


# ---------------------------------------------------------------------------
# G. BME / historical operator
# ---------------------------------------------------------------------------


async def test_bme_raw_text_preserved(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook([_header_row(bme="Header BME Name")], [_line_row(bme="Line BME Name")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    # Line-level BME text takes precedence when present (this row's own event actor).
    assert candidates[0].legacy_bme_name == "Line BME Name"


async def test_absent_bme_does_not_block(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook([_header_row(bme=None)], [_line_row(bme=None)])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    assert candidates[0].legacy_bme_name is None


async def test_no_user_created(db_session: AsyncSession):
    await _seed_equipment(db_session)
    await _seed_ward(db_session, code="Ward 1")
    count_before = (await db_session.execute(select(User.id))).scalars().all()
    content = _build_workbook([_header_row(bme="Some BME Name")], [_line_row(bme="Some BME Name")])
    await _run(db_session, content)
    count_after = (await db_session.execute(select(User.id))).scalars().all()
    assert len(count_after) == len(count_before)


# ---------------------------------------------------------------------------
# H. Privacy (OD-PR21-6)
# ---------------------------------------------------------------------------


def test_candidate_has_no_notes_field():
    field_names = {f.name for f in dataclass_fields(LegacyReceiveCandidate)}
    assert "notes" not in field_names
    assert "หมายเหตุ" not in field_names
    assert not any("note" in name.lower() for name in field_names)


async def test_notes_value_never_appears_in_finding_message(db_session: AsyncSession):
    sensitive_text = "PATIENT-CONFIDENTIAL-XYZ-12345"
    content = _build_workbook(
        [_header_row(notes=sensitive_text)],
        [_line_row(me_code="NOTREGISTERED", notes=sensitive_text)],
    )
    _, _, candidates, findings = await _run(db_session, content)
    assert findings
    for finding in findings:
        assert sensitive_text not in finding.message


# ---------------------------------------------------------------------------
# I. Determinism
# ---------------------------------------------------------------------------


def test_repeated_parse_gives_same_row_order():
    content = _build_workbook(
        [_header_row(order_ref="RET-0001")],
        [_line_row(row_key="1", order_ref="RET-0001"), _line_row(row_key="2", order_ref="RET-0001", me_code="ME002")],
    )
    first_header, first_line = receive_module.parse_workbook(content)
    second_header, second_line = receive_module.parse_workbook(content)
    assert [r.row_number for r in first_header] == [r.row_number for r in second_header]
    assert [r.row_number for r in first_line] == [r.row_number for r in second_line]


async def test_repeated_validation_gives_same_findings_order(db_session: AsyncSession):
    await _seed_equipment(db_session, me_code="ME001", asset_number="AN-A")
    content = _build_workbook([_header_row()], [_line_row(me_code="ME001"), _line_row(row_key="2", me_code="UNREGISTERED")])
    header_records, line_records = receive_module.parse_workbook(content)
    context = await receive_module.preload_business_context(db_session, header_records, line_records)
    _, findings_first = receive_module.validate_and_build_candidates(
        header_records, line_records, context,
        migration_authority_id=MIGRATION_AUTHORITY_ID, import_session_id=IMPORT_SESSION_ID, import_source_id=IMPORT_SOURCE_ID,
    )
    _, findings_second = receive_module.validate_and_build_candidates(
        header_records, line_records, context,
        migration_authority_id=MIGRATION_AUTHORITY_ID, import_session_id=IMPORT_SESSION_ID, import_source_id=IMPORT_SOURCE_ID,
    )
    assert [(f.sheet_name, f.source_row_number, f.error_code) for f in findings_first] == [
        (f.sheet_name, f.source_row_number, f.error_code) for f in findings_second
    ]


# ---------------------------------------------------------------------------
# J. Partial-adapter safety -- unchanged by adding the Receive side
# ---------------------------------------------------------------------------


def test_legacy_dataset_type_adapter_registered_but_still_cannot_execute():
    """Adding the Receive-side canonical parser did not, by itself,
    register the final PR21 dataset adapter while SDC scope remained
    open. Roadmap PR21D1 (design doc §55.2/§55.8, Owner Decision
    Closure Round 3, GitHub PR #106) explicitly supersedes that now
    that SDC is excluded and both canonical sides are merged --
    `legacy_transaction_history` is now expected to have a real,
    registered production adapter. The safety property that mattered --
    a real `ImportSession` cannot reach executed history -- is preserved
    a different way from here on: the registered adapter's own
    `execute()` is still the unimplemented `ImportAdapter` base-class
    default (§21/§55.5/§60/§61 of the PR21D1 task), so
    `import_execution_service.run_execute`'s own
    `type(adapter).execute is ImportAdapter.execute` guard keeps
    `POST .../execute` structurally unreachable regardless of
    registration."""
    from app.services.import_adapter import ImportAdapter

    adapter = get_adapter(DATASET_TYPE)
    assert adapter is not None
    assert type(adapter).execute is ImportAdapter.execute


def test_workbook_authority_matching_checksum_passes():
    receive_module.verify_workbook_authority(workbook_checksum="abc123", approved_workbook_sha256="abc123")


def test_workbook_authority_mismatched_checksum_raises():
    with pytest.raises(InvalidInputError):
        receive_module.verify_workbook_authority(workbook_checksum="abc123", approved_workbook_sha256="different")


# ---------------------------------------------------------------------------
# Structural: unrecognized/extra column in the governed sheets rejected
# ---------------------------------------------------------------------------


def test_unrecognized_column_in_header_sheet_rejected():
    wb = Workbook()
    ws = wb.active
    ws.title = HEADER_SHEET
    ws.append(HEADER_COLUMNS + ["Unexpected Extra Column"])
    ws.append([_header_row().get(col) for col in HEADER_COLUMNS] + ["some value"])
    ws_line = wb.create_sheet(LINE_SHEET)
    ws_line.append(LINE_COLUMNS)
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(InvalidInputError):
        receive_module.parse_workbook(buf.getvalue())


def test_missing_governed_column_rejected():
    wb = Workbook()
    ws = wb.active
    ws.title = HEADER_SHEET
    ws.append([c for c in HEADER_COLUMNS if c != "หมายเหตุ"])  # drop one governed column
    ws_line = wb.create_sheet(LINE_SHEET)
    ws_line.append(LINE_COLUMNS)
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(InvalidInputError):
        receive_module.parse_workbook(buf.getvalue())


# ---------------------------------------------------------------------------
# No Issue<->Receive pairing / no cross-module coupling
# ---------------------------------------------------------------------------


async def test_receive_candidates_built_independently_of_issue_module(db_session: AsyncSession):
    """RECEIVE candidates must be constructible with zero Issue-side
    data present anywhere -- no matching, pairing, or cross-reference to
    any ISSUE candidate/order occurs in this module."""
    await _seed_equipment(db_session)
    await _seed_ward(db_session, code="Ward 1")
    content = _build_workbook([_header_row(order_ref="RET-9999")], [_line_row(order_ref="RET-9999")])
    _, _, candidates, findings = await _run(db_session, content)
    assert not findings
    assert len(candidates) == 1
    assert candidates[0].event_type == "RECEIVE"


# ---------------------------------------------------------------------------
# J (regressions). PR21B Issue module and PR21A schema remain green --
# exercised by running this file alongside
# `test_pr21b_legacy_issue_parser.py` and `test_pr21a_legacy_history_schema.py`
# in the same suite run (see this PR's validation commands); nothing in
# this file imports or monkeypatches either module.
# ---------------------------------------------------------------------------

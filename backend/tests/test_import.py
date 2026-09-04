import io
import uuid
import zipfile

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from app.crud import equipment as equipment_crud
from app.models.audit import AuditLog
from app.models.equipment import Equipment, EquipmentStatus
from app.models.user import ROLE_ADMINISTRATOR, ROLE_EQUIPMENT_POOL_STAFF, ROLE_READ_ONLY
from app.services.import_service import NO_MATCHING_EQUIPMENT_REASON, REQUIRED_HEADERS
from tests.conftest import auth_headers as _auth_headers

pytestmark = pytest.mark.asyncio

# Roadmap PR12 is update-only (Repository Owner decision resolving review
# findings PR12-H1/H1R -- see import_service.py's module docstring): the
# import workflow never creates new equipment and never generates,
# derives, or synthesizes an Asset Number. Every test that needs an
# import row to succeed must first seed a matching equipment record
# directly through app.crud.equipment (never through the import endpoints
# themselves, since import can no longer create anything).


def _build_workbook(rows: list[dict[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(list(REQUIRED_HEADERS))
    for row in rows:
        ws.append([row.get(header, "") for header in REQUIRED_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _base_row(**overrides) -> dict[str, str]:
    row = {
        "Item No.": "",
        "ID CODE": "001",
        "Asset ID": "",
        "Equipment Name": "Infusion Pump",
        "Manufacturer": "Acme",
        "Model": "X1",
        "Serial Number": "",
        "Location": "Ward 3",
        "Receive Date": "2024-01-01",
        "Register Date": "2024-01-02",
        "Purchase Year": "2023",
        "Asset Status": "Active",
    }
    row.update(overrides)
    return row


def _upload(rows: list[dict[str, str]], filename: str = "inventory.xlsx"):
    content = _build_workbook(rows)
    return {"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}


async def _preview(client, headers, rows, *, update_existing: bool = True):
    return await client.post(
        "/api/v1/import/preview",
        headers=headers,
        files=_upload(rows),
        data={"update_existing": "true" if update_existing else "false"},
    )


async def _commit(client, headers, rows, *, update_existing: bool = True):
    return await client.post(
        "/api/v1/import/commit",
        headers=headers,
        files=_upload(rows),
        data={"update_existing": "true" if update_existing else "false"},
    )


async def _seed_equipment(
    db_session,
    *,
    bcm_code: str,
    item_no: str | None = None,
    serial_number: str | None = None,
    asset_id: str | None = None,
    equipment_name: str = "Seed Equipment",
    brand: str | None = None,
    model: str | None = None,
    asset_number: str | None = None,
) -> str:
    """Creates an equipment record directly (never through the import
    endpoints, which can no longer create anything) so import-mode tests
    have an existing record to match by BCM Code and update."""
    equipment = await equipment_crud.create(
        db_session,
        data={
            "asset_number": asset_number or f"AST-{bcm_code}",
            "bcm_code": bcm_code,
            "item_no": item_no,
            "serial_number": serial_number,
            "asset_id": asset_id,
            "equipment_name": equipment_name,
            "brand": brand,
            "model": model,
        },
    )
    await db_session.commit()
    return str(equipment.id)


# ---------------------------------------------------------------------------
# Authorization
# ---------------------------------------------------------------------------


async def test_non_admin_forbidden_from_preview(client, seeded_users):
    for role in (ROLE_EQUIPMENT_POOL_STAFF, ROLE_READ_ONLY):
        headers = await _auth_headers(client, role)
        resp = await _preview(client, headers, [_base_row()])
        assert resp.status_code == 403


async def test_non_admin_forbidden_from_commit(client, seeded_users):
    headers = await _auth_headers(client, ROLE_EQUIPMENT_POOL_STAFF)
    resp = await _commit(client, headers, [_base_row()])
    assert resp.status_code == 403


# ---------------------------------------------------------------------------
# Header validation
# ---------------------------------------------------------------------------


async def test_missing_required_header_rejects_whole_file(client, seeded_users):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    wb = Workbook()
    ws = wb.active
    incomplete_headers = [h for h in REQUIRED_HEADERS if h != "Asset Status"]
    ws.append(incomplete_headers)
    ws.append(["x"] * len(incomplete_headers))
    buf = io.BytesIO()
    wb.save(buf)
    resp = await client.post(
        "/api/v1/import/preview",
        headers=headers,
        files={"file": ("bad.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"update_existing": "true"},
    )
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"
    assert "Asset Status" in resp.json()["detail"]


# ---------------------------------------------------------------------------
# Row-level validation (Part G.4)
# ---------------------------------------------------------------------------


async def test_missing_bcm_code_fails_row(client, seeded_users):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await _preview(client, headers, [_base_row(**{"ID CODE": ""})])
    assert resp.status_code == 200
    body = resp.json()
    assert body["failed"] == 1
    assert body["rows"][0]["status"] == "failed"
    assert "BCM Code" in body["rows"][0]["reason"]


async def test_duplicate_bcm_code_within_file_flags_all_occurrences(client, seeded_users):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    rows = [_base_row(**{"ID CODE": "777"}), _base_row(**{"ID CODE": "777"})]
    resp = await _preview(client, headers, rows)
    body = resp.json()
    assert body["failed"] == 2
    assert all(r["status"] == "failed" and "Duplicate BCM Code" in r["reason"] for r in body["rows"])


async def test_leading_zero_bcm_code_round_trip(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM00042")

    resp = await _commit(client, headers, [_base_row(**{"ID CODE": "00042"})], update_existing=True)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["succeeded"] == 1
    assert body["rows"][0]["bcm_code"] == "BCM00042"


async def test_unrecognized_asset_status_fails_row(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM001")

    resp = await _preview(
        client, headers, [_base_row(**{"Asset Status": "Mystery Value"})], update_existing=True
    )
    body = resp.json()
    assert body["failed"] == 1
    assert "Unrecognized Asset Status" in body["rows"][0]["reason"]


async def test_duplicate_serial_number_within_file_flagged(client, seeded_users):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    rows = [
        _base_row(**{"ID CODE": "101", "Serial Number": "SN-1"}),
        _base_row(**{"ID CODE": "102", "Serial Number": "SN-1"}),
    ]
    resp = await _preview(client, headers, rows)
    body = resp.json()
    assert body["failed"] == 2
    assert all("Duplicate Serial Number" in r["reason"] for r in body["rows"])


async def test_partial_row_failure_still_commits_valid_rows(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM201")
    await _seed_equipment(db_session, bcm_code="BCM202")

    rows = [
        _base_row(**{"ID CODE": "201"}),
        _base_row(**{"ID CODE": ""}),  # invalid, missing BCM
        _base_row(**{"ID CODE": "202"}),
    ]
    resp = await _commit(client, headers, rows, update_existing=True)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["succeeded"] == 2
    assert body["failed"] == 1

    result = await db_session.execute(select(Equipment).where(Equipment.bcm_code.in_(["BCM201", "BCM202"])))
    assert len(result.scalars().all()) == 2


# ---------------------------------------------------------------------------
# Update-only architecture (Repository Owner decision resolving review
# findings PR12-H1/H1R): import never creates new equipment and never
# generates, derives, or synthesizes an Asset Number.
# ---------------------------------------------------------------------------


async def test_non_matching_bcm_code_fails_row_and_creates_nothing(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    before = (await db_session.execute(select(Equipment))).scalars().all()

    resp = await _commit(client, headers, [_base_row(**{"ID CODE": "601"})], update_existing=True)
    body = resp.json()
    assert body["succeeded"] == 0
    assert body["failed"] == 1
    assert body["rows"][0]["reason"] == NO_MATCHING_EQUIPMENT_REASON
    assert body["rows"][0]["equipment_id"] is None

    after = (await db_session.execute(select(Equipment))).scalars().all()
    assert len(after) == len(before), "a non-matching BCM Code must never create a new equipment record"


async def test_non_matching_bcm_code_fails_even_with_update_existing_true(client, seeded_users):
    """update_existing only ever changes whether an *existing* match is
    updated or skipped -- it must never enable row creation."""
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await _commit(client, headers, [_base_row(**{"ID CODE": "602"})], update_existing=True)
    body = resp.json()
    assert body["failed"] == 1
    assert body["rows"][0]["reason"] == NO_MATCHING_EQUIPMENT_REASON


async def test_update_never_changes_asset_number(client, seeded_users, db_session):
    equipment_id = await _seed_equipment(db_session, bcm_code="BCM603", asset_number="AST-ORIGINAL-603")
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)

    resp = await _commit(
        client, headers, [_base_row(**{"ID CODE": "603", "Model": "Revised"})], update_existing=True
    )
    body = resp.json()
    assert body["succeeded"] == 1

    result = await db_session.execute(select(Equipment).where(Equipment.id == uuid.UUID(equipment_id)))
    equipment = result.scalar_one()
    assert equipment.asset_number == "AST-ORIGINAL-603"
    assert equipment.model == "Revised"


# ---------------------------------------------------------------------------
# Database-duplicate handling / update mode
# ---------------------------------------------------------------------------


async def test_update_existing_false_is_rejected_with_clear_error(client, seeded_users, db_session):
    """Review finding PR12-H1R: Roadmap PR12 is update-only, so
    `update_existing=false` no longer has a coherent meaning -- there is
    no create path for it to select instead. The request is rejected
    clearly and immediately (before any row is parsed), rather than
    silently accepted into a batch where nothing could ever succeed."""
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM301")
    row = _base_row(**{"ID CODE": "301"})

    preview_resp = await _preview(client, headers, [row], update_existing=False)
    assert preview_resp.status_code == 400
    assert preview_resp.json()["code"] == "INVALID_INPUT"
    assert "update-only" in preview_resp.json()["detail"]

    commit_resp = await _commit(client, headers, [row], update_existing=False)
    assert commit_resp.status_code == 400
    assert commit_resp.json()["code"] == "INVALID_INPUT"
    assert "update-only" in commit_resp.json()["detail"]

    # Nothing was written -- the rejection happens before any parsing.
    result = await db_session.execute(select(Equipment).where(Equipment.bcm_code == "BCM301"))
    equipment = result.scalar_one()
    assert equipment.equipment_name == "Seed Equipment"


async def test_update_existing_field_omitted_defaults_to_update_only(client, seeded_users, db_session):
    """Review 4782986913 non-blocking comment: the router declares
    `update_existing: bool = Form(default=True)`, so a client that omits
    the field entirely (not just one that sends "true") must still be
    accepted and treated as update-only, matching the explicit-true path."""
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    equipment_id = await _seed_equipment(db_session, bcm_code="BCM399", model="Original")
    row = _base_row(**{"ID CODE": "399", "Model": "Revised"})

    preview_resp = await client.post(
        "/api/v1/import/preview",
        headers=headers,
        files=_upload([row]),
    )
    assert preview_resp.status_code == 200
    assert preview_resp.json()["rows"][0]["action"] == "update"

    commit_resp = await client.post(
        "/api/v1/import/commit",
        headers=headers,
        files=_upload([row]),
    )
    assert commit_resp.status_code == 200
    body = commit_resp.json()
    assert body["succeeded"] == 1
    assert body["rows"][0]["equipment_id"] == equipment_id

    result = await db_session.execute(select(Equipment).where(Equipment.id == uuid.UUID(equipment_id)))
    equipment = result.scalar_one()
    assert equipment.model == "Revised"


async def test_db_duplicate_bcm_updates_master_data_when_update_mode_on(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    equipment_id = await _seed_equipment(db_session, bcm_code="BCM302", model="Original")

    updated_row = _base_row(**{"ID CODE": "302", "Model": "Revised"})
    resp = await _commit(client, headers, [updated_row], update_existing=True)
    body = resp.json()
    assert body["succeeded"] == 1
    assert body["rows"][0]["action"] == "update"
    assert body["rows"][0]["equipment_id"] == equipment_id

    result = await db_session.execute(select(Equipment).where(Equipment.id == uuid.UUID(equipment_id)))
    equipment = result.scalar_one()
    assert equipment.model == "Revised"


async def test_update_mode_never_touches_status_or_legacy_status(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    equipment_id = await _seed_equipment(db_session, bcm_code="BCM303")

    result = await db_session.execute(select(Equipment).where(Equipment.id == uuid.UUID(equipment_id)))
    equipment = result.scalar_one()
    equipment.status = EquipmentStatus.UNAVAILABLE_DEFECTIVE
    equipment.legacy_status = "manually_set"
    await db_session.commit()

    # Re-import the same BCM Code with a *different* Asset Status in update
    # mode -- status/legacy_status must remain exactly as manually set above.
    updated_row = _base_row(**{"ID CODE": "303", "Asset Status": "Decommissioned"})
    resp = await _commit(client, headers, [updated_row], update_existing=True)
    assert resp.json()["succeeded"] == 1

    result2 = await db_session.execute(select(Equipment).where(Equipment.id == uuid.UUID(equipment_id)))
    equipment2 = result2.scalar_one()
    assert equipment2.status.value == "unavailable_defective"
    assert equipment2.legacy_status == "manually_set"


async def test_update_mode_rejects_row_whose_item_no_belongs_to_a_different_record(client, seeded_users, db_session):
    """Review finding PR12-H3: update mode must reject a row whose Item
    No belongs to a *different* equipment record than the one BCM Code
    matched, even though it never writes Item No itself."""
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM901", item_no="ITEM-OTHER")
    await _seed_equipment(db_session, bcm_code="BCM902")

    conflicting_row = _base_row(**{"ID CODE": "902", "Item No.": "ITEM-OTHER"})
    resp = await _commit(client, headers, [conflicting_row], update_existing=True)
    body = resp.json()
    assert body["failed"] == 1
    assert "Item No" in body["rows"][0]["reason"]


async def test_update_mode_rejects_row_whose_serial_number_belongs_to_a_different_record(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM903", serial_number="SN-OTHER")
    await _seed_equipment(db_session, bcm_code="BCM904")

    conflicting_row = _base_row(**{"ID CODE": "904", "Serial Number": "SN-OTHER"})
    resp = await _commit(client, headers, [conflicting_row], update_existing=True)
    body = resp.json()
    assert body["failed"] == 1
    assert "Serial Number" in body["rows"][0]["reason"]


async def test_update_mode_allows_row_whose_item_no_and_serial_number_belong_to_the_same_record(
    client, seeded_users, db_session
):
    """A value that already belongs to the *same* BCM-matched record is
    not a conflict -- re-importing an unchanged row in update mode must
    still succeed."""
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    equipment_id = await _seed_equipment(db_session, bcm_code="BCM905", item_no="ITEM905", serial_number="SN905")

    row = _base_row(**{"ID CODE": "905", "Item No.": "ITEM905", "Serial Number": "SN905"})
    resp = await _commit(client, headers, [row], update_existing=True)
    body = resp.json()
    assert body["succeeded"] == 1
    assert body["rows"][0]["action"] == "update"
    assert body["rows"][0]["equipment_id"] == equipment_id


# ---------------------------------------------------------------------------
# Asset ID conservative duplicate detection (approved treatment)
# ---------------------------------------------------------------------------


async def test_asset_id_duplicate_in_file_flagged_but_others_unaffected(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM403")

    rows = [
        _base_row(**{"ID CODE": "401", "Asset ID": "A-1"}),
        _base_row(**{"ID CODE": "402", "Asset ID": "A-1"}),
        _base_row(**{"ID CODE": "403", "Asset ID": "A-2"}),
    ]
    resp = await _commit(client, headers, rows, update_existing=True)
    body = resp.json()
    assert body["succeeded"] == 1
    assert body["failed"] == 2
    by_bcm = {r["bcm_code"]: r for r in body["rows"]}
    assert by_bcm["BCM401"]["status"] == "failed"
    assert by_bcm["BCM402"]["status"] == "failed"
    assert by_bcm["BCM403"]["status"] == "success"


async def test_asset_id_duplicate_against_existing_db_row_flagged(client, seeded_users, db_session):
    await _seed_equipment(db_session, bcm_code="BCM501", asset_id="DUP-ASSET")
    await _seed_equipment(db_session, bcm_code="BCM502")
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)

    resp = await _commit(
        client, headers, [_base_row(**{"ID CODE": "502", "Asset ID": "DUP-ASSET"})], update_existing=True
    )
    body = resp.json()
    assert body["failed"] == 1
    assert "Asset ID" in body["rows"][0]["reason"]


async def test_no_unique_constraint_added_for_asset_id(db_session):
    """Roadmap PR12 approved treatment: asset_id must never gain a
    database uniqueness constraint (hospital-wide uniqueness unconfirmed).
    Two equipment rows sharing an asset_id must be a valid database
    state -- only application-layer import validation flags it."""
    e1 = await equipment_crud.create(
        db_session,
        data={"asset_number": "AST-D1", "equipment_name": "A", "asset_id": "SHARED"},
    )
    e2 = await equipment_crud.create(
        db_session,
        data={"asset_number": "AST-D2", "equipment_name": "B", "asset_id": "SHARED"},
    )
    await db_session.commit()
    assert e1.asset_id == e2.asset_id == "SHARED"


# ---------------------------------------------------------------------------
# Preview vs. commit isolation
# ---------------------------------------------------------------------------


async def test_preview_performs_zero_database_writes(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM701")

    before_equipment = (await db_session.execute(select(Equipment))).scalars().all()
    before_audit = (await db_session.execute(select(AuditLog))).scalars().all()

    resp = await _preview(client, headers, [_base_row(**{"ID CODE": "701"})], update_existing=True)
    assert resp.status_code == 200
    assert resp.json()["succeeded"] == 1

    after_equipment = (await db_session.execute(select(Equipment))).scalars().all()
    after_audit = (await db_session.execute(select(AuditLog))).scalars().all()
    assert len(after_equipment) == len(before_equipment)
    assert len(after_audit) == len(before_audit)


async def test_commit_does_not_require_a_prior_preview_call(client, seeded_users, db_session):
    """Commit independently re-parses/re-validates the raw uploaded file
    -- it never depends on or trusts a client-supplied preview result."""
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM702")

    resp = await _commit(client, headers, [_base_row(**{"ID CODE": "702"})], update_existing=True)
    assert resp.status_code == 200
    assert resp.json()["succeeded"] == 1


async def test_commit_produces_exactly_one_audit_log_entry_per_batch(client, seeded_users, db_session):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _seed_equipment(db_session, bcm_code="BCM801")
    await _seed_equipment(db_session, bcm_code="BCM802")
    before = (await db_session.execute(select(AuditLog).where(AuditLog.action == "import"))).scalars().all()

    rows = [_base_row(**{"ID CODE": "801"}), _base_row(**{"ID CODE": "802"}), _base_row(**{"ID CODE": ""})]
    resp = await _commit(client, headers, rows, update_existing=True)
    body = resp.json()
    assert body["audit_log_id"] is not None

    after = (await db_session.execute(select(AuditLog).where(AuditLog.action == "import"))).scalars().all()
    assert len(after) == len(before) + 1
    entry = [a for a in after if str(a.id) == body["audit_log_id"]][0]
    assert entry.after_data["succeeded"] == 2
    assert entry.after_data["failed"] == 1


# ---------------------------------------------------------------------------
# Upload bounds (review finding PR12-H2)
# ---------------------------------------------------------------------------


async def test_oversized_upload_rejected_before_parsing(client, seeded_users):
    from app.services import import_service

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    oversized = b"x" * (import_service.MAX_UPLOAD_BYTES + 1)
    resp = await client.post(
        "/api/v1/import/preview",
        headers=headers,
        files={"file": ("big.xlsx", oversized, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"update_existing": "true"},
    )
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"
    assert "maximum allowed size" in resp.json()["detail"]


async def test_too_many_rows_rejected(client, seeded_users):
    from app.services import import_service

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    rows = [_base_row(**{"ID CODE": str(1000 + i)}) for i in range(import_service.MAX_IMPORT_ROWS + 1)]
    resp = await _preview(client, headers, rows)
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"
    assert "maximum a single import batch" in resp.json()["detail"]


async def test_filename_exceeding_max_length_rejected(client, seeded_users):
    from app.services import import_service

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    long_name = ("a" * (import_service.MAX_FILENAME_LENGTH + 1)) + ".xlsx"
    content = _build_workbook([_base_row()])
    resp = await client.post(
        "/api/v1/import/preview",
        headers=headers,
        files={"file": (long_name, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"update_existing": "true"},
    )
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"
    assert "Filename exceeds" in resp.json()["detail"]


async def test_non_xlsx_extension_rejected(client, seeded_users):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    content = _build_workbook([_base_row()])
    resp = await client.post(
        "/api/v1/import/preview",
        headers=headers,
        files={"file": ("inventory.csv", content, "text/csv")},
        data={"update_existing": "true"},
    )
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"
    assert "Only .xlsx files are supported" in resp.json()["detail"]


# ---------------------------------------------------------------------------
# Preview-phase field length validation (review finding PR12-H4)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "field,header",
    [
        ("Asset ID", "Asset ID"),
        ("Serial Number", "Serial Number"),
        ("Equipment Name", "Equipment Name"),
        ("Manufacturer", "Manufacturer"),
        ("Model", "Model"),
        ("Asset Status", "Asset Status"),
    ],
)
async def test_overlength_field_fails_in_preview_not_just_commit(client, seeded_users, field, header):
    from app.services.import_service import FIELD_MAX_LENGTHS

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    overlength_value = "x" * (FIELD_MAX_LENGTHS[field] + 1)
    # Deliberately a non-matching BCM Code -- the length check (Pass 1)
    # happens before any database lookup, so this must fail on length,
    # never on "no matching equipment".
    row = _base_row(**{"ID CODE": "951", header: overlength_value})

    preview_resp = await _preview(client, headers, [row])
    preview_body = preview_resp.json()
    assert preview_body["failed"] == 1
    assert field in preview_body["rows"][0]["reason"]

    # Committing the same row must fail the same way (never a raw DB
    # DataError) -- preview and commit never disagree on whether a row
    # will fit the database.
    commit_resp = await _commit(client, headers, [row])
    commit_body = commit_resp.json()
    assert commit_body["failed"] == 1
    assert field in commit_body["rows"][0]["reason"]


# ---------------------------------------------------------------------------
# Decompressed XLSX archive bounds (review finding PR12-H2R). An .xlsx file
# is a ZIP container -- these prove a small, highly-compressible upload
# (well under MAX_UPLOAD_BYTES on the wire) is still rejected before
# openpyxl is allowed to decompress its content.
# ---------------------------------------------------------------------------


def _build_raw_zip(entries: dict[str, bytes], *, compress: bool = True) -> bytes:
    buf = io.BytesIO()
    compression = zipfile.ZIP_DEFLATED if compress else zipfile.ZIP_STORED
    with zipfile.ZipFile(buf, "w", compression=compression) as archive:
        for name, content in entries.items():
            archive.writestr(name, content)
    return buf.getvalue()


async def _upload_raw_bytes(client, headers, content: bytes, filename: str = "inventory.xlsx"):
    return await client.post(
        "/api/v1/import/preview",
        headers=headers,
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"update_existing": "true"},
    )


async def test_zip_entry_exceeding_uncompressed_size_limit_rejected(client, seeded_users):
    from app.services import import_service

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    # Highly compressible content: tiny on the wire, far larger once
    # expanded than any legitimate inventory-import spreadsheet member.
    oversized_member = b"0" * (import_service.MAX_ZIP_ENTRY_UNCOMPRESSED_BYTES + 1)
    content = _build_raw_zip({"xl/sharedStrings.xml": oversized_member})
    assert len(content) < import_service.MAX_UPLOAD_BYTES

    resp = await _upload_raw_bytes(client, headers, content)
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"


async def test_zip_total_uncompressed_size_exceeding_aggregate_limit_rejected(client, seeded_users):
    from app.services import import_service

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    # Each entry stays under the per-entry cap, but three of them together
    # exceed the aggregate cap.
    per_entry = 21 * 1024 * 1024
    assert per_entry < import_service.MAX_ZIP_ENTRY_UNCOMPRESSED_BYTES
    assert per_entry * 3 > import_service.MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES
    entries = {
        "xl/sharedStrings.xml": b"1" * per_entry,
        "xl/styles.xml": b"2" * per_entry,
        "xl/theme/theme1.xml": b"3" * per_entry,
    }
    content = _build_raw_zip(entries)
    assert len(content) < import_service.MAX_UPLOAD_BYTES

    resp = await _upload_raw_bytes(client, headers, content)
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"


async def test_zip_suspicious_compression_ratio_rejected(client, seeded_users):
    from app.services import import_service

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    # Under the per-entry absolute size cap, but a single repeated byte
    # compresses at a ratio far beyond MAX_ZIP_COMPRESSION_RATIO.
    member_size = min(20 * 1024 * 1024, import_service.MAX_ZIP_ENTRY_UNCOMPRESSED_BYTES - 1)
    content = _build_raw_zip({"xl/sharedStrings.xml": b"\x00" * member_size})
    assert len(content) < import_service.MAX_UPLOAD_BYTES

    resp = await _upload_raw_bytes(client, headers, content)
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"


async def test_zip_entry_with_disallowed_path_rejected(client, seeded_users):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    content = _build_raw_zip({"unexpected/payload.bin": b"hello"})

    resp = await _upload_raw_bytes(client, headers, content)
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"


async def test_zip_path_traversal_entry_rejected(client, seeded_users):
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    content = _build_raw_zip({"../../etc/passwd": b"hello"})

    resp = await _upload_raw_bytes(client, headers, content)
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"


async def test_zip_too_many_entries_rejected(client, seeded_users):
    from app.services import import_service

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    entries = {f"xl/worksheets/sheet{i}.xml": b"x" for i in range(import_service.MAX_ZIP_ENTRIES + 1)}
    content = _build_raw_zip(entries)

    resp = await _upload_raw_bytes(client, headers, content)
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"


async def test_too_many_worksheets_rejected(client, seeded_users):
    from app.services import import_service

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    wb = Workbook()
    ws = wb.active
    ws.append(list(REQUIRED_HEADERS))
    for i in range(import_service.MAX_WORKSHEET_COUNT + 1):
        wb.create_sheet(f"Extra{i}")
    buf = io.BytesIO()
    wb.save(buf)

    resp = await _upload_raw_bytes(client, headers, buf.getvalue())
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"
    assert "worksheets" in resp.json()["detail"]


async def test_too_many_header_columns_rejected(client, seeded_users):
    from app.services import import_service

    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    wb = Workbook()
    ws = wb.active
    padded_headers = list(REQUIRED_HEADERS) + [f"Extra{i}" for i in range(import_service.MAX_HEADER_COLUMNS)]
    ws.append(padded_headers)
    buf = io.BytesIO()
    wb.save(buf)

    resp = await _upload_raw_bytes(client, headers, buf.getvalue())
    assert resp.status_code == 400
    assert resp.json()["code"] == "INVALID_INPUT"
    assert "columns" in resp.json()["detail"]


async def test_valid_small_workbook_still_accepted_after_zip_bounds_checks(client, seeded_users, db_session):
    """Sanity check: the new archive-bounds validation must not reject a
    normal, legitimate import file -- it must still reach full per-row
    processing."""
    await _seed_equipment(db_session, bcm_code="BCM961")
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await _preview(client, headers, [_base_row(**{"ID CODE": "961"})], update_existing=True)
    assert resp.status_code == 200
    assert resp.json()["succeeded"] == 1


# ---------------------------------------------------------------------------
# raw_source_status verbatim preservation (review finding PR12-#2).
# raw_source_status is an audit field and must store the exact source-cell
# text -- business validation (the Asset Status mapping lookup) uses a
# separately normalized copy, never mutating the persisted value.
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "source_value",
    [
        "  Active",  # leading spaces
        "Active  ",  # trailing spaces
        "  Active  ",  # leading and trailing spaces
        "AcTiVe",  # mixed casing
        "aCTIVE",  # mixed casing, different pattern
    ],
)
async def test_raw_source_status_preserves_exact_source_cell_text(client, seeded_users, db_session, source_value):
    equipment_id = await _seed_equipment(db_session, bcm_code="BCM971")
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)

    resp = await _commit(
        client, headers, [_base_row(**{"ID CODE": "971", "Asset Status": source_value})], update_existing=True
    )
    body = resp.json()
    assert body["succeeded"] == 1, body

    result = await db_session.execute(select(Equipment).where(Equipment.id == uuid.UUID(equipment_id)))
    equipment = result.scalar_one()
    assert equipment.raw_source_status == source_value, (
        f"raw_source_status must preserve the source cell text exactly, "
        f"got {equipment.raw_source_status!r} for source {source_value!r}"
    )


async def test_raw_source_status_verbatim_preview_matches_commit(client, seeded_users, db_session):
    """Preview and commit must agree on the verbatim value (both re-parse
    the same raw file independently -- see process_import's docstring)."""
    await _seed_equipment(db_session, bcm_code="BCM972")
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)
    row = _base_row(**{"ID CODE": "972", "Asset Status": "  In Use  "})

    preview_resp = await _preview(client, headers, [row], update_existing=True)
    assert preview_resp.json()["succeeded"] == 1

    commit_resp = await _commit(client, headers, [row], update_existing=True)
    body = commit_resp.json()
    assert body["succeeded"] == 1
    equipment_id = body["rows"][0]["equipment_id"]

    result = await db_session.execute(select(Equipment).where(Equipment.id == uuid.UUID(equipment_id)))
    equipment = result.scalar_one()
    assert equipment.raw_source_status == "  In Use  "


async def test_raw_source_status_verbatim_value_still_maps_correctly_despite_whitespace_and_casing(
    client, seeded_users, db_session
):
    """The verbatim value is stored as-is, but business validation (status
    mapping) must still succeed via a separately normalized copy -- an
    unusual but valid source value must not spuriously fail the row."""
    equipment_id = await _seed_equipment(db_session, bcm_code="BCM973")
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)

    resp = await _commit(
        client, headers, [_base_row(**{"ID CODE": "973", "Asset Status": "  DeFECTIVE  "})], update_existing=True
    )
    body = resp.json()
    assert body["succeeded"] == 1, body

    result = await db_session.execute(select(Equipment).where(Equipment.id == uuid.UUID(equipment_id)))
    equipment = result.scalar_one()
    # Raw value preserved verbatim...
    assert equipment.raw_source_status == "  DeFECTIVE  "
    # ...while the normalized value correctly drove the (unpersisted,
    # since update mode never writes status) mapping to UNAVAILABLE_DEFECTIVE
    # -- proven indirectly by the row succeeding rather than failing as
    # "Unrecognized Asset Status value".
    assert equipment.status.value == "available_at_pool"  # unchanged -- update never writes status


@pytest.mark.parametrize(
    "source_value",
    [
        "  Mystery Value  ",  # leading/trailing whitespace
        "In  Use",  # internal whitespace (double space) -- normalization
        # is strip+lower only, so this legitimately does not match the
        # "in use" mapping key; the point here is that the exact source
        # text still appears in the failure reason, not a collapsed copy.
        "In\tUse",  # internal whitespace (tab)
    ],
)
async def test_unrecognized_asset_status_error_message_shows_verbatim_text(
    client, seeded_users, db_session, source_value
):
    """The per-row failure reason must quote what the user actually typed,
    including any leading/trailing/internal whitespace, not a
    silently-normalized copy."""
    await _seed_equipment(db_session, bcm_code="BCM974")
    headers = await _auth_headers(client, ROLE_ADMINISTRATOR)

    resp = await _preview(
        client, headers, [_base_row(**{"ID CODE": "974", "Asset Status": source_value})], update_existing=True
    )
    body = resp.json()
    assert body["failed"] == 1
    assert f"'{source_value}'" in body["rows"][0]["reason"]

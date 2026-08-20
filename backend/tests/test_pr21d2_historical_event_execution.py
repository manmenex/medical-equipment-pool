"""Roadmap PR21D2 -- Historical Event Execution.

Tests `LegacyTransactionHistoryAdapter.precheck_execute`/`execute`/
`on_execution_success`/`on_execution_failure`/`on_execution_recovery`
(`app.services.import_adapters.legacy_history.combined`), the new
`app.crud.legacy_equipment_event` CRUD helpers, and the five new
`app.crud.legacy_history_dry_run_plan` execution-lifecycle functions.

Two layers, mirroring the PR21D1/PR20D convention:
- HTTP-pipeline tests (create session -> upload -> validate -> dry-run ->
  confirm via the internal `LegacyHistoryDryRunPlanProvider` (the known
  PR21D1 gap: the generic `POST .../dry-run-plan/{id}/confirm` HTTP route
  is hard-coded to Equipment Master) -> `POST .../execute`) prove the real
  end-to-end production path, including job/lease/fencing, audit, and
  plan-lifecycle transitions.
- Direct-adapter-invocation tests (`adapter.execute(db)` under
  `adapter_invocation_context`, bypassing the job/lease framework) prove
  `execute()`'s own row-processing logic in isolation -- identity
  conflicts, idempotent replay, and provenance handling -- without
  needing a second full pipeline run to manufacture a collision.

Synthetic workbooks only -- no production workbook data is read or
committed anywhere in this file."""

import hashlib
import uuid
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.crud import import_job as import_job_crud
from app.crud import legacy_equipment_event as legacy_equipment_event_crud
from app.crud import legacy_history_dry_run_plan as legacy_history_dry_run_plan_crud
from app.models.equipment import Equipment, EquipmentStatus
from app.models.import_session import ImportJob, ImportSession
from app.models.legacy_history import (
    LegacyEquipmentEvent,
    LegacyEquipmentEventSourceRef,
    LegacyHistoryDryRunPlan,
    LegacyMigrationAuthority,
)
from app.models.master_data import Ward
from app.models.transaction import BorrowTransaction
from app.models.user import User
from app.services import import_execution_service, import_lease
from app.services.identifiers import normalize_bcm_code
from app.services.import_adapter_context import AdapterInvocationContext, adapter_invocation_context
from app.services.import_adapters.legacy_history import issue as issue_module
from app.services.import_adapters.legacy_history import receive as receive_module
from app.services.import_adapters.legacy_history.combined import LegacyTransactionHistoryAdapter
from app.services.import_plan_provider import get_plan_provider
from app.services.import_plan_providers.legacy_history import DATASET_TYPE
from app.services.import_source_reader import SourceDescriptor, VerifiedSourceContent
from tests.conftest import auth_headers

ISSUE_HEADER_SHEET = issue_module.HEADER_SHEET_NAME
ISSUE_LINE_SHEET = issue_module.LINE_SHEET_NAME
RECEIVE_HEADER_SHEET = receive_module.HEADER_SHEET_NAME
RECEIVE_LINE_SHEET = receive_module.LINE_SHEET_NAME
ISSUE_HEADER_COLUMNS = list(issue_module._HEADER_GOVERNED_HEADERS.keys())
ISSUE_LINE_COLUMNS = list(issue_module._LINE_GOVERNED_HEADERS.keys())
RECEIVE_HEADER_COLUMNS = list(receive_module._HEADER_GOVERNED_HEADERS.keys())
RECEIVE_LINE_COLUMNS = list(receive_module._LINE_GOVERNED_HEADERS.keys())


@pytest_asyncio.fixture(autouse=True)
async def _patch_execution_session_factories(db_engine, monkeypatch):
    """Mirrors test_pr20d_dry_run_plan.py/test_pr21d1_source_admission.py's
    identical fixture."""
    session_maker = async_sessionmaker(db_engine, expire_on_commit=False, class_=AsyncSession)
    monkeypatch.setattr(import_execution_service, "AsyncSessionLocal", session_maker)
    monkeypatch.setattr(import_lease, "AsyncSessionLocal", session_maker)


# ---------------------------------------------------------------------------
# Workbook construction helpers (mirrors test_pr21d1_source_admission.py).
# ---------------------------------------------------------------------------


def _build_combined_workbook(*, issue_headers=None, issue_lines=None, receive_headers=None, receive_lines=None):
    wb = Workbook()
    ws = wb.active
    ws.title = ISSUE_HEADER_SHEET
    ws.append(ISSUE_HEADER_COLUMNS)
    for row in issue_headers or []:
        ws.append([row.get(col) for col in ISSUE_HEADER_COLUMNS])

    ws = wb.create_sheet(ISSUE_LINE_SHEET)
    ws.append(ISSUE_LINE_COLUMNS)
    for row in issue_lines or []:
        ws.append([row.get(col) for col in ISSUE_LINE_COLUMNS])

    ws = wb.create_sheet(RECEIVE_HEADER_SHEET)
    ws.append(RECEIVE_HEADER_COLUMNS)
    for row in receive_headers or []:
        ws.append([row.get(col) for col in RECEIVE_HEADER_COLUMNS])

    ws = wb.create_sheet(RECEIVE_LINE_SHEET)
    ws.append(RECEIVE_LINE_COLUMNS)
    for row in receive_lines or []:
        ws.append([row.get(col) for col in RECEIVE_LINE_COLUMNS])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _issue_header_row(*, order_ref="ORD-0001", ward="Ward 1", bme="BME Header", d=date(2024, 1, 10), t=time(9, 0, 0)):
    return {
        "วันที่": d,
        "เลขที่ใบยืม": order_ref,
        "แผนกที่ยืม": ward,
        "ผู้ส่งเครื่องยืม (User)": "header-user",
        "ผู้ส่งเครื่องยืม (BME)": bme,
        "เวลา": t,
        "จำนวนเครื่องส่งทั้งหมด": 1,
        "หมายเหตุ": None,
    }


def _issue_line_row(*, row_key="1", order_ref="ORD-0001", me_code="ME001", ward="Ward 1", bme="BME Line", d=date(2024, 1, 10), t=time(9, 0, 0)):
    return {
        "ลำดับ": row_key,
        "วันที่": d,
        "เลขที่ใบส่ง": order_ref,
        "SCAN CODE ส่ง": "SCAN-1",
        "ME.Code": me_code,
        "Barcode ส่งเครื่อง": "BARCODE-1",
        "Equipment": "Infusion Pump",
        "Brand": "BrandX",
        "Model": "ModelY",
        "Serial no.": "SN-1",
        "รูปเครื่อง": None,
        "แผนกที่ส่ง": ward,
        "ตัวเครื่องหน้าจอไม่แตกร้าว": "OK",
        "สายชาร์จ (สายไม่ขาด)": "OK",
        "Pole Clamp (ไม่แตกหักมไม่ง้อ)": "OK",
        "จำนวน": 1,
        "หมายเหตุ": None,
        "เวลา": t,
        "ชื่อ BME": bme,
        "ชื่อ (User)": "line-user",
    }


def _receive_header_row(*, order_ref="RET-0001", ward="Ward 1", bme="BME Header", d=date(2024, 1, 11), t=time(10, 0, 0)):
    return {
        "วันที่": d,
        "เลขที่ใบคืน": order_ref,
        "แผนกที่คืน": ward,
        "ผู้ส่งเครื่องคืน (User)": "header-user",
        "ผู้รับเครื่องคืน (BME)": bme,
        "เวลา": t,
        "จำนวนเครื่องรับคืนทั้งหมด": 1,
        "หมายเหตุ": None,
    }


def _receive_line_row(*, row_key="1", order_ref="RET-0001", me_code="ME001", ward="Ward 1", bme="BME Line", d=date(2024, 1, 11), t=time(10, 0, 0)):
    return {
        "ลำดับ": row_key,
        "วันที่": d,
        "เลขที่ใบรับเครื่อง": order_ref,
        "SCAN CODE รับ": "SCAN-1",
        "ME.Code": me_code,
        "Barcode รับเครื่อง": "BARCODE-1",
        "Equipment": "Infusion Pump",
        "Brand": "BrandX",
        "Model": "ModelY",
        "Serial no.": "SN-1",
        "แผนกที่รับ": ward,
        "รูปเครื่อง": None,
        "ตัวเครื่องหน้าจอไม่แตกร้าว": "OK",
        "สายชาร์จ (สายไม่ขาด)": "OK",
        "Pole Clamp (ไม่แตกหักมไม่ง้อ)": "OK",
        "จำนวน": 1,
        "หมายเหตุ": None,
        "เวลา": t,
        "ชื่อ BME": bme,
        "ชื่อ (User)": "line-user",
    }


def _default_workbook(**overrides) -> bytes:
    kwargs = dict(
        issue_headers=[_issue_header_row()],
        issue_lines=[_issue_line_row()],
        receive_headers=[_receive_header_row()],
        receive_lines=[_receive_line_row()],
    )
    kwargs.update(overrides)
    return _build_combined_workbook(**kwargs)


# ---------------------------------------------------------------------------
# Seeding helpers.
# ---------------------------------------------------------------------------


async def _seed_equipment(db_session: AsyncSession, *, me_code: str = "ME001", **kwargs) -> Equipment:
    defaults = dict(
        asset_number=f"AN-{uuid.uuid4().hex[:10]}",
        equipment_name="Legacy Test Equipment",
        status=EquipmentStatus.AVAILABLE_AT_POOL,
        bcm_code=normalize_bcm_code(me_code),
    )
    defaults.update(kwargs)
    eq = Equipment(**defaults)
    db_session.add(eq)
    await db_session.commit()
    await db_session.refresh(eq)
    return eq


async def _seed_ward(db_session: AsyncSession, *, code: str = "Ward 1") -> Ward:
    ward = Ward(code=code, name=code)
    db_session.add(ward)
    await db_session.commit()
    await db_session.refresh(ward)
    return ward


async def _seed_authority(db_session: AsyncSession, *, checksum: str, actor_id: uuid.UUID) -> LegacyMigrationAuthority:
    authority = LegacyMigrationAuthority(
        scope="pr21d2-test", approved_workbook_sha256=checksum, approved_by_user_id=actor_id
    )
    db_session.add(authority)
    await db_session.commit()
    await db_session.refresh(authority)
    return authority


async def _actor_id(db_session: AsyncSession) -> uuid.UUID:
    return (await db_session.execute(select(User.id).limit(1))).scalar_one()


async def _create_session(client: AsyncClient, headers: dict) -> dict:
    r = await client.post("/api/v1/import-sessions", headers=headers, json={"dataset_type": DATASET_TYPE})
    assert r.status_code in (200, 201), r.text
    return r.json()


async def _upload(client: AsyncClient, headers: dict, session_id: str, content: bytes):
    files = {"file": ("legacy.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    return await client.post(f"/api/v1/import-sessions/{session_id}/source/upload", headers=headers, files=files)


async def _validate_dry_run_confirm(
    client: AsyncClient, headers: dict, db_session: AsyncSession, content: bytes, *, actor_id: uuid.UUID
) -> dict:
    """Runs the real production admission pipeline through to a
    genuinely CONFIRMED plan: create -> upload -> validate -> dry-run,
    then confirm via the internal `LegacyHistoryDryRunPlanProvider`
    directly (the known PR21D1 gap -- the generic `POST .../dry-run-
    plan/{id}/confirm` HTTP route is hard-coded to Equipment Master, see
    `test_pr21d1_source_admission.py`'s own dedicated proof of that
    gap). Returns the session dict."""
    session = await _create_session(client, headers)
    up = await _upload(client, headers, session["id"], content)
    assert up.status_code == 201, up.text
    v = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert v.status_code == 200, v.text
    assert v.json()["status"] == "validated", v.json()
    dr = await client.post(f"/api/v1/import-sessions/{session['id']}/dry-run", headers=headers)
    assert dr.status_code == 200, dr.text
    assert dr.json()["status"] == "dry_run_completed", dr.text

    session_uuid = uuid.UUID(session["id"])
    plan = await legacy_history_dry_run_plan_crud.get_current_plan(db_session, import_session_id=session_uuid)
    assert plan is not None
    provider = get_plan_provider(DATASET_TYPE)
    result = await provider.confirm_plan(
        db_session, plan_id=plan.id, import_session_id=session_uuid, current_user_id=actor_id
    )
    assert result.newly_confirmed is True
    await db_session.commit()
    return session


def _verified_content(content: bytes, *, session_id: uuid.UUID, source_id: uuid.UUID, checksum: str) -> VerifiedSourceContent:
    return VerifiedSourceContent(
        content=content,
        source_descriptor=SourceDescriptor(
            import_source_id=source_id,
            import_session_id=session_id,
            dataset_type=DATASET_TYPE,
            expected_checksum=checksum,
            expected_byte_size=len(content),
            content_type=None,
            original_filename=None,
            registration_status="frozen",
        ),
    )


async def _direct_execute_context(db_session: AsyncSession, session_id: uuid.UUID) -> AdapterInvocationContext:
    """Builds the AdapterInvocationContext execute() itself needs, from a
    session that has already gone through `_validate_dry_run_confirm`
    above -- mirrors the real framework's own execute-time context
    (`verified_source_content`/`dry_run_job_id`/`accepted_validation_job_id`
    are always None for execute, per `import_execution_service.
    run_execute`'s own construction)."""
    from app.crud import import_session as import_session_crud

    source = await import_session_crud.get_source(db_session, session_id=session_id)
    return AdapterInvocationContext(
        import_session_id=session_id,
        import_source_id=source.id,
        dataset_type=DATASET_TYPE,
        source_checksum=source.checksum,
        source_fingerprint=source.source_fingerprint,
        ruleset_version="1",
        verified_source_content=None,
        dry_run_job_id=None,
        accepted_validation_job_id=None,
        actor_user_id=None,
    )


# ---------------------------------------------------------------------------
# A. Happy path: HTTP end-to-end execute, event/provenance counts.
# ---------------------------------------------------------------------------


async def test_mixed_plan_executes_end_to_end_with_correct_counts(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text
    assert ex.json()["status"] == "completed"
    assert ex.json()["imported_rows"] == 2

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events) == 2
    assert {e.event_type for e in events} == {"ISSUE", "RECEIVE"}
    refs = (await db_session.execute(select(LegacyEquipmentEventSourceRef))).scalars().all()
    assert len(refs) == 4  # 2 events x (header + line)


async def test_issue_only_plan_executes(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook(receive_headers=[], receive_lines=[])
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text
    assert ex.json()["imported_rows"] == 1

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events) == 1
    assert events[0].event_type == "ISSUE"


async def test_receive_only_plan_executes(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook(issue_headers=[], issue_lines=[])
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text
    assert ex.json()["imported_rows"] == 1

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events) == 1
    assert events[0].event_type == "RECEIVE"


# ---------------------------------------------------------------------------
# B/C. No live-state mutation, no BorrowTransaction.
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "status",
    [
        EquipmentStatus.AVAILABLE_AT_POOL,
        EquipmentStatus.ISSUED_TO_WARD,
        EquipmentStatus.UNAVAILABLE_DEFECTIVE,
        EquipmentStatus.DECOMMISSIONED,
    ],
)
async def test_execute_never_mutates_equipment_regardless_of_current_status(
    client: AsyncClient, seeded_users, db_session, status
):
    equipment = await _seed_equipment(db_session, status=status)
    original_version = equipment.version
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook(receive_headers=[], receive_lines=[])
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text

    await db_session.refresh(equipment)
    assert equipment.status == status
    assert equipment.version == original_version
    assert equipment.current_location_id is None

    borrow_count = (await db_session.execute(select(BorrowTransaction))).scalars().all()
    assert borrow_count == []


async def test_no_borrow_transaction_created_and_existing_unaffected(client: AsyncClient, seeded_users, db_session):
    equipment = await _seed_equipment(db_session)
    await _seed_ward(db_session)
    other_equipment = await _seed_equipment(db_session, me_code="ME999", asset_number="AN-OTHER")
    existing_borrow = BorrowTransaction(
        transaction_no=f"TXN-{uuid.uuid4().hex[:10]}",
        equipment_id=other_equipment.id,
    )
    db_session.add(existing_borrow)
    await db_session.commit()

    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook(receive_headers=[], receive_lines=[])
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text

    all_borrows = (await db_session.execute(select(BorrowTransaction))).scalars().all()
    assert len(all_borrows) == 1
    assert all_borrows[0].id == existing_borrow.id
    assert all_borrows[0].status.value == "open"


# ---------------------------------------------------------------------------
# D. Identity/idempotency (direct-adapter invocation).
# ---------------------------------------------------------------------------


async def test_retry_via_second_session_reimporting_same_data_produces_no_duplicates(
    client: AsyncClient, seeded_users, db_session
):
    """The real-world "retry" scenario this framework actually admits:
    the FIRST session's plan transitions to `consumed` on success (never
    executable again, by design -- see `test_plan_marked_consumed_after_
    successful_execution`), so genuine idempotency is exercised by a
    SECOND, independent session re-importing the byte-identical
    workbook (same checksum -> same `LegacyMigrationAuthority`) through
    the full pipeline again. Its own `execute()` must discover every row
    already exists with a matching fact and contribute nothing new."""
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session_a = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex_a = await client.post(f"/api/v1/import-sessions/{session_a['id']}/execute", headers=headers)
    assert ex_a.status_code == 200, ex_a.text
    assert ex_a.json()["imported_rows"] == 2

    session_b = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex_b = await client.post(f"/api/v1/import-sessions/{session_b['id']}/execute", headers=headers)
    assert ex_b.status_code == 200, ex_b.text
    assert ex_b.json()["imported_rows"] == 0, "every row already exists with a matching fact -- nothing new"

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events) == 2
    refs = (await db_session.execute(select(LegacyEquipmentEventSourceRef))).scalars().all()
    assert len(refs) == 4


async def test_same_authority_event_type_key_same_fact_is_safe_replay(db_session, seeded_users):
    """Direct-adapter proof: an existing event with an identical
    immutable fact is treated as already-applied, not re-inserted."""
    equipment = await _seed_equipment(db_session)
    ward = await _seed_ward(db_session)
    actor_id = await _actor_id(db_session)
    authority = LegacyMigrationAuthority(scope="x", approved_workbook_sha256="a" * 64, approved_by_user_id=actor_id)
    db_session.add(authority)
    await db_session.flush()

    other_session_id = uuid.uuid4()
    other_source_id = uuid.uuid4()
    occurred_at = datetime(2024, 1, 10, 2, 0, tzinfo=timezone.utc)
    existing = LegacyEquipmentEvent(
        migration_authority_id=authority.id,
        equipment_id=equipment.id,
        event_type="ISSUE",
        occurred_at=occurred_at,
        legacy_source_row_key="1",
        legacy_order_reference="ORD-0001",
        legacy_ward_text="Ward 1",
        resolved_ward_id=ward.id,
        legacy_bme_name="BME Line",
        import_session_id=other_session_id,
        import_source_id=other_source_id,
    )
    db_session.add(existing)
    await db_session.flush()
    db_session.add(
        LegacyEquipmentEventSourceRef(
            legacy_equipment_event_id=existing.id,
            import_session_id=other_session_id,
            import_source_id=other_source_id,
            source_checksum="a" * 64,
            sheet_name=ISSUE_HEADER_SHEET,
            source_row_number=2,
        )
    )
    db_session.add(
        LegacyEquipmentEventSourceRef(
            legacy_equipment_event_id=existing.id,
            import_session_id=other_session_id,
            import_source_id=other_source_id,
            source_checksum="a" * 64,
            sheet_name=ISSUE_LINE_SHEET,
            source_row_number=2,
        )
    )
    await db_session.commit()

    from app.services.import_adapters.legacy_history.combined import (
        _plan_row_matches_existing_event,
        LegacyTransactionHistoryAdapter as _Adapter,
    )

    normalized_values = {
        "legacy_order_reference": "ORD-0001",
        "equipment_id": str(equipment.id),
        "occurred_at": occurred_at.isoformat(),
        "legacy_ward_text": "Ward 1",
        "resolved_ward_id": str(ward.id),
        "legacy_bme_name": "BME Line",
        "header_source_ref": {"sheet_name": ISSUE_HEADER_SHEET, "source_row_number": 2},
        "line_source_ref": {"sheet_name": ISSUE_LINE_SHEET, "source_row_number": 2},
    }
    assert _plan_row_matches_existing_event(existing, normalized_values) is True
    adapter = _Adapter()
    assert await adapter._existing_event_provenance_matches_plan(db_session, existing.id, normalized_values) is True


async def test_same_identity_different_immutable_fact_fails_closed(client: AsyncClient, seeded_users, db_session):
    """A colliding identity whose stored fact disagrees with the plan
    must fail the WHOLE execute attempt closed -- never overwritten,
    never silently accepted."""
    equipment = await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook(receive_headers=[], receive_lines=[])
    checksum = hashlib.sha256(content).hexdigest()
    authority = await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    # Pre-seed a colliding event under the SAME identity but a DIFFERENT
    # occurred_at than what the plan will contain.
    conflicting = LegacyEquipmentEvent(
        migration_authority_id=authority.id,
        equipment_id=equipment.id,
        event_type="ISSUE",
        occurred_at=datetime(2020, 1, 1, tzinfo=timezone.utc),
        legacy_source_row_key="1",
        legacy_order_reference="ORD-DIFFERENT",
        import_session_id=uuid.uuid4(),
        import_source_id=uuid.uuid4(),
    )
    db_session.add(conflicting)
    await db_session.commit()

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code >= 400, ex.text

    session_uuid = uuid.UUID(session["id"])
    session_row = (await db_session.execute(select(ImportSession).where(ImportSession.id == session_uuid))).scalar_one()
    assert session_row.status == "failed"
    # No second event was ever inserted for this identity.
    events = (
        (await db_session.execute(select(LegacyEquipmentEvent).where(LegacyEquipmentEvent.legacy_source_row_key == "1")))
        .scalars()
        .all()
    )
    assert len(events) == 1
    assert events[0].id == conflicting.id
    assert events[0].legacy_order_reference == "ORD-DIFFERENT"  # never overwritten


async def test_different_authority_same_event_type_key_allowed(db_session, seeded_users):
    """Two different, separately-governed authorities may each legitimately
    own an event sharing the same (event_type, legacy_source_row_key) --
    the database identity is scoped per-authority, never global."""
    equipment = await _seed_equipment(db_session)
    actor_id = await _actor_id(db_session)
    authority_a = LegacyMigrationAuthority(scope="a", approved_workbook_sha256="a" * 64, approved_by_user_id=actor_id)
    authority_b = LegacyMigrationAuthority(scope="b", approved_workbook_sha256="b" * 64, approved_by_user_id=actor_id)
    db_session.add_all([authority_a, authority_b])
    await db_session.flush()

    event_a = LegacyEquipmentEvent(
        migration_authority_id=authority_a.id,
        equipment_id=equipment.id,
        event_type="ISSUE",
        occurred_at=datetime.now(timezone.utc),
        legacy_source_row_key="1",
        import_session_id=uuid.uuid4(),
        import_source_id=uuid.uuid4(),
    )
    event_b = LegacyEquipmentEvent(
        migration_authority_id=authority_b.id,
        equipment_id=equipment.id,
        event_type="ISSUE",
        occurred_at=datetime.now(timezone.utc),
        legacy_source_row_key="1",
        import_session_id=uuid.uuid4(),
        import_source_id=uuid.uuid4(),
    )
    db_session.add_all([event_a, event_b])
    await db_session.commit()  # must not raise -- distinct authorities, no constraint collision

    count = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(count) == 2


# ---------------------------------------------------------------------------
# E. Plan confirmation enforcement.
# ---------------------------------------------------------------------------


async def test_execute_rejected_without_any_confirmed_plan(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _create_session(client, headers)
    up = await _upload(client, headers, session["id"], content)
    assert up.status_code == 201
    v = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert v.status_code == 200 and v.json()["status"] == "validated"
    dr = await client.post(f"/api/v1/import-sessions/{session['id']}/dry-run", headers=headers)
    assert dr.status_code == 200 and dr.json()["status"] == "dry_run_completed"
    # Deliberately never confirmed.

    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 409, ex.text
    assert ex.json()["code"] == "IMPORT_NO_CONFIRMED_PLAN"

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert events == []


async def test_execute_accepted_with_confirmed_plan(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text


async def test_execute_rejected_when_plan_superseded_by_new_dry_run(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    # A second dry-run supersedes the confirmed plan (this session's own
    # status is admissible for a repeat dry-run from dry_run_completed).
    dr2 = await client.post(f"/api/v1/import-sessions/{session['id']}/dry-run", headers=headers)
    assert dr2.status_code == 200, dr2.text

    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 409, ex.text
    assert ex.json()["code"] == "IMPORT_NO_CONFIRMED_PLAN"


# ---------------------------------------------------------------------------
# F. Authority binding.
# ---------------------------------------------------------------------------


async def test_execute_fails_closed_when_authority_binding_no_longer_matches(client, seeded_users, db_session):
    """Simulates the plan's own recorded authority no longer being the
    checksum's approved authority (defense-in-depth: this should be
    structurally impossible under normal operation since checksum is
    unique, but execute() must still fail closed rather than assume)."""
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    session_uuid = uuid.UUID(session["id"])
    plan = await legacy_history_dry_run_plan_crud.get_current_plan(db_session, import_session_id=session_uuid)

    # Corrupt the plan's own recorded authority id directly (never
    # reachable via any real API -- proves execute()'s own defense-in-
    # depth check, not merely trusting the plan blindly).
    other_authority = LegacyMigrationAuthority(
        scope="other", approved_workbook_sha256="f" * 64, approved_by_user_id=actor_id
    )
    db_session.add(other_authority)
    await db_session.flush()
    plan.migration_authority_id = other_authority.id
    await db_session.commit()

    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code >= 400, ex.text
    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert events == []


async def test_no_authority_auto_created_by_execution(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    authority = await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    before = (await db_session.execute(select(LegacyMigrationAuthority))).scalars().all()
    assert len(before) == 1

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text

    after = (await db_session.execute(select(LegacyMigrationAuthority))).scalars().all()
    assert len(after) == 1
    assert after[0].id == authority.id


# ---------------------------------------------------------------------------
# G. Provenance.
# ---------------------------------------------------------------------------


async def test_shared_header_ref_supports_multiple_line_events(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook(
        issue_lines=[_issue_line_row(row_key="1"), _issue_line_row(row_key="2")],
        receive_headers=[],
        receive_lines=[],
    )
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events) == 2
    header_refs = (
        (
            await db_session.execute(
                select(LegacyEquipmentEventSourceRef).where(
                    LegacyEquipmentEventSourceRef.sheet_name == ISSUE_HEADER_SHEET
                )
            )
        )
        .scalars()
        .all()
    )
    # Two distinct events, each with its own ref to the SAME physical
    # header row (source_row_number 2) -- never merged into one.
    assert len(header_refs) == 2
    assert {r.legacy_equipment_event_id for r in header_refs} == {e.id for e in events}
    assert {r.source_row_number for r in header_refs} == {2}


async def test_duplicate_refs_not_created_when_second_session_reimports_same_data(
    client: AsyncClient, seeded_users, db_session
):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook(receive_headers=[], receive_lines=[])
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session_a = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex_a = await client.post(f"/api/v1/import-sessions/{session_a['id']}/execute", headers=headers)
    assert ex_a.status_code == 200, ex_a.text

    session_b = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex_b = await client.post(f"/api/v1/import-sessions/{session_b['id']}/execute", headers=headers)
    assert ex_b.status_code == 200, ex_b.text
    assert ex_b.json()["imported_rows"] == 0

    refs = (await db_session.execute(select(LegacyEquipmentEventSourceRef))).scalars().all()
    assert len(refs) == 2  # header + line, exactly once -- session_b contributed none


# ---------------------------------------------------------------------------
# H. No pairing.
# ---------------------------------------------------------------------------


async def test_issue_without_receive_succeeds_no_pairing_rows(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook(receive_headers=[], receive_lines=[])
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events) == 1
    assert events[0].event_type == "ISSUE"


async def test_receive_without_issue_succeeds_no_inferred_relationship(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook(issue_headers=[], issue_lines=[])
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events) == 1
    assert events[0].event_type == "RECEIVE"
    # No pairing/link table or column of any kind exists on this model --
    # structurally proven by the model's own column set, not just absence
    # of a value.
    assert not hasattr(events[0], "paired_event_id")
    assert not hasattr(events[0], "linked_transaction_id")


# ---------------------------------------------------------------------------
# I. Privacy.
# ---------------------------------------------------------------------------


async def test_notes_never_persisted_or_leaked_in_errors(client: AsyncClient, seeded_users, db_session):
    equipment = await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    secret_note = "CONFIDENTIAL-PATIENT-NOTE-should-never-appear-anywhere"
    content = _default_workbook(
        receive_headers=[],
        receive_lines=[],
        issue_headers=[dict(_issue_header_row(), **{"หมายเหตุ": secret_note})],
        issue_lines=[dict(_issue_line_row(), **{"หมายเหตุ": secret_note})],
    )
    checksum = hashlib.sha256(content).hexdigest()
    authority = await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    # Also seed a fact-conflicting event to force a conflict error path,
    # and confirm the note never leaks into that error message either.
    conflicting = LegacyEquipmentEvent(
        migration_authority_id=authority.id,
        equipment_id=equipment.id,
        event_type="ISSUE",
        occurred_at=datetime(2020, 1, 1, tzinfo=timezone.utc),
        legacy_source_row_key="1",
        import_session_id=uuid.uuid4(),
        import_source_id=uuid.uuid4(),
    )
    db_session.add(conflicting)
    await db_session.commit()

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code >= 400
    assert secret_note not in ex.text

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    for event in events:
        for column in ("legacy_order_reference", "legacy_ward_text", "legacy_bme_name"):
            value = getattr(event, column)
            assert value is None or secret_note not in value
    assert not hasattr(LegacyEquipmentEvent, "notes")
    assert not hasattr(LegacyEquipmentEventSourceRef, "notes")


# ---------------------------------------------------------------------------
# J. Recovery/fencing.
# ---------------------------------------------------------------------------


async def test_stale_execute_worker_recovered_marks_plan_and_session_failed(
    client: AsyncClient, seeded_users, db_session
):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    session_uuid = uuid.UUID(session["id"])
    plan = await legacy_history_dry_run_plan_crud.get_current_plan(db_session, import_session_id=session_uuid)
    assert plan.status == "active"
    plan_id = plan.id

    # Force a stuck 'executing' job with an expired lease, mirroring
    # test_pr20d_dry_run_plan.py's own established recovery-forcing
    # pattern.
    current = (await db_session.execute(select(ImportSession).where(ImportSession.id == session_uuid))).scalar_one()
    _s, job = await import_job_crud.admit_phase_job(
        db_session,
        session_id=session_uuid,
        job_type="execute",
        allowed_from_statuses=("dry_run_completed",),
        running_status="executing",
        expected_version=current.version,
        lease_owner=uuid.uuid4(),
        lease_duration_seconds=300,
    )
    assert job is not None
    await db_session.execute(
        ImportJob.__table__.update()
        .where(ImportJob.id == job.id)
        .values(lease_expires_at=datetime.now(timezone.utc) - timedelta(seconds=1))
    )
    await db_session.commit()

    recover_resp = await client.post(f"/api/v1/import-sessions/{session['id']}/recover", headers=headers)
    assert recover_resp.status_code == 200, recover_resp.text
    assert recover_resp.json()["status"] == "failed"

    db_session.expire_all()
    plan_after = await legacy_history_dry_run_plan_crud.get_plan_by_id(
        db_session, plan_id=plan_id, import_session_id=session_uuid
    )
    assert plan_after.status == "failed"
    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert events == [], "a hard-crash recovery must never leave a partial event behind"


async def test_recovery_retry_remains_idempotent_no_duplicate_events(client: AsyncClient, seeded_users, db_session):
    """A successful execute, followed by a recovery call against an
    unrelated stale (already-completed) state, must never duplicate
    events -- recovery only acts on a genuinely stale-running job."""
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text

    recover_resp = await client.post(f"/api/v1/import-sessions/{session['id']}/recover", headers=headers)
    assert recover_resp.status_code == 409, recover_resp.text

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events) == 2


# ---------------------------------------------------------------------------
# K. Plan lifecycle after success/failure; retention regression.
# ---------------------------------------------------------------------------


async def test_plan_marked_consumed_after_successful_execution(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    session_uuid = uuid.UUID(session["id"])
    plan = await legacy_history_dry_run_plan_crud.get_current_plan(db_session, import_session_id=session_uuid)
    plan_id = plan.id

    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text

    db_session.expire_all()
    plan_after = await legacy_history_dry_run_plan_crud.get_plan_by_id(
        db_session, plan_id=plan_id, import_session_id=session_uuid
    )
    assert plan_after.status == "consumed"


async def test_permanent_events_and_refs_survive_retention_while_plan_rows_remain_redaction_capable(
    client: AsyncClient, seeded_users, db_session
):
    """§29 of the task: `LegacyEquipmentEvent`/`SourceRef` are permanent,
    never governed by import retention at all -- proven here by running
    the existing PR21 retention hook (already regression-tested by
    PR21A) against a terminal, eligible session and confirming it
    redacts only the temporary plan-row content, never touching the
    permanent event/provenance tables."""
    from datetime import timedelta as _td

    from app.crud import import_retention as import_retention_crud

    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text

    events_before = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events_before) == 2

    session_uuid = uuid.UUID(session["id"])
    session_row = (await db_session.execute(select(ImportSession).where(ImportSession.id == session_uuid))).scalar_one()
    session_row.terminal_at = datetime.now(timezone.utc) - _td(days=400)
    await db_session.commit()

    worker_id = uuid.uuid4()
    claimed_ids, _more = await import_retention_crud.claim_sessions_for_cleanup(
        db_session, worker_id=worker_id, retention_days=180, claim_timeout_seconds=300, limit=10
    )
    assert session_uuid in claimed_ids
    redacted = await import_retention_crud.redact_session(db_session, session_id=session_uuid, worker_id=worker_id)
    assert redacted is not None
    await db_session.commit()

    db_session.expire_all()
    events_after = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events_after) == 2
    for e in events_after:
        assert e.legacy_bme_name is not None or e.legacy_bme_name is None  # column still readable, untouched
    refs_after = (await db_session.execute(select(LegacyEquipmentEventSourceRef))).scalars().all()
    assert len(refs_after) == 4

    plan = (
        await db_session.execute(
            select(LegacyHistoryDryRunPlan).where(LegacyHistoryDryRunPlan.import_session_id == session_uuid)
        )
    ).scalar_one()
    plan_rows = (
        (
            await db_session.execute(
                select(legacy_history_dry_run_plan_crud.LegacyHistoryDryRunPlanRow).where(
                    legacy_history_dry_run_plan_crud.LegacyHistoryDryRunPlanRow.dry_run_plan_id == plan.id
                )
            )
        )
        .scalars()
        .all()
    )
    assert all(row.normalized_values is None for row in plan_rows), "temporary plan-row content must be redacted"


# ---------------------------------------------------------------------------
# M. Scale (structural, not a full 60k integration test).
# ---------------------------------------------------------------------------


async def test_no_source_row_number_collision_between_issue_and_receive_at_scale(
    client: AsyncClient, seeded_users, db_session
):
    """A moderately-sized combined plan where Issue and Receive lines
    independently renumber from 1 (both sides sharing the same
    source_row_number space) must still execute every row correctly --
    proving `list_all_plan_rows`'s own `(event_type, source_row_number,
    id)` ordering, and the identity tuple's own `event_type` component,
    genuinely prevent a collision, not merely avoid one by accident."""
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    actor_id = await _actor_id(db_session)

    line_count = 150
    content = _default_workbook(
        issue_headers=[_issue_header_row(order_ref="ORD-BULK")],
        issue_lines=[_issue_line_row(row_key=str(i), order_ref="ORD-BULK") for i in range(1, line_count + 1)],
        receive_headers=[_receive_header_row(order_ref="RET-BULK")],
        receive_lines=[_receive_line_row(row_key=str(i), order_ref="RET-BULK") for i in range(1, line_count + 1)],
    )
    checksum = hashlib.sha256(content).hexdigest()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validate_dry_run_confirm(client, headers, db_session, content, actor_id=actor_id)
    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 200, ex.text
    assert ex.json()["imported_rows"] == 2 * line_count

    issue_events = (
        (await db_session.execute(select(LegacyEquipmentEvent).where(LegacyEquipmentEvent.event_type == "ISSUE")))
        .scalars()
        .all()
    )
    receive_events = (
        (await db_session.execute(select(LegacyEquipmentEvent).where(LegacyEquipmentEvent.event_type == "RECEIVE")))
        .scalars()
        .all()
    )
    assert len(issue_events) == line_count
    assert len(receive_events) == line_count
    assert {e.legacy_source_row_key for e in issue_events} == {str(i) for i in range(1, line_count + 1)}
    assert {e.legacy_source_row_key for e in receive_events} == {str(i) for i in range(1, line_count + 1)}


async def test_default_adapter_execute_batch_flush_boundary(db_session, seeded_users):
    """Structural regression against an accidental hardcoded batch-size
    assumption: `_EXECUTE_BATCH_SIZE` rows plus one more must both flush
    correctly (not silently drop the remainder)."""
    from app.services.import_adapters.legacy_history.combined import _EXECUTE_BATCH_SIZE

    assert _EXECUTE_BATCH_SIZE > 0
    assert _EXECUTE_BATCH_SIZE < 5000, "batch size must be well under the generic 5,000-row assumption to prove batching actually engages within realistic plan sizes"

"""Roadmap PR20C (docs/design/PR20_EQUIPMENT_MASTER_IMPORT_PLAN.md §7, §8,
§9 OD-1/OD-2/OD-3/OD-4). Covers the Equipment Master import adapter's own
scope only: workbook/header contract, BCM/Item No. cell extraction and
normalization, the OD-3 blank-identifier matrix and seven-case identity
matrix, OD-4's fail-closed CREATE Asset Number policy, batched (no-N+1)
Equipment lookup, no-Equipment-mutation, and PR19/PR20A framework
integration (adapter dispatch, off-thread parsing, findings/counters,
structural-failure handling). Does not re-test PR19A2's own lease/
fencing/recovery mechanics (see test_import_validation.py) or PR20A's own
source-artifact infrastructure (see
test_pr20a_source_artifact_infrastructure.py) -- both are reused unchanged
by this adapter and are exercised here only incidentally, through a small
number of true end-to-end tests proving the wiring works."""

import io
import uuid
import zipfile

import openpyxl
import pytest
import pytest_asyncio
from httpx import AsyncClient
from sqlalchemy import event, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.exceptions import InvalidInputError
from app.crud import equipment as equipment_crud
from app.models.equipment import Equipment, EquipmentStatus
from app.services import import_adapters
from app.services.import_adapters import equipment_master
from app.services.import_adapters.equipment_master import (
    CODE_ASSET_ID_CONFLICT,
    CODE_ASSET_NUMBER_REQUIRED_FOR_CREATE,
    CODE_BCM_DUPLICATE_IN_SOURCE,
    CODE_BCM_INVALID,
    CODE_BCM_MISSING,
    CODE_EQUIPMENT_NAME_MISSING,
    CODE_FIELD_TOO_LONG,
    CODE_IDENTITY_CONFLICT,
    CODE_ITEM_NO_DUPLICATE_IN_SOURCE,
    CODE_ITEM_NO_INVALID,
    CODE_ITEM_NO_MISSING,
    CODE_SERIAL_NUMBER_CONFLICT,
    CODE_STATUS_MISMATCH,
    CODE_STATUS_MISSING,
    CODE_STATUS_UNMAPPABLE,
    DATASET_TYPE,
    STATUS_MAPPING,
    WORKSHEET_NAME,
    EquipmentMasterAdapter,
    EquipmentMasterContext,
    _extract_bcm_cell,
    _extract_item_no_cell,
    _GOVERNED_HEADERS,
    _load_sheet1,
    _parse_workbook_sync,
    _reject_macro_ooxml_structure,
    _reject_macro_parts,
    _validate_headers,
)
from app.services.import_source_reader import SourceDescriptor, VerifiedSourceContent
from app.services import import_lease
from tests.conftest import auth_headers

# pytest.ini sets asyncio_mode = auto -- async test functions below need no
# explicit marker. This module deliberately mixes sync (workbook-contract,
# cell-extraction) and async (DB-backed) tests, so no blanket
# `pytestmark = pytest.mark.asyncio` is applied here (it would spuriously
# warn on every sync test).
_HEADERS = list(_GOVERNED_HEADERS)


# ---------------------------------------------------------------------------
# Workbook builder -- a genuine OOXML workbook via openpyxl (not a hand-rolled
# zip stub), so parser-contract tests exercise the real `load_workbook` path.
# ---------------------------------------------------------------------------


def _build_workbook_bytes(
    rows: list[dict[str, object]] | None = None,
    *,
    headers: list[str] | None = None,
    sheet_name: str = WORKSHEET_NAME,
    extra_sheets: dict[str, list[list[object]]] | None = None,
    no_header_row: bool = False,
) -> bytes:
    headers = _HEADERS if headers is None else headers
    rows = rows or []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if not no_header_row:
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    for extra_name, extra_rows in (extra_sheets or {}).items():
        extra_ws = wb.create_sheet(extra_name)
        for row in extra_rows:
            extra_ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _valid_row(**overrides: object) -> dict[str, object]:
    row = {
        "ID CODE": overrides.pop("bcm", "BCM100001"),
        "Item No.": overrides.pop("item_no", "100001"),
        "ชื่อไทย": overrides.pop("equipment_name", "เครื่องมือทดสอบ"),
        "สถานะเครื่องมือ": overrides.pop("status", "active"),
    }
    row.update(overrides)
    return row


# ---------------------------------------------------------------------------
# §7/§9 OD-1 Workbook/header contract (direct unit tests -- no DB/HTTP)
# ---------------------------------------------------------------------------


def test_valid_sheet1_parses():
    content = _build_workbook_bytes([_valid_row()])
    records = _parse_workbook_sync(content)
    assert len(records) == 1
    assert records[0].row_number == 2


def test_missing_sheet1_raises():
    content = _build_workbook_bytes([_valid_row()], sheet_name="Sheet2")
    with pytest.raises(InvalidInputError, match="Sheet1"):
        _load_sheet1(content)


def test_additional_worksheet_alongside_sheet1_is_ignored():
    content = _build_workbook_bytes([_valid_row()], extra_sheets={"Notes": [["ignore me"]]})
    ws = _load_sheet1(content)
    assert ws.title == WORKSHEET_NAME
    records = _parse_workbook_sync(content)
    assert len(records) == 1


def test_header_must_be_row_1_not_a_later_row():
    # Header shifted to row 2 by inserting a blank row above it -- row 1 is
    # then something else entirely, so every governed header is "missing".
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = WORKSHEET_NAME
    ws.append(["not a header row"])
    ws.append(_HEADERS)
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(InvalidInputError, match="missing required column"):
        _validate_headers(_load_sheet1(buf.getvalue()))


def test_reordered_columns_still_parse_by_exact_name():
    shuffled = list(reversed(_HEADERS))
    content = _build_workbook_bytes([_valid_row()], headers=shuffled)
    records = _parse_workbook_sync(content)
    assert len(records) == 1
    assert records[0].fields["equipment_name"] == "เครื่องมือทดสอบ"


def test_missing_header_column_raises():
    headers = [h for h in _HEADERS if h != "ID CODE"]
    content = _build_workbook_bytes([], headers=headers)
    with pytest.raises(InvalidInputError, match="missing required column"):
        _load_and_validate(content)


def test_duplicate_header_column_raises():
    headers = _HEADERS + ["ID CODE"]
    content = _build_workbook_bytes([], headers=headers)
    with pytest.raises(InvalidInputError, match="duplicate column"):
        _load_and_validate(content)


def test_extra_unrecognized_header_column_raises():
    headers = _HEADERS + ["Some Extra Column"]
    content = _build_workbook_bytes([], headers=headers)
    with pytest.raises(InvalidInputError, match="unrecognized column"):
        _load_and_validate(content)


def test_empty_workbook_no_header_row_raises():
    content = _build_workbook_bytes(no_header_row=True)
    with pytest.raises(InvalidInputError, match="empty"):
        _load_and_validate(content)


def test_blank_rows_are_skipped_not_counted():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = WORKSHEET_NAME
    ws.append(_HEADERS)
    ws.append([None for _ in _HEADERS])  # fully blank row
    ws.append([_valid_row().get(h) for h in _HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    records = _parse_workbook_sync(buf.getvalue())
    assert len(records) == 1
    assert records[0].row_number == 3


def test_excessive_rows_raises(monkeypatch):
    monkeypatch.setattr(equipment_master, "MAX_IMPORT_ROWS", 3)
    content = _build_workbook_bytes([_valid_row(bcm=f"BCM{i}", item_no=str(i)) for i in range(5)])
    with pytest.raises(InvalidInputError, match="more than 3 data rows"):
        _parse_workbook_sync(content)


def test_corrupt_xlsx_raises():
    with pytest.raises(InvalidInputError, match="valid Excel"):
        _load_sheet1(b"this is not a zip archive at all")


def test_formula_cell_is_never_executed_treated_as_blank():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = WORKSHEET_NAME
    ws.append(_HEADERS)
    row = _valid_row()
    row["ID CODE"] = "=1+1"  # a formula string, no cached value when read back
    ws.append([row.get(h) for h in _HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    records = _parse_workbook_sync(buf.getvalue())
    assert records[0].fields["bcm"].outcome == "blank", (
        "a formula cell with no cached value must never be evaluated -- it is treated as absent, "
        "never interpreted or executed"
    )


def test_excessive_header_columns_raises(monkeypatch):
    monkeypatch.setattr(equipment_master, "MAX_HEADER_COLUMNS", 5)
    content = _build_workbook_bytes([])
    with pytest.raises(InvalidInputError, match="more than 5 columns"):
        _load_and_validate(content)


def test_excessive_worksheet_count_raises(monkeypatch):
    monkeypatch.setattr(equipment_master, "MAX_WORKSHEET_COUNT", 1)
    content = _build_workbook_bytes([], extra_sheets={"Extra": [["x"]]})
    with pytest.raises(InvalidInputError, match="more than 1 worksheet"):
        _load_sheet1(content)


def _load_and_validate(content: bytes) -> dict[str, int]:
    return _validate_headers(_load_sheet1(content))


def test_blank_header_column_with_data_underneath_is_rejected():
    """PR93-H2: `_validate_headers` only inspects non-blank header cells,
    so a 33rd column whose row-1 cell is blank passes structural header
    validation even when the same 32 governed headers are otherwise all
    present exactly once. §7's closed-world contract must still reject
    that column once it carries data -- an unnamed/blank-header data
    column must never silently bypass the approved 32-column schema."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = WORKSHEET_NAME
    ws.append(_HEADERS + [None])  # 33rd header cell deliberately blank
    row = _valid_row()
    ws.append([row.get(h) for h in _HEADERS] + ["smuggled extra value"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(InvalidInputError, match="outside the approved Equipment Master schema"):
        _parse_workbook_sync(buf.getvalue())


def test_blank_header_column_with_no_data_is_still_accepted():
    """The fix for PR93-H2 must not become over-eager: a genuinely empty
    trailing column (blank header, no data in any row) is not itself a
    schema violation -- only a blank-header column that actually carries
    data is rejected."""
    content = _build_workbook_bytes([_valid_row()])
    records = _parse_workbook_sync(content)
    assert len(records) == 1


def test_macro_enabled_vba_project_part_is_rejected():
    """PR93-H3: §21 ("never execute workbook formulas or macros") requires
    more than merely not executing a macro that's present -- a macro-
    bearing OOXML part must be rejected outright, even though it's a
    structurally legitimate path under the shared `xl/` allowlist (so the
    reused PR12 zip-bounds validator alone does not reject it). Simulates
    a macro-enabled workbook masquerading as a plain `.xlsx` by injecting
    `xl/vbaProject.bin` into an otherwise valid, governed-schema workbook."""
    content = _build_workbook_bytes([_valid_row()])
    buf = io.BytesIO(content)
    with zipfile.ZipFile(buf, "a") as archive:
        archive.writestr("xl/vbaProject.bin", b"\x00\x01fake-vba-binary-payload")
    macro_content = buf.getvalue()

    with pytest.raises(InvalidInputError, match="macro-enabled component"):
        _reject_macro_parts(macro_content)
    with pytest.raises(InvalidInputError, match="macro-enabled component"):
        _load_sheet1(macro_content)
    with pytest.raises(InvalidInputError, match="macro-enabled component"):
        _parse_workbook_sync(macro_content)


def test_ordinary_workbook_without_macro_parts_is_unaffected():
    content = _build_workbook_bytes([_valid_row()])
    _reject_macro_parts(content)  # must not raise
    records = _parse_workbook_sync(content)
    assert len(records) == 1


# ---------------------------------------------------------------------------
# PR93-H3R: OOXML macro/VBA STRUCTURE detection ([Content_Types].xml and
# relationship-file inspection), not merely a suspicious part filename.
# ---------------------------------------------------------------------------


def _replace_zip_entry(content: bytes, entry_name: str, new_data: bytes) -> bytes:
    """Rebuilds `content`'s zip archive with `entry_name`'s bytes replaced
    -- `zipfile` has no in-place update, so every other entry is copied
    through unchanged and only the target entry's content differs."""
    src = zipfile.ZipFile(io.BytesIO(content))
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w") as dst:
        for info in src.infolist():
            data = new_data if info.filename == entry_name else src.read(info.filename)
            dst.writestr(info.filename, data)
    return out_buf.getvalue()


def test_normal_plain_xlsx_passes_macro_structure_check():
    content = _build_workbook_bytes([_valid_row()])
    _reject_macro_ooxml_structure(content)  # must not raise
    records = _parse_workbook_sync(content)
    assert len(records) == 1


def test_macro_enabled_content_type_without_vbaproject_named_part_is_rejected():
    """The exact reviewer reproduction: `[Content_Types].xml` declares the
    main workbook part's content type as macro-enabled
    (`application/vnd.ms-excel.sheet.macroEnabled.main+xml`) while no
    part in the archive is named `vbaProject...` anywhere -- the
    filename-substring check (`_reject_macro_parts`) alone would miss
    this; the content-type-structure check must catch it."""
    content = _build_workbook_bytes([_valid_row()])
    src = zipfile.ZipFile(io.BytesIO(content))
    content_types_xml = src.read("[Content_Types].xml").decode()
    assert "vbaproject" not in content_types_xml.lower()
    macro_content_types_xml = content_types_xml.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    assert macro_content_types_xml != content_types_xml, "the fixture must actually change the main part's content type"
    macro_content = _replace_zip_entry(content, "[Content_Types].xml", macro_content_types_xml.encode())

    with pytest.raises(InvalidInputError, match="macro-enabled or VBA-related"):
        _reject_macro_ooxml_structure(macro_content)
    with pytest.raises(InvalidInputError, match="macro-enabled or VBA-related"):
        _load_sheet1(macro_content)
    with pytest.raises(InvalidInputError, match="macro-enabled or VBA-related"):
        _parse_workbook_sync(macro_content)


def test_vba_relationship_type_with_innocuous_target_filename_is_rejected():
    """A `Relationship` declaring the VBA-project relationship `Type`
    (the actual OOXML/Microsoft schema URI) pointing at a target filename
    that itself gives no hint of being a VBA part -- proves detection
    reads the relationship's declared `Type`, not the target filename."""
    content = _build_workbook_bytes([_valid_row()])
    src = zipfile.ZipFile(io.BytesIO(content))
    rels_xml = src.read("xl/_rels/workbook.xml.rels").decode()
    injected = rels_xml.replace(
        "</Relationships>",
        '<Relationship Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
        'Target="innocuous_name.bin" Id="rIdInjected" /></Relationships>',
    )
    assert "vbaproject" not in "innocuous_name.bin", "the target filename itself must carry no 'vbaproject' substring"
    macro_content = _replace_zip_entry(content, "xl/_rels/workbook.xml.rels", injected.encode())

    with pytest.raises(InvalidInputError, match="macro-enabled or VBA-related"):
        _reject_macro_ooxml_structure(macro_content)


def test_macrosheet_content_type_is_rejected():
    content = _build_workbook_bytes([_valid_row()])
    src = zipfile.ZipFile(io.BytesIO(content))
    content_types_xml = src.read("[Content_Types].xml").decode()
    injected = content_types_xml.replace(
        "</Types>",
        '<Override PartName="/xl/macrosheets/macrosheet1.xml" '
        'ContentType="application/vnd.ms-excel.macrosheet+xml" /></Types>',
    )
    macro_content = _replace_zip_entry(content, "[Content_Types].xml", injected.encode())

    with pytest.raises(InvalidInputError, match="macro-enabled or VBA-related"):
        _reject_macro_ooxml_structure(macro_content)


def test_malformed_content_types_xml_fails_closed():
    content = _build_workbook_bytes([_valid_row()])
    malformed_content = _replace_zip_entry(content, "[Content_Types].xml", b"<Types><Override not-well-formed")

    with pytest.raises(InvalidInputError, match="malformed"):
        _reject_macro_ooxml_structure(malformed_content)
    with pytest.raises(InvalidInputError, match="malformed"):
        _load_sheet1(malformed_content)


def test_malformed_workbook_rels_xml_fails_closed():
    content = _build_workbook_bytes([_valid_row()])
    malformed_content = _replace_zip_entry(
        content, "xl/_rels/workbook.xml.rels", b"<Relationships><Relationship not-well-formed"
    )

    with pytest.raises(InvalidInputError, match="malformed"):
        _reject_macro_ooxml_structure(malformed_content)


def test_content_types_xml_with_doctype_is_rejected_fail_closed():
    """Defends against entity-expansion ("billion laughs") and external-
    entity attacks: any DOCTYPE at all in a package-metadata part this
    check parses is rejected outright, since a conformant OOXML
    `[Content_Types].xml` never legitimately carries one."""
    content = _build_workbook_bytes([_valid_row()])
    doctype_content = (
        b'<?xml version="1.0"?><!DOCTYPE Types [<!ENTITY x "y">]>'
        + b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"></Types>'
    )
    malformed_content = _replace_zip_entry(content, "[Content_Types].xml", doctype_content)

    with pytest.raises(InvalidInputError, match="malformed"):
        _reject_macro_ooxml_structure(malformed_content)


def test_macro_structure_decision_does_not_depend_on_filename_or_extension():
    """A macro-enabled OOXML package renamed/declared as `.xlsx` (the
    upload path's own filename/content-type is never consulted by this
    check -- only the archive's internal package metadata) must still be
    rejected, proving the decision is structural, not extension-based."""
    content = _build_workbook_bytes([_valid_row()])
    src = zipfile.ZipFile(io.BytesIO(content))
    content_types_xml = src.read("[Content_Types].xml").decode()
    macro_content_types_xml = content_types_xml.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    macro_content = _replace_zip_entry(content, "[Content_Types].xml", macro_content_types_xml.encode())
    # `_reject_macro_ooxml_structure` takes only raw bytes -- there is no
    # filename/content-type parameter to spoof in the first place.
    with pytest.raises(InvalidInputError, match="macro-enabled or VBA-related"):
        _reject_macro_ooxml_structure(macro_content)


def test_formula_cells_are_not_confused_with_macro_structure():
    """§21's formula non-execution and PR93-H3R's macro-structure
    rejection are two distinct checks -- an ordinary workbook containing
    formula cells (no macro-enabled content type, no VBA part/
    relationship) must pass the macro-structure check cleanly, exactly
    like `test_formula_cell_is_never_executed_treated_as_blank` already
    proves it's never evaluated."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = WORKSHEET_NAME
    ws.append(_HEADERS)
    row = _valid_row()
    row["ID CODE"] = "=1+1"
    ws.append([row.get(h) for h in _HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    _reject_macro_ooxml_structure(content)  # must not raise
    records = _parse_workbook_sync(content)
    assert records[0].fields["bcm"].outcome == "blank"


# ---------------------------------------------------------------------------
# §7 BCM / Item No. cell extraction
# ---------------------------------------------------------------------------


def test_bcm_cell_text_valid():
    assert _extract_bcm_cell(" BCM100001 ") == equipment_master._CellResult("valid", value="BCM100001")


def test_bcm_cell_blank():
    assert _extract_bcm_cell(None).outcome == "blank"
    assert _extract_bcm_cell("   ").outcome == "blank"


def test_bcm_cell_numeric_is_invalid_never_coerced():
    result = _extract_bcm_cell(100001)
    assert result.outcome == "invalid"


def test_item_no_cell_text_valid():
    result = _extract_item_no_cell(" 0134621 ")
    assert result.outcome == "valid"
    assert result.value == "0134621", "a text cell's leading zero must be preserved exactly"


def test_item_no_cell_integer_converted_losslessly_no_invented_leading_zero():
    result = _extract_item_no_cell(134621)
    assert result.outcome == "valid"
    assert result.value == "134621", "must never fabricate a leading zero that was never in the cell"


def test_item_no_cell_blank():
    assert _extract_item_no_cell(None).outcome == "blank"
    assert _extract_item_no_cell("").outcome == "blank"


def test_item_no_cell_fractional_rejected():
    assert _extract_item_no_cell(134621.5).outcome == "invalid"


def test_item_no_cell_bool_rejected():
    assert _extract_item_no_cell(True).outcome == "invalid"


def test_item_no_cell_nan_inf_rejected():
    assert _extract_item_no_cell(float("nan")).outcome == "invalid"
    assert _extract_item_no_cell(float("inf")).outcome == "invalid"


def test_item_no_cell_whole_valued_float_converted_losslessly():
    result = _extract_item_no_cell(134621.0)
    assert result.outcome == "valid"
    assert result.value == "134621"


# ---------------------------------------------------------------------------
# §9 OD-3 blank/null identifier requiredness matrix + duplicate detection
# (direct adapter-level tests: parse -> preload -> validate, no DB rows
# pre-seeded, no HTTP layer -- isolates the adapter's own logic).
# ---------------------------------------------------------------------------


async def _run_adapter(db_session, rows: list[dict[str, object]]) -> list[list]:
    """Parses `rows` through the real adapter pipeline and returns each
    row's findings in row order, exactly as `import_validation_service.
    run_validation` would drive it (parse -> preload -> per-record
    validate)."""
    content = _build_workbook_bytes(rows)
    adapter = EquipmentMasterAdapter()
    records = adapter.parse(VerifiedSourceContent(content=content, source_descriptor=_fake_descriptor()))
    context = await adapter.preload_business_context(db_session, records)
    return [adapter.validate_business_rules(r, context) for r in records]


def _fake_descriptor() -> SourceDescriptor:
    return SourceDescriptor(
        import_source_id=uuid.uuid4(),
        import_session_id=uuid.uuid4(),
        dataset_type=DATASET_TYPE,
        expected_checksum="x",
        expected_byte_size=0,
        content_type=None,
        original_filename=None,
        registration_status="frozen",
    )


def _codes(findings: list) -> set[str]:
    return {f.error_code for f in findings}


async def test_bcm_blank_item_present_is_blocking_error(db_session, seeded_users):
    [findings] = await _run_adapter(db_session, [_valid_row(bcm=None)])
    assert CODE_BCM_MISSING in _codes(findings)
    assert all(f.severity == "error" for f in findings if f.error_code == CODE_BCM_MISSING)


async def test_bcm_present_item_blank_is_blocking_error(db_session, seeded_users):
    [findings] = await _run_adapter(db_session, [_valid_row(item_no=None)])
    assert CODE_ITEM_NO_MISSING in _codes(findings)


async def test_both_identifiers_blank_is_blocking_error(db_session, seeded_users):
    [findings] = await _run_adapter(db_session, [_valid_row(bcm=None, item_no=None)])
    codes = _codes(findings)
    assert CODE_BCM_MISSING in codes
    assert CODE_ITEM_NO_MISSING in codes


async def test_malformed_bcm_is_blocking_error(db_session, seeded_users):
    # "bcm 001" -- whitespace embedded after the stripped prefix is rejected
    # by normalize_bcm_code, not silently squeezed together.
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="bcm 001")])
    assert CODE_BCM_INVALID in _codes(findings)


async def test_bcm_duplicate_within_source_flags_both_rows(db_session, seeded_users):
    rows = [_valid_row(bcm="BCM777", item_no="1"), _valid_row(bcm="BCM777", item_no="2")]
    findings_a, findings_b = await _run_adapter(db_session, rows)
    assert CODE_BCM_DUPLICATE_IN_SOURCE in _codes(findings_a)
    assert CODE_BCM_DUPLICATE_IN_SOURCE in _codes(findings_b)


async def test_item_no_duplicate_within_source_flags_both_rows(db_session, seeded_users):
    rows = [_valid_row(bcm="BCM1", item_no="999"), _valid_row(bcm="BCM2", item_no="999")]
    findings_a, findings_b = await _run_adapter(db_session, rows)
    assert CODE_ITEM_NO_DUPLICATE_IN_SOURCE in _codes(findings_a)
    assert CODE_ITEM_NO_DUPLICATE_IN_SOURCE in _codes(findings_b)


async def test_item_no_fractional_cell_is_blocking_error(db_session, seeded_users):
    content = _build_workbook_bytes([_valid_row(item_no=1.5)])
    adapter = EquipmentMasterAdapter()
    records = adapter.parse(VerifiedSourceContent(content=content, source_descriptor=_fake_descriptor()))
    context = await adapter.preload_business_context(db_session, records)
    findings = adapter.validate_business_rules(records[0], context)
    assert CODE_ITEM_NO_INVALID in _codes(findings)


# ---------------------------------------------------------------------------
# §9 OD-3 seven-case identity matrix
# ---------------------------------------------------------------------------


async def _seed_equipment(db_session, **kwargs) -> Equipment:
    defaults = dict(
        asset_number=f"AN-{uuid.uuid4().hex[:10]}",
        equipment_name="Existing Equipment",
        status=EquipmentStatus.AVAILABLE_AT_POOL,
    )
    defaults.update(kwargs)
    eq = Equipment(**defaults)
    db_session.add(eq)
    await db_session.commit()
    await db_session.refresh(eq)
    return eq


async def test_case1_new_bcm_new_item_is_create_candidate_no_conflict(db_session, seeded_users):
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_NEW1", item_no="NEW-ITEM-1")])
    assert CODE_IDENTITY_CONFLICT not in _codes(findings)
    # §9 OD-4: a CREATE candidate is always blocked pending an authoritative
    # Asset Number, but that is a distinct finding from an identity conflict.
    assert CODE_ASSET_NUMBER_REQUIRED_FOR_CREATE in _codes(findings)


async def test_case2_bcm_and_item_match_same_equipment_is_update_candidate(db_session, seeded_users):
    existing = await _seed_equipment(db_session, bcm_code="BCM_A", item_no="ITEM_A")
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_A", item_no="ITEM_A")])
    assert CODE_IDENTITY_CONFLICT not in _codes(findings)
    assert CODE_ASSET_NUMBER_REQUIRED_FOR_CREATE not in _codes(findings), (
        "an UPDATE candidate must never receive the CREATE-only Asset Number finding"
    )


async def test_case3_bcm_matches_a_item_unexpected_is_conflict(db_session, seeded_users):
    await _seed_equipment(db_session, bcm_code="BCM_B", item_no="ITEM_B")
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_B", item_no="SOME-OTHER-ITEM")])
    assert CODE_IDENTITY_CONFLICT in _codes(findings)


async def test_case4_new_bcm_item_matches_existing_b_is_conflict(db_session, seeded_users):
    await _seed_equipment(db_session, bcm_code="BCM_C", item_no="ITEM_C")
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_NEW_XYZ", item_no="ITEM_C")])
    assert CODE_IDENTITY_CONFLICT in _codes(findings)


async def test_case5_bcm_matches_a_item_matches_different_b_is_conflict(db_session, seeded_users):
    await _seed_equipment(db_session, bcm_code="BCM_D1", item_no="ITEM_D1")
    await _seed_equipment(db_session, bcm_code="BCM_D2", item_no="ITEM_D2")
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_D1", item_no="ITEM_D2")])
    assert CODE_IDENTITY_CONFLICT in _codes(findings)


async def test_case6_duplicate_bcm_never_reaches_identity_resolution(db_session, seeded_users):
    rows = [_valid_row(bcm="BCM_DUP", item_no="1"), _valid_row(bcm="BCM_DUP", item_no="2")]
    findings_a, findings_b = await _run_adapter(db_session, rows)
    for findings in (findings_a, findings_b):
        assert CODE_IDENTITY_CONFLICT not in _codes(findings)
        assert CODE_ASSET_NUMBER_REQUIRED_FOR_CREATE not in _codes(findings), (
            "a row blocked at the blank/duplicate precondition must never also reach identity "
            "resolution/classification"
        )


async def test_case7_duplicate_item_no_never_reaches_identity_resolution(db_session, seeded_users):
    rows = [_valid_row(bcm="BCM_X1", item_no="DUP"), _valid_row(bcm="BCM_X2", item_no="DUP")]
    findings_a, findings_b = await _run_adapter(db_session, rows)
    for findings in (findings_a, findings_b):
        assert CODE_IDENTITY_CONFLICT not in _codes(findings)
        assert CODE_ASSET_NUMBER_REQUIRED_FOR_CREATE not in _codes(findings)


# ---------------------------------------------------------------------------
# §9 OD-4 -- CREATE Asset Number policy (fail-closed, never fabricated)
# ---------------------------------------------------------------------------


async def test_od4_create_candidate_always_blocked_never_fabricated(db_session, seeded_users):
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm="BCM_CREATE_1", item_no="ITEM_CREATE_1", **{"Asset ID": "AST-LEGACY-1"})]
    )
    blocking = [f for f in findings if f.error_code == CODE_ASSET_NUMBER_REQUIRED_FOR_CREATE]
    assert len(blocking) == 1
    assert blocking[0].severity == "error"
    assert blocking[0].field == "asset_number"
    # No substitute identifier is ever produced as a stand-in asset_number --
    # this adapter never writes an asset_number value anywhere; the proof is
    # structural: FieldError carries no value field to smuggle one through,
    # and no Equipment write occurs at all (see no-mutation tests below).


async def test_od4_update_candidate_retains_existing_asset_number_untouched(db_session, seeded_users):
    existing = await _seed_equipment(
        db_session, bcm_code="BCM_UPD_1", item_no="ITEM_UPD_1", asset_number="ORIGINAL-ASSET-NUMBER"
    )
    await _run_adapter(db_session, [_valid_row(bcm="BCM_UPD_1", item_no="ITEM_UPD_1")])
    await db_session.refresh(existing)
    assert existing.asset_number == "ORIGINAL-ASSET-NUMBER"


# ---------------------------------------------------------------------------
# §8/OD-2 field-level validation (length bounds, status, serial/asset_id
# conflicts) for both CREATE and UPDATE candidates.
# ---------------------------------------------------------------------------


async def test_create_candidate_missing_equipment_name_is_blocking(db_session, seeded_users):
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_NM", item_no="1", equipment_name=None)])
    assert CODE_EQUIPMENT_NAME_MISSING in _codes(findings)


async def test_create_candidate_missing_status_is_blocking(db_session, seeded_users):
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_ST1", item_no="1", status=None)])
    assert CODE_STATUS_MISSING in _codes(findings)


async def test_create_candidate_unmappable_status_is_blocking_never_defaulted(db_session, seeded_users):
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_ST2", item_no="1", status="some unknown value")])
    assert CODE_STATUS_UNMAPPABLE in _codes(findings)


async def test_create_candidate_status_never_defaults_to_issued_to_ward():
    assert EquipmentStatus.ISSUED_TO_WARD not in STATUS_MAPPING.values(), (
        "§10/§11: a CREATE candidate must never be initialized directly into ISSUED_TO_WARD"
    )


# ---------------------------------------------------------------------------
# PR93-H1R: authoritative status mapping, verified against the real
# Repository-Owner-supplied `export_template.xlsx` (4,729 source rows).
# Observed vocabulary: Active (3,873), Decommission (734), Defective (84),
# Missing (38) -- no other value was observed. STATUS_MAPPING must contain
# only these, and "Missing" must remain unmappable (blocking ERROR), never
# a new/fifth lifecycle state.
# ---------------------------------------------------------------------------

_AUTHORITATIVE_MAPPED_STATUSES = (
    ("Active", EquipmentStatus.AVAILABLE_AT_POOL),
    ("Defective", EquipmentStatus.UNAVAILABLE_DEFECTIVE),
    ("Decommission", EquipmentStatus.DECOMMISSIONED),
)


@pytest.mark.parametrize("source_value,expected_status", _AUTHORITATIVE_MAPPED_STATUSES)
async def test_authoritative_status_mapping_create_candidate(db_session, seeded_users, source_value, expected_status):
    """Table-driven proof of the exact Repository-Owner-confirmed mapping
    for a CREATE candidate: no CODE_STATUS_* blocking/unmappable finding
    is produced for any of the three mappable authoritative values (the
    row is still blocked overall by OD-4's unrelated
    ASSET_NUMBER_REQUIRED_FOR_CREATE, which is not a status finding)."""
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm=f"BCM_AUTH_{source_value}", item_no=f"ITEM_AUTH_{source_value}", status=source_value)]
    )
    codes = _codes(findings)
    assert CODE_STATUS_MISSING not in codes
    assert CODE_STATUS_UNMAPPABLE not in codes
    assert STATUS_MAPPING[source_value.strip().lower()] == expected_status


async def test_authoritative_status_mapping_missing_is_unmappable_not_a_new_state(db_session, seeded_users):
    """"Missing" is confirmed present in the real source (38 rows) but the
    Repository Owner has resolved it as unmappable -- never a fifth
    lifecycle state, never silently equated with any of the four."""
    assert "missing" not in STATUS_MAPPING
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm="BCM_MISSING_STATUS", item_no="ITEM_MISSING_STATUS", status="Missing")]
    )
    matches = [f for f in findings if f.error_code == CODE_STATUS_UNMAPPABLE]
    assert len(matches) == 1
    assert matches[0].severity == "error"


async def test_authoritative_status_mapping_missing_is_unmappable_for_update_too(db_session, seeded_users):
    existing = await _seed_equipment(
        db_session, bcm_code="BCM_MISSING_UPD", item_no="ITEM_MISSING_UPD", status=EquipmentStatus.AVAILABLE_AT_POOL
    )
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm="BCM_MISSING_UPD", item_no="ITEM_MISSING_UPD", status="Missing")]
    )
    matches = [f for f in findings if f.error_code == CODE_STATUS_UNMAPPABLE]
    assert len(matches) == 1
    assert matches[0].severity == "error"
    await db_session.refresh(existing)
    assert existing.status == EquipmentStatus.AVAILABLE_AT_POOL


@pytest.mark.parametrize("unobserved_value", ["available", "faulty", "broken", "disposed", "written off", "decommissioned"])
async def test_unobserved_illustrative_aliases_are_no_longer_accepted(db_session, seeded_users, unobserved_value):
    """PR93-H1R: the prior "illustrative" mapping accepted these strings
    as aliases (e.g. "decommissioned" -> DECOMMISSIONED, "available" ->
    AVAILABLE_AT_POOL). None of them was ever observed in the real
    4,729-row `export_template.xlsx` -- only the authoritative
    Active/Decommission/Defective/Missing vocabulary is accepted now, so
    each of these must be unmappable, not silently mapped."""
    assert unobserved_value not in STATUS_MAPPING
    identifier_safe = unobserved_value.replace(" ", "_")
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm=f"BCM_UNOBS_{identifier_safe}", item_no=f"ITEM_UNOBS_{identifier_safe}", status=unobserved_value)]
    )
    assert CODE_STATUS_UNMAPPABLE in _codes(findings)


async def test_status_mapping_is_case_and_whitespace_normalized():
    """Approved normalization rule (`.strip().lower()`, unchanged by
    PR93-H1R) still applies to the authoritative vocabulary."""
    assert STATUS_MAPPING["active"] == EquipmentStatus.AVAILABLE_AT_POOL
    for variant in (" Active ", "ACTIVE", "AcTiVe", "active"):
        assert STATUS_MAPPING[variant.strip().lower()] == EquipmentStatus.AVAILABLE_AT_POOL


async def test_status_mapping_contains_only_authoritative_vocabulary():
    """The full observed source vocabulary is exactly four strings;
    STATUS_MAPPING (which only ever holds *mappable* values -- "Missing"
    is deliberately absent, see above) must therefore contain exactly the
    three mappable ones and nothing else."""
    assert set(STATUS_MAPPING.keys()) == {"active", "defective", "decommission"}


async def test_create_candidate_overlength_field_is_blocking(db_session, seeded_users):
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_LEN", item_no="1", **{"ยี่ห้อ": "x" * 101})])
    assert CODE_FIELD_TOO_LONG in _codes(findings)


async def test_create_candidate_serial_number_conflict_with_existing_is_blocking(db_session, seeded_users):
    await _seed_equipment(db_session, serial_number="SN-EXISTING")
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm="BCM_SN1", item_no="ITEM_SN1", **{"S/N": "SN-EXISTING"})]
    )
    assert CODE_SERIAL_NUMBER_CONFLICT in _codes(findings)
    assert all(f.severity == "error" for f in findings if f.error_code == CODE_SERIAL_NUMBER_CONFLICT), (
        "a CREATE candidate has no 'self' record to exempt -- any existing match is blocking"
    )


async def test_update_candidate_serial_number_conflict_with_different_record_is_blocking(db_session, seeded_users):
    await _seed_equipment(db_session, serial_number="SN-OTHER", bcm_code="BCM_OTHER", item_no="ITEM_OTHER")
    await _seed_equipment(db_session, bcm_code="BCM_SELF", item_no="ITEM_SELF")
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm="BCM_SELF", item_no="ITEM_SELF", **{"S/N": "SN-OTHER"})]
    )
    assert CODE_SERIAL_NUMBER_CONFLICT in _codes(findings)


async def test_update_candidate_serial_number_matching_self_is_not_a_conflict(db_session, seeded_users):
    await _seed_equipment(db_session, serial_number="SN-SELF", bcm_code="BCM_SELF2", item_no="ITEM_SELF2")
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm="BCM_SELF2", item_no="ITEM_SELF2", **{"S/N": "SN-SELF"})]
    )
    assert CODE_SERIAL_NUMBER_CONFLICT not in _codes(findings)


async def test_update_candidate_asset_id_conflict_is_a_warning_not_blocking(db_session, seeded_users):
    await _seed_equipment(db_session, asset_id="AID-OTHER", bcm_code="BCM_AID_OTHER", item_no="ITEM_AID_OTHER")
    await _seed_equipment(db_session, bcm_code="BCM_AID_SELF", item_no="ITEM_AID_SELF")
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm="BCM_AID_SELF", item_no="ITEM_AID_SELF", **{"Asset ID": "AID-OTHER"})]
    )
    matches = [f for f in findings if f.error_code == CODE_ASSET_ID_CONFLICT]
    assert len(matches) == 1
    assert matches[0].severity == "warning"


async def test_update_candidate_status_mismatch_is_warning_and_never_applied(db_session, seeded_users):
    existing = await _seed_equipment(
        db_session, bcm_code="BCM_SM", item_no="ITEM_SM", status=EquipmentStatus.UNAVAILABLE_DEFECTIVE
    )
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_SM", item_no="ITEM_SM", status="active")])
    matches = [f for f in findings if f.error_code == CODE_STATUS_MISMATCH]
    assert len(matches) == 1
    assert matches[0].severity == "warning"
    await db_session.refresh(existing)
    assert existing.status == EquipmentStatus.UNAVAILABLE_DEFECTIVE, "legacy status must never overwrite live status"


async def test_update_candidate_unmappable_status_is_blocking_not_silently_ignored(db_session, seeded_users):
    """PR93-H1: §10 states the unmappable-status fallback generally ("any
    legacy status this design cannot safely map produces a blocking
    ERROR"), not only for CREATE candidates. A non-blank UPDATE-row status
    value with no entry in STATUS_MAPPING must receive the same blocking
    CODE_STATUS_UNMAPPABLE a CREATE candidate would -- never silently pass
    through with zero findings."""
    existing = await _seed_equipment(
        db_session, bcm_code="BCM_UNMAP", item_no="ITEM_UNMAP", status=EquipmentStatus.AVAILABLE_AT_POOL
    )
    [findings] = await _run_adapter(
        db_session, [_valid_row(bcm="BCM_UNMAP", item_no="ITEM_UNMAP", status="some totally unrecognized value")]
    )
    matches = [f for f in findings if f.error_code == CODE_STATUS_UNMAPPABLE]
    assert len(matches) == 1
    assert matches[0].severity == "error"
    await db_session.refresh(existing)
    assert existing.status == EquipmentStatus.AVAILABLE_AT_POOL, "an unmappable legacy status is never applied either"


async def test_update_candidate_blank_status_is_not_required_and_produces_no_finding(db_session, seeded_users):
    """§8 row 27: status is "not required for UPDATE" -- a blank cell on
    an UPDATE candidate must not itself become CODE_STATUS_UNMAPPABLE."""
    await _seed_equipment(db_session, bcm_code="BCM_BLANKST", item_no="ITEM_BLANKST")
    [findings] = await _run_adapter(db_session, [_valid_row(bcm="BCM_BLANKST", item_no="ITEM_BLANKST", status=None)])
    assert CODE_STATUS_UNMAPPABLE not in _codes(findings)
    assert CODE_STATUS_MISSING not in _codes(findings)


# ---------------------------------------------------------------------------
# No-Equipment-mutation guarantee (including Equipment.version unchanged)
# ---------------------------------------------------------------------------


async def test_validation_never_mutates_any_equipment_field(db_session, seeded_users):
    existing = await _seed_equipment(
        db_session,
        bcm_code="BCM_NOMUT",
        item_no="ITEM_NOMUT",
        equipment_name="Original Name",
        asset_number="ORIGINAL-AN",
        status=EquipmentStatus.AVAILABLE_AT_POOL,
    )
    before = {
        "id": existing.id,
        "asset_number": existing.asset_number,
        "equipment_name": existing.equipment_name,
        "bcm_code": existing.bcm_code,
        "item_no": existing.item_no,
        "status": existing.status,
        "version": existing.version,
        "updated_at": existing.updated_at,
    }

    await _run_adapter(
        db_session,
        [_valid_row(bcm="BCM_NOMUT", item_no="ITEM_NOMUT", equipment_name="A Totally Different Name", status="defective")],
    )

    db_session.expire_all()
    after = await equipment_crud.get_by_id(db_session, before["id"])
    assert after.asset_number == before["asset_number"]
    assert after.equipment_name == before["equipment_name"]
    assert after.bcm_code == before["bcm_code"]
    assert after.item_no == before["item_no"]
    assert after.status == before["status"]
    assert after.version == before["version"], "PR20C must never bump Equipment.version"
    assert after.updated_at == before["updated_at"]


async def test_no_new_equipment_rows_are_ever_created(db_session, seeded_users):
    count_before = len((await db_session.execute(select(Equipment))).scalars().all())
    await _run_adapter(
        db_session,
        [_valid_row(bcm="BCM_NC1", item_no="1"), _valid_row(bcm="BCM_NC2", item_no="2")],
    )
    count_after = len((await db_session.execute(select(Equipment))).scalars().all())
    assert count_after == count_before


# ---------------------------------------------------------------------------
# Batch lookup / no-N+1 (bounded query count regardless of source row count)
# ---------------------------------------------------------------------------


async def test_preload_business_context_issues_a_bounded_query_count_regardless_of_row_count(db_engine, seeded_users_via_engine):
    session_maker = async_sessionmaker(db_engine, expire_on_commit=False, class_=AsyncSession)

    async def _count_queries_for(n_rows: int) -> int:
        rows = [_valid_row(bcm=f"BCM_BATCH_{i}", item_no=str(i)) for i in range(n_rows)]
        content = _build_workbook_bytes(rows)
        adapter = EquipmentMasterAdapter()
        async with session_maker() as db:
            records = adapter.parse(VerifiedSourceContent(content=content, source_descriptor=_fake_descriptor()))
            statements: list[str] = []

            def _capture(conn, cursor, statement, parameters, context, executemany):
                statements.append(statement)

            event.listen(db_engine.sync_engine, "before_cursor_execute", _capture)
            try:
                await adapter.preload_business_context(db, records)
            finally:
                event.remove(db_engine.sync_engine, "before_cursor_execute", _capture)
            select_statements = [s for s in statements if s.strip().upper().startswith("SELECT")]
            return len(select_statements)

    small_count = await _count_queries_for(20)
    large_count = await _count_queries_for(400)
    assert small_count == large_count, (
        f"Equipment lookup query count must stay bounded as source row count grows "
        f"(20 rows -> {small_count} SELECTs, 400 rows -> {large_count} SELECTs)"
    )
    assert small_count <= 4, "exactly the four bulk IN(...) lookups (bcm/item_no/serial/asset_id), never one per row"


@pytest_asyncio.fixture
async def seeded_users_via_engine(db_engine):
    """A `seeded_users`-equivalent for the query-count test above, which
    needs its own session per query-count measurement (so it manages
    `db_engine` directly rather than depending on the shared `db_session`
    fixture)."""
    from app.core.security import hash_password
    from app.models.user import ALL_ROLES, User, Role

    session_maker = async_sessionmaker(db_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_maker() as db:
        for name in ALL_ROLES:
            db.add(Role(name=name, permissions={}))
        await db.flush()
        result = await db.execute(select(Role))
        roles = {r.name: r for r in result.scalars().all()}
        for role_name in ALL_ROLES:
            db.add(
                User(
                    employee_code=f"{role_name.upper()}001",
                    full_name=f"Test {role_name}",
                    email=f"{role_name}@mep-hospital-test.dev",
                    password_hash=hash_password("Password@123"),
                    role_id=roles[role_name].id,
                )
            )
        await db.commit()


# ---------------------------------------------------------------------------
# Adapter registry / registration
# ---------------------------------------------------------------------------


def test_adapter_is_registered_at_module_import_for_equipment_master_dataset_type():
    from app.services.import_adapter import get_adapter

    adapter = get_adapter(DATASET_TYPE)
    assert adapter is not None
    assert isinstance(adapter, EquipmentMasterAdapter)
    assert adapter.dataset_type == "equipment_master"


def test_adapter_package_marker_performs_no_registration_itself():
    # The package `__init__` module itself must define nothing registry-
    # related -- registration is a side effect of importing the concrete
    # `equipment_master` submodule only (verified by app.main doing exactly
    # that import), never of importing the package.
    assert not hasattr(import_adapters, "register_adapter")


# ---------------------------------------------------------------------------
# PR19/PR20A framework integration -- true end-to-end tests through the HTTP
# API, proving dispatch/off-thread-parsing/findings/counters/structural
# failure all work correctly for this adapter without re-testing the
# generic framework mechanics themselves (already covered by
# test_import_validation.py / test_pr20a_source_artifact_infrastructure.py).
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture(autouse=True)
async def _patch_validation_service_session_factory(db_engine, monkeypatch):
    session_maker = async_sessionmaker(db_engine, expire_on_commit=False, class_=AsyncSession)
    monkeypatch.setattr(import_lease, "AsyncSessionLocal", session_maker)


async def _create_session(client: AsyncClient, headers: dict) -> dict:
    r = await client.post("/api/v1/import-sessions", headers=headers, json={"dataset_type": DATASET_TYPE})
    assert r.status_code in (200, 201), r.text
    return r.json()


async def _upload(client: AsyncClient, headers: dict, session_id: str, content: bytes):
    files = {"file": ("equipment_master.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    return await client.post(f"/api/v1/import-sessions/{session_id}/source/upload", headers=headers, files=files)


async def test_end_to_end_valid_workbook_persists_findings_and_coherent_counters(client: AsyncClient, seeded_users):
    headers = await auth_headers(client)
    session = await _create_session(client, headers)
    content = _build_workbook_bytes(
        [
            _valid_row(bcm="BCM_E2E_1", item_no="1"),  # CREATE candidate -> blocked by OD-4
            _valid_row(bcm=None, item_no="2"),  # blank BCM -> blocked
        ]
    )
    up = await _upload(client, headers, session["id"], content)
    assert up.status_code == 201, up.text

    r = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "validation_failed"
    assert body["total_rows"] == 2
    assert body["invalid_rows"] == 2
    assert body["valid_rows"] == 0

    errors = await client.get(f"/api/v1/import-sessions/{session['id']}/errors", headers=headers)
    assert errors.status_code == 200
    codes = {item["error_code"] for item in errors.json()["items"]}
    assert CODE_ASSET_NUMBER_REQUIRED_FOR_CREATE in codes
    assert CODE_BCM_MISSING in codes


async def test_end_to_end_dispatches_to_equipment_master_adapter_only(client: AsyncClient, seeded_users):
    """Proves dispatch-by-dataset_type: an unrelated dataset_type on the
    same running process must not be routed to this adapter."""
    headers = await auth_headers(client)
    r = await client.post(
        "/api/v1/import-sessions", headers=headers, json={"dataset_type": "some_unrelated_dataset_type"}
    )
    assert r.status_code in (200, 201), r.text
    session = r.json()
    reg = await client.post(
        f"/api/v1/import-sessions/{session['id']}/source",
        headers=headers,
        json={"checksum": "a" * 64, "byte_size": 1},
    )
    # Metadata-only registration is only rejected for equipment_master
    # (§6.2) -- an unrelated dataset_type is unaffected either way; what
    # matters here is that validate for it never reaches
    # EquipmentMasterAdapter (no adapter registered for it at all -> 422).
    v = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    if reg.status_code == 201:
        assert v.status_code == 422
        assert v.json()["code"] == "IMPORT_ADAPTER_NOT_REGISTERED"


def _build_valid_zip_invalid_xlsx_bytes() -> bytes:
    """A structurally valid ZIP archive using an allowed OOXML-internal
    entry path (passes upload-time `_validate_zip_archive_bounds`'s
    allowed-entry-path check, mirroring
    `test_pr20a_source_artifact_infrastructure.py`'s own `_build_xlsx_
    bytes` helper) whose content is not genuine OOXML XML -- isolates
    "upload-time zip-bound checks" from "validate-time `load_workbook`
    parse failure", which are two distinct checks at two distinct times."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as archive:
        archive.writestr("[Content_Types].xml", b"not genuine OOXML content")
    return buf.getvalue()


async def test_end_to_end_off_thread_parsing_a_malformed_workbook_fails_cleanly(client: AsyncClient, seeded_users):
    headers = await auth_headers(client)
    session = await _create_session(client, headers)
    up = await _upload(client, headers, session["id"], _build_valid_zip_invalid_xlsx_bytes())
    assert up.status_code == 201, up.text  # upload only enforces zip-structure bounds, not sheet contract

    r = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "validation_failed"
    assert body["failure_reason"] is not None


async def test_end_to_end_structural_failure_leaves_no_partial_findings(client: AsyncClient, seeded_users, db_session):
    from app.models.import_session import ImportJob, ImportRowError

    headers = await auth_headers(client)
    session = await _create_session(client, headers)
    up = await _upload(client, headers, session["id"], _build_valid_zip_invalid_xlsx_bytes())
    assert up.status_code == 201, up.text

    r = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "validation_failed"

    job = (
        await db_session.execute(select(ImportJob).where(ImportJob.import_session_id == uuid.UUID(session["id"])))
    ).scalar_one()
    findings = (
        (await db_session.execute(select(ImportRowError).where(ImportRowError.import_job_id == job.id))).scalars().all()
    )
    assert findings == [], "a structural parse failure must never leave partial per-row findings behind"


async def test_end_to_end_metadata_only_registration_rejected_for_equipment_master(client: AsyncClient, seeded_users):
    headers = await auth_headers(client)
    session = await _create_session(client, headers)
    r = await client.post(
        f"/api/v1/import-sessions/{session['id']}/source",
        headers=headers,
        json={"checksum": "a" * 64, "byte_size": 1},
    )
    assert r.status_code != 201, "equipment_master must only accept the byte-upload source path (§6.2)"


async def test_end_to_end_valid_update_candidate_produces_no_blocking_findings(
    client: AsyncClient, seeded_users, db_session
):
    await _seed_equipment(db_session, bcm_code="BCM_E2E_UPD", item_no="ITEM_E2E_UPD", equipment_name="Old Name")

    headers = await auth_headers(client)
    session = await _create_session(client, headers)
    content = _build_workbook_bytes([_valid_row(bcm="BCM_E2E_UPD", item_no="ITEM_E2E_UPD")])
    up = await _upload(client, headers, session["id"], content)
    assert up.status_code == 201, up.text

    r = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "validated"
    assert body["invalid_rows"] == 0
    assert body["valid_rows"] == 1

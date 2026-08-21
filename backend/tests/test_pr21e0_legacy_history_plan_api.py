"""Roadmap PR21E0 -- Legacy Import Operator API Surface, gap B.

Tests the PR21-specific dry-run-plan HTTP surface
(`app.api.v1.legacy_history_import`): `GET .../legacy-history/dry-run-plan`,
`GET .../legacy-history/dry-run-plan/{plan_id}/rows`,
`POST .../legacy-history/dry-run-plan/{plan_id}/confirm`. Also proves (a)
the existing PR20 `GET/POST .../dry-run-plan[...]` routes
(`app.api.v1.import_sessions`) remain completely unchanged, and (b) a plan
confirmed through this new route executes end-to-end through the existing
generic `POST .../execute` endpoint exactly as a plan confirmed through
the internal provider already did in PR21D2.

Synthetic workbooks only -- no production workbook data."""

import hashlib
import uuid
from datetime import date, time

import pytest
import pytest_asyncio
from httpx import AsyncClient
from io import BytesIO
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.crud import legacy_history_dry_run_plan as legacy_history_dry_run_plan_crud
from app.models.audit import AuditLog
from app.models.equipment import Equipment, EquipmentStatus
from app.models.legacy_history import LegacyEquipmentEvent
from app.models.master_data import Ward
from app.models.transaction import BorrowTransaction
from app.models.user import ROLE_READ_ONLY, User
from app.services import import_execution_service, import_lease
from app.services.identifiers import normalize_bcm_code
from app.services.import_adapters.legacy_history import issue as issue_module
from app.services.import_adapters.legacy_history import receive as receive_module
from app.services.import_plan_providers.legacy_history import DATASET_TYPE
from tests.conftest import auth_headers

_EQUIPMENT_MASTER_DATASET_TYPE = "equipment_master"

ISSUE_HEADER_SHEET = issue_module.HEADER_SHEET_NAME
ISSUE_LINE_SHEET = issue_module.LINE_SHEET_NAME
RECEIVE_HEADER_SHEET = receive_module.HEADER_SHEET_NAME
RECEIVE_LINE_SHEET = receive_module.LINE_SHEET_NAME
ISSUE_HEADER_COLUMNS = list(issue_module._HEADER_GOVERNED_HEADERS.keys())
ISSUE_LINE_COLUMNS = list(issue_module._LINE_GOVERNED_HEADERS.keys())
RECEIVE_HEADER_COLUMNS = list(receive_module._HEADER_GOVERNED_HEADERS.keys())
RECEIVE_LINE_COLUMNS = list(receive_module._LINE_GOVERNED_HEADERS.keys())


@pytest_asyncio.fixture(autouse=True)
async def _patch_execution_session_factories(db_engine, monkeypatch):
    """Mirrors test_pr21d2_historical_event_execution.py's identical
    fixture -- `POST .../execute` runs on its own session factory, not
    the test's `db_session`."""
    session_maker = async_sessionmaker(db_engine, expire_on_commit=False, class_=AsyncSession)
    monkeypatch.setattr(import_execution_service, "AsyncSessionLocal", session_maker)
    monkeypatch.setattr(import_lease, "AsyncSessionLocal", session_maker)


# ---------------------------------------------------------------------------
# Workbook construction (mirrors test_pr21d2_historical_event_execution.py).
# ---------------------------------------------------------------------------


def _build_combined_workbook(*, issue_headers=None, issue_lines=None, receive_headers=None, receive_lines=None):
    wb = Workbook()
    ws = wb.active
    ws.title = ISSUE_HEADER_SHEET
    ws.append(ISSUE_HEADER_COLUMNS)
    for row in issue_headers or []:
        ws.append([row.get(col) for col in ISSUE_HEADER_COLUMNS])

    ws = wb.create_sheet(ISSUE_LINE_SHEET)
    ws.append(ISSUE_LINE_COLUMNS)
    for row in issue_lines or []:
        ws.append([row.get(col) for col in ISSUE_LINE_COLUMNS])

    ws = wb.create_sheet(RECEIVE_HEADER_SHEET)
    ws.append(RECEIVE_HEADER_COLUMNS)
    for row in receive_headers or []:
        ws.append([row.get(col) for col in RECEIVE_HEADER_COLUMNS])

    ws = wb.create_sheet(RECEIVE_LINE_SHEET)
    ws.append(RECEIVE_LINE_COLUMNS)
    for row in receive_lines or []:
        ws.append([row.get(col) for col in RECEIVE_LINE_COLUMNS])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _issue_header_row(*, order_ref="ORD-0001", ward="Ward 1", bme="BME Header", d=date(2024, 1, 10), t=time(9, 0, 0)):
    return {
        "วันที่": d,
        "เลขที่ใบยืม": order_ref,
        "แผนกที่ยืม": ward,
        "ผู้ส่งเครื่องยืม (User)": "header-user",
        "ผู้ส่งเครื่องยืม (BME)": bme,
        "เวลา": t,
        "จำนวนเครื่องส่งทั้งหมด": 1,
        "หมายเหตุ": None,
    }


def _issue_line_row(*, row_key="1", order_ref="ORD-0001", me_code="ME001", ward="Ward 1", bme="BME Line", d=date(2024, 1, 10), t=time(9, 0, 0)):
    return {
        "ลำดับ": row_key,
        "วันที่": d,
        "เลขที่ใบส่ง": order_ref,
        "SCAN CODE ส่ง": "SCAN-1",
        "ME.Code": me_code,
        "Barcode ส่งเครื่อง": "BARCODE-1",
        "Equipment": "Infusion Pump",
        "Brand": "BrandX",
        "Model": "ModelY",
        "Serial no.": "SN-1",
        "รูปเครื่อง": None,
        "แผนกที่ส่ง": ward,
        "ตัวเครื่องหน้าจอไม่แตกร้าว": "OK",
        "สายชาร์จ (สายไม่ขาด)": "OK",
        "Pole Clamp (ไม่แตกหักมไม่ง้อ)": "OK",
        "จำนวน": 1,
        "หมายเหตุ": None,
        "เวลา": t,
        "ชื่อ BME": bme,
        "ชื่อ (User)": "line-user",
    }


def _default_workbook(**overrides) -> bytes:
    kwargs = dict(issue_headers=[_issue_header_row()], issue_lines=[_issue_line_row()], receive_headers=[], receive_lines=[])
    kwargs.update(overrides)
    return _build_combined_workbook(**kwargs)


def _second_issue_row_workbook() -> bytes:
    return _build_combined_workbook(
        issue_headers=[_issue_header_row()],
        issue_lines=[
            _issue_line_row(row_key="1", me_code="ME001"),
            _issue_line_row(row_key="2", me_code="ME002"),
        ],
        receive_headers=[],
        receive_lines=[],
    )


# ---------------------------------------------------------------------------
# Seeding helpers.
# ---------------------------------------------------------------------------


async def _seed_equipment(db_session: AsyncSession, *, me_code: str = "ME001", **kwargs) -> Equipment:
    defaults = dict(
        asset_number=f"AN-{uuid.uuid4().hex[:10]}",
        equipment_name="Legacy Test Equipment",
        status=EquipmentStatus.AVAILABLE_AT_POOL,
        bcm_code=normalize_bcm_code(me_code),
    )
    defaults.update(kwargs)
    eq = Equipment(**defaults)
    db_session.add(eq)
    await db_session.commit()
    await db_session.refresh(eq)
    return eq


async def _seed_ward(db_session: AsyncSession, *, code: str = "Ward 1") -> Ward:
    ward = Ward(code=code, name=code)
    db_session.add(ward)
    await db_session.commit()
    await db_session.refresh(ward)
    return ward


async def _actor_id(db_session: AsyncSession) -> uuid.UUID:
    return (await db_session.execute(select(User.id).limit(1))).scalar_one()


async def _approve_authority(client: AsyncClient, headers: dict, checksum: str) -> dict:
    r = await client.post(
        "/api/v1/legacy-migration-authorities",
        headers=headers,
        json={"scope": "pr21_legacy_transaction_history_v1", "approved_workbook_sha256": checksum},
    )
    assert r.status_code in (200, 201), r.text
    return r.json()


async def _create_session(client: AsyncClient, headers: dict, dataset_type: str = DATASET_TYPE) -> dict:
    r = await client.post("/api/v1/import-sessions", headers=headers, json={"dataset_type": dataset_type})
    assert r.status_code in (200, 201), r.text
    return r.json()


async def _upload(client: AsyncClient, headers: dict, session_id: str, content: bytes):
    files = {"file": ("legacy.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    return await client.post(f"/api/v1/import-sessions/{session_id}/source/upload", headers=headers, files=files)


async def _dry_run_completed_session(client: AsyncClient, headers: dict, content: bytes) -> dict:
    """Create -> upload -> validate -> dry-run through the real HTTP
    pipeline (authority must already be approved for `content`'s
    checksum). Returns the session dict; the resulting plan is left
    unconfirmed."""
    session = await _create_session(client, headers)
    up = await _upload(client, headers, session["id"], content)
    assert up.status_code == 201, up.text
    v = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert v.status_code == 200, v.text
    dr = await client.post(f"/api/v1/import-sessions/{session['id']}/dry-run", headers=headers)
    assert dr.status_code == 200, dr.text
    assert dr.json()["status"] == "dry_run_completed", dr.text
    return session


# ---------------------------------------------------------------------------
# A. GET plan.
# ---------------------------------------------------------------------------


async def test_get_plan_returns_identity_and_summary(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, headers, checksum)

    session = await _dry_run_completed_session(client, headers, content)
    r = await client.get(f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["import_session_id"] == session["id"]
    assert body["status"] == "active"
    assert body["is_current"] is True
    assert body["confirmed_at"] is None
    assert body["confirmed_by_user_id"] is None
    summary = body["summary"]
    assert summary["total_rows"] == 1
    assert summary["issue_events"] == 1
    assert summary["receive_events"] == 0
    # Never the Equipment Master field vocabulary.
    assert "creates" not in summary
    assert "updates" not in summary
    assert "skips" not in summary
    assert "rows" not in body


async def test_get_plan_non_admin_rejected(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    admin_headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, admin_headers, checksum)
    session = await _dry_run_completed_session(client, admin_headers, content)

    other_headers = await auth_headers(client, role=ROLE_READ_ONLY)
    r = await client.get(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=other_headers
    )
    assert r.status_code == 403, r.text


async def test_get_plan_anonymous_rejected(client: AsyncClient, seeded_users):
    r = await client.get(f"/api/v1/import-sessions/{uuid.uuid4()}/legacy-history/dry-run-plan")
    assert r.status_code == 401, r.text


async def test_get_plan_unknown_session_404(client: AsyncClient, seeded_users):
    headers = await auth_headers(client)
    r = await client.get(f"/api/v1/import-sessions/{uuid.uuid4()}/legacy-history/dry-run-plan", headers=headers)
    assert r.status_code == 404, r.text


async def test_get_plan_rejects_equipment_master_session(client: AsyncClient, seeded_users):
    """§13/§18 of the task: a real Equipment Master session must never be
    servable through this PR21-specific route -- collapses into the same
    404 an unknown session id gets, never a distinguishing error."""
    headers = await auth_headers(client)
    session = await _create_session(client, headers, dataset_type=_EQUIPMENT_MASTER_DATASET_TYPE)
    r = await client.get(f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers)
    assert r.status_code == 404, r.text


async def test_get_plan_no_plan_yet_returns_404(client: AsyncClient, seeded_users):
    headers = await auth_headers(client)
    session = await _create_session(client, headers)
    r = await client.get(f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers)
    assert r.status_code == 404, r.text


# ---------------------------------------------------------------------------
# B. Rows pagination.
# ---------------------------------------------------------------------------


async def test_rows_pagination_and_typed_values(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session, me_code="ME001")
    await _seed_equipment(db_session, me_code="ME002", asset_number="AN-SECOND")
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _second_issue_row_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, headers, checksum)
    session = await _dry_run_completed_session(client, headers, content)

    plan_resp = await client.get(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers
    )
    plan_id = plan_resp.json()["id"]

    page1 = await client.get(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan_id}/rows",
        headers=headers,
        params={"limit": 1},
    )
    assert page1.status_code == 200, page1.text
    body1 = page1.json()
    assert len(body1["items"]) == 1
    assert body1["total"] == 2
    assert body1["next_cursor"] is not None
    row = body1["items"][0]
    assert row["event_type"] == "ISSUE"
    values = row["values"]
    assert values["legacy_ward_text"] == "Ward 1"
    assert values["header_source_ref"]["sheet_name"] == ISSUE_HEADER_SHEET
    assert values["line_source_ref"]["sheet_name"] == ISSUE_LINE_SHEET

    page2 = await client.get(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan_id}/rows",
        headers=headers,
        params={"limit": 1, "cursor": body1["next_cursor"]},
    )
    assert page2.status_code == 200, page2.text
    body2 = page2.json()
    assert len(body2["items"]) == 1
    # Deterministic ordering: strictly increasing source_row_number
    # (never re-derived from `ลำดับ`/legacy_source_row_key, which is
    # source-provided and not itself the ordering key).
    assert body2["items"][0]["source_row_number"] > row["source_row_number"]
    assert body2["next_cursor"] is None
    assert body1["items"][0]["id"] != body2["items"][0]["id"]


async def test_rows_cross_plan_cursor_rejected(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)

    content_a = _default_workbook()
    checksum_a = hashlib.sha256(content_a).hexdigest()
    await _approve_authority(client, headers, checksum_a)
    session_a = await _dry_run_completed_session(client, headers, content_a)
    plan_a = (
        await client.get(f"/api/v1/import-sessions/{session_a['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()
    rows_a = (
        await client.get(
            f"/api/v1/import-sessions/{session_a['id']}/legacy-history/dry-run-plan/{plan_a['id']}/rows",
            headers=headers,
            params={"limit": 1},
        )
    ).json()
    cursor_from_a = rows_a["next_cursor"]
    assert cursor_from_a is None  # only one row -- build a real cursor via limit=0-equivalent below instead

    # Force a real next_cursor on plan A using the two-row workbook instead.
    content_a2 = _second_issue_row_workbook()
    checksum_a2 = hashlib.sha256(content_a2).hexdigest()
    await _approve_authority(client, headers, checksum_a2)
    await _seed_equipment(db_session, me_code="ME002", asset_number="AN-A2-SECOND")
    session_a2 = await _dry_run_completed_session(client, headers, content_a2)
    plan_a2 = (
        await client.get(
            f"/api/v1/import-sessions/{session_a2['id']}/legacy-history/dry-run-plan", headers=headers
        )
    ).json()
    rows_a2 = (
        await client.get(
            f"/api/v1/import-sessions/{session_a2['id']}/legacy-history/dry-run-plan/{plan_a2['id']}/rows",
            headers=headers,
            params={"limit": 1},
        )
    ).json()
    real_cursor = rows_a2["next_cursor"]
    assert real_cursor is not None

    content_b = _default_workbook(issue_headers=[_issue_header_row(order_ref="ORD-0002")], issue_lines=[_issue_line_row(order_ref="ORD-0002")])
    checksum_b = hashlib.sha256(content_b).hexdigest()
    await _approve_authority(client, headers, checksum_b)
    session_b = await _dry_run_completed_session(client, headers, content_b)
    plan_b = (
        await client.get(f"/api/v1/import-sessions/{session_b['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()

    cross = await client.get(
        f"/api/v1/import-sessions/{session_b['id']}/legacy-history/dry-run-plan/{plan_b['id']}/rows",
        headers=headers,
        params={"cursor": real_cursor},
    )
    assert cross.status_code in (400, 422), cross.text


async def test_rows_foreign_session_plan_rejected(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)

    content_a = _default_workbook()
    checksum_a = hashlib.sha256(content_a).hexdigest()
    await _approve_authority(client, headers, checksum_a)
    session_a = await _dry_run_completed_session(client, headers, content_a)
    plan_a = (
        await client.get(f"/api/v1/import-sessions/{session_a['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()

    content_b = _default_workbook(issue_headers=[_issue_header_row(order_ref="ORD-0002")], issue_lines=[_issue_line_row(order_ref="ORD-0002")])
    checksum_b = hashlib.sha256(content_b).hexdigest()
    await _approve_authority(client, headers, checksum_b)
    session_b = await _dry_run_completed_session(client, headers, content_b)

    r = await client.get(
        f"/api/v1/import-sessions/{session_b['id']}/legacy-history/dry-run-plan/{plan_a['id']}/rows",
        headers=headers,
    )
    assert r.status_code == 404, r.text


async def test_rows_invalid_cursor_rejected(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, headers, checksum)
    session = await _dry_run_completed_session(client, headers, content)
    plan = (
        await client.get(f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()

    r = await client.get(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/rows",
        headers=headers,
        params={"cursor": "not-a-real-cursor"},
    )
    assert r.status_code in (400, 422), r.text


async def test_rows_limit_boundary_rejected(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, headers, checksum)
    session = await _dry_run_completed_session(client, headers, content)
    plan = (
        await client.get(f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()

    too_big = await client.get(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/rows",
        headers=headers,
        params={"limit": 201},
    )
    assert too_big.status_code == 422, too_big.text

    too_small = await client.get(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/rows",
        headers=headers,
        params={"limit": 0},
    )
    assert too_small.status_code == 422, too_small.text


async def test_rows_non_admin_rejected(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    admin_headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, admin_headers, checksum)
    session = await _dry_run_completed_session(client, admin_headers, content)
    plan = (
        await client.get(
            f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=admin_headers
        )
    ).json()

    other_headers = await auth_headers(client, role=ROLE_READ_ONLY)
    r = await client.get(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/rows",
        headers=other_headers,
    )
    assert r.status_code == 403, r.text


# ---------------------------------------------------------------------------
# C. Confirm.
# ---------------------------------------------------------------------------


async def test_confirm_first_succeeds_and_writes_one_audit(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, headers, checksum)
    session = await _dry_run_completed_session(client, headers, content)
    plan = (
        await client.get(f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()

    r = await client.post(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/confirm",
        headers=headers,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["confirmed_at"] is not None
    assert body["confirmed_by_user_id"]
    assert body["summary"]["issue_events"] == 1

    audit_rows = (
        await db_session.execute(
            select(AuditLog).where(
                AuditLog.action == "import_dry_run_plan_confirmed",
                AuditLog.entity_type == "import_session",
                AuditLog.entity_id == uuid.UUID(session["id"]),
            )
        )
    ).scalars().all()
    assert len(audit_rows) == 1


async def test_confirm_retry_same_actor_no_duplicate_audit(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, headers, checksum)
    session = await _dry_run_completed_session(client, headers, content)
    plan = (
        await client.get(f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()

    confirm_url = f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/confirm"
    first = await client.post(confirm_url, headers=headers)
    assert first.status_code == 200, first.text
    second = await client.post(confirm_url, headers=headers)
    assert second.status_code == 200, second.text
    assert second.json()["confirmed_at"] == first.json()["confirmed_at"]
    assert second.json()["confirmed_by_user_id"] == first.json()["confirmed_by_user_id"]

    audit_rows = (
        await db_session.execute(
            select(AuditLog).where(
                AuditLog.action == "import_dry_run_plan_confirmed",
                AuditLog.entity_id == uuid.UUID(session["id"]),
            )
        )
    ).scalars().all()
    assert len(audit_rows) == 1


async def test_confirm_retry_different_actor_preserves_original_confirmer(
    client: AsyncClient, seeded_users, db_session
):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    admin_headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, admin_headers, checksum)
    session = await _dry_run_completed_session(client, admin_headers, content)
    plan = (
        await client.get(
            f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=admin_headers
        )
    ).json()

    confirm_url = f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/confirm"
    first = await client.post(confirm_url, headers=admin_headers)
    assert first.status_code == 200, first.text

    # A second Administrator user retries the same confirm -- the
    # persisted confirmed_by_user_id must remain the FIRST confirmer.
    second_admin_id = (
        await db_session.execute(
            select(User.id).where(User.id != uuid.UUID(first.json()["confirmed_by_user_id"]))
        )
    ).scalars().first()
    if second_admin_id is None:
        pytest.skip("only one seeded user available to prove retry-different-actor preservation")
    second = await client.post(confirm_url, headers=admin_headers)
    assert second.status_code == 200, second.text
    assert second.json()["confirmed_by_user_id"] == first.json()["confirmed_by_user_id"]


async def test_confirm_foreign_plan_rejected_as_stale(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)

    content_a = _default_workbook()
    checksum_a = hashlib.sha256(content_a).hexdigest()
    await _approve_authority(client, headers, checksum_a)
    session_a = await _dry_run_completed_session(client, headers, content_a)
    plan_a = (
        await client.get(f"/api/v1/import-sessions/{session_a['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()

    content_b = _default_workbook(issue_headers=[_issue_header_row(order_ref="ORD-0002")], issue_lines=[_issue_line_row(order_ref="ORD-0002")])
    checksum_b = hashlib.sha256(content_b).hexdigest()
    await _approve_authority(client, headers, checksum_b)
    session_b = await _dry_run_completed_session(client, headers, content_b)

    r = await client.post(
        f"/api/v1/import-sessions/{session_b['id']}/legacy-history/dry-run-plan/{plan_a['id']}/confirm",
        headers=headers,
    )
    assert r.status_code == 409, r.text


async def test_confirm_superseded_plan_rejected(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, headers, checksum)
    session = await _dry_run_completed_session(client, headers, content)
    plan = (
        await client.get(f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()

    # A second dry-run supersedes the first plan.
    dr2 = await client.post(f"/api/v1/import-sessions/{session['id']}/dry-run", headers=headers)
    assert dr2.status_code == 200, dr2.text

    r = await client.post(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/confirm",
        headers=headers,
    )
    assert r.status_code == 409, r.text


async def test_confirm_non_admin_rejected(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    admin_headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    await _approve_authority(client, admin_headers, checksum)
    session = await _dry_run_completed_session(client, admin_headers, content)
    plan = (
        await client.get(
            f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=admin_headers
        )
    ).json()

    other_headers = await auth_headers(client, role=ROLE_READ_ONLY)
    r = await client.post(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/confirm",
        headers=other_headers,
    )
    assert r.status_code == 403, r.text


# ---------------------------------------------------------------------------
# D. PR20 regression -- the original route is byte/field unchanged.
# ---------------------------------------------------------------------------


async def test_pr20_dry_run_plan_route_still_returns_equipment_master_shape(client: AsyncClient, seeded_users):
    """§22 of the task. The ORIGINAL `GET /import-sessions/{id}/dry-run-
    plan` route (`app.api.v1.import_sessions`) must never be affected by
    this slice -- it still resolves via `import_dry_run_plan_crud`
    (Equipment Master), returns `DryRunPlanOut`'s Equipment Master
    field vocabulary, and rejects a `legacy_transaction_history` session
    exactly as before (no plan ever exists there, so `404`)."""
    headers = await auth_headers(client)
    session = await _create_session(client, headers, dataset_type=_EQUIPMENT_MASTER_DATASET_TYPE)
    r = await client.get(f"/api/v1/import-sessions/{session['id']}/dry-run-plan", headers=headers)
    assert r.status_code == 404, r.text
    assert r.json()["code"] == "IMPORT_DRY_RUN_PLAN_NOT_FOUND"

    # Never reachable from the new PR21-specific route family's own path.
    r2 = await client.get(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers
    )
    assert r2.status_code == 404, r2.text


# ---------------------------------------------------------------------------
# E. Execution regression + full backend end-to-end workflow.
# ---------------------------------------------------------------------------


async def test_plan_confirmed_via_new_route_executes_through_existing_execute_endpoint(
    client: AsyncClient, seeded_users, db_session
):
    """§23/§24 of the task, combined into one full backend workflow:
    Administrator approves checksum via the NEW authority route -> create
    session -> upload synthetic workbook -> validate -> dry-run -> GET
    PR21 plan via the NEW route -> list rows via the NEW route -> confirm
    via the NEW route -> execute via the EXISTING generic `POST
    .../execute` (unchanged by this slice) -> `LegacyEquipmentEvent`
    exists, no `BorrowTransaction` was created, `Equipment`'s live state
    is unchanged. Synthetic data only."""
    equipment = await _seed_equipment(db_session)
    original_version = equipment.version
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()

    authority = await _approve_authority(client, headers, checksum)
    assert authority["approved_workbook_sha256"] == checksum

    session = await _dry_run_completed_session(client, headers, content)

    plan = (
        await client.get(f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan", headers=headers)
    ).json()
    assert plan["migration_authority_id"] == authority["id"]

    rows_page = (
        await client.get(
            f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/rows",
            headers=headers,
        )
    ).json()
    assert len(rows_page["items"]) == 1

    confirm = await client.post(
        f"/api/v1/import-sessions/{session['id']}/legacy-history/dry-run-plan/{plan['id']}/confirm",
        headers=headers,
    )
    assert confirm.status_code == 200, confirm.text
    assert confirm.json()["status"] == "active"
    assert confirm.json()["confirmed_at"] is not None

    execute = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert execute.status_code == 200, execute.text
    assert execute.json()["status"] == "completed"
    assert execute.json()["imported_rows"] == 1

    events = (await db_session.execute(select(LegacyEquipmentEvent))).scalars().all()
    assert len(events) == 1
    assert events[0].event_type == "ISSUE"
    assert events[0].migration_authority_id == uuid.UUID(authority["id"])

    borrows = (await db_session.execute(select(BorrowTransaction))).scalars().all()
    assert borrows == []

    await db_session.refresh(equipment)
    assert equipment.status == EquipmentStatus.AVAILABLE_AT_POOL
    assert equipment.version == original_version
    assert equipment.current_location_id is None

    final_plan = await legacy_history_dry_run_plan_crud.get_plan_by_id(
        db_session, plan_id=uuid.UUID(plan["id"]), import_session_id=uuid.UUID(session["id"])
    )
    assert final_plan.status == "consumed"

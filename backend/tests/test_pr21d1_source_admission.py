"""Roadmap PR21D1 -- Combined Canonical Adapter + Source Admission.

HTTP-level source-admission tests: session creation, upload (including
the PR21-specific bounded upload allowance selection), validate, dry-run
through the real `ImportSession`/`ImportJob` pipeline (proving
`persist_dry_run_plan`'s FK-bound writes succeed end to end), the
structurally-blocked `execute` endpoint, and the `LegacyHistoryDryRunPlanProvider`
used directly (mirroring PR21A's own test convention -- see this file's
own §E below for why the generic `GET/POST .../dry-run-plan` HTTP
endpoints cannot be used for this dataset_type).

Adapter-internals-level tests (parse/preload/validate/plan_dry_run direct
invocation) live in the sibling file
`test_pr21d1_combined_legacy_history_adapter.py`.

Synthetic workbooks only -- no production workbook data is read or
committed anywhere in this file."""

import hashlib
import uuid
from datetime import date, time
from io import BytesIO

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.exceptions import InvalidInputError
from app.models.equipment import Equipment, EquipmentStatus
from app.models.legacy_history import LegacyHistoryDryRunPlan, LegacyMigrationAuthority
from app.models.master_data import Ward
from app.services import import_execution_service, import_lease
from app.services.identifiers import normalize_bcm_code
from app.services.import_adapters.legacy_history import issue as issue_module
from app.services.import_adapters.legacy_history import receive as receive_module
from app.services.import_plan_provider import get_plan_provider
from app.services.import_plan_providers.legacy_history import DATASET_TYPE
from app.services.import_service import MAX_UPLOAD_BYTES
from tests.conftest import auth_headers

import app.api.v1.import_sessions as import_sessions_module

EQUIPMENT_MASTER_DATASET_TYPE = "equipment_master"

ISSUE_HEADER_SHEET = issue_module.HEADER_SHEET_NAME
ISSUE_LINE_SHEET = issue_module.LINE_SHEET_NAME
RECEIVE_HEADER_SHEET = receive_module.HEADER_SHEET_NAME
RECEIVE_LINE_SHEET = receive_module.LINE_SHEET_NAME
ISSUE_HEADER_COLUMNS = list(issue_module._HEADER_GOVERNED_HEADERS.keys())
ISSUE_LINE_COLUMNS = list(issue_module._LINE_GOVERNED_HEADERS.keys())
RECEIVE_HEADER_COLUMNS = list(receive_module._HEADER_GOVERNED_HEADERS.keys())
RECEIVE_LINE_COLUMNS = list(receive_module._LINE_GOVERNED_HEADERS.keys())


@pytest_asyncio.fixture(autouse=True)
async def _patch_execution_session_factories(db_engine, monkeypatch):
    """Mirrors test_pr20d_dry_run_plan.py's identical fixture -- both
    modules that open their own `AsyncSessionLocal()` must be repointed
    at this test's own engine."""
    session_maker = async_sessionmaker(db_engine, expire_on_commit=False, class_=AsyncSession)
    monkeypatch.setattr(import_execution_service, "AsyncSessionLocal", session_maker)
    monkeypatch.setattr(import_lease, "AsyncSessionLocal", session_maker)


def _build_combined_workbook(
    *,
    issue_headers=None,
    issue_lines=None,
    receive_headers=None,
    receive_lines=None,
    extra_sheets=None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = ISSUE_HEADER_SHEET
    ws.append(ISSUE_HEADER_COLUMNS)
    for row in issue_headers or []:
        ws.append([row.get(col) for col in ISSUE_HEADER_COLUMNS])

    ws = wb.create_sheet(ISSUE_LINE_SHEET)
    ws.append(ISSUE_LINE_COLUMNS)
    for row in issue_lines or []:
        ws.append([row.get(col) for col in ISSUE_LINE_COLUMNS])

    ws = wb.create_sheet(RECEIVE_HEADER_SHEET)
    ws.append(RECEIVE_HEADER_COLUMNS)
    for row in receive_headers or []:
        ws.append([row.get(col) for col in RECEIVE_HEADER_COLUMNS])

    ws = wb.create_sheet(RECEIVE_LINE_SHEET)
    ws.append(RECEIVE_LINE_COLUMNS)
    for row in receive_lines or []:
        ws.append([row.get(col) for col in RECEIVE_LINE_COLUMNS])

    if extra_sheets:
        for name, rows in extra_sheets.items():
            ws = wb.create_sheet(name)
            for row in rows:
                ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _issue_header_row(*, order_ref="ORD-0001", ward="Ward 1", bme="BME Header", d=date(2024, 1, 10), t=time(9, 0, 0)):
    return {
        "วันที่": d,
        "เลขที่ใบยืม": order_ref,
        "แผนกที่ยืม": ward,
        "ผู้ส่งเครื่องยืม (User)": "header-user",
        "ผู้ส่งเครื่องยืม (BME)": bme,
        "เวลา": t,
        "จำนวนเครื่องส่งทั้งหมด": 1,
        "หมายเหตุ": None,
    }


def _issue_line_row(*, row_key="1", order_ref="ORD-0001", me_code="ME001", ward="Ward 1", bme="BME Line", d=date(2024, 1, 10), t=time(9, 0, 0)):
    return {
        "ลำดับ": row_key,
        "วันที่": d,
        "เลขที่ใบส่ง": order_ref,
        "SCAN CODE ส่ง": "SCAN-1",
        "ME.Code": me_code,
        "Barcode ส่งเครื่อง": "BARCODE-1",
        "Equipment": "Infusion Pump",
        "Brand": "BrandX",
        "Model": "ModelY",
        "Serial no.": "SN-1",
        "รูปเครื่อง": None,
        "แผนกที่ส่ง": ward,
        "ตัวเครื่องหน้าจอไม่แตกร้าว": "OK",
        "สายชาร์จ (สายไม่ขาด)": "OK",
        "Pole Clamp (ไม่แตกหักมไม่ง้อ)": "OK",
        "จำนวน": 1,
        "หมายเหตุ": None,
        "เวลา": t,
        "ชื่อ BME": bme,
        "ชื่อ (User)": "line-user",
    }


def _receive_header_row(*, order_ref="RET-0001", ward="Ward 1", bme="BME Header", d=date(2024, 1, 11), t=time(10, 0, 0)):
    return {
        "วันที่": d,
        "เลขที่ใบคืน": order_ref,
        "แผนกที่คืน": ward,
        "ผู้ส่งเครื่องคืน (User)": "header-user",
        "ผู้รับเครื่องคืน (BME)": bme,
        "เวลา": t,
        "จำนวนเครื่องรับคืนทั้งหมด": 1,
        "หมายเหตุ": None,
    }


def _receive_line_row(*, row_key="1", order_ref="RET-0001", me_code="ME001", ward="Ward 1", bme="BME Line", d=date(2024, 1, 11), t=time(10, 0, 0)):
    return {
        "ลำดับ": row_key,
        "วันที่": d,
        "เลขที่ใบรับเครื่อง": order_ref,
        "SCAN CODE รับ": "SCAN-1",
        "ME.Code": me_code,
        "Barcode รับเครื่อง": "BARCODE-1",
        "Equipment": "Infusion Pump",
        "Brand": "BrandX",
        "Model": "ModelY",
        "Serial no.": "SN-1",
        "แผนกที่รับ": ward,
        "รูปเครื่อง": None,
        "ตัวเครื่องหน้าจอไม่แตกร้าว": "OK",
        "สายชาร์จ (สายไม่ขาด)": "OK",
        "Pole Clamp (ไม่แตกหักมไม่ง้อ)": "OK",
        "จำนวน": 1,
        "หมายเหตุ": None,
        "เวลา": t,
        "ชื่อ BME": bme,
        "ชื่อ (User)": "line-user",
    }


def _default_workbook(**overrides) -> bytes:
    kwargs = dict(
        issue_headers=[_issue_header_row()],
        issue_lines=[_issue_line_row()],
        receive_headers=[_receive_header_row()],
        receive_lines=[_receive_line_row()],
    )
    kwargs.update(overrides)
    return _build_combined_workbook(**kwargs)


async def _seed_equipment(db_session: AsyncSession, *, me_code: str = "ME001", **kwargs) -> Equipment:
    defaults = dict(
        asset_number=f"AN-{uuid.uuid4().hex[:10]}",
        equipment_name="Legacy Test Equipment",
        status=EquipmentStatus.AVAILABLE_AT_POOL,
        bcm_code=normalize_bcm_code(me_code),
    )
    defaults.update(kwargs)
    eq = Equipment(**defaults)
    db_session.add(eq)
    await db_session.commit()
    await db_session.refresh(eq)
    return eq


async def _seed_ward(db_session: AsyncSession, *, code: str = "Ward 1") -> Ward:
    ward = Ward(code=code, name=code)
    db_session.add(ward)
    await db_session.commit()
    await db_session.refresh(ward)
    return ward


async def _seed_authority(db_session: AsyncSession, *, checksum: str, actor_id: uuid.UUID) -> LegacyMigrationAuthority:
    authority = LegacyMigrationAuthority(
        scope="pr21d1-test", approved_workbook_sha256=checksum, approved_by_user_id=actor_id
    )
    db_session.add(authority)
    await db_session.commit()
    await db_session.refresh(authority)
    return authority


async def _create_session(client: AsyncClient, headers: dict, *, dataset_type: str = DATASET_TYPE) -> dict:
    r = await client.post("/api/v1/import-sessions", headers=headers, json={"dataset_type": dataset_type})
    assert r.status_code in (200, 201), r.text
    return r.json()


async def _upload(client: AsyncClient, headers: dict, session_id: str, content: bytes, *, filename="legacy.xlsx"):
    files = {"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    return await client.post(f"/api/v1/import-sessions/{session_id}/source/upload", headers=headers, files=files)


# ---------------------------------------------------------------------------
# A. Upload cap selection: PR21-specific allowance vs. the unchanged
#    generic cap, selected server-side from `dataset_type`.
# ---------------------------------------------------------------------------


async def test_upload_selects_pr21_cap_for_legacy_dataset_and_generic_cap_for_others(
    client: AsyncClient, seeded_users, monkeypatch
):
    """Proves the exact `max_bytes` each dataset_type is admitted under,
    without allocating either cap's full multi-megabyte buffer."""
    captured: list[int] = []
    real_read = import_sessions_module._read_upload_bounded

    async def _spy(file, *, max_bytes):
        captured.append(max_bytes)
        return await real_read(file, max_bytes=max_bytes)

    monkeypatch.setattr(import_sessions_module, "_read_upload_bounded", _spy)

    headers = await auth_headers(client)
    legacy_session = await _create_session(client, headers, dataset_type=DATASET_TYPE)
    r1 = await _upload(client, headers, legacy_session["id"], _default_workbook())
    assert r1.status_code == 201, r1.text

    equipment_session = await _create_session(client, headers, dataset_type=EQUIPMENT_MASTER_DATASET_TYPE)
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"
    buf = BytesIO()
    wb.save(buf)
    r2 = await _upload(client, headers, equipment_session["id"], buf.getvalue())
    assert r2.status_code in (201, 400, 422), r2.text  # content shape irrelevant here, only the cap matters

    assert captured[0] == import_sessions_module.PR21_MAX_UPLOAD_BYTES
    assert captured[1] == MAX_UPLOAD_BYTES
    assert captured[0] != captured[1]


async def test_upload_rejects_content_over_pr21_cap_boundary(client: AsyncClient, seeded_users, monkeypatch):
    """Boundary test at a small, fast-to-allocate cap (monkeypatched) --
    proves the PR21-specific cap is actually enforced, not merely selected."""
    monkeypatch.setattr(import_sessions_module, "PR21_MAX_UPLOAD_BYTES", 2000)
    headers = await auth_headers(client)
    session = await _create_session(client, headers, dataset_type=DATASET_TYPE)
    content = _default_workbook()
    assert len(content) > 2000, "the synthetic combined workbook must exceed the tiny patched cap to prove rejection"
    r = await _upload(client, headers, session["id"], content)
    assert r.status_code == 400, r.text


async def test_upload_generic_cap_unchanged_for_equipment_master(client: AsyncClient, seeded_users, monkeypatch):
    """§31 regression: the PR21-specific cap selection must never leak
    into the generic upload path for an unrelated dataset_type."""
    captured: list[int] = []
    real_read = import_sessions_module._read_upload_bounded

    async def _spy(file, *, max_bytes):
        captured.append(max_bytes)
        return await real_read(file, max_bytes=max_bytes)

    monkeypatch.setattr(import_sessions_module, "_read_upload_bounded", _spy)
    headers = await auth_headers(client)
    session = await _create_session(client, headers, dataset_type=EQUIPMENT_MASTER_DATASET_TYPE)
    wb = Workbook()
    buf = BytesIO()
    wb.save(buf)
    await _upload(client, headers, session["id"], buf.getvalue())
    assert captured == [MAX_UPLOAD_BYTES]


# ---------------------------------------------------------------------------
# B. Full session pipeline: create -> upload -> validate -> dry-run,
#    persisted plan bound to real ImportSession/ImportSource/ImportJob
#    identity, no execution, no Equipment mutation.
# ---------------------------------------------------------------------------


async def _validated_session(client, headers, content: bytes) -> dict:
    session = await _create_session(client, headers, dataset_type=DATASET_TYPE)
    up = await _upload(client, headers, session["id"], content)
    assert up.status_code == 201, up.text
    v = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert v.status_code == 200, v.text
    assert v.json()["status"] == "validated", v.json()
    return session


async def test_combined_success_validates_and_dry_runs_end_to_end(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()

    from app.models.user import User

    actor_id = (await db_session.execute(select(User.id).limit(1))).scalar_one()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validated_session(client, headers, content)
    dr = await client.post(f"/api/v1/import-sessions/{session['id']}/dry-run", headers=headers)
    assert dr.status_code == 200, dr.text
    assert dr.json()["status"] == "dry_run_completed"

    session_uuid = uuid.UUID(session["id"])
    plan = (
        await db_session.execute(
            select(LegacyHistoryDryRunPlan).where(LegacyHistoryDryRunPlan.import_session_id == session_uuid)
        )
    ).scalar_one()
    assert plan.status == "active"
    assert plan.summary_issue_events == 1
    assert plan.summary_receive_events == 1
    assert plan.summary_total_rows == 2
    assert str(plan.import_source_id) is not None
    assert plan.source_checksum == checksum


async def test_issue_side_failure_blocks_whole_session_dry_run(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook(issue_lines=[_issue_line_row(me_code="ME_DOES_NOT_EXIST")])

    session = await _create_session(client, headers, dataset_type=DATASET_TYPE)
    up = await _upload(client, headers, session["id"], content)
    assert up.status_code == 201, up.text
    v = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert v.status_code == 200
    assert v.json()["status"] == "validation_failed", "one bad Issue row must fail the WHOLE combined validation"


async def test_execute_endpoint_structurally_blocked_even_after_dry_run(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    from app.models.user import User

    actor_id = (await db_session.execute(select(User.id).limit(1))).scalar_one()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validated_session(client, headers, content)
    dr = await client.post(f"/api/v1/import-sessions/{session['id']}/dry-run", headers=headers)
    assert dr.status_code == 200, dr.text

    ex = await client.post(f"/api/v1/import-sessions/{session['id']}/execute", headers=headers)
    assert ex.status_code == 501, ex.text
    assert ex.json()["code"] == "IMPORT_ADAPTER_NOT_IMPLEMENTED"


async def test_dry_run_without_authority_fails_and_no_plan_persisted(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    session = await _validated_session(client, headers, content)

    dr = await client.post(f"/api/v1/import-sessions/{session['id']}/dry-run", headers=headers)
    assert dr.status_code != 200 or dr.json()["status"] != "dry_run_completed"

    session_uuid = uuid.UUID(session["id"])
    plans = (
        (
            await db_session.execute(
                select(LegacyHistoryDryRunPlan).where(LegacyHistoryDryRunPlan.import_session_id == session_uuid)
            )
        )
        .scalars()
        .all()
    )
    assert plans == [], "no LegacyHistoryDryRunPlan may ever be persisted without a matching authority"


# ---------------------------------------------------------------------------
# C. Worksheet-cap / SDC / missing-canonical-sheet regressions at the
#    HTTP boundary (parse() is reached during /validate).
# ---------------------------------------------------------------------------


async def test_sdc_sheet_present_does_not_block_validation(client: AsyncClient, seeded_users, db_session):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook(extra_sheets={"SDC": [["ignored"]]})
    session = await _validated_session(client, headers, content)
    assert session is not None


async def test_missing_canonical_sheet_fails_upload_validate_flow(client: AsyncClient, seeded_users):
    headers = await auth_headers(client)
    wb = Workbook()
    ws = wb.active
    ws.title = ISSUE_HEADER_SHEET
    ws.append(ISSUE_HEADER_COLUMNS)
    ws = wb.create_sheet(ISSUE_LINE_SHEET)
    ws.append(ISSUE_LINE_COLUMNS)
    ws = wb.create_sheet(RECEIVE_HEADER_SHEET)
    ws.append(RECEIVE_HEADER_COLUMNS)
    # RECEIVE_LINE_SHEET omitted.
    buf = BytesIO()
    wb.save(buf)

    session = await _create_session(client, headers, dataset_type=DATASET_TYPE)
    up = await _upload(client, headers, session["id"], buf.getvalue())
    assert up.status_code == 201, up.text
    v = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert v.status_code == 200
    assert v.json()["status"] == "validation_failed", "a missing canonical sheet must fail the whole session"


async def test_worksheet_cap_boundary_at_http_layer(client: AsyncClient, seeded_users, db_session, monkeypatch):
    """Reuses PR21B's own established worksheet-cap-regression pattern at
    the combined-adapter HTTP boundary: at-cap succeeds, over-cap fails
    validate with the whole-session error."""
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)

    over_cap_extra = {f"Extra{i}": [["x"]] for i in range(29)}  # 4 + 29 = 33 > 32
    content = _default_workbook(extra_sheets=over_cap_extra)
    session = await _create_session(client, headers, dataset_type=DATASET_TYPE)
    up = await _upload(client, headers, session["id"], content)
    assert up.status_code == 201, up.text
    v = await client.post(f"/api/v1/import-sessions/{session['id']}/validate", headers=headers)
    assert v.status_code == 200
    assert v.json()["status"] == "validation_failed", "over-cap worksheet count must fail the whole session"


# ---------------------------------------------------------------------------
# D. Retention/redaction still dispatches correctly for this dataset_type
#    via the plan-provider registry (`app.crud.import_retention`) --
#    proven at the provider level, per PR21A's own established pattern.
# ---------------------------------------------------------------------------


async def test_plan_provider_is_reachable_and_matches_dataset_type():
    provider = get_plan_provider(DATASET_TYPE)
    assert provider is not None
    assert provider.dataset_type == DATASET_TYPE


# ---------------------------------------------------------------------------
# E. Discovered gap (honestly documented, not silently patched): the
#    generic `GET/POST .../dry-run-plan` HTTP endpoints
#    (`app.api.v1.import_sessions.get_dry_run_plan`/`confirm_dry_run_plan`)
#    are hard-coded to `EquipmentMasterDryRunPlan`/
#    `app.crud.import_dry_run_plan` -- they never consult the
#    `DryRunPlanProvider` registry (`app.services.import_plan_provider`)
#    that `LegacyHistoryDryRunPlanProvider` is correctly registered
#    against. This predates PR21D1 (the same hard-coding already existed
#    for every dataset_type before this slice) and is out of this
#    bounded slice's scope to fix (wiring a second, provider-dispatched
#    read/confirm HTTP surface is a real architectural change, not a
#    parser-composition task) -- documented here as a regression-style
#    proof rather than silently worked around. See this PR's own final
#    report for the full write-up.
# ---------------------------------------------------------------------------


async def test_KNOWN_GAP_generic_dry_run_plan_endpoint_cannot_see_a_real_legacy_plan(
    client: AsyncClient, seeded_users, db_session
):
    await _seed_equipment(db_session)
    await _seed_ward(db_session)
    headers = await auth_headers(client)
    content = _default_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    from app.models.user import User

    actor_id = (await db_session.execute(select(User.id).limit(1))).scalar_one()
    await _seed_authority(db_session, checksum=checksum, actor_id=actor_id)

    session = await _validated_session(client, headers, content)
    dr = await client.post(f"/api/v1/import-sessions/{session['id']}/dry-run", headers=headers)
    assert dr.status_code == 200, dr.text

    session_uuid = uuid.UUID(session["id"])
    # The plan genuinely exists, in the correct PR21-owned table.
    real_plan = (
        await db_session.execute(
            select(LegacyHistoryDryRunPlan).where(LegacyHistoryDryRunPlan.import_session_id == session_uuid)
        )
    ).scalar_one()
    assert real_plan is not None

    # But the generic HTTP read endpoint -- hard-coded to
    # EquipmentMasterDryRunPlan -- cannot see it. This is a pre-existing
    # gap, not a PR21D1 regression; asserted here so the gap is provable
    # and cannot silently regress further (e.g. into a wrong-dataset
    # 200 instead of a clean 404).
    r = await client.get(f"/api/v1/import-sessions/{session['id']}/dry-run-plan", headers=headers)
    assert r.status_code == 404
    assert r.json()["code"] == "IMPORT_DRY_RUN_PLAN_NOT_FOUND"

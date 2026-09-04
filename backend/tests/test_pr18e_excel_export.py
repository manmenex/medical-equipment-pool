import asyncio
import io
import threading
import time
import uuid
from datetime import date, datetime, timezone

import openpyxl
import pytest
from sqlalchemy import func, select

from app.core.exceptions import XlsxRenderTimeoutError
from app.models.audit import AuditLog
from app.models.transaction import BorrowTransaction
from app.models.user import ROLE_ADMINISTRATOR, ROLE_EQUIPMENT_POOL_STAFF, ROLE_READ_ONLY
from app.schemas.report_export import (
    ExportColumn,
    ExportDocument,
    ExportFilterSummary,
    ExportMetadata,
    ExportRow,
    ReportIdentity,
)
from app.services import report_export_service, report_xlsx_service
from tests.conftest import auth_headers as _auth_headers
from tests.conftest import create_ward as _create_ward

# ---------------------------------------------------------------------------
# Roadmap PR18E (docs/design/PR18_PRINTING_EXPORT_PLAN.md §11, §16): unit
# tests for app.services.report_xlsx_service (the openpyxl adapter) and API
# tests for GET /reports/{report_id}/xlsx.
# ---------------------------------------------------------------------------

_FIXED_NOW = datetime(2026, 8, 3, 3, 15, 0, tzinfo=timezone.utc)


def _document(
    *,
    rows: list[ExportRow] | None = None,
    columns: list[ExportColumn] | None = None,
    generated_by_display_name: str = "ผู้ทดสอบ",
    applied_filters: list[ExportFilterSummary] | None = None,
    display_name_th: str = "รายงานการรับคืน",
) -> ExportDocument:
    columns = columns if columns is not None else [ExportColumn(key="note", label_th="หมายเหตุ", value_type="string")]
    rows = rows if rows is not None else []
    applied_filters = (
        applied_filters if applied_filters is not None else [ExportFilterSummary(label_th="หอผู้ป่วย", value="ICU ทดสอบ")]
    )
    return ExportDocument(
        metadata=ExportMetadata(
            report_identity=ReportIdentity.RECEIVE_REPORT,
            display_name_th=display_name_th,
            template_version="1",
            generated_at=_FIXED_NOW,
            generated_by_display_name=generated_by_display_name,
            generated_by_user_id="00000000-0000-0000-0000-000000000001",
            timezone="Asia/Bangkok",
            applied_filters=applied_filters,
            row_count=len(rows),
            filename_stem="receive-report_all_all_20260803T031500Z",
        ),
        columns=columns,
        rows=rows,
    )


def _open(xlsx_bytes: bytes):
    assert xlsx_bytes.startswith(b"PK")  # xlsx is a ZIP container
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes))


def _all_rows(ws) -> list[tuple]:
    return list(ws.iter_rows(values_only=True))


def _header_row_index(ws, document: ExportDocument) -> int:
    """Finds the header row by matching the first column's declared Thai
    label -- avoids hardcoding a row index the metadata block's own length
    could shift."""
    first_label = document.columns[0].label_th
    for row in ws.iter_rows():
        if row[0].value == first_label:
            return row[0].row
    raise AssertionError("header row not found")


# ---------------------------------------------------------------------------
# Unit tests: app.services.report_xlsx_service.build_workbook_sync
# ---------------------------------------------------------------------------


def test_build_workbook_produces_a_valid_xlsx():
    document = _document(rows=[ExportRow(values={"note": "x"})])
    xlsx_bytes = report_xlsx_service.build_workbook_sync(document)
    wb = _open(xlsx_bytes)
    assert wb.active is not None


def test_build_workbook_worksheet_title_matches_report_identity():
    document = _document()
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    assert wb.active.title == "Receive Report"


@pytest.mark.parametrize(
    "report_identity,expected_title",
    [
        (ReportIdentity.RECEIVE_REPORT, "Receive Report"),
        (ReportIdentity.ISSUE_REPORT, "Issue Report"),
        (ReportIdentity.EQUIPMENT_VERIFY_CHECKLIST, "Equipment Verify Checklist"),
    ],
)
def test_build_workbook_worksheet_title_per_report_identity(report_identity, expected_title):
    document = _document()
    document = document.model_copy(update={"metadata": document.metadata.model_copy(update={"report_identity": report_identity})})
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    assert wb.active.title == expected_title


def test_build_workbook_metadata_block_contains_generation_context():
    document = _document()
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert "Medical Equipment Pool" in values
    assert document.metadata.display_name_th in values
    assert document.metadata.generated_by_display_name in values
    assert document.metadata.report_identity.value in values
    assert document.metadata.template_version in values
    assert "ICU ทดสอบ" in values  # the applied filter's value


def test_build_workbook_header_row_matches_declared_columns_in_order():
    document = _document(
        columns=[
            ExportColumn(key="a", label_th="คอลัมน์ที่หนึ่ง", value_type="string"),
            ExportColumn(key="b", label_th="คอลัมน์ที่สอง", value_type="string"),
        ],
        rows=[ExportRow(values={"a": "1", "b": "2"})],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    header_values = [cell.value for cell in ws[header_row_index]][:2]
    assert header_values == ["คอลัมน์ที่หนึ่ง", "คอลัมน์ที่สอง"]


def test_build_workbook_empty_result_set_produces_valid_workbook_with_header_only():
    document = _document(rows=[])
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    # No data rows below the header.
    assert ws.cell(row=header_row_index + 1, column=1).value is None
    # Row count is still visible in the metadata block.
    values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert "0" in values


def test_build_workbook_date_and_datetime_columns_use_native_cell_types_not_text():
    document = _document(
        columns=[
            ExportColumn(key="d", label_th="วันที่", value_type="date"),
            ExportColumn(key="dt", label_th="วันเวลา", value_type="datetime"),
        ],
        rows=[
            ExportRow(
                values={
                    "d": date(2026, 7, 1),
                    "dt": datetime(2026, 7, 1, 10, 30, tzinfo=timezone.utc),
                }
            )
        ],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    date_cell = ws.cell(row=header_row_index + 1, column=1)
    datetime_cell = ws.cell(row=header_row_index + 1, column=2)
    assert isinstance(date_cell.value, (date, datetime))
    assert isinstance(datetime_cell.value, datetime)
    # Not a preformatted ISO string.
    assert not isinstance(date_cell.value, str)
    assert not isinstance(datetime_cell.value, str)
    # The UTC 10:30 instant is displayed converted to the document's own
    # declared Asia/Bangkok timezone (UTC+7) -- 17:30, not the bare UTC value.
    assert datetime_cell.value.hour == 17
    assert datetime_cell.value.tzinfo is None


def test_build_workbook_none_values_are_genuinely_blank_cells():
    document = _document(
        columns=[ExportColumn(key="note", label_th="หมายเหตุ", value_type="string")],
        rows=[ExportRow(values={"note": None})],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    assert ws.cell(row=header_row_index + 1, column=1).value is None


def test_build_workbook_numeric_columns_are_native_numbers():
    document = _document(
        columns=[
            ExportColumn(key="i", label_th="จำนวนเต็ม", value_type="integer"),
            ExportColumn(key="f", label_th="ทศนิยม", value_type="decimal"),
        ],
        rows=[ExportRow(values={"i": 42, "f": 3.5})],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    assert ws.cell(row=header_row_index + 1, column=1).value == 42
    assert ws.cell(row=header_row_index + 1, column=2).value == 3.5


def test_build_workbook_boolean_columns_use_thai_labels():
    document = _document(
        columns=[ExportColumn(key="b", label_th="สถานะ", value_type="boolean")],
        rows=[ExportRow(values={"b": True}), ExportRow(values={"b": False})],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    assert ws.cell(row=header_row_index + 1, column=1).value == "ใช่"
    assert ws.cell(row=header_row_index + 2, column=1).value == "ไม่ใช่"


def test_build_workbook_freezes_header_row_and_sets_autofilter():
    document = _document(rows=[ExportRow(values={"note": "x"})])
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    assert ws.freeze_panes == f"A{header_row_index + 1}"
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith(f"A{header_row_index}:")


def test_build_workbook_column_widths_are_set_within_bounds():
    document = _document(
        columns=[ExportColumn(key="note", label_th="ก" * 100, value_type="string")],
        rows=[ExportRow(values={"note": "x"})],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    width = ws.column_dimensions["A"].width
    assert width is not None
    assert width <= report_xlsx_service._MAX_COLUMN_WIDTH


def test_build_workbook_single_worksheet_only():
    document = _document(rows=[ExportRow(values={"note": "x"})])
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    assert wb.sheetnames == [wb.active.title]


# ---------------------------------------------------------------------------
# Unit tests: formula-injection protection (design §11/security)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("raw_value", ["=SUM(A1:A2)", "+1+1", "-1+1", "@SUM(A1)", "=cmd|'/c calc'!A1"])
def test_build_workbook_escapes_formula_injection_prefixes(raw_value):
    document = _document(
        columns=[ExportColumn(key="note", label_th="หมายเหตุ", value_type="string")],
        rows=[ExportRow(values={"note": raw_value})],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    stored_value = ws.cell(row=header_row_index + 1, column=1).value
    # Stored as literal text -- prefixed so the leading character can never
    # be interpreted as a formula operator, but the original content is
    # still fully recoverable (nothing was silently dropped or truncated).
    assert stored_value == f"'{raw_value}"
    assert isinstance(stored_value, str)


@pytest.mark.parametrize("safe_value", ["ปกติ", "AST-0001", "ข้อความ = ปกติ", "10% เสร็จ"])
def test_build_workbook_does_not_escape_safe_string_values(safe_value):
    """A value that merely *contains* a formula-prefix character elsewhere
    (not as its first character) must pass through unchanged -- only a
    leading =, +, -, or @ is a formula-injection risk."""
    document = _document(
        columns=[ExportColumn(key="note", label_th="หมายเหตุ", value_type="string")],
        rows=[ExportRow(values={"note": safe_value})],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    assert ws.cell(row=header_row_index + 1, column=1).value == safe_value


def test_build_workbook_html_escaping_not_needed_but_special_characters_preserved():
    """Excel has no HTML-injection concern (unlike the PDF/print adapters),
    but the literal text must still survive unmodified once the formula-
    injection guard has run."""
    raw_value = "<b>ทดสอบ</b> & \"quoted\""
    document = _document(
        columns=[ExportColumn(key="note", label_th="หมายเหตุ", value_type="string")],
        rows=[ExportRow(values={"note": raw_value})],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    assert ws.cell(row=header_row_index + 1, column=1).value == raw_value


# ---------------------------------------------------------------------------
# Unit tests: workbook-wide formula-injection protection (Roadmap PR18E
# review round 1, H1). The round 1 review found that only report *rows*
# were sanitized -- the metadata block (title, generated-by, applied-filter
# values, all of which can carry operator-editable free text such as a
# user's display name or a ward/category/equipment/operator name) bypassed
# `_sanitize_string_cell` entirely. `report_xlsx_service._write_cell` is
# now the single, centralized call site every string write in the module
# goes through; these tests prove that end to end for every dynamic
# metadata field, not just report rows.
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("dangerous_prefix", ["=", "+", "-", "@"])
def test_build_workbook_sanitizes_generated_by_display_name(dangerous_prefix):
    """`generated_by_display_name` is a user's `full_name` -- operator-
    editable free text, not a developer-authored constant."""
    dangerous_value = f"{dangerous_prefix}HYPERLINK(\"http://evil.example\")"
    document = _document(generated_by_display_name=dangerous_value)
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert f"'{dangerous_value}" in values
    assert dangerous_value not in values


@pytest.mark.parametrize("dangerous_prefix", ["=", "+", "-", "@"])
def test_build_workbook_sanitizes_report_title(dangerous_prefix):
    """`display_name_th` (the report title written into the metadata
    block) is document-supplied, not a hardcoded worksheet label -- must
    still be guarded defensively."""
    dangerous_value = f"{dangerous_prefix}cmd|'/c calc'!A1"
    document = _document(display_name_th=dangerous_value)
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert f"'{dangerous_value}" in values
    assert dangerous_value not in values


@pytest.mark.parametrize("dangerous_prefix", ["=", "+", "-", "@"])
def test_build_workbook_sanitizes_applied_filter_values(dangerous_prefix):
    """Applied-filter values are backend-resolved *display names* -- Ward,
    category, equipment, and operator names are all administrator-editable
    free text (report_export_service._filter_summary), the same class of
    value the round 1 review flagged as unprotected."""
    dangerous_value = f"{dangerous_prefix}SUM(A1:A100)"
    document = _document(applied_filters=[ExportFilterSummary(label_th="หอผู้ป่วย", value=dangerous_value)])
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert f"'{dangerous_value}" in values
    assert dangerous_value not in values


def test_build_workbook_metadata_ordinary_text_remains_unchanged():
    """A metadata value that does not begin with a formula-injection
    prefix must reach the workbook byte-for-byte unchanged."""
    document = _document(
        generated_by_display_name="สมชาย ใจดี",
        display_name_th="รายงานทดสอบปกติ",
        applied_filters=[ExportFilterSummary(label_th="หอผู้ป่วย", value="ICU อาคาร 2")],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert "สมชาย ใจดี" in values
    assert "รายงานทดสอบปกติ" in values
    assert "ICU อาคาร 2" in values


def test_build_workbook_numeric_and_date_metadata_cells_remain_unchanged():
    """Row-count and generated-at are never strings in this module's
    output (row-count is embedded in a formatted label string, matching
    design's Thai-first metadata block; the table's own date/datetime
    *data* cells stay fully native, proven separately above) -- this test
    proves the sanitization pass introduced for H1 does not corrupt those
    already-correct native/numeric-adjacent cells."""
    document = _document(
        columns=[ExportColumn(key="d", label_th="วันที่", value_type="date")],
        rows=[ExportRow(values={"d": date(2026, 7, 1)})],
    )
    wb = _open(report_xlsx_service.build_workbook_sync(document))
    ws = wb.active
    header_row_index = _header_row_index(ws, document)
    date_cell_value = ws.cell(row=header_row_index + 1, column=1).value
    assert isinstance(date_cell_value, (date, datetime))
    assert not isinstance(date_cell_value, str)

    values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert "1" in values  # row_count == 1, embedded in "จำนวนรายการ: 1"


# ---------------------------------------------------------------------------
# Unit tests: admission control (Roadmap PR18E review round 1, H2) --
# report_xlsx_service.build_workbook_bounded. Reuses the exact same proven
# protection model report_pdf_service.render_pdf_bounded implements: one
# total deadline covering both queue wait and active generation, and
# renderer-lifetime (not request-lifetime) concurrency accounting. These
# tests mirror tests/test_pr18d_pdf_export.py's own admission-control
# tests for the PDF adapter, using threading.Event gates (not time.sleep)
# so a background generation is provably still running -- not just
# probably still running due to timing -- at the moment each assertion
# executes.
# ---------------------------------------------------------------------------


async def _wait_until(predicate, *, timeout: float = 5.0, interval: float = 0.01) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        if predicate():
            return
        await asyncio.sleep(interval)
    raise AssertionError(f"condition not met within {timeout}s")


async def test_build_workbook_bounded_succeeds_within_timeout(monkeypatch):
    monkeypatch.setattr(report_xlsx_service, "RENDER_TIMEOUT_SECONDS", 15)
    xlsx_bytes = await report_xlsx_service.build_workbook_bounded(_document(rows=[ExportRow(values={"note": "x"})]))
    assert xlsx_bytes.startswith(b"PK")


async def test_build_workbook_bounded_times_out_while_queued_for_capacity_without_starting_generation(monkeypatch):
    """Required test: "queue timeout" -- a request that never gets a chance
    to acquire the renderer semaphore before its total timeout budget
    elapses must still raise the structured timeout error, and
    `build_workbook_sync` must never have been invoked on its behalf."""
    render_call_count = 0
    hold_gate = threading.Event()
    occupier_started = threading.Event()

    def _holding_render(document):
        nonlocal render_call_count
        render_call_count += 1
        occupier_started.set()
        hold_gate.wait(timeout=5)
        return b"PK\x03\x04fake"

    monkeypatch.setattr(report_xlsx_service, "build_workbook_sync", _holding_render)
    monkeypatch.setattr(report_xlsx_service, "RENDER_TIMEOUT_SECONDS", 0.05)
    monkeypatch.setattr(report_xlsx_service, "_render_semaphore", asyncio.Semaphore(1))

    document = _document()

    occupier_task = asyncio.ensure_future(report_xlsx_service.build_workbook_bounded(document))
    await _wait_until(lambda: occupier_started.is_set())

    with pytest.raises(XlsxRenderTimeoutError):
        await report_xlsx_service.build_workbook_bounded(document)

    assert render_call_count == 1, "the queue-timeout request's generation must never have been started"

    hold_gate.set()
    with pytest.raises(XlsxRenderTimeoutError):
        await occupier_task


async def test_build_workbook_bounded_active_generation_times_out_for_caller(monkeypatch):
    """Required test: "active export timeout" -- a request that has
    acquired renderer capacity and is actively generating must still raise
    the structured timeout error for its caller once the total budget
    elapses, even though the underlying generation keeps running in the
    background."""
    started = threading.Event()
    release_gate = threading.Event()

    def _slow_render(document):
        started.set()
        release_gate.wait(timeout=5)
        return b"PK\x03\x04fake"

    monkeypatch.setattr(report_xlsx_service, "build_workbook_sync", _slow_render)
    monkeypatch.setattr(report_xlsx_service, "RENDER_TIMEOUT_SECONDS", 0.05)
    monkeypatch.setattr(report_xlsx_service, "_render_semaphore", asyncio.Semaphore(2))

    with pytest.raises(XlsxRenderTimeoutError):
        await report_xlsx_service.build_workbook_bounded(_document())

    assert started.is_set(), "generation should have actually started before the timeout fired"
    release_gate.set()


async def test_build_workbook_bounded_timeout_does_not_release_capacity_early(monkeypatch):
    """Required test: "slot released only after generation completes" --
    immediately after a caller-facing timeout, the semaphore must still
    show the slot as held, not freed, because the underlying generation
    has not actually finished yet."""
    release_gate = threading.Event()
    started = threading.Event()

    def _blocking_render(document):
        started.set()
        release_gate.wait(timeout=5)
        return b"PK\x03\x04fake"

    monkeypatch.setattr(report_xlsx_service, "build_workbook_sync", _blocking_render)
    monkeypatch.setattr(report_xlsx_service, "RENDER_TIMEOUT_SECONDS", 0.05)
    test_semaphore = asyncio.Semaphore(2)
    monkeypatch.setattr(report_xlsx_service, "_render_semaphore", test_semaphore)

    with pytest.raises(XlsxRenderTimeoutError):
        await report_xlsx_service.build_workbook_bounded(_document())

    assert started.is_set()
    assert test_semaphore._value == 1, "the slot must remain held after a caller-facing timeout, not be freed early"

    release_gate.set()
    await _wait_until(lambda: test_semaphore._value == 2)


async def test_build_workbook_bounded_next_request_proceeds_after_slot_release(monkeypatch):
    """Required test: "next request proceeds after slot release" -- a
    timed-out slot must not permanently consume its concurrency capacity;
    once the original generation actually finishes, the next request must
    still be admitted and succeed."""
    call_count = 0

    def _slow_then_fast_render(document):
        nonlocal call_count
        call_count += 1
        if call_count == 1:
            time.sleep(0.2)
        return b"PK\x03\x04fake"

    monkeypatch.setattr(report_xlsx_service, "build_workbook_sync", _slow_then_fast_render)
    monkeypatch.setattr(report_xlsx_service, "RENDER_TIMEOUT_SECONDS", 0.05)
    monkeypatch.setattr(report_xlsx_service, "_render_semaphore", asyncio.Semaphore(1))

    document = _document()
    with pytest.raises(XlsxRenderTimeoutError):
        await report_xlsx_service.build_workbook_bounded(document)

    monkeypatch.setattr(report_xlsx_service, "RENDER_TIMEOUT_SECONDS", 15)
    xlsx_bytes = await report_xlsx_service.build_workbook_bounded(document)
    assert xlsx_bytes.startswith(b"PK")


async def test_build_workbook_bounded_timeout_does_not_increase_effective_concurrency(monkeypatch):
    """Required test: "concurrency limit never exceeded" (timeout variant)
    -- a burst of requests that all time out from the caller's perspective
    must not be able to start more simultaneous generations than
    `_render_semaphore` allows."""
    active = 0
    max_active = 0
    count_lock = threading.Lock()
    release_gate = threading.Event()

    def _tracking_blocking_render(document):
        nonlocal active, max_active
        with count_lock:
            active += 1
            max_active = max(max_active, active)
        release_gate.wait(timeout=5)
        with count_lock:
            active -= 1
        return b"PK\x03\x04fake"

    monkeypatch.setattr(report_xlsx_service, "build_workbook_sync", _tracking_blocking_render)
    monkeypatch.setattr(report_xlsx_service, "RENDER_TIMEOUT_SECONDS", 0.02)
    monkeypatch.setattr(report_xlsx_service, "_render_semaphore", asyncio.Semaphore(2))

    document = _document()

    async def _one_timed_out_call():
        with pytest.raises(XlsxRenderTimeoutError):
            await report_xlsx_service.build_workbook_bounded(document)

    await asyncio.gather(*[_one_timed_out_call() for _ in range(6)])

    release_gate.set()
    await _wait_until(lambda: active == 0)

    assert max_active <= 2, (
        f"effective concurrent generations ({max_active}) exceeded the configured limit (2) -- "
        "a caller-facing timeout must never let a new generation start on top of one still running"
    )


async def test_build_workbook_bounded_concurrent_requests_never_exceed_configured_limit(monkeypatch):
    """Required test: "concurrency limit never exceeded" (normal-load
    variant) -- combining normal completions in the same burst, the real
    number of simultaneously-executing generations must never exceed the
    configured limit."""
    active = 0
    max_active = 0
    count_lock = threading.Lock()

    def _tracking_render(document):
        nonlocal active, max_active
        with count_lock:
            active += 1
            max_active = max(max_active, active)
        time.sleep(0.08)
        with count_lock:
            active -= 1
        return b"PK\x03\x04fake"

    monkeypatch.setattr(report_xlsx_service, "build_workbook_sync", _tracking_render)
    monkeypatch.setattr(report_xlsx_service, "RENDER_TIMEOUT_SECONDS", 15)
    monkeypatch.setattr(report_xlsx_service, "_render_semaphore", asyncio.Semaphore(3))

    document = _document()
    results = await asyncio.gather(*[report_xlsx_service.build_workbook_bounded(document) for _ in range(9)])

    assert all(r.startswith(b"PK") for r in results)
    assert max_active <= 3


# ---------------------------------------------------------------------------
# Unit tests: filename generation (reused, not reinvented -- see
# app.utils.export_filename, already covered by its own PR18B tests; this
# module only proves the xlsx route uses the same stem with a .xlsx suffix,
# in the API tests below).
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# API tests: GET /reports/{report_id}/xlsx
# ---------------------------------------------------------------------------


async def _create_equipment(client, headers, asset_number: str, *, equipment_name: str = "Infusion Pump"):
    resp = await client.post(
        "/api/v1/equipment", headers=headers, json={"asset_number": asset_number, "equipment_name": equipment_name}
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


async def _dispatch(client, headers, equipment_id: str, ward_id: str):
    resp = await client.post(
        "/api/v1/borrow",
        headers=headers,
        json={"equipment_id": equipment_id, "ward_id": ward_id, "dispatch_type": "on_demand"},
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


async def _receive(client, headers, transaction_id: str):
    resp = await client.post(
        f"/api/v1/return/{transaction_id}", headers=headers, json={"receipt_outcome": "usable"}
    )
    assert resp.status_code == 200, resp.text
    return resp.json()


async def test_xlsx_authorized_roles_succeed(client, seeded_users):
    for role in (ROLE_ADMINISTRATOR, ROLE_EQUIPMENT_POOL_STAFF, ROLE_READ_ONLY):
        headers = await _auth_headers(client, role)
        resp = await client.get("/api/v1/reports/receive-report/xlsx", headers=headers)
        assert resp.status_code == 200, resp.text
        assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert resp.content.startswith(b"PK")


async def test_xlsx_unauthenticated_rejected_with_401(client, seeded_users):
    resp = await client.get("/api/v1/reports/receive-report/xlsx")
    assert resp.status_code == 401, resp.text


async def test_xlsx_unsupported_report_id_rejected(client, seeded_users):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await client.get("/api/v1/reports/not-a-real-report/xlsx", headers=admin)
    assert resp.status_code == 422, resp.text


async def test_xlsx_reverse_business_date_range_rejected(client, seeded_users):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await client.get(
        "/api/v1/reports/receive-report/xlsx",
        headers=admin,
        params={"business_date_from": "2026-07-15", "business_date_to": "2026-07-01"},
    )
    assert resp.status_code == 400, resp.text
    assert resp.json()["code"] == "INVALID_INPUT"


async def test_xlsx_rejects_inapplicable_filter(client, seeded_users):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await client.get(
        "/api/v1/reports/receive-report/xlsx", headers=admin, params={"dispatch_type": "on_demand"}
    )
    assert resp.status_code == 400, resp.text
    assert resp.json()["code"] == "INVALID_INPUT"


async def test_xlsx_content_disposition_filename_is_ascii_safe_and_ends_with_xlsx(client, seeded_users):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await client.get("/api/v1/reports/equipment-verify-checklist/xlsx", headers=admin)
    assert resp.status_code == 200, resp.text
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith('attachment; filename="equipment-verify-checklist_')
    assert disposition.endswith('.xlsx"')
    filename = disposition.split('filename="')[1].rstrip('"')
    assert filename.encode("ascii")  # raises if not ASCII-safe


async def test_xlsx_row_limit_exceeded_returns_structured_error_not_a_workbook(monkeypatch, client, seeded_users):
    monkeypatch.setattr(report_export_service, "MAX_EXPORT_ROWS", 0)
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    staff = await _auth_headers(client, ROLE_EQUIPMENT_POOL_STAFF)
    ward_id = await _create_ward(client, admin, "W-PR18E-1")
    eq = await _create_equipment(client, admin, "AST-PR18E-0001")
    await _dispatch(client, staff, eq["id"], ward_id)

    resp = await client.get("/api/v1/reports/issue-report/xlsx", headers=admin, params={"ward_id": ward_id})
    assert resp.status_code == 422, resp.text
    assert resp.headers["content-type"].startswith("application/json")
    body = resp.json()
    assert body["code"] == "EXPORT_TOO_LARGE"


async def test_xlsx_does_not_mutate_transaction_state(client, seeded_users, db_session):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    staff = await _auth_headers(client, ROLE_EQUIPMENT_POOL_STAFF)
    ward_id = await _create_ward(client, admin, "W-PR18E-2")
    eq = await _create_equipment(client, admin, "AST-PR18E-0002")
    tx = await _dispatch(client, staff, eq["id"], ward_id)

    resp = await client.get("/api/v1/reports/issue-report/xlsx", headers=admin, params={"ward_id": ward_id})
    assert resp.status_code == 200, resp.text

    row = (
        await db_session.execute(select(BorrowTransaction).where(BorrowTransaction.id == uuid.UUID(tx["id"])))
    ).scalar_one()
    assert row.returned_at is None


async def test_xlsx_does_not_write_persistent_audit_log(client, seeded_users, db_session):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    before = (await db_session.execute(select(func.count()).select_from(AuditLog))).scalar_one()

    resp = await client.get("/api/v1/reports/receive-report/xlsx", headers=admin)
    assert resp.status_code == 200, resp.text

    after = (await db_session.execute(select(func.count()).select_from(AuditLog))).scalar_one()
    assert after == before


async def test_xlsx_receive_report_contains_seeded_asset_number(client, seeded_users):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    staff = await _auth_headers(client, ROLE_EQUIPMENT_POOL_STAFF)
    ward_id = await _create_ward(client, admin, "W-PR18E-3", "หอผู้ป่วยทดสอบ PR18E")
    eq = await _create_equipment(client, admin, "AST-PR18E-0003")
    tx = await _dispatch(client, staff, eq["id"], ward_id)
    await _receive(client, staff, tx["id"])

    resp = await client.get("/api/v1/reports/receive-report/xlsx", headers=admin, params={"ward_id": ward_id})
    assert resp.status_code == 200, resp.text
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    cell_values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert "AST-PR18E-0003" in cell_values


async def test_xlsx_equipment_verify_checklist_contains_thai_status_label(client, seeded_users):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    await _create_equipment(client, admin, "AST-PR18E-0004")

    resp = await client.get("/api/v1/reports/equipment-verify-checklist/xlsx", headers=admin)
    assert resp.status_code == 200, resp.text
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    cell_values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert "พร้อมใช้งาน" in cell_values


async def test_xlsx_equipment_verify_checklist_does_not_expose_item_no(client, seeded_users):
    """Design §12: 'item_no remains excluded from operator-facing Equipment
    Verify Checklist output.' `item_no` is settable at creation
    (`EquipmentCreate.item_no`) but deliberately absent from every
    operator-facing response (`EquipmentOut` never declares it) -- this
    proves the Excel adapter's own boundary
    (report_export_service._EQUIPMENT_VERIFY_CHECKLIST_COLUMNS has no
    item_no key at all) end to end through the actual xlsx route/adapter,
    not just by re-reading the column list."""
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    item_no = "ITEM-PR18E-0005"
    resp = await client.post(
        "/api/v1/equipment",
        headers=admin,
        json={"asset_number": "AST-PR18E-0005", "equipment_name": "Infusion Pump", "item_no": item_no},
    )
    assert resp.status_code == 201, resp.text

    resp = await client.get("/api/v1/reports/equipment-verify-checklist/xlsx", headers=admin)
    assert resp.status_code == 200, resp.text
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    cell_values = {cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)}
    assert item_no not in cell_values


@pytest.mark.parametrize(
    "report_id,expected_title",
    [
        ("receive-report", "Receive Report"),
        ("issue-report", "Issue Report"),
        ("equipment-verify-checklist", "Equipment Verify Checklist"),
    ],
)
async def test_xlsx_worksheet_title_matches_report_id(client, seeded_users, report_id, expected_title):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await client.get(f"/api/v1/reports/{report_id}/xlsx", headers=admin)
    assert resp.status_code == 200, resp.text
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.active.title == expected_title


# ---------------------------------------------------------------------------
# API tests: renderer-failure handling (mirrors PR18D's H3 review fix,
# 4838921407) -- an unexpected generation failure must surface as a
# structured 500 and must log its own distinguishable export-attempt event
# rather than the "success" event above.
# ---------------------------------------------------------------------------


async def _raw_client():
    """Matches tests/test_pr18d_pdf_export.py's own `_raw_client` helper --
    ASGITransport(raise_app_exceptions=False) is needed to inspect the 500
    response the app already sent, instead of letting httpx re-raise the
    underlying exception into the test itself."""
    from httpx import ASGITransport, AsyncClient

    from app.main import app as fastapi_app

    transport = ASGITransport(app=fastapi_app, raise_app_exceptions=False)
    return AsyncClient(transport=transport, base_url="http://test")


async def test_xlsx_unexpected_generation_error_returns_500_and_logs_distinguishable_event(
    monkeypatch, client, seeded_users
):
    logged_calls = []

    def _capture(**kwargs):
        logged_calls.append(kwargs)

    def _boom(document):
        raise RuntimeError("simulated unexpected openpyxl failure")

    monkeypatch.setattr(report_xlsx_service, "build_workbook_sync", _boom)
    monkeypatch.setattr(report_export_service, "log_export_attempt", _capture)
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)

    async with await _raw_client() as raw_client:
        resp = await raw_client.get("/api/v1/reports/receive-report/xlsx", headers=admin)
    assert resp.status_code == 500, resp.text

    assert len(logged_calls) == 1
    assert logged_calls[0]["output_format"] == "xlsx"
    assert logged_calls[0]["outcome"] == "render_error"
    assert logged_calls[0]["outcome"] != "success"


async def test_xlsx_success_and_failure_outcomes_are_distinguishable(monkeypatch, client, seeded_users):
    logged_calls = []

    def _capture(**kwargs):
        logged_calls.append(kwargs)

    monkeypatch.setattr(report_export_service, "log_export_attempt", _capture)
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)

    ok_resp = await client.get("/api/v1/reports/receive-report/xlsx", headers=admin)
    assert ok_resp.status_code == 200, ok_resp.text

    def _boom(document):
        raise RuntimeError("simulated failure")

    monkeypatch.setattr(report_xlsx_service, "build_workbook_sync", _boom)

    async with await _raw_client() as raw_client:
        failed_resp = await raw_client.get("/api/v1/reports/receive-report/xlsx", headers=admin)
    assert failed_resp.status_code == 500, failed_resp.text

    assert len(logged_calls) == 2
    assert logged_calls[0]["outcome"] == "success"
    assert logged_calls[1]["outcome"] == "render_error"
    assert logged_calls[0]["outcome"] != logged_calls[1]["outcome"]


# ---------------------------------------------------------------------------
# Regression tests: Browser Print, PDF, and ExportDocument are unchanged by
# this PR (scope protection, per the assigned task).
# ---------------------------------------------------------------------------


async def test_print_data_route_still_works_unchanged(client, seeded_users):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await client.get("/api/v1/reports/receive-report/print-data", headers=admin)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["metadata"]["report_identity"] == "receive-report"


async def test_pdf_route_still_works_unchanged(client, seeded_users):
    admin = await _auth_headers(client, ROLE_ADMINISTRATOR)
    resp = await client.get("/api/v1/reports/receive-report/pdf", headers=admin)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")
